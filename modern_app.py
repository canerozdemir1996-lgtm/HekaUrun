from __future__ import annotations

import json
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
import uuid
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
from urllib.parse import unquote, urlparse
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps
from PIL.ImageQt import ImageQt
from PySide6.QtCore import (
    QElapsedTimer,
    QEvent,
    QPoint,
    QObject,
    QRunnable,
    QRect,
    QSize,
    Qt,
    QThreadPool,
    QTimer,
    Signal,
    Slot,
)
from PySide6.QtGui import (
    QAction,
    QBrush,
    QClipboard,
    QColor,
    QFont,
    QIcon,
    QLinearGradient,
    QPainter,
    QPen,
    QPixmap,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QScrollArea,
    QSizePolicy,
    QSizeGrip,
    QSplitter,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app import (
    ALL_BREAKDOWNS_LABEL,
    APP_DIR,
    APP_ICON_ICO,
    APP_ICON_PNG,
    APP_LOGO_PNG,
    APP_TITLE,
    ASSETS_DIR,
    CHANNEL_FOLDERS,
    DATA_DIR,
    DEFAULT_SETTINGS,
    DEFAULT_YDK_LABEL_LAYOUT,
    INDEX_FILE,
    LOGODATA_SQL_DEFAULTS,
    LOGODATA_SQL_PROFILE_VERSION,
    RENAME_FILTER_OPTIONS,
    RENAME_LOG_FILE,
    SETTINGS_FILE,
    YDK_LABEL_ICON_PNG,
    YDK_LABEL_LOGO_PNG,
    YDK_LABEL_SIZE,
    YDK_LOGO_PNG,
    ProductRecord,
    RenameAction,
    YdkProduct,
    apply_logodata_sql_profile,
    digits_only,
    ean13_check_digit,
    ean13_font_text,
    ean13_full_code,
    ensure_prefixed_extensions,
    first_ean_value,
    first_matching_value,
    first_value_matching_regex,
    is_generic_header,
    load_excel_headers,
    normalize_data_source,
    normalize_text,
    open_in_explorer,
    open_path_with_default_app,
    parse_csv,
    path_exists,
    path_is_dir,
    path_is_file,
    prettify_feature_name,
    resolve_column_index,
    resolve_header_first_index,
    safe_path_part,
    safe_rglob_files,
    safe_walk,
    split_multivalue_config,
    stringify_cell,
    unique_paths,
    unique_preserve_order,
)

COMBO_ARROW_SVG = ASSETS_DIR / "combo_arrow.svg"
QR_CODE_ROOT_DEFAULT = "K:/HEKA/000_HEKATasar\u0131mlar/009_\u00dcretim Tasar\u0131mlar\u0131/QR KOD BARKOD"
QR_URL_BASE_DEFAULT = "https://hekastore.com/products/"
QR_SITEMAP_URL_DEFAULT = "https://hekastore.com/sitemap_products_1.xml?from=8093858103481&to=8481939325113"
CE_MEASUREMENTS_XLSX_DEFAULT = "O:/HEKA/LABORATUVAR/ARGE Ã–LÃ‡ÃœMLER/ARGE Ã–LÃ‡ÃœMLER.xlsx"
CE_OUTPUT_ROOT_DEFAULT = DATA_DIR / "ce_exports"
BARTENDER_EXE_DEFAULT = "C:/Program Files/Seagull/BarTender 11.8/BarTend.exe"
BARTENDER_PRINTER_DEFAULT = "Argox CP-3140EX PPLB"
CE_BARTENDER_TEMPLATE_OLD_DEFAULT = "K:/HEKA/000_HEKATasarÄ±mlar/009_Ãœretim TasarÄ±mlarÄ±/CE BARKOD/KÄ±sa Etiket.btw"
CE_BARTENDER_TEMPLATE_DEFAULT = "K:/HEKA/000_HEKATasarÄ±mlar/009_Ãœretim TasarÄ±mlarÄ±/CE BARKOD/Barkodlar/Excel DEneme.btw"
QR_BARTENDER_TEMPLATE_DEFAULT = "K:/HEKA/000_HEKATasarÄ±mlar/009_Ãœretim TasarÄ±mlarÄ±/QR KOD BARKOD/QR.btw"
LEGACY_YDK_LABEL_LAYOUT = {
    "logo_x": 0,
    "logo_y": 0,
    "logo_w": 350,
    "logo_h": 145,
    "photo_x": 360,
    "photo_y": 0,
    "photo_w": 330,
    "photo_h": 330,
    "model_x": 28,
    "model_y": 226,
    "model_font": 45,
    "tr_x": 31,
    "tr_y": 354,
    "tr_font": 36,
    "en_x": 31,
    "en_y": 478,
    "en_font": 35,
    "barcode_x": 0,
    "barcode_y": 610,
    "barcode_w": 390,
    "barcode_h": 150,
    "info_x": 407,
    "heading_y": 633,
    "code_y": 688,
    "producer_box_x": 398,
    "producer_box_y": 723,
    "producer_box_w": 314,
    "producer_box_h": 86,
    "producer_font": 46,
    "footer_y": 815,
    "footer_font": 27,
}
PREVIOUS_YDK_LABEL_LAYOUT = {
    "logo_x": 0,
    "logo_y": 0,
    "logo_w": 350,
    "logo_h": 145,
    "photo_x": 360,
    "photo_y": 0,
    "photo_w": 330,
    "photo_h": 330,
    "model_x": 28,
    "model_y": 214,
    "model_font": 40,
    "tr_x": 31,
    "tr_y": 336,
    "tr_font": 32,
    "en_x": 31,
    "en_y": 455,
    "en_font": 29,
    "barcode_x": 0,
    "barcode_y": 586,
    "barcode_w": 390,
    "barcode_h": 138,
    "info_x": 407,
    "heading_y": 606,
    "code_y": 654,
    "producer_box_x": 398,
    "producer_box_y": 704,
    "producer_box_w": 314,
    "producer_box_h": 78,
    "producer_font": 38,
    "footer_y": 812,
    "footer_font": 24,
}
RENAME_SCOPE_ALL = "all"
RENAME_SCOPE_FLAT_IMAGES = "flat_images"
RENAME_SCOPE_OPTIONS = (
    (RENAME_SCOPE_ALL, "Tum uygun gorseller"),
    ("b2b", "Sadece B2B"),
    ("web", "Sadece WEB / Katalog"),
    ("instagram", "Sadece Instagram"),
    ("beymen", "Sadece Beymen"),
    ("technical", "Sadece Teknik Cizim"),
)
def turkish_upper(text: str) -> str:
    value = str(text or "")
    return value.replace("i", "\u0130").replace("\u0131", "I").upper()


@dataclass
class QrCatalogEntry:
    stock_code: str
    handle: str = ""
    title: str = ""
    image_url: str = ""


@dataclass
class CeLabelEntry:
    stock_code: str
    product_name: str = ""
    mancode: str = ""
    model: str = ""
    product_group: str = ""
    dimming: str = ""
    voltage_110: str = ""
    voltage_220: str = ""
    driver_model: str = ""
    cable_type: str = ""
    kelvin: str = ""
    lumen: str = ""
    watt: str = ""
    duy: str = ""
    label_kind: str = ""


def pil_to_pixmap(image: Image.Image) -> QPixmap:
    qimage = ImageQt(image.convert("RGBA"))
    return QPixmap.fromImage(qimage)


def load_scaled_pixmap(path: Path | None, size: QSize, fill: str = "#101820") -> QPixmap | None:
    if path is None or not path_is_file(path):
        return None
    try:
        image = Image.open(path)
        image = ImageOps.exif_transpose(image).convert("RGB")
        image.thumbnail((size.width(), size.height()), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (size.width(), size.height()), fill)
        canvas.paste(image, ((size.width() - image.width) // 2, (size.height() - image.height) // 2))
        return pil_to_pixmap(canvas)
    except Exception:
        return None


def resolve_runtime_path(value: str) -> Path:
    text = str(value or "").strip()
    if not text:
        return Path()
    path = Path(text).expanduser()
    if path.is_absolute():
        return path
    candidates = [DATA_DIR / path, APP_DIR / path, Path.cwd() / path]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def alpha_sort_key(value: str) -> tuple[list[Any], str]:
    text = normalize_text(value)
    parts: list[Any] = [(0, int(part)) if part.isdigit() else (1, part) for part in re.split(r"(\d+)", text) if part]
    return parts, str(value).casefold()


def slugify_handle(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def resolve_first_header_index(headers: list[str], candidates: list[str]) -> int | None:
    for candidate in candidates:
        resolved = resolve_header_first_index(candidate, headers)
        if resolved is not None:
            return resolved
    return None


def merge_text(preferred: str, fallback: str) -> str:
    left = str(preferred or "").strip()
    if left and normalize_text(left) not in {"-", "n/a", "na", "none", "null"}:
        return left
    return str(fallback or "").strip()


class WorkerSignals(QObject):
    result = Signal(object)
    error = Signal(str)


class Worker(QRunnable):
    def __init__(self, job: Callable[[], Any]):
        super().__init__()
        self.job = job
        self.signals = WorkerSignals()

    @Slot()
    def run(self) -> None:
        try:
            self.signals.result.emit(self.job())
        except Exception as exc:  # noqa: BLE001
            self.signals.error.emit(str(exc))


class LoadingOverlay(QWidget):
    def __init__(self, parent: QWidget) -> None:
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setAutoFillBackground(False)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.hide()
        self._message = "Yukleniyor..."
        self._elapsed = QElapsedTimer()
        self._timer = QTimer(self)
        self._timer.setInterval(16)
        self._timer.timeout.connect(self._tick)
        self._logo_pixmap = self._load_logo_pixmap()

    def _load_logo_pixmap(self) -> QPixmap:
        for path in (APP_ICON_PNG, APP_LOGO_PNG):
            if path_is_file(path):
                pixmap = QPixmap(str(path))
                if not pixmap.isNull():
                    return pixmap
        return QPixmap()

    def start(self, message: str, theme_key: str = "light") -> None:
        del theme_key
        self._message = str(message or "Yukleniyor...")
        parent = self.parentWidget()
        if parent is not None:
            self.setGeometry(parent.rect())
        self._elapsed.restart()
        self.show()
        self.raise_()
        self._timer.start()
        self.update()

    def stop(self) -> None:
        self._timer.stop()
        self.hide()

    def set_message(self, message: str) -> None:
        self._message = str(message or "Yukleniyor...")
        if self.isVisible():
            self.update()

    def set_theme_key(self, theme_key: str) -> None:
        del theme_key
        if self.isVisible():
            self.update()

    def sync_to_parent(self) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.setGeometry(parent.rect())

    def _tick(self) -> None:
        if not self.isVisible():
            self._timer.stop()
            return
        self.sync_to_parent()
        self.update()

    def _elapsed_s(self) -> float:
        if not self._elapsed.isValid():
            return 0.0
        return self._elapsed.elapsed() / 1000.0

    def _animated_message(self) -> str:
        base = re.sub(r"\.+$", "", self._message.strip()) or "Yukleniyor"
        dot_count = int((self._elapsed.elapsed() if self._elapsed.isValid() else 0) / 450) % 4
        return f"{base}{'.' * dot_count}"

    def paintEvent(self, event: QEvent) -> None:
        del event
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)

        width = max(1, self.width())
        height = max(1, self.height())
        elapsed_s = self._elapsed_s()
        pulse = 0.5 + 0.5 * math.sin(elapsed_s * 2.2)

        painter.fillRect(self.rect(), QColor(248, 250, 252, 236))
        bg = QLinearGradient(0.0, 0.0, float(width), float(height))
        bg.setColorAt(0.0, QColor(255, 255, 255, 246))
        bg.setColorAt(0.52, QColor(241, 247, 255, 232))
        bg.setColorAt(1.0, QColor(232, 240, 250, 242))
        painter.fillRect(self.rect(), QBrush(bg))

        panel_w = min(int(width * 0.56), 560)
        panel_h = 188
        panel = QRect(int((width - panel_w) / 2), int(height * 0.5 - panel_h / 2), panel_w, panel_h)
        shadow = QRect(panel.left() + 0, panel.top() + 14, panel.width(), panel.height())
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(15, 23, 42, 24))
        painter.drawRoundedRect(shadow, 22.0, 22.0)
        painter.setBrush(QColor(255, 255, 255, 248))
        painter.drawRoundedRect(panel, 22.0, 22.0)
        painter.setPen(QPen(QColor(207, 216, 229), 1.0))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(panel, 22.0, 22.0)

        logo_rect = QRect(panel.left() + 28, panel.top() + 34, 74, 74)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(226, 232, 240))
        painter.drawRoundedRect(logo_rect, 18.0, 18.0)
        if not self._logo_pixmap.isNull():
            logo = self._logo_pixmap.scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            painter.drawPixmap(logo_rect.center().x() - logo.width() // 2, logo_rect.center().y() - logo.height() // 2, logo)
        else:
            painter.setFont(QFont("Bahnschrift", 19, QFont.Weight.Black))
            painter.setPen(QColor("#0F172A"))
            painter.drawText(logo_rect, Qt.AlignmentFlag.AlignCenter, "H")

        title_font = QFont("Bahnschrift", max(18, int(height * 0.025)), QFont.Weight.Bold)
        sub_font = QFont("Segoe UI", max(10, int(height * 0.014)), QFont.Weight.Medium)
        title_rect = QRect(panel.left() + 122, panel.top() + 38, panel.width() - 154, 34)
        sub_rect = QRect(panel.left() + 122, panel.top() + 72, panel.width() - 154, 30)
        painter.setFont(title_font)
        painter.setPen(QColor("#111827"))
        painter.drawText(title_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, self._animated_message())
        painter.setFont(sub_font)
        painter.setPen(QColor("#64748B"))
        painter.drawText(sub_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, "Veriler hazirlaniyor, lutfen bekleyin.")

        track = QRect(panel.left() + 28, panel.bottom() - 48, panel.width() - 56, 9)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor("#E2E8F0"))
        painter.drawRoundedRect(track, 5.0, 5.0)
        runner_w = max(72, int(track.width() * (0.20 + pulse * 0.08)))
        travel = (math.sin(elapsed_s * 2.0) + 1.0) * 0.5
        runner_x = track.left() + int((track.width() - runner_w) * travel)
        fill = QLinearGradient(track.left(), track.top(), track.right(), track.top())
        fill.setColorAt(0.0, QColor("#2563EB"))
        fill.setColorAt(1.0, QColor("#14B8A6"))
        painter.setBrush(QBrush(fill))
        painter.drawRoundedRect(QRect(runner_x, track.top(), runner_w, track.height()), 5.0, 5.0)


class DataEngine:
    def __init__(self) -> None:
        self.settings = self.load_settings()
        self.records: list[ProductRecord] = []
        self.ydk_products: dict[str, YdkProduct] = {}
        self.product_data_warning = ""
        self.search_folder_cache: dict[str, Path | None] = {}
        self.channel_folder_cache: dict[str, Path | None] = {}
        self.preview_image_cache: dict[str, Path | None] = {}
        self.preview_image_list_cache: dict[str, list[Path]] = {}
        self.preview_search_roots_cache: list[Path] | None = None
        self.qr_catalog_cache: dict[str, QrCatalogEntry] | None = None
        self.qr_file_index_cache: dict[str, Path] | None = None
        self.qr_file_cache: dict[str, Path | None] = {}
        self.qr_sitemap_cache: dict[str, str] | None = None
        self.ce_label_cache: dict[str, CeLabelEntry] | None = None
        self.rename_selection: dict[str, bool] = self.load_json_setting("rename_selection_state", {})
        self.rename_manual_targets: dict[str, str] = self.load_json_setting("rename_manual_targets", {})
        self.ydk_label_overrides: dict[str, dict[str, str]] = self.load_json_setting("ydk_label_overrides", {})
        self.ydk_logo_image: Image.Image | None = None
        self.ydk_icon_image: Image.Image | None = None
        self.ensure_index_schema()

    def load_settings(self) -> dict[str, str]:
        if not path_exists(SETTINGS_FILE):
            settings = apply_logodata_sql_profile(DEFAULT_SETTINGS.copy())
            settings = self.apply_modern_defaults(settings)
            self.save_settings(settings)
            return settings
        try:
            loaded = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return self.apply_modern_defaults(apply_logodata_sql_profile(DEFAULT_SETTINGS.copy()))
        settings = DEFAULT_SETTINGS.copy()
        settings.update({str(key): str(value) for key, value in loaded.items()})
        settings = apply_logodata_sql_profile(settings)
        settings = self.apply_modern_defaults(settings)
        return settings

    def apply_modern_defaults(self, settings: dict[str, str]) -> dict[str, str]:
        settings["data_source"] = "sql"
        qr_root = settings.get("qr_code_root", "").strip() or QR_CODE_ROOT_DEFAULT
        settings["qr_code_root"] = qr_root
        if not settings.get("qr_output_root", "").strip():
            settings["qr_output_root"] = str(Path(qr_root) / "EXPORTS")
        if not settings.get("qr_url_base", "").strip():
            settings["qr_url_base"] = QR_URL_BASE_DEFAULT
        if not settings.get("qr_sitemap_url", "").strip():
            settings["qr_sitemap_url"] = QR_SITEMAP_URL_DEFAULT
        if not settings.get("ce_measurement_excel_path", "").strip():
            settings["ce_measurement_excel_path"] = CE_MEASUREMENTS_XLSX_DEFAULT
        if not settings.get("ce_output_root", "").strip():
            settings["ce_output_root"] = str(CE_OUTPUT_ROOT_DEFAULT)
        bartender_exe = self.resolve_bartender_exe_path(settings.get("bartender_exe_path", ""))
        if bartender_exe is not None:
            settings["bartender_exe_path"] = str(bartender_exe)
        elif not settings.get("bartender_exe_path", "").strip():
            settings["bartender_exe_path"] = BARTENDER_EXE_DEFAULT
        if not settings.get("bartender_printer_name", "").strip():
            settings["bartender_printer_name"] = BARTENDER_PRINTER_DEFAULT
        ce_template_text = settings.get("ce_bartender_template_path", "").strip()
        if not ce_template_text or normalize_text(ce_template_text) == normalize_text(CE_BARTENDER_TEMPLATE_OLD_DEFAULT):
            settings["ce_bartender_template_path"] = CE_BARTENDER_TEMPLATE_DEFAULT
        if not settings.get("qr_bartender_template_path", "").strip():
            settings["qr_bartender_template_path"] = QR_BARTENDER_TEMPLATE_DEFAULT
        if not settings.get("bartender_print_timeout_seconds", "").strip():
            settings["bartender_print_timeout_seconds"] = "180"
        settings.setdefault("label_reset_enabled", "1")
        settings.setdefault("label_reset_before_each_label", "0")
        settings.setdefault("label_reset_feed_count", "1")
        settings.setdefault("label_reset_pause_ms", "1200")
        settings.setdefault("label_reset_raw_command", "\\x0c")
        settings["ui_theme"] = "light"
        settings["ydk_label_layout"] = self.normalize_ydk_label_layout_json(settings.get("ydk_label_layout", ""))
        required_technical = ["teknik", "teknik cizim", "technical", "cizim", "olcu"]
        technical_items = split_multivalue_config(settings.get("technical_keywords", ""))
        technical_seen = {normalize_text(item) for item in technical_items}
        for keyword in required_technical:
            if normalize_text(keyword) not in technical_seen:
                technical_items.append(keyword)
                technical_seen.add(normalize_text(keyword))
        settings["technical_keywords"] = ",".join(technical_items)
        if settings.get("rename_scope", "").strip() == RENAME_SCOPE_FLAT_IMAGES:
            settings["rename_scope"] = RENAME_SCOPE_ALL
            settings["rename_flat_output_enabled"] = "1"
        if not settings.get("rename_scope", "").strip():
            settings["rename_scope"] = RENAME_SCOPE_ALL
        if not settings.get("rename_flat_output_enabled", "").strip():
            settings["rename_flat_output_enabled"] = settings.get("rename_flat_images_enabled", "0")
        if not settings.get("rename_flat_images_enabled", "").strip():
            settings["rename_flat_images_enabled"] = "0"
        settings["rename_flat_images_enabled"] = settings.get("rename_flat_output_enabled", "0")
        return settings

    def normalize_ydk_label_layout_json(self, value: str) -> str:
        layout = DEFAULT_YDK_LABEL_LAYOUT.copy()
        saved: dict[str, Any] = {}
        try:
            parsed = json.loads(value or "{}")
            if isinstance(parsed, dict):
                saved = parsed
        except json.JSONDecodeError:
            saved = {}

        def saved_int(key: str) -> int | None:
            try:
                text = str(saved.get(key, "")).strip()
                return int(float(text)) if text else None
            except (TypeError, ValueError):
                return None

        migratable_match = bool(saved) and any(
            all(saved_int(key) == expected for key, expected in candidate.items())
            for candidate in (LEGACY_YDK_LABEL_LAYOUT, PREVIOUS_YDK_LABEL_LAYOUT)
        )
        if saved and not migratable_match:
            for key in layout:
                if key not in saved:
                    continue
                try:
                    layout[key] = int(float(saved[key]))
                except (TypeError, ValueError):
                    pass
        return json.dumps(layout, ensure_ascii=False)

    def resolve_bartender_exe_path(self, configured: str = "") -> Path | None:
        candidates: list[Path] = []
        text = str(configured or "").strip()
        if text:
            candidates.append(Path(text).expanduser())
        candidates.append(Path(BARTENDER_EXE_DEFAULT))
        for env_key in ("ProgramFiles", "ProgramFiles(x86)"):
            base = os.environ.get(env_key)
            if not base:
                continue
            seagull = Path(base) / "Seagull"
            if path_is_dir(seagull):
                candidates.extend(sorted(seagull.glob("BarTender*/BarTend.exe"), reverse=True))
        which = shutil.which("bartend.exe") or shutil.which("BarTend.exe")
        if which:
            candidates.append(Path(which))
        for candidate in unique_paths(candidates):
            if path_is_file(candidate):
                return candidate
        return None

    def save_settings(self, settings: dict[str, str] | None = None) -> None:
        payload = settings or self.settings
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        SETTINGS_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        self.settings = payload

    def load_json_setting(self, key: str, default: dict[str, Any]) -> dict[str, Any]:
        try:
            value = json.loads(self.settings.get(key, ""))
            return value if isinstance(value, dict) else default.copy()
        except json.JSONDecodeError:
            return default.copy()

    def save_runtime_state(self) -> None:
        self.settings["rename_selection_state"] = json.dumps(self.rename_selection, ensure_ascii=False)
        self.settings["rename_manual_targets"] = json.dumps(self.rename_manual_targets, ensure_ascii=False)
        self.save_settings()

    def get_data_source_key(self) -> str:
        return "sql"

    def parse_feature_aliases(self, headers: list[str]) -> dict[int, str]:
        alias_map: dict[int, str] = {}
        for item in split_multivalue_config(self.settings.get("feature_aliases", "")):
            separator = "=" if "=" in item else ":" if ":" in item else None
            if separator is None:
                continue
            left, right = [part.strip() for part in item.split(separator, 1)]
            left_index = resolve_header_first_index(left, headers)
            right_index = resolve_header_first_index(right, headers)
            if left_index is not None and right:
                alias_map[left_index] = prettify_feature_name(right)
            elif right_index is not None and left:
                alias_map[right_index] = prettify_feature_name(left)
        return alias_map

    def build_feature_label(self, column_index: int, headers: list[str], aliases: dict[int, str]) -> str:
        alias = prettify_feature_name(aliases.get(column_index, ""))
        if alias:
            return alias
        header = prettify_feature_name(headers[column_index])
        return header if header and not is_generic_header(header) else f"Alan {column_index + 1}"

    def load_product_records(self) -> list[ProductRecord]:
        self.product_data_warning = ""
        records = self.load_sql_product_records() if self.get_data_source_key() == "sql" else self.load_excel_product_records()
        self.records = records
        self.ydk_products = self.parse_ydk_records(records)
        return records

    def load_excel_product_records(self) -> list[ProductRecord]:
        excel_text = self.settings.get("excel_path", "").strip()
        if not excel_text:
            return []
        excel_path = resolve_runtime_path(excel_text)
        if not path_is_file(excel_path):
            self.product_data_warning = f"Excel dosyasina erisilemiyor: {excel_path}"
            return []
        header_row = int(self.settings.get("header_row", "1") or "1")
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        try:
            sheet_name = self.settings.get("sheet_name", "").strip()
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            headers = load_excel_headers(excel_path, sheet_name, header_row)
            family_index = resolve_column_index(self.settings["family_column"], headers)
            breakdown_index = resolve_column_index(self.settings["breakdown_column"], headers)
            stock_index = resolve_column_index(self.settings["stock_column"], headers)
            if family_index is None or breakdown_index is None or stock_index is None:
                raise ValueError("Aile, kirilim veya stok sutunu cozulmedi.")
            reserved = {family_index, breakdown_index, stock_index}
            feature_specs = parse_csv(self.settings.get("feature_columns", ""))
            if feature_specs:
                feature_indexes = [
                    resolved
                    for spec in feature_specs
                    if (resolved := resolve_column_index(spec, headers)) is not None and resolved not in reserved
                ]
            else:
                feature_indexes = [index for index in range(len(headers)) if index not in reserved]
            aliases = self.parse_feature_aliases(headers)
            records: list[ProductRecord] = []
            for row_values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                row = [stringify_cell(value) for value in row_values]
                if not any(row):
                    continue
                while len(row) < len(headers):
                    row.append("")
                stock = row[stock_index].strip()
                if not stock:
                    continue
                features = {
                    self.build_feature_label(index, headers, aliases): row[index]
                    for index in feature_indexes
                    if index < len(row) and row[index]
                }
                raw = {
                    (self.build_feature_label(index, headers, aliases) if index in feature_indexes else headers[index]): row[index]
                    for index in range(len(headers))
                    if index < len(row) and row[index]
                }
                records.append(ProductRecord(row[family_index].strip(), row[breakdown_index].strip(), stock, features, raw))
            return records
        finally:
            workbook.close()

    def open_sql_connection(self) -> tuple[Any, str]:
        connection_text = self.settings.get("sql_connection_string", "").strip()
        if not connection_text:
            raise ValueError("SQL baglanti bilgisi girilmedi.")
        lowered = connection_text.casefold()
        looks_like_sqlite = lowered.startswith("file:") or lowered.startswith("sqlite:///") or lowered.endswith((".sqlite", ".sqlite3", ".db", ".db3"))
        if looks_like_sqlite:
            sqlite_target = connection_text[10:] if lowered.startswith("sqlite:///") else connection_text
            sqlite_uri = sqlite_target.startswith("file:")
            if not sqlite_uri:
                sqlite_target = str(resolve_runtime_path(sqlite_target))
            return sqlite3.connect(sqlite_target, uri=sqlite_uri), "sqlite"
        import pyodbc  # type: ignore[import-not-found]
        try:
            return pyodbc.connect(connection_text), "odbc"
        except Exception as exc:
            fallback = self.build_trusted_odbc_fallback(connection_text)
            if fallback:
                try:
                    return pyodbc.connect(fallback), "odbc"
                except Exception:
                    pass
            raise exc

    def build_trusted_odbc_fallback(self, connection_text: str) -> str:
        parts = [part.strip() for part in connection_text.split(";") if part.strip()]
        lowered_keys = {part.split("=", 1)[0].strip().casefold() for part in parts if "=" in part}
        if "pwd" in lowered_keys or "password" in lowered_keys or "trusted_connection" in lowered_keys:
            return ""
        filtered = [part for part in parts if part.split("=", 1)[0].strip().casefold() not in {"uid", "user", "user id"}]
        filtered.append("Trusted_Connection=yes")
        return ";".join(filtered)

    def load_sql_product_records(self) -> list[ProductRecord]:
        connection, driver_kind = self.open_sql_connection()
        try:
            cursor = connection.cursor()
            query = self.settings.get("sql_query", "").strip()
            if not query:
                table_name = self.settings.get("sql_table", "").strip()
                if not table_name:
                    raise ValueError("SQL sorgusu veya tablo/gorunum girilmeli.")
                query = f'SELECT * FROM "{table_name}"' if driver_kind == "sqlite" else f"SELECT * FROM {table_name}"
            cursor.execute(query)
            headers = [stringify_cell(column[0]) or f"Alan {index + 1}" for index, column in enumerate(cursor.description or [])]
            if not headers:
                return []
            family_index = resolve_column_index(self.settings["family_column"], headers)
            breakdown_index = resolve_column_index(self.settings["breakdown_column"], headers)
            stock_index = resolve_column_index(self.settings["stock_column"], headers)
            if family_index is None or breakdown_index is None or stock_index is None:
                raise ValueError("Aile, kirilim veya stok sutunu SQL sonucunda cozulmedi.")
            reserved = {family_index, breakdown_index, stock_index}
            feature_specs = parse_csv(self.settings.get("feature_columns", ""))
            if feature_specs:
                feature_indexes = [
                    resolved
                    for spec in feature_specs
                    if (resolved := resolve_column_index(spec, headers)) is not None and resolved not in reserved
                ]
            else:
                feature_indexes = [index for index in range(len(headers)) if index not in reserved]
            aliases = self.parse_feature_aliases(headers)
            records: list[ProductRecord] = []
            for row_values in cursor.fetchall():
                row = [stringify_cell(value) for value in row_values]
                if not any(row):
                    continue
                while len(row) < len(headers):
                    row.append("")
                stock = row[stock_index].strip()
                if not stock:
                    continue
                features = {
                    self.build_feature_label(index, headers, aliases): row[index]
                    for index in feature_indexes
                    if index < len(row) and row[index]
                }
                raw = {
                    (self.build_feature_label(index, headers, aliases) if index in feature_indexes else headers[index]): row[index]
                    for index in range(len(headers))
                    if index < len(row) and row[index]
                }
                records.append(ProductRecord(row[family_index].strip(), row[breakdown_index].strip(), stock, features, raw))
            return records
        except Exception as exc:  # noqa: BLE001
            self.product_data_warning = f"SQL verisi okunamadi: {exc}"
            return []
        finally:
            try:
                connection.close()
            except Exception:
                pass

    def should_skip_fast_scan_folder(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return compact in {"2d", "3d", "arsiv", "archive", "backup", "yedek", "temp", "tmp", "nodemodules"}

    def is_product_images_folder_name(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return "urungorselleri" in compact

    def ensure_index_schema(self) -> None:
        with sqlite3.connect(INDEX_FILE) as connection:
            connection.execute(
                "CREATE TABLE IF NOT EXISTS product_index (kind TEXT, stock_key TEXT, channel_key TEXT DEFAULT '', path TEXT, updated_at TEXT, PRIMARY KEY(kind, stock_key, channel_key, path))"
            )
            connection.execute("CREATE TABLE IF NOT EXISTS index_meta (key TEXT PRIMARY KEY, value TEXT)")
            connection.commit()

    def lookup_index_paths(self, kind: str, stock_keys: set[str], channel_key: str = "") -> dict[str, Path | None]:
        if not stock_keys or not path_exists(INDEX_FILE):
            return {}
        found: dict[str, Path | None] = {}
        with sqlite3.connect(INDEX_FILE) as connection:
            for key in stock_keys:
                rows = connection.execute(
                    "SELECT path FROM product_index WHERE kind=? AND stock_key=? AND channel_key=?",
                    (kind, key, channel_key),
                ).fetchall()
                candidates = [Path(row[0]) for row in rows if row and path_exists(Path(row[0]))]
                if not candidates:
                    continue
                if kind == "preview":
                    found[key] = min(candidates, key=lambda path: self.preview_image_rank(path, key))
                elif kind == "channel":
                    found[key] = min(candidates, key=lambda path: self.channel_path_rank(path, key, channel_key))
                else:
                    found[key] = min(candidates, key=lambda path: self.folder_path_rank(path, key))
        return found

    def get_index_last_rebuild_text(self) -> str:
        if not path_exists(INDEX_FILE):
            return "Indeks yok"
        try:
            with sqlite3.connect(INDEX_FILE) as connection:
                row = connection.execute("SELECT value FROM index_meta WHERE key='last_rebuild'").fetchone()
            return row[0] if row else "Indeks hazir"
        except sqlite3.Error:
            return "Indeks okunamadi"

    def rebuild_product_index(self) -> dict[str, int]:
        self.ensure_index_schema()
        stock_keys = sorted({normalize_text(record.stock_code) for record in self.records if normalize_text(record.stock_code)}, key=len, reverse=True)
        if not stock_keys:
            return {"preview": 0, "folder": 0, "channel": 0}
        lookup = re.compile("|".join(re.escape(key) for key in stock_keys))
        rows: set[tuple[str, str, str, str, str]] = set()
        now = datetime.now().isoformat(timespec="seconds")
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))

        preview_roots = unique_paths([
            Path(self.settings.get("preview_image_root", "")).expanduser(),
            Path(self.settings.get("ydk_image_root", "")).expanduser(),
        ])
        for preview_root in [root for root in preview_roots if path_is_dir(root)]:
            for current, dirs, files in safe_walk(preview_root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if path.suffix.lower() in valid and (match := lookup.search(normalize_text(str(path)))):
                        rows.add(("preview", match.group(0), "", str(path), now))

        search_root = Path(self.settings.get("search_root", "")).expanduser()
        if path_is_dir(search_root):
            for current, dirs, files in safe_walk(search_root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                current_path = Path(current)
                root_match = lookup.search(normalize_text(str(current_path)))
                for directory in dirs:
                    path = current_path / directory
                    match = lookup.search(normalize_text(str(path)))
                    if match:
                        rows.add(("folder", match.group(0), "", str(path), now))
                    channel_stock = match.group(0) if match else root_match.group(0) if root_match else ""
                    if channel_stock:
                        for channel_key in CHANNEL_FOLDERS:
                            if self.path_matches_channel(path, channel_key):
                                rows.add(("channel", channel_stock, channel_key, str(path), now))
                for file_name in files:
                    path = current_path / file_name
                    match = lookup.search(normalize_text(str(path)))
                    if not match:
                        continue
                    if path.suffix.lower() in valid:
                        rows.add(("folder", match.group(0), "", str(current_path), now))
                    for channel_key in CHANNEL_FOLDERS:
                        if self.path_matches_channel(current_path, channel_key) or self.path_matches_channel(path, channel_key):
                            rows.add(("channel", match.group(0), channel_key, str(current_path), now))

        with sqlite3.connect(INDEX_FILE) as connection:
            connection.execute("DELETE FROM product_index")
            connection.executemany("INSERT OR REPLACE INTO product_index VALUES (?, ?, ?, ?, ?)", list(rows))
            connection.execute("INSERT OR REPLACE INTO index_meta VALUES ('last_rebuild', ?)", (now,))
            connection.commit()
        return {
            "preview": len([row for row in rows if row[0] == "preview"]),
            "folder": len([row for row in rows if row[0] == "folder"]),
            "channel": len([row for row in rows if row[0] == "channel"]),
        }

    def get_preview_search_roots(self) -> list[Path]:
        if self.preview_search_roots_cache is not None:
            return self.preview_search_roots_cache
        configured = [
            Path(self.settings.get("preview_image_root", "")).expanduser(),
            Path(self.settings.get("ydk_image_root", "")).expanduser(),
            Path(self.settings.get("search_root", "")).expanduser(),
        ]
        self.preview_search_roots_cache = unique_paths([path for path in configured if path_is_dir(path)])
        return self.preview_search_roots_cache

    def get_image_repository_roots(self) -> list[Path]:
        configured = [
            Path(self.settings.get("preview_image_root", "")).expanduser(),
            Path(self.settings.get("ydk_image_root", "")).expanduser(),
        ]
        return unique_paths([path for path in configured if path_is_dir(path)])

    def preview_image_rank(self, path: Path, stock_key: str, root_order: int = 0) -> tuple[int, int, int, str]:
        stem = normalize_text(path.stem)
        full = normalize_text(str(path))
        parts = [normalize_text(part) for part in path.parts]
        if stem == stock_key:
            stock_score = 0
        elif stem.startswith(stock_key):
            stock_score = 1
        elif any(part.startswith(stock_key) for part in parts):
            stock_score = 2
        elif stock_key in stem:
            stock_score = 3
        elif stock_key in full:
            stock_score = 4
        else:
            stock_score = 9
        channel_score = 0 if any(word in full for word in ("katalog", "web", "etsy")) else 1
        if any(word in full for word in ("teknik", "cizim", "technical")):
            channel_score += 4
        if "b2b" in full:
            channel_score += 1
        return (root_order, stock_score + channel_score, len(path.parts), str(path).casefold())

    def scan_image_repository_for_keys(self, stock_keys: set[str]) -> dict[str, Path]:
        if not stock_keys:
            return {}
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        lookup = re.compile("|".join(re.escape(key) for key in sorted(stock_keys, key=len, reverse=True)))
        candidates: dict[str, tuple[tuple[int, int, int, str], Path]] = {}
        grouped: dict[str, list[Path]] = defaultdict(list)
        for root_order, root in enumerate(self.get_image_repository_roots()):
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if valid and path.suffix.lower() not in valid:
                        continue
                    match = lookup.search(normalize_text(str(path)))
                    if not match:
                        continue
                    key = match.group(0)
                    rank = self.preview_image_rank(path, key, root_order)
                    grouped[key].append(path)
                    current_best = candidates.get(key)
                    if current_best is None or rank < current_best[0]:
                        candidates[key] = (rank, path)
        for key, paths in grouped.items():
            self.preview_image_list_cache[key] = sorted(set(paths), key=lambda path: self.preview_image_rank(path, key))
        return {key: value[1] for key, value in candidates.items()}

    def find_preview_images(self, stock_codes: list[str]) -> dict[str, Path | None]:
        requested = {normalize_text(code): code for code in stock_codes if normalize_text(code)}
        results: dict[str, Path | None] = {}
        missing = set(requested)
        indexed = self.lookup_index_paths("preview", missing)
        for key, path in indexed.items():
            self.preview_image_cache[key] = path
            results[key] = path
        missing -= set(indexed)
        for key in list(missing):
            if key in self.preview_image_cache:
                results[key] = self.preview_image_cache[key]
                missing.remove(key)
        if not missing:
            return results
        direct_matches = self.scan_image_repository_for_keys(missing)
        for key, path in direct_matches.items():
            self.preview_image_cache[key] = path
            results[key] = path
        missing -= set(direct_matches)
        if not missing:
            return results
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        candidates: dict[str, tuple[tuple[int, int, int, str], Path]] = {}
        if len(missing) <= 40:
            for root_order, root in enumerate(self.get_preview_search_roots()):
                for current, dirs, files in safe_walk(root):
                    dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                    for file_name in files:
                        path = Path(current) / file_name
                        if path.suffix.lower() not in valid:
                            continue
                        norm = normalize_text(str(path))
                        stem = normalize_text(path.stem)
                        for key in missing:
                            if stem == key or stem.startswith(key) or key in norm:
                                rank = self.preview_image_rank(path, key, root_order)
                                current_best = candidates.get(key)
                                if current_best is None or rank < current_best[0]:
                                    candidates[key] = (rank, path)
                    if all(key in candidates and candidates[key][0][1] <= 2 for key in missing):
                        break
        for key in missing:
            image = candidates[key][1] if key in candidates else None
            self.preview_image_cache[key] = image
            results[key] = image
        return results

    def warm_preview_cache(self, records: list[ProductRecord]) -> dict[str, int]:
        requested = {normalize_text(record.stock_code) for record in records if normalize_text(record.stock_code)}
        if not requested:
            return {"requested": 0, "found": 0, "indexed": 0, "scanned": 0}

        indexed = self.lookup_index_paths("preview", requested)
        for key, path in indexed.items():
            self.preview_image_cache[key] = path

        missing = {
            key
            for key in requested - set(indexed)
            if not isinstance(self.preview_image_cache.get(key), Path)
        }
        if not missing:
            return {"requested": len(requested), "found": len(indexed), "indexed": len(indexed), "scanned": 0}

        scanned = 0
        candidates = self.scan_image_repository_for_keys(missing)

        for key in missing:
            image = candidates.get(key)
            self.preview_image_cache[key] = image
        scanned = len(candidates)
        found = sum(1 for key in requested if isinstance(self.preview_image_cache.get(key), Path))
        return {"requested": len(requested), "found": found, "indexed": len(indexed), "scanned": scanned}

    def get_cached_preview_image(self, stock_code: str) -> Path | None:
        stock_key = normalize_text(stock_code)
        if not stock_key:
            return None
        cached = self.preview_image_cache.get(stock_key)
        if isinstance(cached, Path) and path_is_file(cached):
            return cached
        indexed = self.lookup_index_paths("preview", {stock_key})
        if stock_key in indexed:
            self.preview_image_cache[stock_key] = indexed[stock_key]
            return indexed[stock_key]
        return None

    def get_cached_preview_image_list(self, stock_code: str) -> list[Path]:
        stock_key = normalize_text(stock_code)
        if not stock_key:
            return []
        cached = self.preview_image_list_cache.get(stock_key)
        if cached:
            return [path for path in cached if path_is_file(path)]
        image = self.get_cached_preview_image(stock_code)
        return [image] if image is not None else []

    def get_qr_code_roots(self) -> list[Path]:
        root = Path(self.settings.get("qr_code_root", QR_CODE_ROOT_DEFAULT)).expanduser()
        return unique_paths([path for path in [root, root / "QRCODES"] if path_is_dir(path)])

    def find_sheet_by_name(self, workbook: Any, candidates: list[str]) -> Any | None:
        normalized = {normalize_text(candidate) for candidate in candidates}
        for sheet in workbook.worksheets:
            if normalize_text(sheet.title) in normalized:
                return sheet
        return None

    def build_qr_catalog_entry_from_record(self, record: ProductRecord) -> QrCatalogEntry:
        values = {str(key): str(value) for key, value in record.raw_values.items() if str(value).strip()}
        handle = first_matching_value(
            values,
            [("handle",), ("url", "handle"), ("shopify", "handle"), ("slug",)],
        )
        title = first_matching_value(
            values,
            [("title",), ("urun", "adi"), ("adi",), ("name",), ("aciklama",)],
            excludes=("image", "src", "url"),
        )
        image_url = first_matching_value(
            values,
            [("image", "src"), ("image", "url"), ("gorsel", "url"), ("resim", "url"), ("image",)],
        )
        return QrCatalogEntry(
            stock_code=record.stock_code,
            handle=slugify_handle(handle),
            title=(title or record.breakdown or record.family or record.stock_code).strip(),
            image_url=(image_url or "").strip(),
        )

    def load_qr_catalog(self) -> dict[str, QrCatalogEntry]:
        if self.qr_catalog_cache is not None:
            return self.qr_catalog_cache
        catalog: dict[str, QrCatalogEntry] = {}
        for record in self.records:
            if not record.stock_code:
                continue
            key = normalize_text(record.stock_code)
            if key in catalog:
                continue
            catalog[key] = self.build_qr_catalog_entry_from_record(record)
        self.qr_catalog_cache = catalog
        return catalog

    def get_qr_catalog_entry(self, record: ProductRecord) -> QrCatalogEntry | None:
        return self.load_qr_catalog().get(normalize_text(record.stock_code))

    def get_qr_title(self, record: ProductRecord) -> str:
        entry = self.get_qr_catalog_entry(record)
        return (entry.title if entry and entry.title else record.breakdown or record.family or record.stock_code).strip()

    def get_qr_handle(self, record: ProductRecord, qr_path: Path | None = None) -> str:
        entry = self.get_qr_catalog_entry(record)
        if entry and entry.handle:
            return entry.handle
        for key, value in record.raw_values.items():
            if normalize_text(key) in {"handle", "url handle", "shopify handle"} and value:
                return slugify_handle(value)
        if qr_path is not None:
            return qr_path.stem
        return slugify_handle(self.get_qr_title(record))

    def get_qr_file_index(self) -> dict[str, Path]:
        if self.qr_file_index_cache is not None:
            return self.qr_file_index_cache
        index: dict[str, Path] = {}
        for root in self.get_qr_code_roots():
            for path in safe_rglob_files(root):
                if path.suffix.lower() != ".png":
                    continue
                key = normalize_text(path.stem)
                if key and key not in index:
                    index[key] = path
        self.qr_file_index_cache = index
        return index

    def find_qr_code(self, record: ProductRecord) -> Path | None:
        stock_key = normalize_text(record.stock_code)
        if stock_key in self.qr_file_cache:
            return self.qr_file_cache[stock_key]
        index = self.get_qr_file_index()
        candidates = [
            normalize_text(self.get_qr_handle(record)),
            normalize_text(record.stock_code),
            normalize_text(slugify_handle(record.breakdown)),
        ]
        for candidate in [item for item in candidates if item]:
            if candidate in index:
                self.qr_file_cache[stock_key] = index[candidate]
                return index[candidate]
        scored: list[tuple[int, int, str, Path]] = []
        for candidate in [item for item in candidates if item]:
            for stem, path in index.items():
                if stem.startswith(candidate):
                    scored.append((0, len(stem), stem, path))
                elif candidate in stem:
                    scored.append((1, len(stem), stem, path))
        result = min(scored, default=None)
        self.qr_file_cache[stock_key] = result[3] if result else None
        return self.qr_file_cache[stock_key]

    def get_qr_web_url(self, record: ProductRecord, handle: str) -> str:
        for value in record.raw_values.values():
            text = str(value or "").strip()
            if text.casefold().startswith("http") and "/products/" in text.casefold():
                return text
        sitemap_url = self.get_qr_sitemap_product_url(handle)
        if sitemap_url:
            return sitemap_url
        base = self.settings.get("qr_url_base", QR_URL_BASE_DEFAULT).strip() or QR_URL_BASE_DEFAULT
        if not base.endswith("/"):
            base += "/"
        return f"{base}{handle}" if handle else base

    def extract_product_handle_from_url(self, value: str) -> str:
        try:
            path = unquote(urlparse(value).path).strip("/")
        except Exception:
            return ""
        if not path.startswith("products/"):
            return ""
        handle = path.split("products/", 1)[1].strip("/").split("/", 1)[0]
        return slugify_handle(handle)

    def load_qr_sitemap_urls(self) -> dict[str, str]:
        if self.qr_sitemap_cache is not None:
            return self.qr_sitemap_cache
        mapping: dict[str, str] = {}
        sitemap_url = self.settings.get("qr_sitemap_url", "").strip()
        if not sitemap_url:
            self.qr_sitemap_cache = mapping
            return mapping
        try:
            with urlopen(sitemap_url, timeout=30) as response:
                payload = response.read()
            root = ET.fromstring(payload)
            namespace = "{http://www.sitemaps.org/schemas/sitemap/0.9}"
            for loc in root.findall(f".//{namespace}loc"):
                url = stringify_cell(loc.text)
                handle = self.extract_product_handle_from_url(url)
                if handle and handle not in mapping:
                    mapping[normalize_text(handle)] = url
        except Exception as exc:  # noqa: BLE001
            self.product_data_warning = f"QR sitemap okunamadi: {exc}"
        self.qr_sitemap_cache = mapping
        return mapping

    def get_qr_sitemap_product_url(self, handle: str) -> str:
        if not handle:
            return ""
        return self.load_qr_sitemap_urls().get(normalize_text(handle), "")

    def build_qr_export_rows(self, records: list[ProductRecord]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for record in records:
            qr_path = self.find_qr_code(record)
            handle = self.get_qr_handle(record, qr_path)
            safe_handle = handle or slugify_handle(record.stock_code) or slugify_handle(self.get_qr_title(record))
            rows.append({
                "sku": record.stock_code,
                "title": self.get_qr_title(record),
                "handle": safe_handle,
                "qr_file": qr_path.name if qr_path else f"{safe_handle}.png" if safe_handle else "",
                "qr_path": str(qr_path) if qr_path else "",
                "qr_missing": "1" if qr_path is None else "0",
                "web_url": self.get_qr_web_url(record, safe_handle),
            })
        return rows

    def get_qr_output_root(self) -> Path:
        output = Path(self.settings.get("qr_output_root", "")).expanduser()
        return output if str(output).strip() else DATA_DIR / "qr_exports"

    def get_ce_measurement_workbook_path(self) -> Path:
        value = self.settings.get("ce_measurement_excel_path", "").strip() or CE_MEASUREMENTS_XLSX_DEFAULT
        return resolve_runtime_path(value)

    def get_ce_output_root(self) -> Path:
        output = Path(self.settings.get("ce_output_root", "")).expanduser()
        return output if str(output).strip() else CE_OUTPUT_ROOT_DEFAULT

    def find_header_row(self, worksheet: Any, required_headers: list[str], max_scan_rows: int = 8) -> tuple[int, list[str]] | None:
        for row_index, row_values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
            start=1,
        ):
            headers = [stringify_cell(value).strip() for value in row_values]
            normalized = {normalize_text(header) for header in headers if header}
            if all(normalize_text(header) in normalized for header in required_headers):
                return row_index, headers
        return None

    def merge_ce_label_entry(self, current: CeLabelEntry | None, incoming: CeLabelEntry) -> CeLabelEntry:
        if current is None:
            return incoming
        return CeLabelEntry(
            stock_code=current.stock_code or incoming.stock_code,
            product_name=merge_text(current.product_name, incoming.product_name),
            mancode=merge_text(current.mancode, incoming.mancode),
            model=merge_text(current.model, incoming.model),
            product_group=merge_text(current.product_group, incoming.product_group),
            dimming=merge_text(current.dimming, incoming.dimming),
            voltage_110=merge_text(current.voltage_110, incoming.voltage_110),
            voltage_220=merge_text(current.voltage_220, incoming.voltage_220),
            driver_model=merge_text(current.driver_model, incoming.driver_model),
            cable_type=merge_text(current.cable_type, incoming.cable_type),
            kelvin=merge_text(current.kelvin, incoming.kelvin),
            lumen=merge_text(current.lumen, incoming.lumen),
            watt=merge_text(current.watt, incoming.watt),
            duy=merge_text(current.duy, incoming.duy),
            label_kind=merge_text(current.label_kind, incoming.label_kind),
        )

    def build_ce_entry_from_row(self, headers: list[str], row: list[str], sheet_name: str) -> CeLabelEntry | None:
        stock_index = resolve_first_header_index(headers, ["STOK KODU", "SKU", "STOCK CODE"])
        if stock_index is None or stock_index >= len(row):
            return None
        stock_code = row[stock_index].strip()
        if not stock_code:
            return None
        product_name_index = resolve_first_header_index(headers, ["ÃœRÃœN AÃ‡IKLAMA", "ÃœRÃœN", "ÃœRÃœN ADI", "NAME", "ADI"])
        mancode_index = resolve_first_header_index(headers, ["ÃœRETÄ°CÄ° KODU", "MANCODE", "MAN CODE"])
        model_index = resolve_first_header_index(headers, ["MODEL"])
        product_group_index = resolve_first_header_index(headers, ["ÃœRÃœN GRUBU"])
        dimming_index = resolve_first_header_index(headers, ["DÄ°MLENME", "DIMM", "DIMLENME"])
        voltage_110_index = resolve_first_header_index(headers, ["110 V Ã‡ALIÅMA"])
        voltage_220_index = resolve_first_header_index(headers, ["220 V Ã‡ALIÅMA"])
        driver_model_index = resolve_first_header_index(headers, ["LED DRÄ°VER MODELÄ°", "LED DRIVER MODEL"])
        cable_type_index = resolve_first_header_index(headers, ["KABLO TÃœRÃœ"])
        kelvin_index = resolve_first_header_index(headers, ["KELVÄ°N", "KELVIN", "RENK SICAKLIÄI", "RENK SICAKLIGI", "IÅIK RENGÄ°", "ISIK RENGI"])
        lumen_index = resolve_first_header_index(headers, ["LÃœMEN", "LUMEN", "LM", "IÅIK AKISI", "ISIK AKISI"])
        watt_index = resolve_first_header_index(headers, ["Ã–LÃ‡ÃœLEN WATT DEÄERÄ°", "WATT", "WATT DEÄERÄ°", "DRÄ°VER WATT DEÄERÄ°", "GÃœÃ‡", "GUC"])
        duy_index = resolve_first_header_index(headers, ["DUY / LED TÄ°PÄ°", "DUY TÄ°PÄ°", "DUY TIPI", "DUY", "LED", "DUY/LED TÄ°PÄ°", "DUY/LED TIPI"])

        def get_value(index: int | None) -> str:
            return row[index].strip() if index is not None and index < len(row) else ""

        duy_value = get_value(duy_index)
        label_kind = "LED" if "led" in normalize_text(duy_value) else "DUY" if duy_value else ""
        return CeLabelEntry(
            stock_code=stock_code,
            product_name=get_value(product_name_index),
            mancode=get_value(mancode_index),
            model=get_value(model_index),
            product_group=get_value(product_group_index),
            dimming=get_value(dimming_index),
            voltage_110=get_value(voltage_110_index),
            voltage_220=get_value(voltage_220_index),
            driver_model=get_value(driver_model_index),
            cable_type=get_value(cable_type_index),
            kelvin=get_value(kelvin_index),
            lumen=get_value(lumen_index),
            watt=get_value(watt_index),
            duy=duy_value,
            label_kind=label_kind,
        )

    def build_ce_fallback_entry(self, record: ProductRecord) -> CeLabelEntry:
        values = {normalize_text(str(key)): str(value).strip() for key, value in record.raw_values.items() if str(value).strip()}

        def pick(*keys: str) -> str:
            for key in keys:
                value = values.get(normalize_text(key), "")
                if value:
                    return value
            return ""

        duy_value = pick("DUY / LED TÄ°PÄ°", "DUY TÄ°PÄ°", "DUY TIPI", "DUY", "LED", "DUY/LED TÄ°PÄ°", "DUY/LED TIPI")
        label_kind = "LED" if "led" in normalize_text(duy_value) else "DUY" if duy_value else ""
        return CeLabelEntry(
            stock_code=record.stock_code,
            product_name=record.breakdown or pick("ÃœRÃœN AÃ‡IKLAMA", "ÃœRÃœN", "ÃœRÃœN ADI", "ADI", "ADI2", "NAME"),
            mancode=pick("ÃœRETÄ°CÄ° KODU", "URETICI KODU", "MANCODE", "MAN CODE", "ESKIKOD", "ESKÄ°KOD"),
            model=pick("MODEL"),
            product_group=pick("ÃœRÃœN GRUBU"),
            dimming=pick("DÄ°MLENME", "DIMM", "DIMLENME"),
            voltage_110=pick("110 V Ã‡ALIÅMA"),
            voltage_220=pick("220 V Ã‡ALIÅMA"),
            driver_model=pick("LED DRÄ°VER MODELÄ°", "LED DRIVER MODEL"),
            cable_type=pick("KABLO TÃœRÃœ"),
            kelvin=pick("KELVÄ°N", "KELVIN", "RENK SICAKLIÄI", "RENK SICAKLIGI", "IÅIK RENGÄ°", "ISIK RENGI"),
            lumen=pick("LÃœMEN", "LUMEN", "LM", "IÅIK AKISI", "ISIK AKISI"),
            watt=pick("Ã–LÃ‡ÃœLEN WATT DEÄERÄ°", "WATT", "WATT DEÄERÄ°", "DRÄ°VER WATT DEÄERÄ°", "GÃœÃ‡", "GUC"),
            duy=duy_value,
            label_kind=label_kind,
        )

    def load_ce_label_catalog(self) -> dict[str, CeLabelEntry]:
        if self.ce_label_cache is not None:
            return self.ce_label_cache
        workbook_path = self.get_ce_measurement_workbook_path()
        if not path_is_file(workbook_path):
            raise FileNotFoundError(f"CE kaynak Excel bulunamadi: {workbook_path}")
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        catalog: dict[str, CeLabelEntry] = {}
        try:
            sheets: list[Any] = []
            for title_candidates in (
                ["YENÄ° ÃœRETÄ°M TESTLERÄ°", "YENI URETIM TESTLERI"],
                ["Sayfa1", "SAYFA1"],
            ):
                sheet = self.find_sheet_by_name(workbook, title_candidates)
                if sheet is not None and sheet not in sheets:
                    sheets.append(sheet)
            for sheet in sheets:
                header_result = self.find_header_row(sheet, ["STOK KODU"])
                if header_result is None:
                    continue
                header_row, headers = header_result
                for row_values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    row = [stringify_cell(value) for value in row_values]
                    if not any(row):
                        continue
                    while len(row) < len(headers):
                        row.append("")
                    entry = self.build_ce_entry_from_row(headers, row, sheet.title)
                    if entry is None:
                        continue
                    key = normalize_text(entry.stock_code)
                    catalog[key] = self.merge_ce_label_entry(catalog.get(key), entry)
        finally:
            workbook.close()
        self.ce_label_cache = catalog
        return catalog

    def get_ce_label_entry(self, record: ProductRecord) -> CeLabelEntry:
        try:
            entry = self.load_ce_label_catalog().get(normalize_text(record.stock_code))
        except Exception as exc:
            self.product_data_warning = f"CE kaynak Excel okunamadi: {exc}"
            entry = None
        return self.merge_ce_label_entry(entry, self.build_ce_fallback_entry(record))

    def infer_ce_duy_text(self, entry: CeLabelEntry, record: ProductRecord) -> str:
        explicit = str(entry.duy or "").strip()
        explicit_key = normalize_text(explicit)
        if explicit and explicit_key not in {"duy", "duyledtipi", "duyledtipi"}:
            return explicit
        candidates = [
            entry.driver_model,
            entry.cable_type,
            entry.model,
            entry.product_name,
            record.breakdown,
            record.stock_code,
        ]
        for candidate in candidates:
            text = str(candidate or "").upper().replace(" ", "")
            if not text:
                continue
            for token in ("E27", "E14", "G9", "GU10", "GX53", "MR16", "E40"):
                if token in text:
                    return token
            if "LED" in text:
                return "LED"
            if "DUY" in text:
                return "DUY"
        driver_model_text = str(entry.driver_model or "").upper()
        if driver_model_text and "YOK" not in driver_model_text:
            return "LED"
        return ""

    def infer_ce_label_kind(self, entry: CeLabelEntry, record: ProductRecord, duy_text: str) -> str:
        explicit_kind = normalize_text(entry.label_kind)
        if explicit_kind == "led":
            return "LED"
        if explicit_kind == "duy":
            return "DUY"

        normalized_duy = normalize_text(duy_text)
        if normalized_duy == "led":
            return "LED"
        if normalized_duy:
            return "DUY"

        candidates = [
            entry.product_name,
            record.breakdown,
            entry.driver_model,
            entry.cable_type,
            entry.model,
            record.stock_code,
        ]
        for candidate in candidates:
            text = normalize_text(candidate)
            if not text:
                continue
            if "led" in text:
                return "LED"
            if "duy" in text:
                return "DUY"
            if any(token in text for token in ("e27", "e14", "g9", "gu10", "gx53", "mr16", "e40")):
                return "DUY"
        return ""

    def ce_text_pool(self, entry: CeLabelEntry, record: ProductRecord) -> str:
        ignored_raw_keys = {"kodu", "stok kodu", "sku", "stock code", "eskikod", "eski kod", "uretici kodu", "Ã¼retici kodu", "mancode"}
        values = [
            entry.product_name,
            entry.model,
            entry.product_group,
            entry.driver_model,
            entry.cable_type,
            entry.dimming,
            record.family,
            record.breakdown,
            *[
                str(value)
                for key, value in record.raw_values.items()
                if normalize_text(str(key)) not in ignored_raw_keys
            ],
        ]
        return " ".join(str(value or "") for value in values if str(value or "").strip())

    def infer_ce_kelvin_value(self, entry: CeLabelEntry, record: ProductRecord) -> str:
        if str(entry.kelvin or "").strip():
            return str(entry.kelvin).strip()
        text = self.ce_text_pool(entry, record)
        match = re.search(r"\b([1-9]\d{3,4})\s*(?:K|KELVIN)\b", text, re.IGNORECASE)
        return f"{match.group(1)} K" if match else ""

    def infer_ce_lumen_value(self, entry: CeLabelEntry, record: ProductRecord) -> str:
        if str(entry.lumen or "").strip():
            return str(entry.lumen).strip()
        text = self.ce_text_pool(entry, record)
        match = re.search(r"\b(\d+(?:[.,]\d+)?(?:\s*-\s*\d+(?:[.,]\d+)?)?)\s*(?:LM|LUMEN)\b", text, re.IGNORECASE)
        return f"{match.group(1).replace(' ', '')} lm" if match else ""

    def infer_ce_watt_value(self, entry: CeLabelEntry, record: ProductRecord) -> str:
        if str(entry.watt or "").strip():
            return str(entry.watt).strip()
        text = self.ce_text_pool(entry, record)
        match = re.search(r"\b(\d+(?:[.,]\d+)?)\s*(?:W|WATT)\b", text, re.IGNORECASE)
        return f"{match.group(1)} W" if match else ""

    def ce_export_text(self, value: str, placeholder: str = "-") -> str:
        text = str(value or "").strip()
        return turkish_upper(text) if text else placeholder

    def build_ce_export_rows(self, items: list[tuple[ProductRecord, int]], gap_labels: int = 0) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        effective_items = [(record, quantity) for record, quantity in items if quantity > 0]
        for index, (record, quantity) in enumerate(effective_items):
            if quantity <= 0:
                continue
            entry = self.get_ce_label_entry(record)
            duy_text = self.infer_ce_duy_text(entry, record)
            label_kind = self.infer_ce_label_kind(entry, record, duy_text)
            kelvin_text = self.infer_ce_kelvin_value(entry, record)
            lumen_text = self.infer_ce_lumen_value(entry, record)
            watt_text = self.infer_ce_watt_value(entry, record)
            rows.append({
                "TÃœR": self.ce_export_text(label_kind),
                "NAME": self.ce_export_text(entry.product_name or record.breakdown or record.stock_code),
                "SKU": record.stock_code,
                "MANCODE": self.ce_export_text(entry.mancode),
                "KELVIN": self.ce_export_text(kelvin_text),
                "LUMEN": self.ce_export_text(lumen_text),
                "WATT": self.ce_export_text(watt_text),
                "DUY": self.ce_export_text(duy_text),
                "ADT": str(quantity),
            })
            if gap_labels > 0 and index < len(effective_items) - 1:
                rows.append({
                    "TÃœR": "",
                    "NAME": "",
                    "SKU": "",
                    "MANCODE": "",
                    "KELVIN": "",
                    "LUMEN": "",
                    "WATT": "",
                    "DUY": "",
                    "ADT": str(gap_labels),
                })
        return rows

    def save_ce_bartender_workbook(self, rows: list[dict[str, str]], path: Path) -> Path:
        workbook = Workbook()
        final_sheet = workbook.active
        final_sheet.title = "BARKOD FÄ°NAL"
        headers = ["TÃœR", "NAME", "SKU", "MANCODE", "KELVIN", "LUMEN", "WATT", "DUY", "ADT"]
        final_sheet.append(headers)
        for row in rows:
            final_sheet.append([row.get(header, "") for header in headers])
        widths = {
            "A": 8.8,
            "B": 53.3,
            "C": 18.6,
            "D": 10.1,
            "E": 11.0,
            "F": 7.7,
            "G": 8.1,
            "H": 8.3,
            "I": 4.9,
        }
        for column, width in widths.items():
            final_sheet.column_dimensions[column].width = width
        final_sheet.freeze_panes = "A2"
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        workbook.close()
        return path

    def export_ce_bartender_excel(self, items: list[tuple[ProductRecord, int]], gap_labels: int = 0) -> Path:
        rows = self.build_ce_export_rows(items, gap_labels)
        if not rows:
            raise ValueError("Secili urun yok.")
        output_root = self.get_ce_output_root()
        path = output_root / f"CE_Barkod_Listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.save_ce_bartender_workbook(rows, path)

    def bartender_print_timeout(self) -> int:
        try:
            return max(15, int(float(self.settings.get("bartender_print_timeout_seconds", "180"))))
        except ValueError:
            return 180

    def truthy_setting(self, key: str, default: str = "0") -> bool:
        value = str(self.settings.get(key, default)).strip().casefold()
        return value in {"1", "true", "evet", "yes", "on", "aktif"}

    def label_reset_feed_count(self) -> int:
        try:
            return max(1, min(6, int(float(self.settings.get("label_reset_feed_count", "1")))))
        except ValueError:
            return 1

    def label_reset_pause_seconds(self) -> float:
        try:
            return max(0.0, min(5.0, float(self.settings.get("label_reset_pause_ms", "1200")) / 1000.0))
        except ValueError:
            return 1.2

    def parse_raw_printer_command(self, value: str) -> bytes:
        text = str(value or "").strip()
        if not text:
            return b""
        if re.fullmatch(r"(?:[0-9A-Fa-f]{2}[\s,;:-]*)+", text):
            pairs = re.findall(r"[0-9A-Fa-f]{2}", text)
            return bytes(int(pair, 16) for pair in pairs)
        try:
            return text.encode("utf-8").decode("unicode_escape").encode("latin1")
        except (UnicodeDecodeError, UnicodeEncodeError, ValueError):
            return text.encode("latin1", errors="ignore")

    def build_label_reset_payload(self) -> bytes:
        command = self.parse_raw_printer_command(self.settings.get("label_reset_raw_command", "\\x0c"))
        if not command:
            command = b"\x0c"
        return command * self.label_reset_feed_count()

    def send_raw_to_printer(self, printer_name: str, payload: bytes, document_name: str) -> int:
        if os.name != "nt":
            raise RuntimeError("Yazici sifirlama sadece Windows yazici kuyruÄŸu uzerinden destekleniyor.")
        if not printer_name.strip():
            raise ValueError("Barkod yazici adi bos. Ayarlardan yazici adini gir.")
        if not payload:
            raise ValueError("Yazici sifirlama komutu bos.")

        import ctypes
        from ctypes import wintypes

        class DOC_INFO_1(ctypes.Structure):
            _fields_ = [
                ("pDocName", wintypes.LPWSTR),
                ("pOutputFile", wintypes.LPWSTR),
                ("pDatatype", wintypes.LPWSTR),
            ]

        winspool = ctypes.WinDLL("winspool.drv", use_last_error=True)
        winspool.OpenPrinterW.argtypes = [wintypes.LPWSTR, ctypes.POINTER(wintypes.HANDLE), wintypes.LPVOID]
        winspool.OpenPrinterW.restype = wintypes.BOOL
        winspool.ClosePrinter.argtypes = [wintypes.HANDLE]
        winspool.ClosePrinter.restype = wintypes.BOOL
        winspool.StartDocPrinterW.argtypes = [wintypes.HANDLE, wintypes.DWORD, wintypes.LPVOID]
        winspool.StartDocPrinterW.restype = wintypes.DWORD
        winspool.EndDocPrinter.argtypes = [wintypes.HANDLE]
        winspool.EndDocPrinter.restype = wintypes.BOOL
        winspool.StartPagePrinter.argtypes = [wintypes.HANDLE]
        winspool.StartPagePrinter.restype = wintypes.BOOL
        winspool.EndPagePrinter.argtypes = [wintypes.HANDLE]
        winspool.EndPagePrinter.restype = wintypes.BOOL
        winspool.WritePrinter.argtypes = [wintypes.HANDLE, wintypes.LPVOID, wintypes.DWORD, ctypes.POINTER(wintypes.DWORD)]
        winspool.WritePrinter.restype = wintypes.BOOL

        handle = wintypes.HANDLE()
        if not winspool.OpenPrinterW(printer_name, ctypes.byref(handle), None):
            raise ctypes.WinError(ctypes.get_last_error())
        doc_started = False
        page_started = False
        try:
            doc_info = DOC_INFO_1(document_name, None, "RAW")
            job_id = winspool.StartDocPrinterW(handle, 1, ctypes.byref(doc_info))
            if not job_id:
                raise ctypes.WinError(ctypes.get_last_error())
            doc_started = True
            if not winspool.StartPagePrinter(handle):
                raise ctypes.WinError(ctypes.get_last_error())
            page_started = True
            buffer = ctypes.create_string_buffer(payload)
            written = wintypes.DWORD()
            if not winspool.WritePrinter(handle, buffer, len(payload), ctypes.byref(written)):
                raise ctypes.WinError(ctypes.get_last_error())
            if written.value != len(payload):
                raise RuntimeError(f"Yazici sifirlama komutu eksik gonderildi: {written.value}/{len(payload)} bayt.")
            return int(job_id)
        finally:
            if page_started:
                winspool.EndPagePrinter(handle)
            if doc_started:
                winspool.EndDocPrinter(handle)
            winspool.ClosePrinter(handle)

    def reset_label_printer(self, reason: str = "Etiket sifirlama", force: bool = False) -> int | None:
        if not force and not self.truthy_setting("label_reset_enabled", "1"):
            return None
        printer_name = self.settings.get("bartender_printer_name", "").strip()
        payload = self.build_label_reset_payload()
        job_id = self.send_raw_to_printer(printer_name, payload, reason)
        pause = self.label_reset_pause_seconds()
        if pause > 0:
            time.sleep(pause)
        return job_id

    def build_bartender_args(self, data_path: Path, template_key: str, job_name: str) -> list[str]:
        bartender_exe = self.resolve_bartender_exe_path(self.settings.get("bartender_exe_path", ""))
        if bartender_exe is None:
            raise FileNotFoundError("BarTender EXE bulunamadi. Ayarlardan BarTend.exe yolunu sec.")
        template_path = resolve_runtime_path(self.settings.get(template_key, ""))
        if not path_is_file(template_path):
            raise FileNotFoundError("BarTender template bulunamadi. Ayarlardan ilgili .btw dosyasini sec.")
        if not path_is_file(data_path):
            raise FileNotFoundError(f"BarTender veri dosyasi bulunamadi: {data_path}")
        printer_name = self.settings.get("bartender_printer_name", "").strip()
        args = [
            str(bartender_exe),
            f"/AF={template_path}",
            f"/D={data_path}",
            f"/PrintJobName={job_name}",
            "/FP",
            "/X",
        ]
        if printer_name:
            args.insert(3, f"/PRN={printer_name}")
        return args

    def build_bartender_command(self, data_path: Path, template_key: str, job_name: str) -> str:
        return subprocess.list2cmdline(self.build_bartender_args(data_path, template_key, job_name))

    def run_bartender_print(self, data_path: Path, template_key: str, job_name: str) -> Path:
        args = self.build_bartender_args(data_path, template_key, job_name)
        completed = subprocess.run(
            args,
            capture_output=True,
            text=True,
            errors="replace",
            timeout=self.bartender_print_timeout(),
            shell=False,
        )
        if completed.returncode != 0:
            details = "\n".join(part for part in [completed.stdout.strip(), completed.stderr.strip()] if part)
            raise RuntimeError(f"BarTender print basarisiz. Kod: {completed.returncode}\n{details}")
        return data_path

    def print_ce_bartender_excel(self, items: list[tuple[ProductRecord, int]], gap_labels: int = 0) -> Path:
        rows = self.build_ce_export_rows(items, gap_labels)
        if not rows:
            raise ValueError("Secili urun yok.")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_root = self.get_ce_output_root() / f"CE_Direkt_Baski_{timestamp}"
        reset_each_label = self.truthy_setting("label_reset_before_each_label", "0")
        if not reset_each_label:
            self.reset_label_printer(f"HEKA CE etiket sifirlama {timestamp}")
        for index, row in enumerate(rows, start=1):
            sku = safe_path_part(row.get("SKU") or "bos-etiket", f"etiket-{index:03d}")
            data_path = output_root / f"{index:03d}_{sku}.xlsx"
            self.save_ce_bartender_workbook([row], data_path)
            if reset_each_label:
                self.reset_label_printer(f"HEKA CE etiket sifirlama {timestamp} {index:03d}")
            self.run_bartender_print(data_path, "ce_bartender_template_path", f"HEKA CE {timestamp} {index:03d}")
            time.sleep(0.2)
        return output_root

    def export_bartender_qr_excel(self, records: list[ProductRecord]) -> Path:
        rows = self.build_qr_export_rows(records)
        if not rows:
            raise ValueError("Secili urun yok.")

        output_root = self.get_qr_output_root()
        path = output_root / f"Bartender_Baski_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.save_qr_bartender_workbook(rows, path)

    def save_qr_bartender_workbook(self, rows: list[dict[str, str]], path: Path) -> Path:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "BARTENDER"
        headers = ["SKU", "\u00dcR\u00dcN ADI", "QR DOSYA ADI"]
        worksheet.append(headers)
        for row_data in rows:
            worksheet.append([row_data["sku"], turkish_upper(row_data["title"]), row_data["qr_file"]])
        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 54
        worksheet.column_dimensions["C"].width = 72
        worksheet.freeze_panes = "A2"
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        workbook.close()
        return path

    def print_bartender_qr_excel(self, records: list[ProductRecord]) -> Path:
        rows = self.build_qr_export_rows(records)
        if not rows:
            raise ValueError("Secili urun yok.")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_root = self.get_qr_output_root() / f"QR_Direkt_Baski_{timestamp}"
        reset_each_label = self.truthy_setting("label_reset_before_each_label", "0")
        if not reset_each_label:
            self.reset_label_printer(f"HEKA QR etiket sifirlama {timestamp}")
        for index, row in enumerate(rows, start=1):
            sku = safe_path_part(row.get("sku") or "qr-etiket", f"etiket-{index:03d}")
            data_path = output_root / f"{index:03d}_{sku}.xlsx"
            self.save_qr_bartender_workbook([row], data_path)
            if reset_each_label:
                self.reset_label_printer(f"HEKA QR etiket sifirlama {timestamp} {index:03d}")
            self.run_bartender_print(data_path, "qr_bartender_template_path", f"HEKA QR {timestamp} {index:03d}")
            time.sleep(0.2)
        return output_root

    def export_qrcodechimp_excel(self, records: list[ProductRecord]) -> Path:
        rows = self.build_qr_export_rows(records)
        if not rows:
            raise ValueError("Secili urun yok.")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Worksheet"
        worksheet.cell(1, 1).value = "QR Name*"
        worksheet.cell(1, 2).value = "Web URL*"
        for row in range(2, max(worksheet.max_row, len(rows) + 1) + 1):
            worksheet.cell(row, 1).value = None
            worksheet.cell(row, 2).value = None
        for index, row_data in enumerate(rows, start=2):
            worksheet.cell(index, 1).value = row_data["handle"] or slugify_handle(row_data["sku"])
            worksheet.cell(index, 2).value = row_data["web_url"]
        worksheet.column_dimensions["A"].width = 42
        worksheet.column_dimensions["B"].width = 78
        output_root = self.get_qr_output_root()
        output_root.mkdir(parents=True, exist_ok=True)
        path = output_root / f"QRCodeChimp_Bulk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(path)
        workbook.close()
        return path

    def folder_path_rank(self, path: Path, stock_key: str) -> tuple[int, int, int, str]:
        parts = [normalize_text(part) for part in path.parts]
        name = normalize_text(path.name)
        media_score = 0 if any(self.is_product_images_folder_name(part) for part in path.parts) else 2
        stock_score = 0 if name.startswith(stock_key) else 1 if any(part.startswith(stock_key) for part in parts) else 3
        channel_score = 0 if any(word in normalize_text(str(path)) for word in ("katalog", "web", "etsy", "b2b")) else 1
        return (media_score, stock_score + channel_score, len(path.parts), str(path).casefold())

    def find_search_folders(self, stock_codes: list[str]) -> dict[str, Path | None]:
        requested = {normalize_text(code): code for code in stock_codes if normalize_text(code)}
        results: dict[str, Path | None] = {}
        missing = set(requested)
        indexed = self.lookup_index_paths("folder", missing)
        for key, path in indexed.items():
            self.search_folder_cache[key] = path
            results[key] = path
        missing -= set(indexed)
        for key in list(missing):
            if key in self.search_folder_cache:
                results[key] = self.search_folder_cache[key]
                missing.remove(key)
        if not missing:
            return results
        root = Path(self.settings.get("search_root", "")).expanduser()
        found: dict[str, tuple[tuple[int, int, int, str], Path]] = {}
        if path_is_dir(root):
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for directory in dirs:
                    path = Path(current) / directory
                    norm = normalize_text(str(path))
                    name_norm = normalize_text(directory)
                    for key in missing:
                        if name_norm.startswith(key) or key in norm:
                            rank = self.folder_path_rank(path, key)
                            current_best = found.get(key)
                            if current_best is None or rank < current_best[0]:
                                found[key] = (rank, path)
                for file_name in files:
                    path = Path(current) / file_name
                    norm = normalize_text(str(path))
                    for key in missing:
                        if key in norm:
                            parent = Path(current)
                            rank = self.folder_path_rank(parent, key)
                            current_best = found.get(key)
                            if current_best is None or rank < current_best[0]:
                                found[key] = (rank, parent)
                if all(key in found and found[key][0][0] == 0 and found[key][0][1] <= 1 for key in missing):
                    break
        for key in missing:
            folder = found[key][1] if key in found else None
            self.search_folder_cache[key] = folder
            results[key] = folder
        return results

    def collect_product_images(self, stock_code: str, image_path: Path | None, folder_path: Path | None) -> list[Path]:
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        stock_key = normalize_text(stock_code)
        matches: list[Path] = []
        seen: set[str] = set()

        def add_match(path: Path) -> None:
            key = str(path).casefold()
            if key not in seen and path_is_file(path):
                seen.add(key)
                matches.append(path)

        for cached_image in self.get_cached_preview_image_list(stock_code):
            add_match(cached_image)
        if image_path is not None:
            add_match(image_path)

        roots = unique_paths([path for path in [folder_path] if isinstance(path, Path) and path_is_dir(path)])
        for root in roots:
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if valid and path.suffix.lower() not in valid:
                        continue
                    norm = normalize_text(str(path))
                    stem = normalize_text(path.stem)
                    if stem == stock_key or stem.startswith(stock_key) or stock_key in norm:
                        add_match(path)
        return sorted(matches, key=lambda path: self.preview_image_rank(path, stock_key))

    def path_matches_channel(self, path: Path, channel_key: str) -> bool:
        text = normalize_text(str(path))
        keywords = list(CHANNEL_FOLDERS.get(channel_key, ("", ()))[1])
        if channel_key == "technical":
            keywords.extend(split_multivalue_config(self.settings.get("technical_keywords", "")))
        return any(normalize_text(keyword) in text for keyword in keywords if keyword)

    def path_contains_stock(self, path: Path, stock_key: str) -> bool:
        return any(normalize_text(part).startswith(stock_key) for part in path.parts)

    def channel_path_rank(self, path: Path, stock_key: str, channel_key: str) -> tuple[int, int, int, int, str]:
        parts = [normalize_text(part) for part in path.parts]
        name = normalize_text(path.name)
        keywords = list(CHANNEL_FOLDERS.get(channel_key, ("", ()))[1])
        if channel_key == "technical":
            keywords.extend(split_multivalue_config(self.settings.get("technical_keywords", "")))
        channel_exact = 0 if any(normalize_text(keyword) in name for keyword in keywords if keyword) else 1
        stock_score = 0 if name.startswith(stock_key) else 1 if any(part.startswith(stock_key) for part in parts) else 4
        media_score = 0 if any(self.is_product_images_folder_name(part) for part in path.parts) else 2
        return (media_score, stock_score, channel_exact, len(path.parts), str(path).casefold())

    def channel_search_roots(self, folder_hint: Path | None, channel_key: str = "") -> list[Path]:
        roots: list[Path] = []
        search_root = Path(self.settings.get("search_root", "")).expanduser()
        if folder_hint and path_is_dir(folder_hint):
            roots.append(folder_hint)
            current = folder_hint
            max_depth = 9 if channel_key == "technical" else 6
            for _ in range(max_depth):
                parent = current.parent
                if parent == current or str(parent) == parent.anchor:
                    break
                roots.append(parent)
                current = parent
                if path_is_dir(search_root) and current == search_root:
                    break
                if channel_key != "technical" and self.is_product_images_folder_name(current.name):
                    break
        if path_is_dir(search_root):
            roots.append(search_root)
        return unique_paths(roots)

    def find_channel_folder(self, stock_code: str, channel_key: str, folder_hint: Path | None = None) -> Path | None:
        key = f"{normalize_text(stock_code)}::{channel_key}"
        stock_key = normalize_text(stock_code)
        if key in self.channel_folder_cache:
            return self.channel_folder_cache[key]
        indexed = self.lookup_index_paths("channel", {stock_key}, channel_key)
        if stock_key in indexed:
            self.channel_folder_cache[key] = indexed[stock_key]
            return indexed[stock_key]
        candidates: list[Path] = []
        search_root = Path(self.settings.get("search_root", "")).expanduser()
        for root in self.channel_search_roots(folder_hint, channel_key):
            root_candidates: list[Path] = []
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                current_path = Path(current)
                root_channel = self.path_matches_channel(current_path, channel_key)
                root_stock = self.path_contains_stock(current_path, stock_key)
                file_stock = any(stock_key in normalize_text(file_name) for file_name in files)
                for directory in dirs:
                    path = current_path / directory
                    dir_stock = normalize_text(directory).startswith(stock_key)
                    dir_channel = self.path_matches_channel(path, channel_key)
                    if dir_stock and (root_channel or dir_channel):
                        root_candidates.append(path)
                    elif dir_channel and root_stock:
                        root_candidates.append(path)
                if root_channel and root_stock:
                    root_candidates.append(current_path)
                elif root_channel and file_stock:
                    root_candidates.append(current_path)
                if root_candidates and root != search_root:
                    dirs.clear()
            if root_candidates:
                candidates.extend(root_candidates)
                if root != search_root:
                    break
        result = min(candidates, key=lambda path: self.channel_path_rank(path, stock_key, channel_key)) if candidates else None
        self.channel_folder_cache[key] = result
        return result

    def prepare_product_items(self, matched: list[ProductRecord]) -> list[tuple[ProductRecord, list[Path], Path | None]]:
        stock_codes = [record.stock_code for record in matched]
        stock_keys = {normalize_text(code) for code in stock_codes if normalize_text(code)}
        images = self.lookup_index_paths("preview", stock_keys)
        folders = self.lookup_index_paths("folder", stock_keys)
        for key in stock_keys:
            if key not in images and key in self.preview_image_cache:
                images[key] = self.preview_image_cache[key]
            if key not in folders and key in self.search_folder_cache:
                folders[key] = self.search_folder_cache[key]
        items: list[tuple[ProductRecord, list[Path], Path | None]] = []
        for record in matched:
            key = normalize_text(record.stock_code)
            image_path = images.get(key)
            folder_path = folders.get(key)
            card_images = self.collect_product_images(record.stock_code, image_path, folder_path)
            items.append((record, card_images or ([image_path] if image_path else []), folder_path))
        return items

    def build_product_copy_text(self, record: ProductRecord, image_path: Path | None, folder_path: Path | None) -> str:
        lines = [
            f"Urun Ailesi: {record.family or '-'}",
            f"Kirilim: {record.breakdown or '-'}",
            f"Stok Kodu: {record.stock_code or '-'}",
            f"Fotograf: {image_path if image_path else 'Bulunamadi'}",
            f"Klasor: {folder_path if folder_path else 'Bulunamadi'}",
        ]
        if record.features:
            lines.append("")
            lines.append("Ozellikler:")
            lines.extend(f"{name}: {value}" for name, value in record.features.items())
        return "\n".join(lines)

    def product_display_name(self, record: ProductRecord) -> str:
        name = first_matching_value(
            record.raw_values,
            [
                ("urun", "adi"),
                ("urun", "aciklama"),
                ("malzeme", "adi"),
                ("malzeme", "aciklama"),
                ("name",),
                ("description",),
                ("aciklama",),
                ("adi",),
            ],
            excludes=("stok", "kod", "mancode", "barcode", "barkod", "ean"),
        )
        return str(name or record.breakdown or record.family or "").strip()

    def build_stock_product_copy_text(self, record: ProductRecord) -> str:
        stock = str(record.stock_code or "").strip()
        name = self.product_display_name(record)
        return " - ".join(part for part in [stock, name] if part)

    def describe_path(self, path: Path | None, max_parts: int = 4) -> str:
        if path is None:
            return "Bulunamadi"
        parts = path.parts[-max_parts:]
        return ".../" + "/".join(parts) if len(path.parts) > max_parts else str(path)

    def is_under_product_images_folder(self, path: Path, root_path: Path) -> bool:
        if self.is_product_images_folder_name(path.name) or self.is_product_images_folder_name(root_path.name):
            return True
        try:
            parts = path.relative_to(root_path).parts
        except ValueError:
            parts = path.parts
        return any(self.is_product_images_folder_name(part) for part in parts)

    def is_non_product_media_folder(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return compact in {"2d", "3d"} or compact.startswith(("2d", "3d"))

    def infer_rename_group_name(self, root_path: Path, stock_folder: Path) -> str:
        try:
            parts = stock_folder.relative_to(root_path).parts
        except ValueError:
            parts = stock_folder.parts
        for index, part in enumerate(parts):
            if self.is_product_images_folder_name(part) and index > 0:
                return parts[index - 1]
        return parts[-2] if len(parts) >= 2 else parts[0] if parts else "Diger"

    def build_rename_plan(self, root_path: Path, output_root: Path | None = None, flat_output: bool = False) -> list[RenameAction]:
        pattern = re.compile(self.settings.get("stock_regex", DEFAULT_SETTINGS["stock_regex"]), re.IGNORECASE)
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        open_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("open_keywords", ""))]
        closed_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("closed_keywords", ""))]
        technical_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("technical_keywords", ""))]
        plan: list[RenameAction] = []
        planned_targets: set[str] = set()
        planned_sources: set[str] = set()
        for current, dirs, _files in safe_walk(root_path):
            current_path = Path(current)
            if not self.is_under_product_images_folder(current_path, root_path):
                dirs[:] = [item for item in dirs if not self.is_non_product_media_folder(item)]
                continue
            match = pattern.match(current_path.name.strip())
            if not match:
                continue
            stock_code = match.group(1).upper()
            group_name = self.infer_rename_group_name(root_path, current_path)
            plan.extend(
                self.build_stock_folder_actions(
                    current_path,
                    root_path,
                    stock_code,
                    group_name,
                    output_root,
                    valid,
                    open_keywords,
                    closed_keywords,
                    technical_keywords,
                    planned_targets,
                    flat_output,
                )
            )
            planned_sources.update(str(action.source).casefold() for action in plan)
            dirs.clear()
        plan.extend(
            self.build_loose_technical_actions(
                root_path,
                output_root,
                pattern,
                technical_keywords,
                planned_targets,
                planned_sources,
                flat_output,
            )
        )
        return sorted(plan, key=lambda item: (normalize_text(item.group_name), item.stock_code, str(item.source).casefold()))

    def infer_stock_code_from_path(self, path: Path, pattern: re.Pattern[str]) -> str:
        for text in [path.stem, path.name, *reversed(path.parts)]:
            value = str(text or "").strip().upper()
            broad_matches = re.findall(r"(?=[A-Z0-9-]*[A-Z])(?=[A-Z0-9-]*\d)([A-Z0-9]+(?:-[A-Z0-9]+)*)", value)
            if broad_matches:
                return max(broad_matches, key=len).upper()
            match = pattern.search(value)
            if match:
                return match.group(1).upper()
        return ""

    def build_loose_technical_actions(
        self,
        root_path: Path,
        output_root: Path | None,
        pattern: re.Pattern[str],
        technical_keywords: list[str],
        planned_targets: set[str],
        planned_sources: set[str],
        flat_output: bool = False,
    ) -> list[RenameAction]:
        actions: list[RenameAction] = []
        for current, dirs, files in safe_walk(root_path):
            current_path = Path(current)
            dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
            path_norm = normalize_text(str(current_path))
            if not any(keyword in path_norm for keyword in technical_keywords if keyword) and not self.path_matches_channel(current_path, "technical"):
                continue
            group_name = self.infer_rename_group_name(root_path, current_path)
            for file_name in files:
                file_path = current_path / file_name
                source_key = str(file_path).casefold()
                if source_key in planned_sources:
                    continue
                extension = file_path.suffix.lower()
                stock_code = self.infer_stock_code_from_path(file_path, pattern)
                if not stock_code:
                    continue
                if extension == ".ai":
                    actions.append(RenameAction(file_path, None, "Teknik cizimde AI atlandi", stock_code, "Atlandi", group_name))
                    planned_sources.add(source_key)
                    continue
                if extension not in {".jpg", ".jpeg"}:
                    continue
                operation = "export" if output_root is not None else "rename"
                base_name = f"{stock_code}_t"
                target_parent = self.build_rename_export_parent(output_root, root_path, file_path, flat_output)
                target = self.make_unique_target(file_path, base_name, planned_targets, target_parent)
                rule = "Teknik cizim"
                if target.stem != base_name:
                    rule = f"{rule} - isim cakismasi"
                status = "Cikarilacak" if operation == "export" else "Ayni Ad" if target == file_path else "Hazir"
                actions.append(RenameAction(file_path, target, f"Cikartma - {rule}" if operation == "export" else rule, stock_code, status, group_name, operation))
                planned_targets.add(str(target).casefold())
                planned_sources.add(source_key)
        return actions

    def build_rename_export_parent(self, output_root: Path | None, root_path: Path, source_file: Path, flat_output: bool = False) -> Path | None:
        if output_root is None:
            return None
        if flat_output:
            return output_root
        try:
            relative_parent = source_file.parent.relative_to(root_path)
        except ValueError:
            relative_parent = Path(safe_path_part(source_file.parent.name, "Diger"))
        return output_root / relative_parent

    def build_stock_folder_actions(
        self,
        stock_folder: Path,
        root_path: Path,
        stock_code: str,
        group_name: str,
        output_root: Path | None,
        valid_extensions: set[str],
        open_keywords: list[str],
        closed_keywords: list[str],
        technical_keywords: list[str],
        planned_targets: set[str],
        flat_output: bool = False,
        recursive: bool = True,
    ) -> list[RenameAction]:
        actions: list[RenameAction] = []
        sequence = 1
        if recursive:
            source_files = safe_rglob_files(stock_folder)
        else:
            try:
                source_files = [path for path in stock_folder.iterdir() if path_is_file(path)]
            except OSError:
                source_files = []
        for file_path in sorted(source_files, key=lambda path: str(path).casefold()):
            extension = file_path.suffix.lower()
            path_norm = normalize_text(str(file_path))
            stem_norm = normalize_text(file_path.stem)
            is_technical = any(keyword in path_norm for keyword in technical_keywords if keyword)
            if is_technical:
                if extension == ".ai":
                    actions.append(RenameAction(file_path, None, "Teknik cizimde AI atlandi", stock_code, "Atlandi", group_name))
                    continue
                if extension not in {".jpg", ".jpeg"}:
                    continue
                base_name = f"{stock_code}_t"
                rule = "Teknik cizim"
            else:
                if extension not in valid_extensions:
                    continue
                if any(keyword in stem_norm for keyword in open_keywords if keyword):
                    base_name = f"{stock_code}_acik"
                    rule = "Acik gorsel"
                elif any(keyword in stem_norm for keyword in closed_keywords if keyword):
                    base_name = f"{stock_code}_kapali"
                    rule = "Kapali gorsel"
                else:
                    base_name = f"{stock_code}_{sequence}"
                    rule = "Sirali gorsel"
                    sequence += 1
            operation = "export" if output_root is not None else "rename"
            target_parent = self.build_rename_export_parent(output_root, root_path, file_path, flat_output)
            target = self.make_unique_target(file_path, base_name, planned_targets, target_parent)
            if target.stem != base_name:
                rule = f"{rule} - isim cakismasi"
            status = "Cikarilacak" if operation == "export" else "Ayni Ad" if target == file_path else "Hazir"
            actions.append(RenameAction(file_path, target, f"Cikartma - {rule}" if operation == "export" else rule, stock_code, status, group_name, operation))
            planned_targets.add(str(target).casefold())
        return actions

    def make_unique_target(self, source: Path, base_name: str, planned_targets: set[str], target_parent: Path | None = None) -> Path:
        parent = target_parent or source.parent
        allow_same_source = target_parent is None
        candidate = parent / f"{base_name}{source.suffix}"
        if self.target_is_available(candidate, source, planned_targets, allow_same_source):
            return candidate
        counter = 2
        while True:
            candidate = parent / f"{base_name}_{counter}{source.suffix}"
            if self.target_is_available(candidate, source, planned_targets, allow_same_source):
                return candidate
            counter += 1

    def target_is_available(self, candidate: Path, source: Path, planned_targets: set[str], allow_same_source: bool = True) -> bool:
        key = str(candidate).casefold()
        if key in planned_targets:
            return False
        if allow_same_source and candidate == source:
            return True
        return not path_exists(candidate)

    def get_rename_action_key(self, action: RenameAction) -> str:
        return str(action.source).casefold()

    def is_rename_action_selectable(self, action: RenameAction) -> bool:
        return action.status in {"Hazir", "Cikarilacak"} and action.target is not None

    def set_action_target_name(self, action: RenameAction, new_name: str, save_state: bool = True) -> bool:
        if action.target is None:
            return False
        clean = Path(new_name.strip()).name
        if not clean or clean in {".", ".."}:
            return False
        stem = Path(clean).stem if Path(clean).suffix else clean
        if not stem:
            return False
        target = action.target.parent / f"{stem}{action.source.suffix}"
        action.target = target
        action.reason = "Manuel"
        if action.operation == "export":
            action.status = "Cikarilacak" if not path_exists(target) else "Cakisma"
        elif target == action.source:
            action.status = "Ayni Ad"
        elif path_exists(target):
            action.status = "Cakisma"
        else:
            action.status = "Hazir"
        if save_state:
            self.rename_manual_targets[self.get_rename_action_key(action)] = target.name
            self.save_runtime_state()
        return True

    def apply_saved_manual_target(self, action: RenameAction) -> None:
        saved = self.rename_manual_targets.get(self.get_rename_action_key(action))
        if saved and action.target is not None:
            self.set_action_target_name(action, saved, save_state=False)

    def find_duplicate_targets(self, actions: list[RenameAction]) -> list[str]:
        seen: set[str] = set()
        duplicates: list[str] = []
        for action in actions:
            if action.target is None:
                continue
            key = str(action.target).casefold()
            if key in seen:
                duplicates.append(str(action.target))
            seen.add(key)
        return duplicates

    def execute_rename_actions(self, actions: list[RenameAction]) -> int:
        batch_id = datetime.now().strftime("%Y%m%d%H%M%S") + f"-{uuid.uuid4().hex[:8]}"
        applied = 0
        export_actions = [action for action in actions if action.operation == "export"]
        rename_actions = [action for action in actions if action.operation != "export"]
        for action in export_actions:
            if action.target is None:
                continue
            if path_exists(action.target):
                raise FileExistsError(f"Hedef dosya zaten var: {action.target}")
            action.target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(action.source, action.target)
            applied += 1
            self.append_rename_log(batch_id, action, "copy")
        temp_moves: list[tuple[Path, Path, Path]] = []
        for action in rename_actions:
            if action.target is None or action.source == action.target:
                continue
            if path_exists(action.target):
                raise FileExistsError(f"Hedef dosya zaten var: {action.target}")
            temp = action.source.with_name(f"__codex_tmp__{uuid.uuid4().hex}{action.source.suffix}")
            action.source.rename(temp)
            temp_moves.append((action.source, temp, action.target))
        for original, temp, target in temp_moves:
            target.parent.mkdir(parents=True, exist_ok=True)
            temp.rename(target)
            applied += 1
            self.append_rename_log(batch_id, RenameAction(original, target, "", "", ""), "rename")
        return applied

    def append_rename_log(self, batch_id: str, action: RenameAction, operation: str) -> None:
        entry = {
            "batch_id": batch_id,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "source": str(action.source),
            "target": str(action.target or ""),
            "operation": operation,
        }
        with RENAME_LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry, ensure_ascii=False) + "\n")

    def parse_ydk_records(self, records: list[ProductRecord]) -> dict[str, YdkProduct]:
        products: dict[str, YdkProduct] = {}
        for record in records:
            product = self.build_ydk_product_from_record(record)
            if product.code:
                products[normalize_text(product.code)] = product
        return products

    def build_ydk_product_from_record(self, record: ProductRecord) -> YdkProduct:
        values: dict[str, str] = {}
        values.update(record.raw_values)
        values.update(record.features)
        values.setdefault("Urun Ailesi", record.family)
        values.setdefault("Kirilim", record.breakdown)
        values.setdefault("Stok Kodu", record.stock_code)
        description_tr = first_matching_value(
            values,
            [("adi",), ("tr", "aciklama"), ("turkce",), ("aciklama",), ("urun", "adi"), ("kirilim",)],
            excludes=("adi2", "ingilizce", "english", "description", "barkod", "ean", "koli"),
        )
        description_en = first_matching_value(
            values,
            [("adi2",), ("ad2",), ("name2",), ("en", "aciklama"), ("ingilizce",), ("english",), ("description",)],
            excludes=("tr", "turkce", "barkod", "ean", "koli"),
        )
        model = first_matching_value(values, [("ozelkod3",), ("model",), ("koleksiyon",), ("aile",)], excludes=("kirilim",)) or record.family
        producer_code = first_matching_value(
            values,
            [("uretici_kodu",), ("uretici", "kod"), ("ukod",), ("u_kod",), ("hka",), ("producer", "code"), ("supplier", "code"), ("tedarikci", "kod")],
            excludes=("stok", "urun", "barkod", "ean", "koli"),
        ) or first_value_matching_regex(values, r"\bHKA[-\s]?\d+\b")
        unit_barcode = first_ean_value(values, [("urun_barkod",), ("urun", "barkod"), ("unit", "barcode"), ("ean",), ("gtin",), ("barkod",), ("barcode",)], excludes=("koli", "carton", "case"))
        carton_barcode = ean13_full_code(first_matching_value(values, [("koli_barkod",), ("koli", "barkod"), ("carton", "barcode"), ("case", "barcode"), ("box", "barcode")]))
        carton_quantity = first_matching_value(values, [("koli", "adet"), ("koli", "ici"), ("carton", "quantity"), ("case", "qty"), ("box", "qty")], excludes=("barkod", "barcode", "ean"))
        return YdkProduct(
            code=record.stock_code,
            description_tr=description_tr or record.breakdown,
            description_en=description_en,
            model=model or record.family,
            producer_code=producer_code,
            brand=first_matching_value(values, [("marka",), ("brand",)]),
            unit=first_matching_value(values, [("birim",), ("unit",)]),
            product_type=record.breakdown,
            unit_barcode=unit_barcode,
            carton_barcode=carton_barcode,
            carton_quantity=carton_quantity,
        )

    def find_ydk_product_image(self, stock_code: str) -> Path | None:
        stock_key = normalize_text(stock_code)
        if not stock_key:
            return None
        cached = self.get_cached_preview_image(stock_code)
        if cached is not None:
            return cached
        roots = [Path(self.settings.get("ydk_image_root", "")).expanduser(), Path(self.settings.get("preview_image_root", "")).expanduser()]
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        for root in unique_paths([path for path in roots if path_is_dir(path)]):
            for extension in valid:
                direct = root / f"{stock_code}{extension}"
                if path_is_file(direct):
                    self.preview_image_cache[stock_key] = direct
                    return direct
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if path.suffix.lower() not in valid:
                        continue
                    text = normalize_text(str(path))
                    if normalize_text(path.stem).startswith(stock_key) or stock_key in text:
                        self.preview_image_cache[stock_key] = path
                        return path
        self.preview_image_cache[stock_key] = None
        return None

    def build_ydk_summary_text(self, product: YdkProduct, image_path: Path | None = None) -> str:
        return "\n".join([
            f"Stok Kodu: {product.code}",
            f"Model: {product.model or '-'}",
            f"Uretici Kodu: {product.producer_code or '-'}",
            f"TR Aciklama: {product.description_tr or '-'}",
            f"EN Description: {product.description_en or '-'}",
            f"Urun Barkodu: {product.unit_barcode or '-'}",
            f"Koli Barkodu: {product.carton_barcode or '-'}",
            f"Koli Ici Adet: {product.carton_quantity or '-'}",
            f"Gorsel: {image_path if image_path else 'Bulunamadi'}",
            f"EAN Font Metni: {ean13_font_text(product.unit_barcode) or '-'}",
        ])

    def get_ydk_label_layout(self) -> dict[str, int]:
        layout = DEFAULT_YDK_LABEL_LAYOUT.copy()
        try:
            saved = json.loads(self.settings.get("ydk_label_layout", "{}"))
            if isinstance(saved, dict):
                for key, value in saved.items():
                    try:
                        layout[key] = int(float(value))
                    except (TypeError, ValueError):
                        pass
        except json.JSONDecodeError:
            pass
        return layout

    def load_ydk_logo(self) -> Image.Image | None:
        if self.ydk_logo_image is not None:
            return self.ydk_logo_image
        logo_path = YDK_LABEL_LOGO_PNG if path_is_file(YDK_LABEL_LOGO_PNG) else YDK_LOGO_PNG
        if not path_is_file(logo_path):
            return None
        try:
            self.ydk_logo_image = Image.open(logo_path).convert("RGBA")
        except Exception:
            self.ydk_logo_image = None
        return self.ydk_logo_image

    def load_ydk_icon(self) -> Image.Image | None:
        if self.ydk_icon_image is not None:
            return self.ydk_icon_image
        if not path_is_file(YDK_LABEL_ICON_PNG):
            return None
        try:
            self.ydk_icon_image = Image.open(YDK_LABEL_ICON_PNG).convert("RGBA")
        except Exception:
            self.ydk_icon_image = None
        return self.ydk_icon_image

    def get_ydk_font(self, size: int, bold: bool = False, italic: bool = False) -> ImageFont.ImageFont:
        if bold and italic:
            file_names = ("arialbi.ttf", "segoeuiz.ttf")
        elif bold:
            file_names = ("arialbd.ttf", "segoeuib.ttf")
        elif italic:
            file_names = ("ariali.ttf", "segoeuii.ttf")
        else:
            file_names = ("arial.ttf", "segoeui.ttf")
        for file_name in file_names:
            candidate = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / file_name
            if path_is_file(candidate):
                try:
                    return ImageFont.truetype(str(candidate), size)
                except OSError:
                    continue
        return ImageFont.load_default()

    def wrap_ydk_text(self, draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int, max_lines: int = 2) -> list[str]:
        words = str(text or "").split()
        lines: list[str] = []
        current = ""
        for word in words:
            trial = f"{current} {word}".strip()
            if draw.textlength(trial, font=font) <= max_width or not current:
                current = trial
            else:
                lines.append(current)
                current = word
            if len(lines) >= max_lines:
                break
        if current and len(lines) < max_lines:
            lines.append(current)
        if len(lines) == max_lines and " ".join(lines).casefold() != " ".join(words).casefold():
            while draw.textlength(lines[-1] + "...", font=font) > max_width and len(lines[-1]) > 3:
                lines[-1] = lines[-1][:-1].rstrip()
            lines[-1] += "..."
        return lines or [""]

    def create_ean13_barcode_image(self, value: str, width: int, height: int) -> Image.Image | None:
        code = ean13_full_code(value)
        if len(code) != 13:
            return None
        l_codes = {"0": "0001101", "1": "0011001", "2": "0010011", "3": "0111101", "4": "0100011", "5": "0110001", "6": "0101111", "7": "0111011", "8": "0110111", "9": "0001011"}
        g_codes = {"0": "0100111", "1": "0110011", "2": "0011011", "3": "0100001", "4": "0011101", "5": "0111001", "6": "0000101", "7": "0010001", "8": "0001001", "9": "0010111"}
        r_codes = {"0": "1110010", "1": "1100110", "2": "1101100", "3": "1000010", "4": "1011100", "5": "1001110", "6": "1010000", "7": "1000100", "8": "1001000", "9": "1110100"}
        parity = {"0": "LLLLLL", "1": "LLGLGG", "2": "LLGGLG", "3": "LLGGGL", "4": "LGLLGG", "5": "LGGLLG", "6": "LGGGLL", "7": "LGLGLG", "8": "LGLGGL", "9": "LGGLGL"}
        sequence = "101"
        for digit, side in zip(code[1:7], parity[code[0]], strict=False):
            sequence += l_codes[digit] if side == "L" else g_codes[digit]
        sequence += "01010"
        for digit in code[7:]:
            sequence += r_codes[digit]
        sequence += "101"
        quiet_left = 40
        quiet_right = 18
        module_width = (width - quiet_left - quiet_right) / len(sequence)
        barcode = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(barcode)
        bar_height = height - 44
        guard_indexes = set(range(3)) | set(range(45, 50)) | set(range(92, 95))
        x = float(quiet_left)
        for index, bit in enumerate(sequence):
            if bit == "1":
                extra = 18 if index in guard_indexes else 0
                draw.rectangle((round(x), 0, round(x + module_width), bar_height + extra), fill="black")
            x += module_width
        font = self.get_ydk_font(35)
        text_y = height - 43
        draw.text((2, text_y), code[0], fill="black", font=font)
        left_digits = code[1:7]
        right_digits = code[7:]
        left_start = quiet_left + 3 * module_width
        left_width = 42 * module_width
        right_start = quiet_left + 50 * module_width
        right_width = 42 * module_width
        draw.text((left_start + (left_width - draw.textlength(left_digits, font=font)) / 2, text_y), left_digits, fill="black", font=font)
        draw.text((right_start + (right_width - draw.textlength(right_digits, font=font)) / 2, text_y), right_digits, fill="black", font=font)
        return barcode

    def render_ydk_label_image(self, product: YdkProduct, label_type: str, image_path: Path | None = None) -> Image.Image:
        width, height = YDK_LABEL_SIZE
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)
        layout = self.get_ydk_label_layout()

        def fitted_font(
            text: str,
            size: int,
            max_width: int,
            max_height: int | None = None,
            bold: bool = False,
            italic: bool = False,
        ) -> ImageFont.ImageFont:
            font = self.get_ydk_font(size, bold=bold, italic=italic)
            while size > 12:
                bbox = draw.textbbox((0, 0), text or "Ay", font=font)
                too_wide = draw.textlength(text, font=font) > max_width
                too_tall = max_height is not None and (bbox[3] - bbox[1]) > max_height
                if not too_wide and not too_tall:
                    break
                size -= 2
                font = self.get_ydk_font(size, bold=bold, italic=italic)
            return font

        def fitted_wrapped_lines(
            text: str,
            size: int,
            max_width: int,
            max_height: int,
            max_lines: int,
            bold: bool = False,
            italic: bool = False,
        ) -> tuple[ImageFont.ImageFont, list[str], int]:
            clean_text = text or "-"
            size = max(12, int(size))
            while size > 12:
                font = self.get_ydk_font(size, bold=bold, italic=italic)
                lines = self.wrap_ydk_text(draw, clean_text, font, max_width, max_lines)
                line_step = max(22, size + 8)
                heights = [draw.textbbox((0, 0), line or "Ay", font=font)[3] - draw.textbbox((0, 0), line or "Ay", font=font)[1] for line in lines]
                total_height = (len(lines) - 1) * line_step + (max(heights) if heights else 0)
                if total_height <= max_height:
                    return font, lines, line_step
                size -= 2
            font = self.get_ydk_font(size, bold=bold, italic=italic)
            return font, self.wrap_ydk_text(draw, clean_text, font, max_width, max_lines), max(22, size + 8)

        logo = self.load_ydk_logo()
        if logo is not None:
            logo_copy = logo.copy()
            logo_copy.thumbnail((layout["logo_w"], layout["logo_h"]))
            image.paste(logo_copy.convert("RGB"), (layout["logo_x"], layout["logo_y"]))
        else:
            draw.text((layout["logo_x"] + 12, layout["logo_y"] + 24), "HEKA", fill="black", font=self.get_ydk_font(70, True))

        product_box = (layout["photo_x"], layout["photo_y"], layout["photo_x"] + layout["photo_w"], layout["photo_y"] + layout["photo_h"])
        if image_path is not None and path_is_file(image_path):
            try:
                product_image = Image.open(image_path)
                product_image = ImageOps.exif_transpose(product_image)
                product_image = ImageOps.fit(product_image.convert("RGB"), (product_box[2] - product_box[0], product_box[3] - product_box[1]), method=Image.Resampling.LANCZOS)
                image.paste(product_image, (product_box[0], product_box[1]))
            except Exception:
                draw.rectangle(product_box, outline="#d8dee6", width=2)
        else:
            draw.rectangle(product_box, outline="#d8dee6", width=2)
            placeholder = "URUN GORSELI"
            placeholder_font = self.get_ydk_font(24, True)
            draw.text((product_box[0] + ((product_box[2] - product_box[0]) - draw.textlength(placeholder, font=placeholder_font)) / 2, 150), placeholder, fill="#8892a0", font=placeholder_font)

        model_text = (product.model or product.code or "-").upper()
        draw.text(
            (layout["model_x"], layout["model_y"]),
            model_text,
            fill="black",
            font=fitted_font(model_text, layout["model_font"], 300, max_height=54, bold=True),
        )
        tr_limit = max(layout["tr_y"] + 44, layout["en_y"] - 18)
        tr_font, tr_lines, tr_step = fitted_wrapped_lines(
            (product.description_tr or product.product_type or "-").upper(),
            layout["tr_font"],
            660,
            max(36, tr_limit - layout["tr_y"]),
            2,
            bold=True,
        )
        y = layout["tr_y"]
        for line in tr_lines:
            draw.text((layout["tr_x"], y), line, fill="black", font=tr_font)
            y += tr_step
        en_limit = max(layout["en_y"] + 40, layout["barcode_y"] - 24)
        en_font, en_lines, en_step = fitted_wrapped_lines(
            (product.description_en or "").upper(),
            layout["en_font"],
            660,
            max(34, en_limit - layout["en_y"]),
            2,
            italic=True,
        )
        y = layout["en_y"]
        for line in en_lines:
            draw.text((layout["en_x"], y), line, fill="black", font=en_font)
            y += en_step
        barcode = product.unit_barcode if label_type == "unit" else product.carton_barcode
        barcode_image = self.create_ean13_barcode_image(barcode, layout["barcode_w"], layout["barcode_h"])
        if barcode_image is not None:
            image.paste(barcode_image, (layout["barcode_x"], layout["barcode_y"]))
        else:
            draw.rectangle((layout["barcode_x"], layout["barcode_y"], layout["barcode_x"] + layout["barcode_w"], layout["barcode_y"] + layout["barcode_h"]), outline="#b8c2cf", width=2)
            draw.text((layout["barcode_x"] + 95, layout["barcode_y"] + 55), "BARKOD YOK", fill="#9b1c1c", font=self.get_ydk_font(30, True))
        heading = "URUN KODU" if label_type == "unit" else "KOLI KODU"
        heading_font = self.get_ydk_font(30, True)
        draw.text((layout["info_x"], layout["heading_y"]), heading, fill="black", font=heading_font)
        heading_width = draw.textlength(heading, font=heading_font)
        draw.line((layout["info_x"], layout["heading_y"] + 36, layout["info_x"] + heading_width, layout["heading_y"] + 36), fill="black", width=3)
        code_max_height = max(24, layout["producer_box_y"] - layout["code_y"] - 12)
        draw.text((layout["info_x"], layout["code_y"]), product.code, fill="black", font=fitted_font(product.code, 29, 295, max_height=code_max_height, bold=True))
        producer = product.producer_code or "-"
        box = (layout["producer_box_x"], layout["producer_box_y"], layout["producer_box_x"] + layout["producer_box_w"], layout["producer_box_y"] + layout["producer_box_h"])
        draw.rectangle(box, outline="black", width=4)
        producer_font = fitted_font(producer, layout["producer_font"], box[2] - box[0] - 24, max_height=box[3] - box[1] - 16, bold=True)
        producer_bbox = draw.textbbox((0, 0), producer, font=producer_font)
        producer_x = box[0] + ((box[2] - box[0]) - (producer_bbox[2] - producer_bbox[0])) / 2
        producer_y = box[1] + ((box[3] - box[1]) - (producer_bbox[3] - producer_bbox[1])) / 2 - 3
        draw.text((producer_x, producer_y), producer, fill="black", font=producer_font)
        icon = self.load_ydk_icon()
        bottom_font = self.get_ydk_font(layout["footer_font"])
        if icon is not None:
            left_icon = icon.copy()
            left_icon.thumbnail((35, 45))
            image.paste(left_icon, (36, layout["footer_y"] - 7), left_icon)
            right_icon = icon.copy()
            right_icon.thumbnail((35, 45))
            image.paste(right_icon, (437, layout["footer_y"] - 7), right_icon)
        draw.text((84, layout["footer_y"]), "www.hekalighting.com", fill="black", font=bottom_font)
        draw.text((480, layout["footer_y"]), "+90 850 711 47 45", fill="black", font=bottom_font)
        return image

    def get_ydk_output_root(self) -> Path:
        text = self.settings.get("ydk_output_root", "").strip()
        return Path(text).expanduser() if text else DATA_DIR / "ydk_exports"

    def build_ydk_pdf_path(self, product: YdkProduct, label_type: str) -> Path:
        date_folder = "YURTDISI ETIKET_" + datetime.now().strftime("%d.%m.%Y")
        folder = self.get_ydk_output_root() / date_folder
        suffix = "_KOLI_BARKOD" if label_type == "carton" else ""
        name = safe_path_part(f"{product.producer_code or product.code} - {product.code}_{product.model}{suffix}.pdf", "ydk_etiket.pdf")
        return folder / name

    def export_ydk_label_pdf(self, product: YdkProduct, label_type: str) -> Path:
        barcode = product.unit_barcode if label_type == "unit" else product.carton_barcode
        if not ean13_full_code(barcode):
            raise ValueError("Bu urun icin secilen etiket tipinde barkod bulunamadi.")
        image_path = self.find_ydk_product_image(product.code)
        label_image = self.render_ydk_label_image(product, label_type, image_path)
        pdf_path = self.build_ydk_pdf_path(product, label_type)
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        label_image.save(pdf_path, "PDF", resolution=300.0)
        return pdf_path



class SettingsDialog(QDialog):
    def __init__(self, engine: DataEngine, parent: QWidget | None = None):
        super().__init__(parent)
        self.engine = engine
        self.setWindowTitle("Ayarlar")
        self.resize(820, 720)
        self.fields: dict[str, QWidget] = {}
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(12)
        title = QLabel("Veri ve klasor ayarlari")
        title.setObjectName("DialogTitle")
        layout.addWidget(title)
        form_host = QWidget()
        form = QGridLayout(form_host)
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(10)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(form_host)
        layout.addWidget(scroll, 1)
        specs = [
            ("sql_connection_string", "SQL Baglanti", "multi"),
            ("sql_table", "SQL Tablo", "text"),
            ("sql_query", "SQL Sorgu", "multi"),
            ("family_column", "Aile Sutunu", "text"),
            ("breakdown_column", "Kirilim Sutunu", "text"),
            ("stock_column", "Stok Sutunu", "text"),
            ("feature_columns", "Ozellik Sutunlari", "text"),
            ("feature_aliases", "Ozellik Baslik Esleme", "multi"),
            ("search_root", "Explorer Ana Klasor", "dir"),
            ("preview_image_root", "Vitrin Gorsel Klasoru", "dir"),
            ("ydk_image_root", "YDK Gorsel Klasoru", "dir"),
            ("ydk_output_root", "YDK PDF Cikis", "dir"),
            ("qr_code_root", "QR PNG Klasoru", "dir"),
            ("qr_output_root", "QR Excel Cikis", "dir"),
            ("qr_url_base", "QR URL Taban", "text"),
            ("qr_sitemap_url", "QR Sitemap URL", "text"),
            ("bartender_exe_path", "BarTender EXE", "exe"),
            ("bartender_printer_name", "Barkod Yazici Adi", "text"),
            ("ce_bartender_template_path", "CE BarTender Template", "btw"),
            ("qr_bartender_template_path", "QR BarTender Template", "btw"),
            ("bartender_print_timeout_seconds", "BarTender Timeout Sn", "text"),
            ("label_reset_enabled", "Baski Oncesi Etiket Sifirla (1/0)", "text"),
            ("label_reset_before_each_label", "Her Etiket Oncesi Sifirla (1/0)", "text"),
            ("label_reset_feed_count", "Sifirlama Feed Adedi", "text"),
            ("label_reset_pause_ms", "Sifirlama Bekleme Ms", "text"),
            ("label_reset_raw_command", "Sifirlama Raw Komut", "text"),
        ]
        for row, (key, label_text, kind) in enumerate(specs):
            label = QLabel(label_text)
            form.addWidget(label, row, 0, alignment=Qt.AlignmentFlag.AlignTop)
            value = engine.settings.get(key, "")
            if kind == "combo":
                widget = QComboBox()
                widget.addItems(["sql", "excel"])
                widget.setCurrentText(normalize_data_source(value))
            elif kind == "multi":
                widget = QPlainTextEdit()
                widget.setPlainText(value)
                widget.setMinimumHeight(76)
            else:
                widget = QLineEdit(value)
            self.fields[key] = widget
            form.addWidget(widget, row, 1)
            if kind in {"file", "dir", "btw", "exe"}:
                button = QPushButton("Sec")
                button.clicked.connect(lambda _checked=False, field=widget, mode=kind: self.pick_path(field, mode))
                form.addWidget(button, row, 2)
        buttons = QHBoxLayout()
        buttons.addStretch(1)
        reset = QPushButton("LOGODATA Varsayilan")
        reset.clicked.connect(self.apply_logodata_defaults)
        cancel = QPushButton("Iptal")
        cancel.clicked.connect(self.reject)
        save = QPushButton("Kaydet")
        save.setObjectName("PrimaryButton")
        save.clicked.connect(self.accept)
        for button in [reset, cancel, save]:
            buttons.addWidget(button)
        layout.addLayout(buttons)


    def pick_path(self, field: QWidget, mode: str) -> None:
        if not isinstance(field, QLineEdit):
            return
        if mode == "btw":
            selected, _ = QFileDialog.getOpenFileName(self, "BarTender template sec", field.text(), "BarTender (*.btw);;Tum Dosyalar (*.*)")
        elif mode == "exe":
            selected, _ = QFileDialog.getOpenFileName(self, "BarTender EXE sec", field.text(), "BarTender (BarTend.exe);;EXE (*.exe);;Tum Dosyalar (*.*)")
        elif mode == "file":
            selected, _ = QFileDialog.getOpenFileName(self, "Dosya sec", field.text(), "Excel (*.xlsx *.xlsm *.xls);;Tum Dosyalar (*.*)")
        else:
            selected = QFileDialog.getExistingDirectory(self, "Klasor sec", field.text())
        if selected:
            field.setText(selected)

    def apply_logodata_defaults(self) -> None:
        for key, value in LOGODATA_SQL_DEFAULTS.items():
            widget = self.fields.get(key)
            if isinstance(widget, QLineEdit):
                widget.setText(value)
            elif isinstance(widget, QPlainTextEdit):
                widget.setPlainText(value)
            elif isinstance(widget, QComboBox):
                widget.setCurrentText(value)

    def values(self) -> dict[str, str]:
        data = self.engine.settings.copy()
        for key, widget in self.fields.items():
            if isinstance(widget, QLineEdit):
                data[key] = widget.text().strip()
            elif isinstance(widget, QPlainTextEdit):
                data[key] = widget.toPlainText().strip()
            elif isinstance(widget, QComboBox):
                data[key] = widget.currentText()
        data["data_source"] = "sql"
        data["excel_path"] = ""
        data["sheet_name"] = ""
        data["header_row"] = "1"
        data["sql_auto_profile_version"] = LOGODATA_SQL_PROFILE_VERSION
        return self.engine.apply_modern_defaults(data)


class StatCard(QFrame):
    def __init__(self, title: str, value: str, accent: str):
        super().__init__()
        self.setObjectName("StatCard")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 16, 18, 16)
        title_label = QLabel(title)
        title_label.setObjectName("StatTitle")
        value_label = QLabel(value)
        value_label.setObjectName("StatValue")
        value_label.setStyleSheet(f"color: {accent};")
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        self.value_label = value_label

    def set_value(self, value: str) -> None:
        self.value_label.setText(value)


class ProductCard(QFrame):
    def __init__(self, engine: DataEngine, record: ProductRecord, images: list[Path], folder_path: Path | None, parent: QWidget | None = None):
        super().__init__(parent)
        self.engine = engine
        self.record = record
        self.images = images
        self.folder_path = folder_path
        self.image_index = 0
        self.setObjectName("ProductCard")
        self.setMinimumWidth(350)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)
        self.image_label = QLabel("Gorsel yok")
        self.image_label.setObjectName("ImageStage")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setFixedHeight(230)
        layout.addWidget(self.image_label)
        nav = QHBoxLayout()
        prev = QPushButton("<")
        prev.setObjectName("IconButton")
        next_button = QPushButton(">")
        next_button.setObjectName("IconButton")
        self.count_label = QLabel("0/0")
        self.count_label.setObjectName("Pill")
        nav.addWidget(prev)
        nav.addWidget(self.count_label)
        nav.addStretch(1)
        nav.addWidget(next_button)
        layout.addLayout(nav)
        prev.clicked.connect(lambda: self.change_image(-1))
        next_button.clicked.connect(lambda: self.change_image(1))
        prev.setEnabled(len(images) > 1)
        next_button.setEnabled(len(images) > 1)
        family = QLabel(record.family or "-")
        family.setObjectName("CardTitle")
        breakdown = QLabel(record.breakdown or "-")
        breakdown.setObjectName("CardSubTitle")
        stock = QLabel(record.stock_code)
        stock.setObjectName("StockPill")
        layout.addWidget(family)
        layout.addWidget(breakdown)
        layout.addWidget(stock, alignment=Qt.AlignmentFlag.AlignLeft)
        copy_row = QHBoxLayout()
        for text, handler in [
            ("Stok", self.copy_stock),
            ("Stok+Urun", self.copy_stock_product),
            ("Kart", self.copy_card),
            ("Klasor", self.open_folder),
            ("Gorsel", self.open_image),
        ]:
            button = QPushButton(text)
            button.clicked.connect(handler)
            copy_row.addWidget(button)
        copy_row.addStretch(1)
        layout.addLayout(copy_row)
        channels = QHBoxLayout()
        for channel_key, (label, _keywords) in CHANNEL_FOLDERS.items():
            button = QPushButton(label)
            button.setObjectName("GhostButton")
            button.clicked.connect(lambda _checked=False, key=channel_key: self.open_channel(key))
            channels.addWidget(button)
        channels.addStretch(1)
        layout.addLayout(channels)
        path_label = QLabel(self.engine.describe_path(folder_path))
        path_label.setObjectName("Muted")
        path_label.setWordWrap(True)
        layout.addWidget(path_label)
        features_title = QLabel("Ozellikler")
        features_title.setObjectName("SectionTitle")
        layout.addWidget(features_title)
        for name, value in list(record.features.items())[:6]:
            feature = QLabel(f"<b>{name}</b><br>{value}")
            feature.setObjectName("FeatureLine")
            feature.setWordWrap(True)
            layout.addWidget(feature)
        layout.addStretch(1)
        self.refresh_image()

    def current_image(self) -> Path | None:
        return self.images[self.image_index] if self.images else None

    def refresh_image(self) -> None:
        pixmap = load_scaled_pixmap(self.current_image(), QSize(340, 230), "#eef5fb")
        if pixmap is None:
            self.image_label.setPixmap(QPixmap())
            self.image_label.setText("Gorsel yok")
        else:
            self.image_label.setText("")
            self.image_label.setPixmap(pixmap)
        self.count_label.setText(f"{self.image_index + 1}/{len(self.images)}" if self.images else "0/0")

    def change_image(self, step: int) -> None:
        if not self.images:
            return
        self.image_index = (self.image_index + step) % len(self.images)
        self.refresh_image()

    def copy_stock(self) -> None:
        QApplication.clipboard().setText(self.record.stock_code)

    def copy_stock_product(self) -> None:
        QApplication.clipboard().setText(self.engine.build_stock_product_copy_text(self.record))

    def copy_card(self) -> None:
        QApplication.clipboard().setText(self.engine.build_product_copy_text(self.record, self.current_image(), self.folder_path))

    def open_folder(self) -> None:
        reveal_path(self.folder_path, self)

    def open_image(self) -> None:
        path = self.current_image()
        if path and path_exists(path):
            open_path_with_default_app(path)

    def open_channel(self, channel_key: str) -> None:
        path = self.engine.find_channel_folder(self.record.stock_code, channel_key, self.folder_path)
        if path:
            reveal_path(path, self)
        else:
            QMessageBox.warning(self, "Kanal Yok", f"{self.record.stock_code} icin {CHANNEL_FOLDERS[channel_key][0]} klasoru bulunamadi.")


class QrProductCard(QFrame):
    def __init__(
        self,
        engine: DataEngine,
        record: ProductRecord,
        images: list[Path],
        qr_path: Path | None,
        selected: bool,
        on_open: Callable[[ProductRecord, list[Path]], None],
        on_toggle: Callable[[ProductRecord], int],
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.engine = engine
        self.record = record
        self.images = images
        self.qr_path = qr_path
        self.on_open = on_open
        self.on_toggle = on_toggle
        self.selected = selected
        self.selected_count = 1 if selected else 0
        self.setObjectName("ProductCard")
        self.setMinimumWidth(240)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # â”€â”€ card header: family + QR status badge â”€â”€
        header = QFrame()
        header.setObjectName("CardHeader")
        header.setFixedHeight(40)
        header_row = QHBoxLayout(header)
        header_row.setContentsMargins(12, 0, 10, 0)
        header_row.setSpacing(8)
        family_lbl = QLabel(record.family or "â€”")
        family_lbl.setObjectName("CardSubTitle")
        family_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        qr_badge = QLabel("â— QR" if qr_path else "â—‹ QR YOK")
        qr_badge.setObjectName("SuccessPill" if qr_path else "DangerPill")
        header_row.addWidget(family_lbl, 1)
        header_row.addWidget(qr_badge)
        outer.addWidget(header)

        # â”€â”€ product image â”€â”€
        self.image_label = QLabel("GÃ¶rsel yok")
        self.image_label.setObjectName("ImageStage")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setFixedHeight(185)
        outer.addWidget(self.image_label)

        # â”€â”€ info section â”€â”€
        info_widget = QWidget()
        info_layout = QVBoxLayout(info_widget)
        info_layout.setContentsMargins(13, 10, 13, 6)
        info_layout.setSpacing(5)
        breakdown_lbl = QLabel(record.breakdown or "â€”")
        breakdown_lbl.setObjectName("CardTitle")
        breakdown_lbl.setWordWrap(True)
        stock_pill = QLabel(f"# {record.stock_code}")
        stock_pill.setObjectName("StockPill")
        info_layout.addWidget(breakdown_lbl)
        info_layout.addWidget(stock_pill)
        outer.addWidget(info_widget)

        # â”€â”€ action row â”€â”€
        actions_widget = QWidget()
        actions_row = QHBoxLayout(actions_widget)
        actions_row.setContentsMargins(13, 4, 13, 13)
        actions_row.setSpacing(8)
        open_button = QPushButton("QR GÃ¶ster")
        open_button.setObjectName("GhostButton")
        open_button.clicked.connect(lambda: self.on_open(self.record, self.images))
        self.select_button = QPushButton("SeÃ§")
        self.select_button.setObjectName("PrimaryButton")
        self.select_button.clicked.connect(self.toggle_selected)
        actions_row.addWidget(open_button)
        actions_row.addWidget(self.select_button, 1)
        outer.addWidget(actions_widget)

        self.refresh_image()
        self.update_selected(selected, self.selected_count)

    def refresh_image(self) -> None:
        image = self.images[0] if self.images else None
        pixmap = load_scaled_pixmap(image, QSize(290, 185))
        if pixmap is None:
            self.image_label.setPixmap(QPixmap())
            self.image_label.setText("GÃ¶rsel yok")
        else:
            self.image_label.setText("")
            self.image_label.setPixmap(pixmap)

    def update_selected(self, selected: bool, count: int = 0) -> None:
        self.selected = selected
        self.selected_count = count if selected else 0
        if self.selected_count >= 2:
            self.select_button.setText(f"{self.selected_count} adet âœ“")
        elif self.selected_count == 1:
            self.select_button.setText("1 adet âœ“")
        else:
            self.select_button.setText("SeÃ§")
        self.setProperty("selected", "true" if selected else "false")
        self.setProperty("qrmissing", "true" if (not selected and self.qr_path is None) else "false")
        self.style().unpolish(self)
        self.style().polish(self)

    def toggle_selected(self) -> None:
        self.on_toggle(self.record)

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            self.on_open(self.record, self.images)
        super().mousePressEvent(event)


class CeProductCard(QFrame):
    def __init__(
        self,
        engine: DataEngine,
        record: ProductRecord,
        images: list[Path],
        selected: bool,
        on_open: Callable[[ProductRecord, list[Path]], None],
        on_toggle: Callable[[ProductRecord], None],
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.engine = engine
        self.record = record
        self.images = images
        self.on_open = on_open
        self.on_toggle = on_toggle
        self.selected = selected
        self.selected_count = 1 if selected else 0
        self.setObjectName("ProductCard")
        self.setMinimumWidth(240)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        header = QFrame()
        header.setObjectName("CardHeader")
        header.setFixedHeight(40)
        header_row = QHBoxLayout(header)
        header_row.setContentsMargins(12, 0, 10, 0)
        header_row.setSpacing(8)
        family_lbl = QLabel(record.family or "â€”")
        family_lbl.setObjectName("CardSubTitle")
        family_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        ce_entry = self.engine.get_ce_label_entry(record)
        kind_text = ce_entry.label_kind or "CE"
        badge = QLabel(f"â— {kind_text}")
        badge.setObjectName("SuccessPill")
        header_row.addWidget(family_lbl, 1)
        header_row.addWidget(badge)
        outer.addWidget(header)

        self.image_label = QLabel("Gorsel yok")
        self.image_label.setObjectName("ImageStage")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setFixedHeight(185)
        outer.addWidget(self.image_label)

        info_widget = QWidget()
        info_layout = QVBoxLayout(info_widget)
        info_layout.setContentsMargins(13, 10, 13, 6)
        info_layout.setSpacing(5)
        breakdown_lbl = QLabel(record.breakdown or "â€”")
        breakdown_lbl.setObjectName("CardTitle")
        breakdown_lbl.setWordWrap(True)
        stock_pill = QLabel(f"# {record.stock_code}")
        stock_pill.setObjectName("StockPill")
        info_layout.addWidget(breakdown_lbl)
        info_layout.addWidget(stock_pill)
        outer.addWidget(info_widget)

        actions_widget = QWidget()
        actions_row = QHBoxLayout(actions_widget)
        actions_row.setContentsMargins(13, 4, 13, 13)
        actions_row.setSpacing(8)
        open_button = QPushButton("CE Detay")
        open_button.setObjectName("GhostButton")
        open_button.clicked.connect(lambda: self.on_open(self.record, self.images))
        self.select_button = QPushButton("Sec")
        self.select_button.setObjectName("PrimaryButton")
        self.select_button.clicked.connect(self.toggle_selected)
        actions_row.addWidget(open_button)
        actions_row.addWidget(self.select_button, 1)
        outer.addWidget(actions_widget)

        self.refresh_image()
        self.update_selected(selected, self.selected_count)

    def refresh_image(self) -> None:
        image = self.images[0] if self.images else None
        pixmap = load_scaled_pixmap(image, QSize(290, 185))
        if pixmap is None:
            self.image_label.setPixmap(QPixmap())
            self.image_label.setText("Gorsel yok")
        else:
            self.image_label.setText("")
            self.image_label.setPixmap(pixmap)

    def update_selected(self, selected: bool, count: int = 0) -> None:
        self.selected = selected
        self.selected_count = count if selected else 0
        if self.selected_count >= 1:
            self.select_button.setText(f"x{self.selected_count} âœ“")
        else:
            self.select_button.setText("Sec")
        self.setProperty("selected", "true" if selected else "false")
        self.style().unpolish(self)
        self.style().polish(self)

    def toggle_selected(self) -> None:
        self.on_toggle(self.record)

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            self.on_open(self.record, self.images)
        super().mousePressEvent(event)


def reveal_path(path: Path | None, parent: QWidget | None = None) -> None:
    if path is None:
        QMessageBox.warning(parent, "Konum Yok", "Konum bulunamadi.")
        return
    try:
        open_in_explorer(path)
    except Exception as exc:  # noqa: BLE001
        QMessageBox.critical(parent, "Explorer Hatasi", str(exc))


class WindowTitleBar(QFrame):
    def __init__(self, window: "ModernMainWindow"):
        super().__init__(window)
        self.window = window
        self.drag_offset: QPoint | None = None
        self.setObjectName("CustomTitleBar")
        self.setFixedHeight(54)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 0, 10, 0)
        layout.setSpacing(10)

        icon_label = QLabel()
        icon_label.setObjectName("TitleIcon")
        icon_label.setFixedSize(24, 24)
        if path_is_file(APP_ICON_PNG):
            pixmap = QIcon(str(APP_ICON_PNG)).pixmap(24, 24)
            icon_label.setPixmap(pixmap)
        layout.addWidget(icon_label)

        title_box = QVBoxLayout()
        title_box.setContentsMargins(0, 0, 0, 0)
        title_box.setSpacing(1)
        title = QLabel(APP_TITLE)
        title.setObjectName("TitleText")
        subtitle = QLabel("frameless Qt dashboard")
        subtitle.setObjectName("TitleHint")
        title_box.addWidget(title)
        title_box.addWidget(subtitle)
        layout.addLayout(title_box)
        layout.addStretch(1)

        self.minimize_button = QPushButton("_")
        self.maximize_button = QPushButton("[]")
        self.close_button = QPushButton("X")
        for button in [self.minimize_button, self.maximize_button, self.close_button]:
            button.setFixedSize(38, 34)
            button.setObjectName("WindowButton")
            layout.addWidget(button)
        self.close_button.setObjectName("CloseButton")
        self.minimize_button.clicked.connect(window.showMinimized)
        self.maximize_button.clicked.connect(window.toggle_max_restore)
        self.close_button.clicked.connect(window.close)

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            self.drag_offset = event.globalPosition().toPoint() - self.window.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event) -> None:  # type: ignore[override]
        if self.drag_offset is not None and event.buttons() & Qt.MouseButton.LeftButton:
            if self.window.isMaximized():
                self.window.showNormal()
                self.window.sync_frame_state()
                self.drag_offset = QPoint(self.window.width() // 2, 24)
            self.window.move(event.globalPosition().toPoint() - self.drag_offset)
            event.accept()

    def mouseReleaseEvent(self, event) -> None:  # type: ignore[override]
        self.drag_offset = None
        event.accept()

    def mouseDoubleClickEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            self.window.toggle_max_restore()
            event.accept()

    def sync_max_button(self) -> None:
        self.maximize_button.setText("][" if self.window.isMaximized() else "[]")


THEMES: dict[str, dict[str, str]] = {
    "light": {
        "SHELL_BG": "#F6F8FB",
        "SHELL_BORDER": "#D8E0EA",
        "SHELL_RADIUS": "22px",
        "TITLEBAR_BG": "#FFFFFF",
        "TITLEBAR_RADIUS": "22px",
        "TITLETEXT_COLOR": "#0F172A",
        "TITLEHINT_COLOR": "#64748B",
        "TITLEICON_COLOR": "transparent",
        "WINBTN_BG": "#F1F5F9",
        "WINBTN_BG_HOVER": "#E2E8F0",
        "WINBTN_COLOR": "#334155",
        "WINBTN_BORDER": "#D8E0EA",
        "CLOSEBTN_BG": "#FFF1F2",
        "CLOSEBTN_BG_HOVER": "#E11D48",
        "CLOSEBTN_COLOR": "#BE123C",
        "CLOSEBTN_BORDER": "#FFE4E6",
        "SIDEBAR_BG": "#FFFFFF",
        "SIDEBAR_BORDER": "#D8E0EA",
        "SIDEBAR_RADIUS": "18px",
        "BRANDCARD_BG": "#F8FAFC",
        "BRANDCARD_BORDER": "#E2E8F0",
        "BRANDMARK_BG1": "#0F172A",
        "BRANDMARK_BG2": "#2563EB",
        "BRAND_COLOR": "#111827",
        "BRANDKICKER_COLOR": "#2563EB",
        "SIDEBAR_HINT_COLOR": "#64748B",
        "SIDEBAR_SECTION_COLOR": "#94A3B8",
        "SIDEBARFOOTER_BG": "#F8FAFC",
        "SIDEBARFOOTER_BORDER": "#E2E8F0",
        "NAVBTN_COLOR": "#475569",
        "NAVBTN_BG_HOVER": "#F1F5F9",
        "NAVBTN_BORDER_HOVER": "#D8E0EA",
        "NAVBTN_COLOR_HOVER": "#0F172A",
        "NAVBTN_BG_ACTIVE": "#EAF2FF",
        "NAVBTN_BORDER_ACTIVE": "#BBD7FF",
        "NAVBTN_COLOR_ACTIVE": "#1D4ED8",
        "TOPBAR_BG": "#FFFFFF",
        "TOPBAR_BORDER": "#E2E8F0",
        "TOPBAR_ACCENT_BORDER": "#2563EB",
        "CARD_BG": "#FFFFFF",
        "CARD_BORDER": "#E2E8F0",
        "CARD_BG_HOVER": "#FBFDFF",
        "CARD_BORDER_HOVER": "#BBD7FF",
        "INNERPANEL_BG": "#F8FAFC",
        "INNERPANEL_BORDER": "#E2E8F0",
        "HERO_BG1": "#FFFFFF",
        "HERO_BG2": "#F2F7FF",
        "HERO_BG3": "#EEFDF9",
        "HERO_TITLE_COLOR": "#0F172A",
        "HERO_TEXT_COLOR": "#64748B",
        "HERO_CHIP_BG": "#EAF2FF",
        "HERO_CHIP_BORDER": "#C7DDFF",
        "HERO_CHIP_COLOR": "#1D4ED8",
        "EYEBROW_COLOR": "#2563EB",
        "PAGE_TITLE_COLOR": "#0F172A",
        "PAGE_HINT_COLOR": "#64748B",
        "STAT_TITLE_COLOR": "#64748B",
        "STAT_VALUE_COLOR": "#0F172A",
        "WIDGET_TEXT": "#1F2937",
        "CARD_TITLE_COLOR": "#0F172A",
        "CARD_SUBTITLE_COLOR": "#64748B",
        "SECTION_TITLE_COLOR": "#334155",
        "MINITITLE_COLOR": "#0F172A",
        "PANEL_HINT_COLOR": "#64748B",
        "MUTED_COLOR": "#64748B",
        "PILL_STOCK_BG": "#EAF2FF",
        "PILL_STOCK_COLOR": "#1D4ED8",
        "PILL_STOCK_BORDER": "#C7DDFF",
        "PILL_BG": "#F1F5F9",
        "PILL_COLOR": "#475569",
        "PILL_BORDER": "#D8E0EA",
        "PILL_SUCCESS_BG": "#DCFCE7",
        "PILL_SUCCESS_COLOR": "#15803D",
        "PILL_SUCCESS_BORDER": "#BBF7D0",
        "PILL_DANGER_BG": "#FFE4E6",
        "PILL_DANGER_COLOR": "#BE123C",
        "PILL_DANGER_BORDER": "#FECDD3",
        "PILL_WARN_BG": "#FEF3C7",
        "PILL_WARN_COLOR": "#B45309",
        "PILL_WARN_BORDER": "#FDE68A",
        "PILL_COUNTER_BG": "#E2E8F0",
        "PILL_COUNTER_COLOR": "#0F172A",
        "SELROW_BG": "#FFFFFF",
        "SELROW_BORDER": "#E2E8F0",
        "SELROW_BG_HOVER": "#F8FAFC",
        "SELROW_BORDER_HOVER": "#CBD5E1",
        "SELROW_STOCK_COLOR": "#0F172A",
        "SELROW_META_COLOR": "#64748B",
        "SELROW_HANDLE_COLOR": "#2563EB",
        "SELLIST_SELECTED_BG": "#EAF2FF",
        "SELLIST_SELECTED_COLOR": "#0F172A",
        "IMAGESTAGE_BG": "#F8FAFC",
        "IMAGESTAGE_BORDER": "#D8E0EA",
        "IMAGESTAGE_COLOR": "#94A3B8",
        "FEATURELINE_BG": "#F8FAFC",
        "FEATURELINE_COLOR": "#334155",
        "BTN_BG": "#FFFFFF",
        "BTN_BORDER": "#D8E0EA",
        "BTN_COLOR": "#334155",
        "BTN_BG_HOVER": "#F1F5F9",
        "BTN_BORDER_HOVER": "#BBD7FF",
        "BTN_BG_PRESSED": "#E2E8F0",
        "PRIMARY_BTN_BG": "#2563EB",
        "PRIMARY_BTN_COLOR": "#FFFFFF",
        "PRIMARY_BTN_BORDER": "#2563EB",
        "PRIMARY_BTN_BG_HOVER": "#1D4ED8",
        "PRIMARY_BTN_BORDER_HOVER": "#1D4ED8",
        "GHOST_BTN_BG": "#F8FAFC",
        "GHOST_BTN_BORDER": "#D8E0EA",
        "GHOST_BTN_COLOR": "#475569",
        "GHOST_BTN_BG_HOVER": "#F1F5F9",
        "GHOST_BTN_BORDER_HOVER": "#CBD5E1",
        "DANGER_BTN_BG": "#FFF1F2",
        "DANGER_BTN_COLOR": "#BE123C",
        "DANGER_BTN_BORDER": "#FECDD3",
        "DANGER_BTN_BG_HOVER": "#FFE4E6",
        "ICON_BTN_BG": "#EAF2FF",
        "ICON_BTN_COLOR": "#1D4ED8",
        "INPUT_BG": "#FFFFFF",
        "INPUT_BORDER": "#CBD5E1",
        "INPUT_COLOR": "#0F172A",
        "INPUT_SELECTION_BG": "#BFDBFE",
        "INPUT_BG_FOCUS": "#FFFFFF",
        "INPUT_BORDER_FOCUS": "#2563EB",
        "COMBO_DROPDOWN_BG": "#F8FAFC",
        "COMBO_DROPDOWN_BG_HOVER": "#EAF2FF",
        "SCROLL_HANDLE": "#CBD5E1",
        "SCROLL_HANDLE_HOVER": "#94A3B8",
        "TABLE_BG": "#FFFFFF",
        "TABLE_BORDER": "#D8E0EA",
        "TABLE_GRID": "#E2E8F0",
        "TABLE_ALT": "#F8FAFC",
        "TABLE_HEADER_BG": "#F1F5F9",
        "TABLE_HEADER_COLOR": "#475569",
        "TABLE_ITEM_PADDING": "8px",
        "TABLE_SELECTED_BG": "#EAF2FF",
        "TABLE_SELECTED_COLOR": "#0F172A",
        "DIALOG_BG": "#FFFFFF",
        "DIALOG_TITLE_COLOR": "#0F172A",
        "SIZEKIT_COLOR": "#CBD5E1",
        "SELECTED_CARD_BG": "#F0F7FF",
        "CARD_HEADER_BG": "#F8FAFC",
        "CARD_HEADER_BORDER": "#E2E8F0",
    },
}
QSS_TEMPLATE = """
    QMainWindow { background: transparent; }
    QWidget { font-family: "Segoe UI Variable", Aptos, Segoe UI; color: {WIDGET_TEXT}; }
    #AppShell { background: {SHELL_BG}; border: 1px solid {SHELL_BORDER}; border-radius: {SHELL_RADIUS}; }
    #AppShell[maximized="true"] { border-radius: 0; border: 0; }
    #CustomTitleBar { background: {TITLEBAR_BG}; border-top-left-radius: {TITLEBAR_RADIUS}; border-top-right-radius: {TITLEBAR_RADIUS}; }
    #CustomTitleBar[maximized="true"] { border-top-left-radius: 0; border-top-right-radius: 0; }
    #TitleText { color: {TITLETEXT_COLOR}; font: 900 13px Bahnschrift; letter-spacing: 1px; }
    #TitleHint { color: {TITLEHINT_COLOR}; font-size: 11px; }
    #TitleIcon { background: {TITLEICON_COLOR}; }
    QPushButton#WindowButton { background: {WINBTN_BG}; color: {WINBTN_COLOR}; border: 1px solid {WINBTN_BORDER}; border-radius: 10px; padding: 0; font: 900 12px Bahnschrift; }
    QPushButton#WindowButton:hover { background: {WINBTN_BG_HOVER}; color: white; }
    QPushButton#CloseButton { background: {CLOSEBTN_BG}; color: {CLOSEBTN_COLOR}; border: 1px solid {CLOSEBTN_BORDER}; border-radius: 10px; padding: 0; font: 900 12px Bahnschrift; }
    QPushButton#CloseButton:hover { background: {CLOSEBTN_BG_HOVER}; color: white; }
    #Workspace { background: transparent; }
    #Sidebar { background: {SIDEBAR_BG}; border: 1px solid {SIDEBAR_BORDER}; border-radius: {SIDEBAR_RADIUS}; }
    #BrandCard { background: {BRANDCARD_BG}; border: 1px solid {BRANDCARD_BORDER}; border-radius: 8px; }
    #BrandMark { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 {BRANDMARK_BG1}, stop:1 {BRANDMARK_BG2}); border-radius: 6px; color: #FFFFFF; font: 900 18px Bahnschrift; }
    #Brand { color: {BRAND_COLOR}; font: 900 24px Bahnschrift; letter-spacing: 1.8px; }
    #BrandKicker { color: {BRANDKICKER_COLOR}; font: 900 10px Bahnschrift; letter-spacing: 1.8px; }
    #SidebarHint { color: {SIDEBAR_HINT_COLOR}; font-size: 12px; }
    #SidebarSectionLabel { color: {SIDEBAR_SECTION_COLOR}; font: 900 10px Bahnschrift; letter-spacing: 1.8px; padding-left: 8px; }
    #SidebarFooter { background: {SIDEBARFOOTER_BG}; border: 1px solid {SIDEBARFOOTER_BORDER}; border-radius: 8px; }
    #NavButton { background: transparent; color: {NAVBTN_COLOR}; border: 1px solid transparent; border-radius: 6px; padding: 10px 12px; text-align: left; font: 800 12px Bahnschrift; }
    #NavButton:hover { background: {NAVBTN_BG_HOVER}; border-color: {NAVBTN_BORDER_HOVER}; color: {NAVBTN_COLOR_HOVER}; }
    #NavButton[active="true"] { background: {NAVBTN_BG_ACTIVE}; border: 1px solid {NAVBTN_BORDER_ACTIVE}; color: {NAVBTN_COLOR_ACTIVE}; }
    #MainSurface { background: transparent; }
    #Topbar { background: {TOPBAR_BG}; border: 1px solid {TOPBAR_BORDER}; border-left: 5px solid {TOPBAR_ACCENT_BORDER}; border-radius: 8px; }
    #Hero, #Panel, #ProductCard, #StatCard, #CommandPanel { background: {CARD_BG}; border: 1px solid {CARD_BORDER}; border-radius: 8px; }
    #ProductCard:hover, #Panel:hover, #CommandPanel:hover { border: 1px solid {CARD_BORDER_HOVER}; background: {CARD_BG_HOVER}; }
    #InnerPanel { background: {INNERPANEL_BG}; border: 1px solid {INNERPANEL_BORDER}; border-radius: 8px; }
    #Hero { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 {HERO_BG1}, stop:.52 {HERO_BG2}, stop:1 {HERO_BG3}); border: 0; border-radius: 8px; }
    #HeroTitle { color: {HERO_TITLE_COLOR}; font: 900 31px Bahnschrift; }
    #HeroText { color: {HERO_TEXT_COLOR}; font-size: 13px; }
    #HeroChip { color: {HERO_CHIP_COLOR}; background: {HERO_CHIP_BG}; border: 1px solid {HERO_CHIP_BORDER}; border-radius: 4px; padding: 8px 12px; font: 900 12px Bahnschrift; }
    #TopbarEyebrow { color: {EYEBROW_COLOR}; font: 900 10px Bahnschrift; letter-spacing: 1.8px; }
    #PageTitle { font: 900 28px Bahnschrift; color: {PAGE_TITLE_COLOR}; }
    #PageHint, #Muted { color: {MUTED_COLOR}; font-size: 12px; }
    #StatTitle { color: {STAT_TITLE_COLOR}; font: 900 12px Bahnschrift; }
    #StatValue { font: 900 25px Bahnschrift; color: {STAT_VALUE_COLOR}; }
    #CardTitle { font: 900 18px Bahnschrift; color: {CARD_TITLE_COLOR}; }
    #CardSubTitle { color: {CARD_SUBTITLE_COLOR}; font: 700 13px Bahnschrift; }
    #StockPill, #Pill, #HeroChip, #SectionTitle, #CounterPill, #SuccessPill, #DangerPill, #WarningPill { border-radius: 4px; }
    #StockPill { background: {PILL_STOCK_BG}; color: {PILL_STOCK_COLOR}; border: 1px solid {PILL_STOCK_BORDER}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #Pill { background: {PILL_BG}; color: {PILL_COLOR}; border: 1px solid {PILL_BORDER}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #SuccessPill { background: {PILL_SUCCESS_BG}; color: {PILL_SUCCESS_COLOR}; border: 1px solid {PILL_SUCCESS_BORDER}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #DangerPill { background: {PILL_DANGER_BG}; color: {PILL_DANGER_COLOR}; border: 1px solid {PILL_DANGER_BORDER}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #WarningPill { background: {PILL_WARN_BG}; color: {PILL_WARN_COLOR}; border: 1px solid {PILL_WARN_BORDER}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #CounterPill { background: {PILL_COUNTER_BG}; color: {PILL_COUNTER_COLOR}; padding: 6px 10px; font: 900 12px Bahnschrift; }
    #SectionTitle { color: {SECTION_TITLE_COLOR}; font: 900 12px Bahnschrift; margin-top: 6px; }
    #MiniTitle { color: {MINITITLE_COLOR}; font: 900 14px Bahnschrift; }
    #PanelHint { color: {PANEL_HINT_COLOR}; font-size: 11px; }
    #SelectionRow { background: {SELROW_BG}; border: 1px solid {SELROW_BORDER}; border-radius: 6px; }
    #SelectionRow:hover { background: {SELROW_BG_HOVER}; border-color: {SELROW_BORDER_HOVER}; }
    #SelectionStock { color: {SELROW_STOCK_COLOR}; font: 900 13px Bahnschrift; }
    #SelectionMeta { color: {SELROW_META_COLOR}; font-size: 11px; }
    #SelectionHandle { color: {SELROW_HANDLE_COLOR}; font: 800 11px Bahnschrift; }
    #ImageStage { background: {IMAGESTAGE_BG}; border: 1px dashed {IMAGESTAGE_BORDER}; border-radius: 8px; color: {IMAGESTAGE_COLOR}; font: 900 14px Bahnschrift; }
    #FeatureLine { background: {FEATURELINE_BG}; border-radius: 6px; padding: 8px; color: {FEATURELINE_COLOR}; }
    QPushButton { background: {BTN_BG}; border: 1px solid {BTN_BORDER}; border-radius: 6px; padding: 10px 14px; font: 900 12px Bahnschrift; color: {BTN_COLOR}; }
    QPushButton:hover { background: {BTN_BG_HOVER}; border-color: {BTN_BORDER_HOVER}; }
    QPushButton:pressed { background: {BTN_BG_PRESSED}; }
    QPushButton#PrimaryButton { background: {PRIMARY_BTN_BG}; color: {PRIMARY_BTN_COLOR}; border: 1px solid {PRIMARY_BTN_BORDER}; }
    QPushButton#PrimaryButton:hover { background: {PRIMARY_BTN_BG_HOVER}; border-color: {PRIMARY_BTN_BORDER_HOVER}; }
    QPushButton#GhostButton { background: {GHOST_BTN_BG}; border: 1px solid {GHOST_BTN_BORDER}; color: {GHOST_BTN_COLOR}; }
    QPushButton#GhostButton:hover { background: {GHOST_BTN_BG_HOVER}; border-color: {GHOST_BTN_BORDER_HOVER}; }
    QPushButton#DangerButton { background: {DANGER_BTN_BG}; color: {DANGER_BTN_COLOR}; border: 1px solid {DANGER_BTN_BORDER}; }
    QPushButton#DangerButton:hover { background: {DANGER_BTN_BG_HOVER}; }
    QPushButton#IconButton { background: {ICON_BTN_BG}; color: {ICON_BTN_COLOR}; min-width: 34px; border: 0; }
    QLineEdit, QComboBox, QPlainTextEdit, QTextEdit { background: {INPUT_BG}; border: 1px solid {INPUT_BORDER}; border-radius: 6px; padding: 10px; selection-background-color: {INPUT_SELECTION_BG}; color: {INPUT_COLOR}; }
    QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus, QTextEdit:focus { border: 1px solid {INPUT_BORDER_FOCUS}; background: {INPUT_BG_FOCUS}; }
    QComboBox { padding-right: 38px; }
    QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 36px; border-left: 1px solid {INPUT_BORDER}; border-top-right-radius: 6px; border-bottom-right-radius: 6px; background: {COMBO_DROPDOWN_BG}; }
    QComboBox::drop-down:hover { background: {COMBO_DROPDOWN_BG_HOVER}; }
    QComboBox::down-arrow { image: url("__COMBO_ARROW__"); width: 14px; height: 9px; }
    QScrollArea { border: 0; background: transparent; }
    QScrollBar:vertical { background: transparent; width: 10px; margin: 6px 0; }
    QScrollBar::handle:vertical { background: {SCROLL_HANDLE}; border-radius: 5px; min-height: 36px; }
    QScrollBar::handle:vertical:hover { background: {SCROLL_HANDLE_HOVER}; }
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
    QScrollBar:horizontal { background: transparent; height: 10px; margin: 0 6px; }
    QScrollBar::handle:horizontal { background: {SCROLL_HANDLE}; border-radius: 5px; min-width: 36px; }
    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }
    QListWidget#SelectionList { background: transparent; border: 0; outline: 0; }
    QListWidget#SelectionList::item { margin: 4px 0; border-radius: 6px; }
    QListWidget#SelectionList::item:selected { background: {SELLIST_SELECTED_BG}; color: {SELLIST_SELECTED_COLOR}; }
    QTableWidget, QTreeWidget { background: {TABLE_BG}; border: 1px solid {TABLE_BORDER}; border-radius: 8px; gridline-color: {TABLE_GRID}; alternate-background-color: {TABLE_ALT}; }
    QHeaderView::section { background: {TABLE_HEADER_BG}; color: {TABLE_HEADER_COLOR}; border: 0; padding: 10px; font: 900 12px Bahnschrift; }
    QTreeWidget::item, QTableWidget::item { padding: {TABLE_ITEM_PADDING}; }
    QTreeWidget::item:selected, QTableWidget::item:selected { background: {TABLE_SELECTED_BG}; color: {TABLE_SELECTED_COLOR}; }
    QSplitter::handle { background: transparent; width: 8px; }
    QDialog { background: {DIALOG_BG}; }
    #DialogTitle { font: 800 24px Bahnschrift; color: {DIALOG_TITLE_COLOR}; }
    QSizeGrip { width: 18px; height: 18px; background: {SIZEKIT_COLOR}; }
    #CardHeader { background: {CARD_HEADER_BG}; border-bottom: 1px solid {CARD_HEADER_BORDER}; border-top-left-radius: 8px; border-top-right-radius: 8px; }
    #ProductCard[selected="true"] { background: {SELECTED_CARD_BG}; border: 2px solid {TOPBAR_ACCENT_BORDER}; }
    #ProductCard[qrmissing="true"] { border: 2px solid {PILL_DANGER_COLOR}; }
"""


class ModernMainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Window)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.engine = DataEngine()
        self.thread_pool = QThreadPool.globalInstance()
        self.current_items: list[tuple[ProductRecord, list[Path], Path | None]] = []
        self.qr_current_items: list[tuple[ProductRecord, list[Path], Path | None]] = []
        self.qr_selected_records: dict[str, ProductRecord] = {}
        self.qr_selected_quantities: dict[str, int] = {}
        self.qr_card_widgets: dict[str, QrProductCard] = {}
        self.ce_current_items: list[tuple[ProductRecord, list[Path], Path | None]] = []
        self.ce_selected_records: dict[str, ProductRecord] = {}
        self.ce_selected_quantities: dict[str, int] = {}
        self.ce_card_widgets: dict[str, CeProductCard] = {}
        self.rename_plan: list[RenameAction] = []
        self.rename_plan_flat_output_mode = False
        self.rename_items: dict[int, RenameAction] = {}
        self._populating_rename = False
        self.active_workers: list[Worker] = []
        self.visual_cache_started = False
        self.pending_ydk_images: set[str] = set()
        self._loading_overlay: LoadingOverlay | None = None
        self._loading_overlay_depth = 0
        self.central_layout: QVBoxLayout | None = None
        self.app_shell: QFrame | None = None
        self.setWindowTitle(APP_TITLE)
        self.resize(1480, 920)
        self.setMinimumSize(1040, 680)
        if path_is_file(APP_ICON_ICO):
            self.setWindowIcon(QIcon(str(APP_ICON_ICO)))
        elif path_is_file(APP_ICON_PNG):
            self.setWindowIcon(QIcon(str(APP_ICON_PNG)))
        self.apply_theme()
        self.build_ui()
        self._update_theme_branding()
        self.load_records_async()

    def apply_theme(self, theme_key: str | None = None) -> None:
        if theme_key is None:
            theme_key = self.engine.settings.get("ui_theme", "light")
        t = THEMES.get(theme_key, THEMES["light"])
        combo_arrow = str(COMBO_ARROW_SVG).replace("\\", "/")
        qss = QSS_TEMPLATE.replace("__COMBO_ARROW__", combo_arrow)
        for k, v in t.items():
            qss = qss.replace(f"{{{k}}}", v)
        self.setStyleSheet(qss)

    def switch_theme(self, key: str) -> None:
        self.apply_theme(key)
        self._update_theme_branding()
        if self._loading_overlay is not None:
            self._loading_overlay.set_theme_key(key)

    def _ensure_loading_overlay(self) -> LoadingOverlay:
        if self._loading_overlay is None:
            self._loading_overlay = LoadingOverlay(self)
        self._loading_overlay.sync_to_parent()
        return self._loading_overlay

    def _show_loading_overlay(self, message: str) -> None:
        overlay = self._ensure_loading_overlay()
        self._loading_overlay_depth += 1
        overlay.start(message, self.engine.settings.get("ui_theme", "light"))

    def _hide_loading_overlay(self) -> None:
        if self._loading_overlay_depth > 0:
            self._loading_overlay_depth -= 1
        if self._loading_overlay_depth > 0:
            return
        if self._loading_overlay is not None:
            self._loading_overlay.stop()

    def _sync_active_overlays(self) -> None:
        if self._loading_overlay is not None and self._loading_overlay.isVisible():
            self._loading_overlay.sync_to_parent()
            self._loading_overlay.raise_()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._sync_active_overlays()

    def moveEvent(self, event) -> None:
        super().moveEvent(event)
        self._sync_active_overlays()


    def _update_theme_branding(self) -> None:
        if not hasattr(self, "brand_mark"):
            return
        self.side_layout.setContentsMargins(14, 14, 14, 14)
        self.brand_mark.setVisible(True)
        self.brand_mark.setStyleSheet("")
        self.brand_mark.setText("H")
        self.brand_kicker.setText("PRODUCT OPS")
        self.brand_kicker.setStyleSheet("")
        self.brand_logo.setStyleSheet("")
        if path_is_file(APP_ICON_PNG):
            brand_pixmap = QPixmap(str(APP_ICON_PNG)).scaled(
                30, 30,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
            self.brand_mark.setPixmap(brand_pixmap)

    def build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        self.central_layout = QVBoxLayout(central)
        self.central_layout.setContentsMargins(8, 8, 8, 8)
        self.central_layout.setSpacing(0)

        self.app_shell = QFrame()
        self.app_shell.setObjectName("AppShell")
        shell_layout = QVBoxLayout(self.app_shell)
        shell_layout.setContentsMargins(0, 0, 0, 0)
        shell_layout.setSpacing(0)

        self.title_bar = WindowTitleBar(self)
        shell_layout.addWidget(self.title_bar)

        content = QWidget()
        content.setObjectName("Workspace")
        root = QHBoxLayout(content)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)
        shell_layout.addWidget(content, 1)
        self.central_layout.addWidget(self.app_shell, 1)

        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(262)
        side_layout = QVBoxLayout(sidebar)
        side_layout.setContentsMargins(14, 14, 14, 14)
        side_layout.setSpacing(10)
        self.side_layout = side_layout

        brand_card = QFrame()
        brand_card.setObjectName("BrandCard")
        self.brand_card = brand_card
        brand_layout = QHBoxLayout(brand_card)
        brand_layout.setContentsMargins(12, 12, 12, 12)
        brand_layout.setSpacing(10)
        brand_mark = QLabel("H")
        brand_mark.setObjectName("BrandMark")
        brand_mark.setFixedSize(46, 46)
        brand_mark.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if path_is_file(APP_ICON_PNG):
            brand_pixmap = QPixmap(str(APP_ICON_PNG)).scaled(34, 34, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            brand_mark.setPixmap(brand_pixmap)
        self.brand_mark = brand_mark
        brand_text = QVBoxLayout()
        kicker = QLabel("PRODUCT OPS")
        kicker.setObjectName("BrandKicker")
        self.brand_kicker = kicker
        logo = QLabel("HEKA")
        logo.setObjectName("Brand")
        self.brand_logo = logo
        hint = QLabel("Command Center")
        hint.setObjectName("SidebarHint")
        brand_text.addWidget(kicker)
        brand_text.addWidget(logo)
        brand_text.addWidget(hint)
        brand_layout.addWidget(brand_mark)
        brand_layout.addLayout(brand_text, 1)
        side_layout.addWidget(brand_card)

        section = QLabel("MODULLER")
        section.setObjectName("SidebarSectionLabel")
        side_layout.addWidget(section)
        self.nav_buttons: dict[str, QPushButton] = {}
        for key, title, subtitle in [
            ("products", "Urun Vitrini", "Sorgula, kopyala, klasor ac"),
            ("qr", "\u00dcr\u00fcn QR Code", "QR goster, sec, Excel olustur"),
            ("ce", "CE Etiketi", "Sec, adet ver, Bartender listesi olustur"),
            ("rename", "Toplu Isim", "Onizle ve kontrollu uygula"),
            ("labels", "Etiket Atolyesi", "Yurtdisi barkod tasarimi"),
        ]:
            button = QPushButton(f"{title}\n{subtitle}")
            button.setObjectName("NavButton")
            button.setProperty("active", "false")
            button.setMinimumHeight(58)
            button.clicked.connect(lambda _checked=False, page=key: self.show_page(page))
            side_layout.addWidget(button)
            self.nav_buttons[key] = button
        side_layout.addStretch(1)

        footer = QFrame()
        footer.setObjectName("SidebarFooter")
        footer_layout = QVBoxLayout(footer)
        footer_layout.setContentsMargins(14, 12, 14, 12)
        footer_layout.setSpacing(5)
        footer_title = QLabel("SISTEM DURUMU")
        footer_title.setObjectName("BrandKicker")
        footer_layout.addWidget(footer_title)
        self.settings_summary = QLabel("")
        self.settings_summary.setObjectName("SidebarHint")
        self.settings_summary.setWordWrap(True)
        footer_layout.addWidget(self.settings_summary)
        side_layout.addWidget(footer)
        root.addWidget(sidebar)

        main = QWidget()
        main.setObjectName("MainSurface")
        main_layout = QVBoxLayout(main)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(16)
        topbar = QFrame()
        topbar.setObjectName("Topbar")
        top = QHBoxLayout(topbar)
        top.setContentsMargins(22, 16, 22, 16)
        eyebrow = QLabel("HEKA OPERASYON PANELI")
        eyebrow.setObjectName("TopbarEyebrow")
        self.page_title = QLabel("Urun Vitrini")
        self.page_title.setObjectName("PageTitle")
        self.page_hint = QLabel("SQL veri, gorsel ve kanal klasorleri tek ekranda.")
        self.page_hint.setObjectName("PageHint")
        title_box = QVBoxLayout()
        title_box.setSpacing(2)
        title_box.addWidget(eyebrow)
        title_box.addWidget(self.page_title)
        title_box.addWidget(self.page_hint)
        top.addLayout(title_box, 1)
        for text, handler, primary in [
            ("Ayarlar", self.open_settings, False),
            ("Veriyi Yenile", self.load_records_async, False),
            ("Indeksi Yenile", self.rebuild_index_async, True),
            ("Indeks Klasoru", self.open_index_location, False),
        ]:
            button = QPushButton(text)
            if primary:
                button.setObjectName("PrimaryButton")
            else:
                button.setObjectName("GhostButton")
            button.clicked.connect(handler)
            top.addWidget(button)
        main_layout.addWidget(topbar)
        self.pages = QStackedWidget()
        self.page_keys = ["products", "qr", "ce", "rename", "labels"]
        self.products_page = self.build_products_page()
        self.qr_page = self.build_qr_page()
        self.ce_page = self.build_ce_page()
        self.rename_page = self.build_rename_page()
        self.labels_page = self.build_labels_page()
        for page in [self.products_page, self.qr_page, self.ce_page, self.rename_page, self.labels_page]:
            self.pages.addWidget(page)
        main_layout.addWidget(self.pages, 1)
        self.status = QLabel("Hazir.")
        self.status.setObjectName("Muted")
        status_row = QHBoxLayout()
        status_row.setContentsMargins(0, 0, 0, 0)
        status_row.addWidget(self.status, 1)
        status_row.addWidget(QSizeGrip(self), 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        main_layout.addLayout(status_row)
        root.addWidget(main, 1)
        self.refresh_settings_summary()
        self.show_page("products")
        self.sync_frame_state()

    def toggle_max_restore(self) -> None:
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()
        self.sync_frame_state()

    def sync_frame_state(self) -> None:
        if self.central_layout is not None:
            margin = 0 if self.isMaximized() else 8
            self.central_layout.setContentsMargins(margin, margin, margin, margin)
        if hasattr(self, "title_bar"):
            self.title_bar.setProperty("maximized", "true" if self.isMaximized() else "false")
            self.title_bar.style().unpolish(self.title_bar)
            self.title_bar.style().polish(self.title_bar)
            self.title_bar.sync_max_button()
        if self.app_shell is not None:
            self.app_shell.setProperty("maximized", "true" if self.isMaximized() else "false")
            self.app_shell.style().unpolish(self.app_shell)
            self.app_shell.style().polish(self.app_shell)

    def changeEvent(self, event) -> None:  # type: ignore[override]
        super().changeEvent(event)
        if event.type() == QEvent.Type.WindowStateChange:
            self.sync_frame_state()

    def build_hero(self, title: str, text: str, chips: list[str]) -> QFrame:
        hero = QFrame()
        hero.setObjectName("Hero")
        layout = QHBoxLayout(hero)
        layout.setContentsMargins(24, 22, 24, 22)
        copy = QVBoxLayout()
        label = QLabel(title)
        label.setObjectName("HeroTitle")
        desc = QLabel(text)
        desc.setObjectName("HeroText")
        desc.setWordWrap(True)
        copy.addWidget(label)
        copy.addWidget(desc)
        layout.addLayout(copy, 1)
        chip_row = QHBoxLayout()
        for chip in chips:
            c = QLabel(chip)
            c.setObjectName("HeroChip")
            chip_row.addWidget(c)
        layout.addLayout(chip_row)
        return hero

    def build_products_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        layout.addWidget(self.build_hero("HEKA Urun Vitrini", "Aile, kirilim, stok, fotograf ve kanal klasorlerini tek modern vitrinde yonet.", ["SQL", "Carousel", "Explorer"]))
        stats = QHBoxLayout()
        self.record_card = StatCard("Kayit", "0", "#ff8a3d")
        self.family_card = StatCard("Aile", "0", "#2fd1b1")
        self.index_card = StatCard("Indeks", "-", "#4a83ff")
        for card in [self.record_card, self.family_card, self.index_card]:
            stats.addWidget(card)
        layout.addLayout(stats)
        control = QFrame()
        control.setObjectName("Panel")
        controls = QGridLayout(control)
        controls.setContentsMargins(18, 18, 18, 18)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Stok, aile, aciklama veya ozellik ara...")
        self.family_combo = QComboBox()
        self.family_combo.setEditable(True)
        self.breakdown_combo = QComboBox()
        self.breakdown_combo.setEditable(True)
        search_btn = QPushButton("Ara")
        search_btn.setObjectName("PrimaryButton")
        fetch_btn = QPushButton("Getir")
        fetch_btn.setObjectName("PrimaryButton")
        family_btn = QPushButton("Tum Aile")
        controls.addWidget(QLabel("Genel Arama"), 0, 0)
        controls.addWidget(self.search_edit, 1, 0)
        controls.addWidget(search_btn, 1, 1)
        controls.addWidget(QLabel("Urun Ailesi"), 0, 2)
        controls.addWidget(self.family_combo, 1, 2)
        controls.addWidget(QLabel("Kirilim"), 0, 3)
        controls.addWidget(self.breakdown_combo, 1, 3)
        controls.addWidget(fetch_btn, 1, 4)
        controls.addWidget(family_btn, 1, 5)
        controls.setColumnStretch(0, 2)
        controls.setColumnStretch(2, 1)
        controls.setColumnStretch(3, 1)
        layout.addWidget(control)
        self.family_combo.currentTextChanged.connect(self.update_breakdown_options)
        self.search_edit.returnPressed.connect(self.search_products)
        search_btn.clicked.connect(self.search_products)
        fetch_btn.clicked.connect(self.fetch_selected_products)
        family_btn.clicked.connect(self.fetch_family_products)
        self.product_scroll = QScrollArea()
        self.product_scroll.setWidgetResizable(True)
        self.product_grid_host = QWidget()
        self.product_grid = QGridLayout(self.product_grid_host)
        self.product_grid.setContentsMargins(0, 0, 0, 0)
        self.product_grid.setSpacing(14)
        self.product_scroll.setWidget(self.product_grid_host)
        layout.addWidget(self.product_scroll, 1)
        return page

    def build_qr_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        layout.addWidget(self.build_hero("\u00dcr\u00fcn QR Code", "Urunu ara, vitrinden sec, QR kodunu kontrol et ve Bartender / QRCodeChimp Excel ciktisi olustur.", ["QR PNG", "Bartender", "QRCodeChimp"]))

        panel = QFrame()
        panel.setObjectName("Panel")
        panel_vbox = QVBoxLayout(panel)
        panel_vbox.setContentsMargins(18, 14, 18, 14)
        panel_vbox.setSpacing(10)
        search_eyebrow = QLabel("ARAMA VE FÄ°LTRE")
        search_eyebrow.setObjectName("TopbarEyebrow")
        panel_vbox.addWidget(search_eyebrow)
        controls = QGridLayout()
        controls.setHorizontalSpacing(12)
        controls.setVerticalSpacing(6)
        self.qr_search_edit = QLineEdit()
        self.qr_search_edit.setPlaceholderText("Stok kodu, Ã¼rÃ¼n adÄ± veya aile ara...")
        self.qr_family_combo = QComboBox()
        self.qr_family_combo.setEditable(True)
        self.qr_breakdown_combo = QComboBox()
        self.qr_breakdown_combo.setEditable(True)
        search_btn = QPushButton("Ara")
        search_btn.setObjectName("PrimaryButton")
        fetch_btn = QPushButton("Getir")
        fetch_btn.setObjectName("PrimaryButton")
        family_btn = QPushButton("TÃ¼m Aile")
        clear_btn = QPushButton("Temizle")
        search_label = QLabel("Genel Arama")
        search_label.setObjectName("SectionTitle")
        family_label = QLabel("ÃœrÃ¼n Ailesi")
        family_label.setObjectName("SectionTitle")
        breakdown_label = QLabel("KÄ±rÄ±lÄ±m")
        breakdown_label.setObjectName("SectionTitle")
        controls.addWidget(search_label, 0, 0)
        controls.addWidget(self.qr_search_edit, 1, 0)
        controls.addWidget(search_btn, 1, 1)
        controls.addWidget(family_label, 0, 2)
        controls.addWidget(self.qr_family_combo, 1, 2)
        controls.addWidget(breakdown_label, 0, 3)
        controls.addWidget(self.qr_breakdown_combo, 1, 3)
        controls.addWidget(fetch_btn, 1, 4)
        controls.addWidget(family_btn, 1, 5)
        controls.addWidget(clear_btn, 1, 6)
        controls.setColumnStretch(0, 2)
        controls.setColumnStretch(2, 1)
        controls.setColumnStretch(3, 1)
        panel_vbox.addLayout(controls)
        layout.addWidget(panel)

        search_btn.clicked.connect(self.search_qr_products)
        fetch_btn.clicked.connect(self.fetch_qr_selected_products)
        family_btn.clicked.connect(self.fetch_qr_family_products)
        clear_btn.clicked.connect(self.clear_qr_selection)
        self.qr_search_edit.returnPressed.connect(self.search_qr_products)
        self.qr_family_combo.currentTextChanged.connect(self.update_qr_breakdown_options)

        splitter = QSplitter()
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        self.qr_scroll = QScrollArea()
        self.qr_scroll.setWidgetResizable(True)
        self.qr_grid_host = QWidget()
        self.qr_grid = QGridLayout(self.qr_grid_host)
        self.qr_grid.setContentsMargins(0, 0, 0, 0)
        self.qr_grid.setSpacing(14)
        self.qr_scroll.setWidget(self.qr_grid_host)
        left_layout.addWidget(self.qr_scroll, 1)
        splitter.addWidget(left)

        side = QFrame()
        side.setObjectName("CommandPanel")
        side.setMinimumWidth(460)
        side_layout = QVBoxLayout(side)
        side_layout.setContentsMargins(16, 16, 16, 16)
        side_layout.setSpacing(12)
        command_header = QVBoxLayout()
        command_header.setSpacing(3)
        command_kicker = QLabel("QR KOMUT MERKEZÄ°")
        command_kicker.setObjectName("TopbarEyebrow")
        command_title = QLabel("SeÃ§ & Ã‡Ä±kart")
        command_title.setObjectName("PageTitle")
        command_hint = QLabel("ÃœrÃ¼nÃ¼ Ã¶nizle, adet belirle, toplu stok ekle â€” Bartender / QRCodeChimp Ã§Ä±ktÄ±sÄ± tek tÄ±kla.")
        command_hint.setObjectName("PanelHint")
        command_hint.setWordWrap(True)
        command_header.addWidget(command_kicker)
        command_header.addWidget(command_title)
        command_header.addWidget(command_hint)
        side_layout.addLayout(command_header)
        side_scroll = QScrollArea()
        side_scroll.setObjectName("CommandScroll")
        side_scroll.setWidgetResizable(True)
        side_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        side_content = QWidget()
        side_content_layout = QVBoxLayout(side_content)
        side_content_layout.setContentsMargins(0, 0, 0, 0)
        side_content_layout.setSpacing(12)

        preview_panel = QFrame()
        preview_panel.setObjectName("InnerPanel")
        preview_layout = QVBoxLayout(preview_panel)
        preview_layout.setContentsMargins(14, 14, 14, 14)
        preview_layout.setSpacing(9)
        preview_header = QHBoxLayout()
        preview_header.setSpacing(6)
        preview_title = QLabel("QR Ã–NÄ°ZLEME")
        preview_title.setObjectName("TopbarEyebrow")
        preview_header.addWidget(preview_title)
        preview_header.addStretch(1)
        self.qr_preview = QLabel("QR\nseÃ§imi\nyok")
        self.qr_preview.setObjectName("ImageStage")
        self.qr_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.qr_preview.setFixedSize(290, 290)
        self.qr_info = QLabel("Kartlardan bir Ã¼rÃ¼ne tÄ±kla; QR kod burada gÃ¶rÃ¼necek.")
        self.qr_info.setObjectName("Muted")
        self.qr_info.setWordWrap(True)
        preview_layout.addLayout(preview_header)
        preview_layout.addWidget(self.qr_preview, alignment=Qt.AlignmentFlag.AlignCenter)
        preview_layout.addWidget(self.qr_info)

        selected_panel = QFrame()
        selected_panel.setObjectName("InnerPanel")
        selected_layout = QVBoxLayout(selected_panel)
        selected_layout.setContentsMargins(14, 14, 14, 14)
        selected_layout.setSpacing(8)
        selected_header = QHBoxLayout()
        selected_title = QLabel("Secilenler Listesi")
        selected_title.setObjectName("MiniTitle")
        self.qr_selected_count_label = QLabel("0 adet")
        self.qr_selected_count_label.setObjectName("CounterPill")
        selected_header.addWidget(selected_title)
        selected_header.addStretch(1)
        selected_header.addWidget(self.qr_selected_count_label)
        self.qr_selected_list = QListWidget()
        self.qr_selected_list.setObjectName("SelectionList")
        self.qr_selected_list.setMinimumHeight(240)
        self.qr_selected_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.qr_selected_list.itemClicked.connect(self.show_qr_selected_list_item)
        remove_btn = QPushButton("Secileni Kaldir")
        remove_btn.setObjectName("DangerButton")
        remove_btn.clicked.connect(self.remove_selected_qr_list_item)
        selected_layout.addLayout(selected_header)
        selected_layout.addWidget(self.qr_selected_list)
        selected_layout.addWidget(remove_btn)

        bulk_panel = QFrame()
        bulk_panel.setObjectName("InnerPanel")
        bulk_layout = QVBoxLayout(bulk_panel)
        bulk_layout.setContentsMargins(14, 14, 14, 14)
        bulk_layout.setSpacing(8)
        bulk_title = QLabel("Toplu Stok Kodu Ekle")
        bulk_title.setObjectName("MiniTitle")
        bulk_hint = QLabel("Stok kodlarini alt alta, virgulle veya Excel'den kopyalayip yapistir.")
        bulk_hint.setObjectName("Muted")
        bulk_hint.setWordWrap(True)
        self.qr_bulk_stock_text = QTextEdit()
        self.qr_bulk_stock_text.setPlaceholderText("Ornek:\nH1W6JMEECB0-040\nH1C1JMEJMER-070")
        self.qr_bulk_stock_text.setFixedHeight(96)
        bulk_actions = QHBoxLayout()
        bulk_add_btn = QPushButton("Listeye Ekle")
        bulk_add_btn.setObjectName("PrimaryButton")
        bulk_clear_btn = QPushButton("Metni Temizle")
        bulk_add_btn.clicked.connect(self.add_bulk_qr_stock_codes)
        bulk_clear_btn.clicked.connect(self.qr_bulk_stock_text.clear)
        bulk_actions.addWidget(bulk_add_btn)
        bulk_actions.addWidget(bulk_clear_btn)
        bulk_layout.addWidget(bulk_title)
        bulk_layout.addWidget(bulk_hint)
        bulk_layout.addWidget(self.qr_bulk_stock_text)
        bulk_layout.addLayout(bulk_actions)

        export_panel = QFrame()
        export_panel.setObjectName("InnerPanel")
        export_layout = QVBoxLayout(export_panel)
        export_layout.setContentsMargins(14, 14, 14, 14)
        export_layout.setSpacing(8)
        export_title = QLabel("Cikti Islemleri")
        export_title.setObjectName("MiniTitle")
        bartender_btn = QPushButton("Bartender Excel Olustur")
        bartender_btn.setObjectName("PrimaryButton")
        label_reset_btn = QPushButton("Etiketi Sifirla")
        label_reset_btn.setObjectName("GhostButton")
        bartender_print_btn = QPushButton("Barkod Makinesine Direkt Yazdir")
        bartender_print_btn.setObjectName("PrimaryButton")
        chimp_btn = QPushButton("QRCodeChimp Excel Olustur")
        chimp_btn.setObjectName("PrimaryButton")
        bartender_btn.clicked.connect(self.export_qr_bartender)
        label_reset_btn.clicked.connect(self.reset_label_printer_from_ui)
        bartender_print_btn.clicked.connect(self.print_qr_bartender)
        chimp_btn.clicked.connect(self.export_qr_chimp)
        export_layout.addWidget(export_title)
        export_layout.addWidget(bartender_btn)
        export_layout.addWidget(label_reset_btn)
        export_layout.addWidget(bartender_print_btn)
        export_layout.addWidget(chimp_btn)
        side_content_layout.addWidget(preview_panel)
        side_content_layout.addWidget(selected_panel, 1)
        side_content_layout.addWidget(bulk_panel)
        side_content_layout.addWidget(export_panel)
        side_content_layout.addStretch(1)
        side_scroll.setWidget(side_content)
        side_layout.addWidget(side_scroll)
        splitter.addWidget(side)
        splitter.setSizes([920, 500])
        layout.addWidget(splitter, 1)
        return page

    def build_ce_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        layout.addWidget(self.build_hero("CE Etiketi", "Urunu sec, adet belirle ve CE Bartender baski listesini hazirla.", ["CE", "Bartender", "ARGE Excel"]))

        panel = QFrame()
        panel.setObjectName("Panel")
        panel_vbox = QVBoxLayout(panel)
        panel_vbox.setContentsMargins(18, 14, 18, 14)
        panel_vbox.setSpacing(10)
        search_eyebrow = QLabel("ARAMA VE FILTRE")
        search_eyebrow.setObjectName("TopbarEyebrow")
        panel_vbox.addWidget(search_eyebrow)
        controls = QGridLayout()
        controls.setHorizontalSpacing(12)
        controls.setVerticalSpacing(6)
        self.ce_search_edit = QLineEdit()
        self.ce_search_edit.setPlaceholderText("Stok kodu, urun adi veya aile ara...")
        self.ce_family_combo = QComboBox()
        self.ce_family_combo.setEditable(True)
        self.ce_breakdown_combo = QComboBox()
        self.ce_breakdown_combo.setEditable(True)
        search_btn = QPushButton("Ara")
        search_btn.setObjectName("PrimaryButton")
        fetch_btn = QPushButton("Getir")
        fetch_btn.setObjectName("PrimaryButton")
        family_btn = QPushButton("Tum Aile")
        clear_btn = QPushButton("Temizle")
        controls.addWidget(QLabel("Genel Arama"), 0, 0)
        controls.addWidget(self.ce_search_edit, 1, 0)
        controls.addWidget(search_btn, 1, 1)
        controls.addWidget(QLabel("Urun Ailesi"), 0, 2)
        controls.addWidget(self.ce_family_combo, 1, 2)
        controls.addWidget(QLabel("Kirilim"), 0, 3)
        controls.addWidget(self.ce_breakdown_combo, 1, 3)
        controls.addWidget(fetch_btn, 1, 4)
        controls.addWidget(family_btn, 1, 5)
        controls.addWidget(clear_btn, 1, 6)
        controls.setColumnStretch(0, 2)
        controls.setColumnStretch(2, 1)
        controls.setColumnStretch(3, 1)
        panel_vbox.addLayout(controls)
        layout.addWidget(panel)

        search_btn.clicked.connect(self.search_ce_products)
        fetch_btn.clicked.connect(self.fetch_ce_selected_products)
        family_btn.clicked.connect(self.fetch_ce_family_products)
        clear_btn.clicked.connect(self.clear_ce_selection)
        self.ce_search_edit.returnPressed.connect(self.search_ce_products)
        self.ce_family_combo.currentTextChanged.connect(self.update_ce_breakdown_options)

        splitter = QSplitter()
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        self.ce_scroll = QScrollArea()
        self.ce_scroll.setWidgetResizable(True)
        self.ce_grid_host = QWidget()
        self.ce_grid = QGridLayout(self.ce_grid_host)
        self.ce_grid.setContentsMargins(0, 0, 0, 0)
        self.ce_grid.setSpacing(14)
        self.ce_scroll.setWidget(self.ce_grid_host)
        left_layout.addWidget(self.ce_scroll, 1)
        splitter.addWidget(left)

        side = QFrame()
        side.setObjectName("CommandPanel")
        side.setMinimumWidth(460)
        side_layout = QVBoxLayout(side)
        side_layout.setContentsMargins(16, 16, 16, 16)
        side_layout.setSpacing(12)
        command_header = QVBoxLayout()
        command_header.setSpacing(3)
        command_kicker = QLabel("CE KOMUT MERKEZI")
        command_kicker.setObjectName("TopbarEyebrow")
        command_title = QLabel("Sec & Cikart")
        command_title.setObjectName("PageTitle")
        command_hint = QLabel("Urunu onizle, adet belirle ve CE Bartender ciktisini tek tikla olustur.")
        command_hint.setObjectName("PanelHint")
        command_hint.setWordWrap(True)
        command_header.addWidget(command_kicker)
        command_header.addWidget(command_title)
        command_header.addWidget(command_hint)
        side_layout.addLayout(command_header)
        side_scroll = QScrollArea()
        side_scroll.setObjectName("CommandScroll")
        side_scroll.setWidgetResizable(True)
        side_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        side_content = QWidget()
        side_content_layout = QVBoxLayout(side_content)
        side_content_layout.setContentsMargins(0, 0, 0, 0)
        side_content_layout.setSpacing(12)

        preview_panel = QFrame()
        preview_panel.setObjectName("InnerPanel")
        preview_layout = QVBoxLayout(preview_panel)
        preview_layout.setContentsMargins(14, 14, 14, 14)
        preview_layout.setSpacing(9)
        preview_title = QLabel("URUN GORSELI")
        preview_title.setObjectName("TopbarEyebrow")
        preview_layout.addWidget(preview_title)
        self.ce_preview = QLabel("Urun\ngorseli\nyok")
        self.ce_preview.setObjectName("ImageStage")
        self.ce_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ce_preview.setFixedSize(290, 220)
        self.ce_info = QLabel("Kartlardan bir urune tikla; urun gorseli ve CE alanlari burada gorunecek.")
        self.ce_info.setObjectName("Muted")
        self.ce_info.setWordWrap(True)
        preview_layout.addWidget(self.ce_preview, alignment=Qt.AlignmentFlag.AlignCenter)
        preview_layout.addWidget(self.ce_info)

        selected_panel = QFrame()
        selected_panel.setObjectName("InnerPanel")
        selected_layout = QVBoxLayout(selected_panel)
        selected_layout.setContentsMargins(14, 14, 14, 14)
        selected_layout.setSpacing(8)
        selected_header = QHBoxLayout()
        selected_title = QLabel("Secilenler Listesi")
        selected_title.setObjectName("MiniTitle")
        self.ce_selected_count_label = QLabel("0 adet")
        self.ce_selected_count_label.setObjectName("CounterPill")
        selected_header.addWidget(selected_title)
        selected_header.addStretch(1)
        selected_header.addWidget(self.ce_selected_count_label)
        self.ce_selected_list = QListWidget()
        self.ce_selected_list.setObjectName("SelectionList")
        self.ce_selected_list.setMinimumHeight(240)
        self.ce_selected_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.ce_selected_list.itemClicked.connect(self.show_ce_selected_list_item)
        remove_btn = QPushButton("Secileni Kaldir")
        remove_btn.setObjectName("DangerButton")
        remove_btn.clicked.connect(self.remove_selected_ce_list_item)
        selected_layout.addLayout(selected_header)
        selected_layout.addWidget(self.ce_selected_list)
        selected_layout.addWidget(remove_btn)

        bulk_panel = QFrame()
        bulk_panel.setObjectName("InnerPanel")
        bulk_layout = QVBoxLayout(bulk_panel)
        bulk_layout.setContentsMargins(14, 14, 14, 14)
        bulk_layout.setSpacing(8)
        bulk_title = QLabel("Toplu Stok Kodu Ekle")
        bulk_title.setObjectName("MiniTitle")
        bulk_hint = QLabel("Stok kodlarini alt alta, virgulle veya Excel'den kopyalayip yapistir.")
        bulk_hint.setObjectName("Muted")
        bulk_hint.setWordWrap(True)
        self.ce_bulk_stock_text = QTextEdit()
        self.ce_bulk_stock_text.setPlaceholderText("Ornek:\nH1W6JMKECS0-030\nH1C6WRBJMER-020")
        self.ce_bulk_stock_text.setFixedHeight(96)
        bulk_actions = QHBoxLayout()
        bulk_add_btn = QPushButton("Listeye Ekle")
        bulk_add_btn.setObjectName("PrimaryButton")
        bulk_clear_btn = QPushButton("Metni Temizle")
        bulk_add_btn.clicked.connect(self.add_bulk_ce_stock_codes)
        bulk_clear_btn.clicked.connect(self.ce_bulk_stock_text.clear)
        bulk_actions.addWidget(bulk_add_btn)
        bulk_actions.addWidget(bulk_clear_btn)
        bulk_layout.addWidget(bulk_title)
        bulk_layout.addWidget(bulk_hint)
        bulk_layout.addWidget(self.ce_bulk_stock_text)
        bulk_layout.addLayout(bulk_actions)

        export_panel = QFrame()
        export_panel.setObjectName("InnerPanel")
        export_layout = QVBoxLayout(export_panel)
        export_layout.setContentsMargins(14, 14, 14, 14)
        export_layout.setSpacing(8)
        export_title = QLabel("Cikti Islemleri")
        export_title.setObjectName("MiniTitle")
        export_hint = QLabel(f"Kaynak: {self.engine.settings.get('ce_measurement_excel_path', CE_MEASUREMENTS_XLSX_DEFAULT)}")
        export_hint.setObjectName("Muted")
        export_hint.setWordWrap(True)
        bartender_btn = QPushButton("CE Bartender Excel Olustur")
        bartender_btn.setObjectName("PrimaryButton")
        label_reset_btn = QPushButton("Etiketi Sifirla")
        label_reset_btn.setObjectName("GhostButton")
        bartender_print_btn = QPushButton("CE Barkod Makinesine Direkt Yazdir")
        bartender_print_btn.setObjectName("PrimaryButton")
        bartender_btn.clicked.connect(self.export_ce_bartender)
        label_reset_btn.clicked.connect(self.reset_label_printer_from_ui)
        bartender_print_btn.clicked.connect(self.print_ce_bartender)
        export_layout.addWidget(export_title)
        export_layout.addWidget(export_hint)
        export_layout.addWidget(bartender_btn)
        export_layout.addWidget(label_reset_btn)
        export_layout.addWidget(bartender_print_btn)
        side_content_layout.addWidget(preview_panel)
        side_content_layout.addWidget(selected_panel, 1)
        side_content_layout.addWidget(bulk_panel)
        side_content_layout.addWidget(export_panel)
        side_content_layout.addStretch(1)
        side_scroll.setWidget(side_content)
        side_layout.addWidget(side_scroll)
        splitter.addWidget(side)
        splitter.setSizes([920, 500])
        layout.addWidget(splitter, 1)
        return page

    def build_rename_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        layout.addWidget(self.build_hero("Toplu Isim Degistirme", "01_ Urun Gorselleri altindaki stok klasorlerini tarar, secili dosyalari uygular ve klasor agacini korur.", ["Preview", "Checkbox", "Export"]))
        panel = QFrame()
        panel.setObjectName("Panel")
        form = QGridLayout(panel)
        form.setContentsMargins(18, 18, 18, 18)
        self.rename_root = QLineEdit(self.engine.settings.get("last_rename_path", ""))
        self.rename_output = QLineEdit(self.engine.settings.get("rename_output_path", ""))
        self.rename_export = QCheckBox("Farkli klasore cikart - orijinale dokunma")
        self.rename_export.setChecked(self.engine.settings.get("rename_export_enabled", "0") == "1")
        self.rename_flat_output = QCheckBox("Tek klasor cikti - alt klasor agaci olusturma")
        self.rename_flat_output.setChecked(self.engine.settings.get("rename_flat_output_enabled", self.engine.settings.get("rename_flat_images_enabled", "0")) == "1")
        root_btn = QPushButton("Sec")
        output_btn = QPushButton("Sec")
        preview_btn = QPushButton("Onizleme")
        preview_btn.setObjectName("PrimaryButton")
        apply_btn = QPushButton("Secilenleri Uygula")
        apply_btn.setObjectName("PrimaryButton")
        self.rename_filter = QComboBox()
        self.rename_filter.addItems(RENAME_FILTER_OPTIONS)
        self.rename_scope = QComboBox()
        for scope_key, scope_label in RENAME_SCOPE_OPTIONS:
            self.rename_scope.addItem(scope_label, scope_key)
        scope_index = self.rename_scope.findData(self.engine.settings.get("rename_scope", RENAME_SCOPE_ALL))
        self.rename_scope.setCurrentIndex(scope_index if scope_index >= 0 else 0)
        form.addWidget(QLabel("Tarama Klasoru"), 0, 0)
        form.addWidget(self.rename_root, 0, 1)
        form.addWidget(root_btn, 0, 2)
        form.addWidget(self.rename_export, 1, 1)
        form.addWidget(QLabel("Cikis Klasoru"), 2, 0)
        form.addWidget(self.rename_output, 2, 1)
        form.addWidget(output_btn, 2, 2)
        form.addWidget(QLabel("Cikti Modu"), 3, 0)
        form.addWidget(self.rename_flat_output, 3, 1)
        form.addWidget(QLabel("Otomatik Secim"), 4, 0)
        form.addWidget(self.rename_scope, 4, 1)
        form.addWidget(QLabel("Filtre"), 5, 0)
        form.addWidget(self.rename_filter, 5, 1)
        actions = QHBoxLayout()
        actions.addWidget(preview_btn)
        for text, handler in [("Tumunu Sec", self.select_all_rename), ("Tumunu Kaldir", self.clear_all_rename), ("Tumunu Ac", lambda: self.rename_tree.expandAll()), ("Tumunu Kapat", lambda: self.rename_tree.collapseAll())]:
            button = QPushButton(text)
            button.clicked.connect(handler)
            actions.addWidget(button)
        actions.addWidget(apply_btn)
        actions.addStretch(1)
        form.addLayout(actions, 6, 1, 1, 2)
        layout.addWidget(panel)
        root_btn.clicked.connect(lambda: self.pick_directory(self.rename_root))
        output_btn.clicked.connect(lambda: self.pick_directory(self.rename_output))
        preview_btn.clicked.connect(self.preview_rename)
        apply_btn.clicked.connect(self.apply_rename)
        self.rename_flat_output.toggled.connect(self.on_rename_flat_output_changed)
        self.rename_scope.currentIndexChanged.connect(self.on_rename_scope_changed)
        self.rename_filter.currentTextChanged.connect(lambda: self.populate_rename_tree())
        splitter = QSplitter()
        self.rename_tree = QTreeWidget()
        self.rename_tree.setColumnCount(7)
        self.rename_tree.setHeaderLabels(["Sec", "Durum", "Stok", "Mevcut", "Yeni / Hedef", "Kural", "Kokten Sonra"])
        self.rename_tree.header().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.rename_tree.header().setStretchLastSection(True)
        self.rename_tree.itemChanged.connect(self.on_rename_item_changed)
        self.rename_tree.currentItemChanged.connect(self.on_rename_current_changed)
        splitter.addWidget(self.rename_tree)
        preview = QFrame()
        preview.setObjectName("ProductCard")
        pv = QVBoxLayout(preview)
        self.rename_preview_image = QLabel("Secim yok")
        self.rename_preview_image.setObjectName("ImageStage")
        self.rename_preview_image.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.rename_preview_image.setFixedSize(330, 240)
        self.rename_preview_text = QLabel("Listeden bir gorsel sec.")
        self.rename_preview_text.setWordWrap(True)
        self.rename_preview_text.setObjectName("Muted")
        open_btn = QPushButton("Explorer'da Goster")
        open_btn.clicked.connect(self.open_selected_rename_file)
        pv.addWidget(QLabel("Gorsel Onizleme"))
        pv.addWidget(self.rename_preview_image)
        pv.addWidget(self.rename_preview_text)
        pv.addWidget(open_btn)
        pv.addStretch(1)
        splitter.addWidget(preview)
        splitter.setSizes([900, 360])
        layout.addWidget(splitter, 1)
        return page

    def build_labels_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)
        layout.addWidget(self.build_hero("Yurtdisi Etiket Atolyesi", "Ana veri kaynagindan barkod bilgilerini ceker, urun bilgilerini duzenlenebilir kartta gosterir.", ["Barkod", "PDF hazirlik", "Kopyala"]))
        panel = QFrame()
        panel.setObjectName("Panel")
        tools = QHBoxLayout(panel)
        tools.setContentsMargins(18, 18, 18, 18)
        self.ydk_search = QLineEdit()
        self.ydk_search.setPlaceholderText("Stok, model, HKA veya barkod ara...")
        search = QPushButton("Ara")
        search.setObjectName("PrimaryButton")
        clear = QPushButton("Temizle")
        tools.addWidget(QLabel("Etiket Arama"))
        tools.addWidget(self.ydk_search, 1)
        tools.addWidget(search)
        tools.addWidget(clear)
        layout.addWidget(panel)
        search.clicked.connect(self.search_ydk)
        clear.clicked.connect(lambda: (self.ydk_search.clear(), self.search_ydk()))
        self.ydk_search.returnPressed.connect(self.search_ydk)
        splitter = QSplitter()
        self.ydk_table = QTableWidget(0, 4)
        self.ydk_table.setHorizontalHeaderLabels(["Stok", "Model", "HKA", "Barkod"])
        self.ydk_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.ydk_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.ydk_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.ydk_table.setAlternatingRowColors(True)
        self.ydk_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.ydk_table.itemSelectionChanged.connect(self.on_ydk_selection_changed)
        self.ydk_table.cellClicked.connect(lambda row, _column: self.show_ydk_product(row))
        splitter.addWidget(self.ydk_table)
        detail = QFrame()
        detail.setObjectName("ProductCard")
        dv = QVBoxLayout(detail)
        self.ydk_image = QLabel("Urun gorseli")
        self.ydk_image.setObjectName("ImageStage")
        self.ydk_image.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ydk_image.setFixedHeight(220)
        self.ydk_summary = QTextEdit()
        self.ydk_summary.setReadOnly(False)
        copy = QPushButton("Bilgiyi Kopyala")
        copy.clicked.connect(lambda: QApplication.clipboard().setText(self.ydk_summary.toPlainText()))
        export_row = QHBoxLayout()
        unit_pdf = QPushButton("Urun PDF")
        carton_pdf = QPushButton("Koli PDF")
        both_pdf = QPushButton("Ikisini Kaydet")
        both_pdf.setObjectName("PrimaryButton")
        unit_pdf.clicked.connect(lambda: self.export_current_ydk("unit"))
        carton_pdf.clicked.connect(lambda: self.export_current_ydk("carton"))
        both_pdf.clicked.connect(self.export_both_current_ydk)
        export_row.addWidget(unit_pdf)
        export_row.addWidget(carton_pdf)
        export_row.addWidget(both_pdf)
        export_row.addStretch(1)
        dv.addWidget(self.ydk_image)
        dv.addWidget(self.ydk_summary, 1)
        dv.addWidget(copy)
        dv.addLayout(export_row)
        splitter.addWidget(detail)
        splitter.setSizes([680, 520])
        layout.addWidget(splitter, 1)
        return page

    def show_page(self, key: str) -> None:
        index = self.page_keys.index(key)
        self.pages.setCurrentIndex(index)
        titles = {
            "products": ("Urun Vitrini", "SQL veri, gorsel ve kanal klasorleri tek ekranda."),
            "qr": ("\u00dcr\u00fcn QR Code", "Urun vitrini, QR onizleme ve iki Excel cikti akisi."),
            "ce": ("CE Etiketi", "Secili stoklardan CE Bartender Excel listesi hazirla."),
            "rename": ("Toplu Isim", "Secimli, onizlemeli ve cikti klasorlu isimlendirme."),
            "labels": ("Etiket Atolyesi", "Yurtdisi etiket verilerini modern panelde incele."),
        }
        self.page_title.setText(titles[key][0])
        self.page_hint.setText(titles[key][1])
        for name, button in self.nav_buttons.items():
            button.setProperty("active", "true" if name == key else "false")
            button.style().unpolish(button)
            button.style().polish(button)

    def set_status(self, text: str) -> None:
        self.status.setText(text)

    def reset_label_printer_from_ui(self) -> None:
        printer = self.engine.settings.get("bartender_printer_name", "").strip() or "ayarlardaki barkod yazicisi"
        if QMessageBox.question(
            self,
            "Etiketi Sifirla",
            f"{printer} yazicisina etiket baslangicini sifirlama/feed komutu gonderilecek.\n\n"
            "Yazici 1 bos etiketi ilerletebilir. Devam edelim mi?",
        ) != QMessageBox.StandardButton.Yes:
            return
        self.run_worker(
            lambda: self.engine.reset_label_printer("HEKA manuel etiket sifirlama", force=True),
            lambda job_id: QMessageBox.information(self, "Etiket Sifirlandi", f"Yazici sifirlama komutu gonderildi.\nJob ID: {job_id or '-'}"),
            "Etiket baslangici sifirlaniyor...",
            show_loading_overlay=True,
        )

    def run_worker(
        self,
        job: Callable[[], Any],
        on_success: Callable[[Any], None],
        description: str,
        show_loading_overlay: bool = False,
    ) -> None:
        self.set_status(description)
        if show_loading_overlay:
            self._show_loading_overlay(description)
        worker = Worker(job)
        self.active_workers.append(worker)

        def cleanup() -> None:
            if worker in self.active_workers:
                self.active_workers.remove(worker)

        def handle_result(result: Any) -> None:
            cleanup()
            if show_loading_overlay:
                self._hide_loading_overlay()
            try:
                on_success(result)
            except Exception as exc:  # noqa: BLE001
                self.write_log(f"Worker success handler error: {exc}")
                self.show_error("Islem Hatasi", str(exc))

        def handle_error(error: str) -> None:
            cleanup()
            if show_loading_overlay:
                self._hide_loading_overlay()
            self.write_log(f"Worker error: {error}")
            self.show_error("Islem Hatasi", error)

        worker.signals.result.connect(handle_result)
        worker.signals.error.connect(handle_error)
        self.thread_pool.start(worker)

    def write_log(self, message: str) -> None:
        try:
            log_path = DATA_DIR / "modern_app.log"
            with log_path.open("a", encoding="utf-8") as handle:
                handle.write(f"{datetime.now().isoformat(timespec='seconds')} {message}\n")
        except OSError:
            pass

    def show_error(self, title: str, message: str) -> None:
        self.set_status(message)
        QMessageBox.critical(self, title, message)

    def refresh_settings_summary(self) -> None:
        source = self.engine.get_data_source_key().upper()
        index = self.engine.get_index_last_rebuild_text()
        self.settings_summary.setText(f"Veri: {source}\nArama: {self.engine.settings.get('search_root', '-')}\n{index}")
        if hasattr(self, "index_card"):
            self.index_card.set_value(index)

    def open_settings(self) -> None:
        dialog = SettingsDialog(self.engine, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.engine.save_settings(dialog.values())
            self.engine.preview_search_roots_cache = None
            self.engine.preview_image_cache.clear()
            self.engine.preview_image_list_cache.clear()
            self.engine.search_folder_cache.clear()
            self.engine.channel_folder_cache.clear()
            self.engine.qr_catalog_cache = None
            self.engine.qr_file_index_cache = None
            self.engine.qr_sitemap_cache = None
            self.engine.qr_file_cache.clear()
            self.engine.ce_label_cache = None
            self.refresh_settings_summary()
            self.load_records_async()

    def open_index_location(self) -> None:
        reveal_path(INDEX_FILE if path_exists(INDEX_FILE) else DATA_DIR, self)

    def rebuild_index_async(self) -> None:
        if not self.engine.records:
            QMessageBox.warning(self, "Veri Yok", "Indeks icin once veriyi yukle.")
            return

        def on_success(stats: dict[str, int]) -> None:
            self.engine.preview_image_cache.clear()
            self.engine.preview_image_list_cache.clear()
            self.engine.search_folder_cache.clear()
            self.engine.channel_folder_cache.clear()
            self.refresh_settings_summary()
            QMessageBox.information(
                self,
                "Indeks Yenilendi",
                f"Fotograf: {stats['preview']}\nKlasor: {stats['folder']}\nKanal: {stats['channel']}",
            )
            self.set_status("Indeks yenilendi. Simdi Getir/Ara cok daha hizli calisir.")

        self.run_worker(
            self.engine.rebuild_product_index,
            on_success,
            "Gorsel ve klasor indeksi yenileniyor...",
            show_loading_overlay=True,
        )

    def load_records_async(self) -> None:
        self.visual_cache_started = False
        self.run_worker(
            self.engine.load_product_records,
            self.on_records_loaded,
            "Veri kaynagi yukleniyor...",
            show_loading_overlay=True,
        )

    def on_records_loaded(self, records: list[ProductRecord]) -> None:
        self.records = records
        families = sorted(unique_preserve_order([record.family for record in records if record.family]), key=alpha_sort_key)
        self.family_combo.blockSignals(True)
        self.family_combo.clear()
        self.family_combo.addItems(families)
        self.family_combo.blockSignals(False)
        self.qr_family_combo.blockSignals(True)
        self.qr_family_combo.clear()
        self.qr_family_combo.addItems(families)
        self.qr_family_combo.blockSignals(False)
        self.ce_family_combo.blockSignals(True)
        self.ce_family_combo.clear()
        self.ce_family_combo.addItems(families)
        self.ce_family_combo.blockSignals(False)
        self.record_card.set_value(str(len(records)))
        self.family_card.set_value(str(len(families)))
        self.populate_ydk_table(sorted(self.engine.ydk_products.values(), key=lambda item: item.code.casefold())[:500])
        if families:
            self.update_breakdown_options(families[0])
            self.update_qr_breakdown_options(families[0])
            self.update_ce_breakdown_options(families[0])
        message = f"{len(records)} kayit yuklendi." if records else self.engine.product_data_warning or "Veri bekleniyor."
        self.set_status(message)
        if self.engine.product_data_warning:
            self.write_log(self.engine.product_data_warning)
        if not records:
            self.clear_product_grid(message)
        else:
            self.start_initial_visual_fetch(records)

    def start_initial_visual_fetch(self, records: list[ProductRecord]) -> None:
        if self.visual_cache_started or self.engine.get_data_source_key() != "sql":
            return
        self.visual_cache_started = True
        self.run_worker(
            lambda: self.engine.warm_preview_cache(records),
            self.on_initial_visual_fetch_done,
            "SQL verisi yuklendi. Gorseller arka planda fetch ediliyor...",
        )

    def on_initial_visual_fetch_done(self, stats: dict[str, int]) -> None:
        self.set_status(f"Gorsel fetch tamamlandi: {stats.get('found', 0)}/{stats.get('requested', 0)} bulundu.")
        if getattr(self, "ydk_table", None) is not None and self.ydk_table.currentRow() >= 0:
            self.show_ydk_product(self.ydk_table.currentRow(), force=True)
        if not self.current_items:
            return
        visible_records = [record for record, _images, _folder in self.current_items]
        self.run_worker(
            lambda: self.engine.prepare_product_items(visible_records),
            lambda items: self.render_product_items(items, "Gorsel cache guncellendi"),
            "Guncel gorseller karta baglaniyor...",
        )

    def update_breakdown_options(self, family: str) -> None:
        breakdowns = sorted(
            unique_preserve_order([record.breakdown for record in self.engine.records if normalize_text(record.family) == normalize_text(family) and record.breakdown]),
            key=alpha_sort_key,
        )
        if breakdowns:
            breakdowns = [ALL_BREAKDOWNS_LABEL, *breakdowns]
        self.breakdown_combo.clear()
        self.breakdown_combo.addItems(breakdowns)

    def update_qr_breakdown_options(self, family: str) -> None:
        breakdowns = sorted(
            unique_preserve_order([record.breakdown for record in self.engine.records if normalize_text(record.family) == normalize_text(family) and record.breakdown]),
            key=alpha_sort_key,
        )
        if breakdowns:
            breakdowns = [ALL_BREAKDOWNS_LABEL, *breakdowns]
        self.qr_breakdown_combo.clear()
        self.qr_breakdown_combo.addItems(breakdowns)

    def update_ce_breakdown_options(self, family: str) -> None:
        breakdowns = sorted(
            unique_preserve_order([record.breakdown for record in self.engine.records if normalize_text(record.family) == normalize_text(family) and record.breakdown]),
            key=alpha_sort_key,
        )
        if breakdowns:
            breakdowns = [ALL_BREAKDOWNS_LABEL, *breakdowns]
        self.ce_breakdown_combo.clear()
        self.ce_breakdown_combo.addItems(breakdowns)

    def fetch_qr_family_products(self) -> None:
        if self.qr_breakdown_combo.findText(ALL_BREAKDOWNS_LABEL) >= 0:
            self.qr_breakdown_combo.setCurrentText(ALL_BREAKDOWNS_LABEL)
        self.fetch_qr_selected_products()

    def fetch_qr_selected_products(self) -> None:
        family = self.qr_family_combo.currentText().strip()
        breakdown = self.qr_breakdown_combo.currentText().strip()
        if not family:
            QMessageBox.warning(self, "Eksik Secim", "Lutfen QR icin aile sec.")
            return
        matched = [
            record
            for record in self.engine.records
            if normalize_text(record.family) == normalize_text(family)
            and (not breakdown or normalize_text(breakdown) == normalize_text(ALL_BREAKDOWNS_LABEL) or normalize_text(record.breakdown) == normalize_text(breakdown))
        ]
        self.render_qr_products_async(matched[:300], f"QR secim: {family} / {breakdown}")

    def search_qr_products(self) -> None:
        query = self.qr_search_edit.text().strip()
        if not query:
            QMessageBox.warning(self, "Arama Bos", "QR aramasi icin bir deger yaz.")
            return
        normalized = normalize_text(query)
        matched: list[ProductRecord] = []
        catalog = self.engine.load_qr_catalog()
        for record in self.engine.records:
            entry = catalog.get(normalize_text(record.stock_code))
            catalog_text = " ".join([entry.handle, entry.title]) if entry else ""
            if (
                normalized in normalize_text(record.family)
                or normalized in normalize_text(record.breakdown)
                or normalized in normalize_text(record.stock_code)
                or normalized in normalize_text(catalog_text)
                or any(normalized in normalize_text(f"{key} {value}") for key, value in record.raw_values.items())
            ):
                matched.append(record)
        self.render_qr_products_async(matched[:300], f"QR arama: {query}")

    def render_qr_products_async(self, matched: list[ProductRecord], context: str) -> None:
        if not matched:
            self.clear_qr_grid("QR icin kayit bulunamadi.")
            return
        immediate_items = [(record, [], None) for record in matched]
        self.render_qr_product_items(immediate_items, f"{context} - gorseller arka planda")
        self.run_worker(
            lambda: self.engine.prepare_product_items(matched),
            lambda items: self.render_qr_product_items(items, context),
            "QR urun gorselleri hazirlaniyor...",
        )

    def clear_qr_grid(self, message: str = "") -> None:
        while self.qr_grid.count():
            item = self.qr_grid.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.qr_card_widgets.clear()
        if message:
            label = QLabel(message)
            label.setObjectName("Muted")
            self.qr_grid.addWidget(label, 0, 0)

    def render_qr_product_items(self, items: list[tuple[ProductRecord, list[Path], Path | None]], context: str) -> None:
        self.qr_current_items = items
        self.clear_qr_grid()
        avail = self.width() - 262 - 480 - 36
        columns = max(1, avail // 270)
        for index, (record, images, _folder) in enumerate(items):
            qr_path = self.engine.find_qr_code(record)
            key = normalize_text(record.stock_code)
            selected = key in self.qr_selected_records
            card = QrProductCard(self.engine, record, images, qr_path, selected, self.show_qr_product, self.toggle_qr_selection)
            card.update_selected(selected, self.qr_selected_quantities.get(key, 0))
            self.qr_grid.addWidget(card, index // columns, index % columns)
            self.qr_card_widgets[key] = card
        self.qr_grid.setRowStretch((len(items) + columns - 1) // columns, 1)
        self.set_status(f"{len(items)} QR urunu listelendi. {context}")

    def show_qr_product(self, record: ProductRecord, images: list[Path] | None = None) -> None:
        qr_path = self.engine.find_qr_code(record)
        pixmap = load_scaled_pixmap(qr_path, QSize(290, 290), "#ffffff")
        if pixmap:
            self.qr_preview.setText("")
            self.qr_preview.setPixmap(pixmap)
        else:
            self.qr_preview.setPixmap(QPixmap())
            self.qr_preview.setText("QR yok")
        title = self.engine.get_qr_title(record)
        handle = self.engine.get_qr_handle(record, qr_path)
        safe_handle = handle or slugify_handle(record.stock_code) or slugify_handle(title)
        url = self.engine.get_qr_web_url(record, safe_handle)
        self.qr_info.setText(
            f"Stok: {record.stock_code}\n"
            f"Urun: {title}\n"
            f"QR Name: {safe_handle or '-'}\n"
            f"Web URL: {url or '-'}\n"
            f"QR Dosyasi: {qr_path if qr_path else 'Bulunamadi - QRCodeChimp exportuna yine dahil edilir'}"
        )

    def toggle_qr_selection(self, record: ProductRecord) -> int:
        key = normalize_text(record.stock_code)
        current_count = self.qr_selected_quantities.get(key, 0)
        next_count = current_count + 1 if current_count < 2 else 0
        self.set_qr_record_count(record, next_count)
        self.refresh_qr_selected_list()
        return next_count

    def set_qr_record_selected(self, record: ProductRecord, selected: bool) -> None:
        self.set_qr_record_count(record, 1 if selected else 0)

    def set_qr_record_count(self, record: ProductRecord, count: int) -> None:
        key = normalize_text(record.stock_code)
        count = max(0, min(2, count))
        if count:
            self.qr_selected_records[key] = record
            self.qr_selected_quantities[key] = count
        else:
            self.qr_selected_records.pop(key, None)
            self.qr_selected_quantities.pop(key, None)
        card = self.qr_card_widgets.get(key)
        if card is not None:
            card.update_selected(count > 0, count)

    def refresh_qr_selected_list(self) -> None:
        self.qr_selected_list.clear()
        total_count = 0
        for key, record in self.qr_selected_records.items():
            quantity = self.qr_selected_quantities.get(key, 1)
            total_count += quantity
            qr_path = self.engine.find_qr_code(record)
            handle = self.engine.get_qr_handle(record, qr_path)
            safe_handle = handle or slugify_handle(record.stock_code) or slugify_handle(self.engine.get_qr_title(record))
            title = self.engine.get_qr_title(record)
            item = QListWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, key)
            row = QWidget()
            row.setObjectName("SelectionRow")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(12, 10, 12, 10)
            row_layout.setSpacing(12)
            text_layout = QVBoxLayout()
            text_layout.setContentsMargins(0, 0, 0, 0)
            text_layout.setSpacing(4)
            stock_label = QLabel(record.stock_code)
            stock_label.setObjectName("SelectionStock")
            meta_label = QLabel(title)
            meta_label.setObjectName("SelectionMeta")
            meta_label.setWordWrap(True)
            handle_label = QLabel(f"QRCodeChimp: {safe_handle or '-'}")
            handle_label.setObjectName("SelectionHandle")
            file_label = QLabel(f"Dosya: {qr_path.name if qr_path else 'QR dosyasi yok, link satiri export edilir'}")
            file_label.setObjectName("SelectionMeta")
            file_label.setWordWrap(True)
            text_layout.addWidget(stock_label)
            text_layout.addWidget(meta_label)
            text_layout.addWidget(handle_label)
            text_layout.addWidget(file_label)

            status_layout = QVBoxLayout()
            status_layout.setContentsMargins(0, 0, 0, 0)
            status_layout.setSpacing(6)
            status_label = QLabel("QR HAZIR" if qr_path else "QR YOK")
            status_label.setObjectName("SuccessPill" if qr_path else "WarningPill")
            quantity_label = QLabel(f"x{quantity}")
            quantity_label.setObjectName("CounterPill")
            status_layout.addWidget(status_label, 0, Qt.AlignmentFlag.AlignRight)
            status_layout.addWidget(quantity_label, 0, Qt.AlignmentFlag.AlignRight)
            status_layout.addStretch(1)

            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setSpacing(6)
            minus_btn = QPushButton("-")
            minus_btn.setObjectName("GhostButton")
            minus_btn.setFixedWidth(36)
            minus_btn.clicked.connect(lambda _checked=False, item_key=key: self.change_qr_selection_quantity(item_key, -1))
            plus_btn = QPushButton("+")
            plus_btn.setObjectName("PrimaryButton")
            plus_btn.setFixedWidth(36)
            plus_btn.clicked.connect(lambda _checked=False, item_key=key: self.change_qr_selection_quantity(item_key, 1))
            remove_item_btn = QPushButton("Kaldir")
            remove_item_btn.setObjectName("DangerButton")
            remove_item_btn.setFixedWidth(72)
            remove_item_btn.clicked.connect(lambda _checked=False, item_key=key: self.remove_qr_selection_by_key(item_key))
            action_layout.addWidget(minus_btn)
            action_layout.addWidget(plus_btn)
            action_layout.addWidget(remove_item_btn)

            row_layout.addLayout(text_layout, 1)
            row_layout.addLayout(status_layout)
            row_layout.addLayout(action_layout)
            item.setSizeHint(QSize(0, 116))
            self.qr_selected_list.addItem(item)
            self.qr_selected_list.setItemWidget(item, row)
        if hasattr(self, "qr_selected_count_label"):
            self.qr_selected_count_label.setText(f"{total_count} adet")
        self.set_status(f"QR secili urun: {len(self.qr_selected_records)} stok / {total_count} adet")

    def remove_selected_qr_list_item(self) -> None:
        item = self.qr_selected_list.currentItem()
        if item is None:
            return
        key = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(key, str):
            self.remove_qr_selection_by_key(key)

    def remove_qr_selection_by_key(self, key: str) -> None:
        record = self.qr_selected_records.pop(key, None)
        self.qr_selected_quantities.pop(key, None)
        card = self.qr_card_widgets.get(key)
        if card is not None:
            card.update_selected(False, 0)
        if record is not None:
            self.set_status(f"{record.stock_code} QR listesinden kaldirildi.")
        self.refresh_qr_selected_list()

    def clear_qr_selection(self) -> None:
        self.qr_selected_records.clear()
        self.qr_selected_quantities.clear()
        for card in self.qr_card_widgets.values():
            card.update_selected(False, 0)
        self.refresh_qr_selected_list()

    def change_qr_selection_quantity(self, key: str, delta: int) -> None:
        record = self.qr_selected_records.get(key)
        if record is None:
            return
        current_count = self.qr_selected_quantities.get(key, 1)
        self.set_qr_record_count(record, current_count + delta)
        self.refresh_qr_selected_list()

    def show_qr_selected_list_item(self, item: QListWidgetItem) -> None:
        key = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(key, str) and key in self.qr_selected_records:
            self.show_qr_product(self.qr_selected_records[key])

    def get_qr_record_by_stock_code(self, stock_code: str) -> ProductRecord | None:
        key = normalize_text(stock_code)
        if not key:
            return None
        for record in self.engine.records:
            if normalize_text(record.stock_code) == key:
                return record
        entry = self.engine.load_qr_catalog().get(key)
        if entry is None:
            return None
        raw_values = {"Handle": entry.handle, "Title": entry.title, "Image Src": entry.image_url}
        return ProductRecord("", entry.title or entry.stock_code, entry.stock_code, {}, raw_values)

    def extract_bulk_qr_stock_codes(self, text: str) -> tuple[list[str], list[str]]:
        found: list[str] = []
        unresolved: list[str] = []
        for raw_line in re.split(r"[\r\n,;\t]+", text):
            line = raw_line.strip().strip("'\"")
            if not line:
                continue
            if normalize_text(line) in {"sku", "stok", "stok kodu", "stock code", "variant sku"}:
                continue
            line_matches: list[str] = []
            direct = self.get_qr_record_by_stock_code(line)
            if direct is not None:
                line_matches.append(direct.stock_code)
            else:
                for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_-]{2,}", line):
                    record = self.get_qr_record_by_stock_code(token)
                    if record is not None:
                        line_matches.append(record.stock_code)
            if line_matches:
                for stock_code in line_matches:
                    found.append(stock_code)
            else:
                unresolved.append(line)
        return found, unresolved

    def add_bulk_qr_stock_codes(self) -> None:
        text = self.qr_bulk_stock_text.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "Toplu Ekleme", "Listeye eklemek icin stok kodu yapistir.")
            return
        stock_codes, unresolved = self.extract_bulk_qr_stock_codes(text)
        added = 0
        maxed_out = 0
        for stock_code in stock_codes:
            record = self.get_qr_record_by_stock_code(stock_code)
            if record is None:
                unresolved.append(stock_code)
                continue
            key = normalize_text(record.stock_code)
            current_count = self.qr_selected_quantities.get(key, 0)
            next_count = min(2, current_count + 1)
            if next_count > current_count:
                added += 1
            else:
                maxed_out += 1
            self.set_qr_record_count(record, next_count)
        self.refresh_qr_selected_list()
        if stock_codes:
            self.qr_bulk_stock_text.clear()
        parts = [f"{added} adet listeye eklendi."]
        if maxed_out:
            parts.append(f"{maxed_out} satir zaten 2 adet limitindeydi.")
        if unresolved:
            preview = ", ".join(unresolved[:12])
            more = f" (+{len(unresolved) - 12})" if len(unresolved) > 12 else ""
            parts.append(f"Bulunamayan: {preview}{more}")
        QMessageBox.information(self, "Toplu QR Ekleme", "\n".join(parts))

    def selected_qr_records(self) -> list[ProductRecord]:
        records: list[ProductRecord] = []
        for key, record in self.qr_selected_records.items():
            records.extend([record] * max(1, self.qr_selected_quantities.get(key, 1)))
        return records

    def export_qr_bartender(self) -> None:
        records = self.selected_qr_records()
        if not records:
            QMessageBox.information(self, "Secim Yok", "Bartender ciktisi icin once urun sec.")
            return
        self.run_worker(
            lambda: self.engine.export_bartender_qr_excel(records),
            lambda path: QMessageBox.information(self, "Bartender Excel Hazir", str(path)),
            "Bartender Excel olusturuluyor...",
        )

    def print_qr_bartender(self) -> None:
        records = self.selected_qr_records()
        if not records:
            QMessageBox.information(self, "Secim Yok", "Direkt baski icin once QR urunu sec.")
            return
        printer = self.engine.settings.get("bartender_printer_name", "").strip() or "template icindeki yazici"
        reset_notice = (
            "Baski oncesi etiket baslangici sifirlanacak/feed edilecek."
            if self.engine.truthy_setting("label_reset_enabled", "1")
            else "Etiket sifirlama ayarlardan kapali."
        )
        if QMessageBox.question(
            self,
            "Direkt Yazdir",
            f"{len(records)} QR etiketi BarTender uzerinden direkt yaziciya gonderilecek.\n"
            f"{reset_notice}\n\n"
            f"Yazici: {printer}\nDevam edelim mi?",
        ) != QMessageBox.StandardButton.Yes:
            return
        self.run_worker(
            lambda: self.engine.print_bartender_qr_excel(records),
            lambda path: QMessageBox.information(self, "Baski Gonderildi", f"QR barkod baskisi BarTender'a gonderildi.\nVeri dosyasi: {path}"),
            "QR barkodlar BarTender ile yazdiriliyor...",
            show_loading_overlay=True,
        )

    def export_qr_chimp(self) -> None:
        records = self.selected_qr_records()
        if not records:
            QMessageBox.information(self, "Secim Yok", "QRCodeChimp ciktisi icin once urun sec.")
            return
        self.run_worker(
            lambda: self.engine.export_qrcodechimp_excel(records),
            lambda path: QMessageBox.information(
                self,
                "QRCodeChimp Excel Hazir",
                f"{path}\n\nQR dosyasi olmayan secimler de Web URL satiri olarak dahil edildi.",
            ),
            "QRCodeChimp Excel olusturuluyor...",
        )

    def fetch_ce_family_products(self) -> None:
        if self.ce_breakdown_combo.findText(ALL_BREAKDOWNS_LABEL) >= 0:
            self.ce_breakdown_combo.setCurrentText(ALL_BREAKDOWNS_LABEL)
        self.fetch_ce_selected_products()

    def fetch_ce_selected_products(self) -> None:
        family = self.ce_family_combo.currentText().strip()
        breakdown = self.ce_breakdown_combo.currentText().strip()
        if not family:
            QMessageBox.warning(self, "Eksik Secim", "Lutfen CE icin aile sec.")
            return
        matched = [
            record
            for record in self.engine.records
            if normalize_text(record.family) == normalize_text(family)
            and (not breakdown or normalize_text(breakdown) == normalize_text(ALL_BREAKDOWNS_LABEL) or normalize_text(record.breakdown) == normalize_text(breakdown))
        ]
        self.render_ce_products_async(matched[:300], f"CE secim: {family} / {breakdown}")

    def search_ce_products(self) -> None:
        query = self.ce_search_edit.text().strip()
        if not query:
            QMessageBox.warning(self, "Arama Bos", "CE aramasi icin bir deger yaz.")
            return
        normalized = normalize_text(query)
        matched: list[ProductRecord] = []
        ce_catalog: dict[str, CeLabelEntry]
        try:
            ce_catalog = self.engine.load_ce_label_catalog()
        except Exception:
            ce_catalog = {}
        for record in self.engine.records:
            entry = ce_catalog.get(normalize_text(record.stock_code))
            ce_text = " ".join([
                entry.product_name if entry else "",
                entry.mancode if entry else "",
                entry.kelvin if entry else "",
                entry.lumen if entry else "",
            ])
            if (
                normalized in normalize_text(record.family)
                or normalized in normalize_text(record.breakdown)
                or normalized in normalize_text(record.stock_code)
                or normalized in normalize_text(ce_text)
                or any(normalized in normalize_text(f"{key} {value}") for key, value in record.raw_values.items())
            ):
                matched.append(record)
        self.render_ce_products_async(matched[:300], f"CE arama: {query}")

    def render_ce_products_async(self, matched: list[ProductRecord], context: str) -> None:
        if not matched:
            self.clear_ce_grid("CE icin kayit bulunamadi.")
            return
        immediate_items = [(record, [], None) for record in matched]
        self.render_ce_product_items(immediate_items, f"{context} - gorseller arka planda")
        self.run_worker(
            lambda: self.engine.prepare_product_items(matched),
            lambda items: self.render_ce_product_items(items, context),
            "CE urun gorselleri hazirlaniyor...",
        )

    def clear_ce_grid(self, message: str = "") -> None:
        while self.ce_grid.count():
            item = self.ce_grid.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.ce_card_widgets.clear()
        if message:
            label = QLabel(message)
            label.setObjectName("Muted")
            self.ce_grid.addWidget(label, 0, 0)

    def render_ce_product_items(self, items: list[tuple[ProductRecord, list[Path], Path | None]], context: str) -> None:
        self.ce_current_items = items
        self.clear_ce_grid()
        avail = self.width() - 262 - 480 - 36
        columns = max(1, avail // 270)
        for index, (record, images, _folder) in enumerate(items):
            key = normalize_text(record.stock_code)
            selected = key in self.ce_selected_records
            card = CeProductCard(self.engine, record, images, selected, self.show_ce_product, self.toggle_ce_selection)
            card.update_selected(selected, self.ce_selected_quantities.get(key, 0))
            self.ce_grid.addWidget(card, index // columns, index % columns)
            self.ce_card_widgets[key] = card
        self.ce_grid.setRowStretch((len(items) + columns - 1) // columns, 1)
        self.set_status(f"{len(items)} CE urunu listelendi. {context}")

    def show_ce_product(self, record: ProductRecord, images: list[Path] | None = None) -> None:
        image_list = images or self.engine.get_cached_preview_image_list(record.stock_code)
        image = image_list[0] if image_list else self.engine.get_cached_preview_image(record.stock_code)
        pixmap = load_scaled_pixmap(image, QSize(290, 220), "#ffffff")
        if pixmap:
            self.ce_preview.setText("")
            self.ce_preview.setPixmap(pixmap)
        else:
            self.ce_preview.setPixmap(QPixmap())
            self.ce_preview.setText("Urun\ngorseli\nyok")
        entry = self.engine.get_ce_label_entry(record)
        duy_text = self.engine.infer_ce_duy_text(entry, record)
        self.ce_info.setText(
            f"ARGE Kaynak: {self.engine.settings.get('ce_measurement_excel_path', CE_MEASUREMENTS_XLSX_DEFAULT)}\n"
            f"Stok: {record.stock_code}\n"
            f"Urun: {entry.product_name or record.breakdown or '-'}\n"
            f"Model: {entry.model or '-'}\n"
            f"Grup: {entry.product_group or '-'}\n"
            f"Tur: {entry.label_kind or '-'}\n"
            f"ManCode: {entry.mancode or '-'}\n"
            f"Dimlenme: {entry.dimming or '-'}\n"
            f"110V: {entry.voltage_110 or '-'}\n"
            f"220V: {entry.voltage_220 or '-'}\n"
            f"Kelvin: {entry.kelvin or '-'}\n"
            f"Lumen: {entry.lumen or '-'}\n"
            f"Watt: {entry.watt or '-'}\n"
            f"Duy: {duy_text or '-'}"
        )

    def toggle_ce_selection(self, record: ProductRecord) -> None:
        key = normalize_text(record.stock_code)
        current_count = self.ce_selected_quantities.get(key, 0)
        self.set_ce_record_count(record, 0 if current_count else 1)
        self.refresh_ce_selected_list()

    def set_ce_record_count(self, record: ProductRecord, count: int) -> None:
        key = normalize_text(record.stock_code)
        count = max(0, min(999, count))
        if count:
            self.ce_selected_records[key] = record
            self.ce_selected_quantities[key] = count
        else:
            self.ce_selected_records.pop(key, None)
            self.ce_selected_quantities.pop(key, None)
        card = self.ce_card_widgets.get(key)
        if card is not None:
            card.update_selected(count > 0, count)

    def refresh_ce_selected_list(self) -> None:
        self.ce_selected_list.clear()
        total_count = 0
        for key, record in self.ce_selected_records.items():
            quantity = self.ce_selected_quantities.get(key, 1)
            total_count += quantity
            entry = self.engine.get_ce_label_entry(record)
            item = QListWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, key)
            row = QWidget()
            row.setObjectName("SelectionRow")
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(12, 10, 12, 10)
            row_layout.setSpacing(12)
            text_layout = QVBoxLayout()
            text_layout.setContentsMargins(0, 0, 0, 0)
            text_layout.setSpacing(4)
            stock_label = QLabel(record.stock_code)
            stock_label.setObjectName("SelectionStock")
            meta_label = QLabel(entry.product_name or record.breakdown or record.family)
            meta_label.setObjectName("SelectionMeta")
            meta_label.setWordWrap(True)
            detail_label = QLabel(
                f"{entry.label_kind or '-'} | {entry.mancode or '-'} | {entry.kelvin or '-'} | {entry.lumen or '-'} | {entry.watt or '-'}"
            )
            detail_label.setObjectName("SelectionHandle")
            detail_label.setWordWrap(True)
            text_layout.addWidget(stock_label)
            text_layout.addWidget(meta_label)
            text_layout.addWidget(detail_label)

            status_layout = QVBoxLayout()
            status_layout.setContentsMargins(0, 0, 0, 0)
            status_layout.setSpacing(6)
            status_label = QLabel(entry.label_kind or "CE")
            status_label.setObjectName("SuccessPill")
            quantity_label = QLabel(f"x{quantity}")
            quantity_label.setObjectName("CounterPill")
            status_layout.addWidget(status_label, 0, Qt.AlignmentFlag.AlignRight)
            status_layout.addWidget(quantity_label, 0, Qt.AlignmentFlag.AlignRight)
            status_layout.addStretch(1)

            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setSpacing(6)
            minus_btn = QPushButton("-")
            minus_btn.setObjectName("GhostButton")
            minus_btn.setFixedWidth(36)
            minus_btn.clicked.connect(lambda _checked=False, item_key=key: self.change_ce_selection_quantity(item_key, -1))
            plus_btn = QPushButton("+")
            plus_btn.setObjectName("PrimaryButton")
            plus_btn.setFixedWidth(36)
            plus_btn.clicked.connect(lambda _checked=False, item_key=key: self.change_ce_selection_quantity(item_key, 1))
            remove_item_btn = QPushButton("Kaldir")
            remove_item_btn.setObjectName("DangerButton")
            remove_item_btn.setFixedWidth(72)
            remove_item_btn.clicked.connect(lambda _checked=False, item_key=key: self.remove_ce_selection_by_key(item_key))
            action_layout.addWidget(minus_btn)
            action_layout.addWidget(plus_btn)
            action_layout.addWidget(remove_item_btn)

            row_layout.addLayout(text_layout, 1)
            row_layout.addLayout(status_layout)
            row_layout.addLayout(action_layout)
            item.setSizeHint(QSize(0, 108))
            self.ce_selected_list.addItem(item)
            self.ce_selected_list.setItemWidget(item, row)
        self.ce_selected_count_label.setText(f"{total_count} adet")
        self.set_status(f"CE secili urun: {len(self.ce_selected_records)} stok / {total_count} adet")

    def remove_selected_ce_list_item(self) -> None:
        item = self.ce_selected_list.currentItem()
        if item is None:
            return
        key = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(key, str):
            self.remove_ce_selection_by_key(key)

    def remove_ce_selection_by_key(self, key: str) -> None:
        record = self.ce_selected_records.pop(key, None)
        self.ce_selected_quantities.pop(key, None)
        card = self.ce_card_widgets.get(key)
        if card is not None:
            card.update_selected(False, 0)
        if record is not None:
            self.set_status(f"{record.stock_code} CE listesinden kaldirildi.")
        self.refresh_ce_selected_list()

    def clear_ce_selection(self) -> None:
        self.ce_selected_records.clear()
        self.ce_selected_quantities.clear()
        for card in self.ce_card_widgets.values():
            card.update_selected(False, 0)
        self.refresh_ce_selected_list()

    def change_ce_selection_quantity(self, key: str, delta: int) -> None:
        record = self.ce_selected_records.get(key)
        if record is None:
            return
        current_count = self.ce_selected_quantities.get(key, 1)
        self.set_ce_record_count(record, current_count + delta)
        self.refresh_ce_selected_list()

    def show_ce_selected_list_item(self, item: QListWidgetItem) -> None:
        key = item.data(Qt.ItemDataRole.UserRole)
        if isinstance(key, str) and key in self.ce_selected_records:
            self.show_ce_product(self.ce_selected_records[key])

    def get_ce_record_by_stock_code(self, stock_code: str) -> ProductRecord | None:
        key = normalize_text(stock_code)
        if not key:
            return None
        for record in self.engine.records:
            if normalize_text(record.stock_code) == key:
                return record
        return None

    def extract_bulk_ce_stock_codes(self, text: str) -> tuple[list[str], list[str]]:
        found: list[str] = []
        unresolved: list[str] = []
        for raw_line in re.split(r"[\r\n,;\t]+", text):
            line = raw_line.strip().strip("'\"")
            if not line:
                continue
            if normalize_text(line) in {"sku", "stok", "stok kodu", "stock code"}:
                continue
            line_matches: list[str] = []
            direct = self.get_ce_record_by_stock_code(line)
            if direct is not None:
                line_matches.append(direct.stock_code)
            else:
                for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_-]{2,}", line):
                    record = self.get_ce_record_by_stock_code(token)
                    if record is not None:
                        line_matches.append(record.stock_code)
            if line_matches:
                found.extend(line_matches)
            else:
                unresolved.append(line)
        return found, unresolved

    def add_bulk_ce_stock_codes(self) -> None:
        text = self.ce_bulk_stock_text.toPlainText().strip()
        if not text:
            QMessageBox.information(self, "Toplu Ekleme", "Listeye eklemek icin stok kodu yapistir.")
            return
        stock_codes, unresolved = self.extract_bulk_ce_stock_codes(text)
        added = 0
        for stock_code in stock_codes:
            record = self.get_ce_record_by_stock_code(stock_code)
            if record is None:
                unresolved.append(stock_code)
                continue
            key = normalize_text(record.stock_code)
            current_count = self.ce_selected_quantities.get(key, 0)
            self.set_ce_record_count(record, current_count + 1 if current_count else 1)
            added += 1
        self.refresh_ce_selected_list()
        if stock_codes:
            self.ce_bulk_stock_text.clear()
        parts = [f"{added} adet listeye eklendi."]
        if unresolved:
            preview = ", ".join(unresolved[:12])
            more = f" (+{len(unresolved) - 12})" if len(unresolved) > 12 else ""
            parts.append(f"Bulunamayan: {preview}{more}")
        QMessageBox.information(self, "Toplu CE Ekleme", "\n".join(parts))

    def selected_ce_items(self) -> list[tuple[ProductRecord, int]]:
        return [
            (record, max(1, self.ce_selected_quantities.get(key, 1)))
            for key, record in self.ce_selected_records.items()
        ]

    def export_ce_bartender(self) -> None:
        items = self.selected_ce_items()
        if not items:
            QMessageBox.information(self, "Secim Yok", "CE ciktisi icin once urun sec.")
            return
        gap_labels = 0
        if len(items) > 1:
            gap_labels, ok = QInputDialog.getInt(
                self,
                "Bos Etiket",
                "Urunler arasinda kac bos etiket birakmak istersin?",
                0,
                0,
                999,
                1,
            )
            if not ok:
                return
        self.run_worker(
            lambda: self.engine.export_ce_bartender_excel(items, gap_labels),
            lambda path: QMessageBox.information(self, "CE Bartender Excel Hazir", str(path)),
            "CE Bartender Excel olusturuluyor...",
        )

    def print_ce_bartender(self) -> None:
        items = self.selected_ce_items()
        if not items:
            QMessageBox.information(self, "Secim Yok", "Direkt CE baskisi icin once urun sec.")
            return
        gap_labels = 0
        if len(items) > 1:
            gap_labels, ok = QInputDialog.getInt(
                self,
                "Bos Etiket",
                "Urunler arasinda kac bos etiket birakmak istersin?",
                0,
                0,
                999,
                1,
            )
            if not ok:
                return
        total_count = sum(quantity for _record, quantity in items) + max(0, len(items) - 1) * gap_labels
        printer = self.engine.settings.get("bartender_printer_name", "").strip() or "template icindeki yazici"
        reset_notice = (
            "Baski oncesi etiket baslangici sifirlanacak/feed edilecek."
            if self.engine.truthy_setting("label_reset_enabled", "1")
            else "Etiket sifirlama ayarlardan kapali."
        )
        if QMessageBox.question(
            self,
            "Direkt Yazdir",
            f"{total_count} CE etiketi BarTender uzerinden direkt yaziciya gonderilecek.\n"
            f"{reset_notice}\n\n"
            f"Yazici: {printer}\nDevam edelim mi?",
        ) != QMessageBox.StandardButton.Yes:
            return
        self.run_worker(
            lambda: self.engine.print_ce_bartender_excel(items, gap_labels),
            lambda path: QMessageBox.information(self, "Baski Gonderildi", f"CE barkod baskisi BarTender'a gonderildi.\nVeri dosyasi: {path}"),
            "CE barkodlar BarTender ile yazdiriliyor...",
            show_loading_overlay=True,
        )

    def fetch_family_products(self) -> None:
        if self.breakdown_combo.findText(ALL_BREAKDOWNS_LABEL) >= 0:
            self.breakdown_combo.setCurrentText(ALL_BREAKDOWNS_LABEL)
        self.fetch_selected_products()

    def fetch_selected_products(self) -> None:
        family = self.family_combo.currentText().strip()
        breakdown = self.breakdown_combo.currentText().strip()
        if not family:
            QMessageBox.warning(self, "Eksik Secim", "Lutfen aile sec.")
            return
        matched = [
            record
            for record in self.engine.records
            if normalize_text(record.family) == normalize_text(family)
            and (not breakdown or normalize_text(breakdown) == normalize_text(ALL_BREAKDOWNS_LABEL) or normalize_text(record.breakdown) == normalize_text(breakdown))
        ]
        self.render_products_async(matched[:300], f"Secim: {family} / {breakdown}")

    def search_products(self) -> None:
        query = self.search_edit.text().strip()
        if not query:
            QMessageBox.warning(self, "Arama Bos", "Aramak icin bir deger yaz.")
            return
        normalized = normalize_text(query)
        matched = [
            record
            for record in self.engine.records
            if normalized in normalize_text(record.family)
            or normalized in normalize_text(record.breakdown)
            or normalized in normalize_text(record.stock_code)
            or any(normalized in normalize_text(f"{key} {value}") for key, value in record.raw_values.items())
        ]
        self.render_products_async(matched[:300], f"Arama: {query}")

    def render_products_async(self, matched: list[ProductRecord], context: str) -> None:
        if not matched:
            self.clear_product_grid("Kayit bulunamadi.")
            return
        immediate_items = [(record, [], None) for record in matched]
        self.render_product_items(immediate_items, f"{context} - gorseller arka planda araniyor")
        self.run_worker(
            lambda: self.engine.prepare_product_items(matched),
            lambda items: self.render_product_items(items, context),
            "Gorseller ve klasorler arka planda araniyor...",
        )

    def clear_product_grid(self, message: str = "") -> None:
        while self.product_grid.count():
            item = self.product_grid.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        if message:
            label = QLabel(message)
            label.setObjectName("Muted")
            self.product_grid.addWidget(label, 0, 0)

    def render_product_items(self, items: list[tuple[ProductRecord, list[Path], Path | None]], context: str) -> None:
        self.current_items = items
        self.clear_product_grid()
        columns = 3 if self.width() >= 1320 else 2
        for index, (record, images, folder) in enumerate(items):
            card = ProductCard(self.engine, record, images, folder)
            self.product_grid.addWidget(card, index // columns, index % columns)
        self.product_grid.setRowStretch((len(items) + columns - 1) // columns, 1)
        self.set_status(f"{len(items)} urun listelendi. {context}")

    def pick_directory(self, field: QLineEdit) -> None:
        selected = QFileDialog.getExistingDirectory(self, "Klasor sec", field.text())
        if selected:
            field.setText(selected)

    def current_rename_scope_key(self) -> str:
        if not hasattr(self, "rename_scope"):
            return self.engine.settings.get("rename_scope", RENAME_SCOPE_ALL)
        data = self.rename_scope.currentData()
        return str(data or RENAME_SCOPE_ALL)

    def is_flat_rename_output_mode(self) -> bool:
        if hasattr(self, "rename_flat_output"):
            return self.rename_flat_output.isChecked()
        return self.engine.settings.get("rename_flat_output_enabled", self.engine.settings.get("rename_flat_images_enabled", "0")) == "1"

    def on_rename_flat_output_changed(self) -> None:
        enabled = "1" if self.is_flat_rename_output_mode() else "0"
        self.engine.settings["rename_flat_output_enabled"] = enabled
        self.engine.settings["rename_flat_images_enabled"] = enabled
        self.engine.save_settings()
        if not self.rename_plan:
            return
        if self.is_flat_rename_output_mode() != self.rename_plan_flat_output_mode:
            self.set_status("Cikti modu degisti. Yeni hedefleri gormek icin Onizleme'ye tekrar bas.")

    def on_rename_scope_changed(self) -> None:
        scope = self.current_rename_scope_key()
        self.engine.settings["rename_scope"] = scope
        self.engine.save_settings()
        if not self.rename_plan:
            return
        self.apply_rename_scope_selection(refresh=True)

    def get_rename_output_root(self) -> Path | None:
        if not self.rename_export.isChecked():
            return None
        text = self.rename_output.text().strip()
        return Path(text).expanduser() if text else None

    def preview_rename(self) -> None:
        root = Path(self.rename_root.text().strip()).expanduser()
        if not path_is_dir(root):
            QMessageBox.warning(self, "Gecersiz Klasor", "Lutfen gecerli tarama klasoru sec.")
            return
        output_root = self.get_rename_output_root()
        if self.rename_export.isChecked() and output_root is None:
            QMessageBox.warning(self, "Cikis Eksik", "Cikartma icin cikis klasoru sec.")
            return
        self.engine.settings["last_rename_path"] = str(root)
        self.engine.settings["rename_output_path"] = self.rename_output.text().strip()
        self.engine.settings["rename_export_enabled"] = "1" if self.rename_export.isChecked() else "0"
        self.engine.settings["rename_scope"] = self.current_rename_scope_key()
        flat_output_enabled = "1" if self.is_flat_rename_output_mode() else "0"
        self.engine.settings["rename_flat_output_enabled"] = flat_output_enabled
        self.engine.settings["rename_flat_images_enabled"] = flat_output_enabled
        self.engine.save_settings()
        flat_output = self.is_flat_rename_output_mode()
        self.run_worker(lambda: self.engine.build_rename_plan(root, output_root, flat_output=flat_output), self.on_rename_plan_ready, "Klasorler taraniyor...")

    def on_rename_plan_ready(self, plan: list[RenameAction]) -> None:
        for action in plan:
            self.engine.apply_saved_manual_target(action)
        self.rename_plan = plan
        self.rename_plan_flat_output_mode = self.is_flat_rename_output_mode()
        self.apply_rename_scope_selection(refresh=False)
        self.populate_rename_tree()
        self.set_status(f"{len(plan)} rename kaydi hazir. {len(self.selected_rename_actions())} dosya secili.")

    def rename_action_matches_scope(self, action: RenameAction, scope: str) -> bool:
        if scope == RENAME_SCOPE_ALL:
            return True
        if scope not in CHANNEL_FOLDERS:
            return True
        return self.engine.path_matches_channel(action.source, scope) or self.engine.path_matches_channel(action.source.parent, scope)

    def apply_rename_scope_selection(self, refresh: bool = True) -> None:
        scope = self.current_rename_scope_key()
        for action in self.rename_plan:
            key = self.engine.get_rename_action_key(action)
            selectable = self.engine.is_rename_action_selectable(action)
            self.engine.rename_selection[key] = selectable and self.rename_action_matches_scope(action, scope)
        self.engine.save_runtime_state()
        if refresh:
            self.populate_rename_tree()
            selected_count = len(self.selected_rename_actions())
            label = self.rename_scope.currentText() if hasattr(self, "rename_scope") else "Tum uygun gorseller"
            self.set_status(f"{label}: {selected_count} dosya otomatik secildi.")

    def filter_rename_plan(self) -> list[RenameAction]:
        selected = self.rename_filter.currentText() or RENAME_FILTER_OPTIONS[0]
        plan = self.rename_plan
        if selected == "Tum Satirlar":
            return plan
        if selected == "Sadece Secili":
            return [action for action in plan if self.engine.is_rename_action_selectable(action) and self.engine.rename_selection.get(self.engine.get_rename_action_key(action), True)]
        if selected == "Sadece Hazir":
            return [action for action in plan if action.status in {"Hazir", "Cikarilacak"}]
        if selected == "Sadece Cakismalar":
            return [action for action in plan if action.status == "Cakisma" or "cakisma" in normalize_text(action.reason)]
        if selected == "Sadece Manuel":
            return [action for action in plan if normalize_text(action.reason) == "manuel" or self.engine.get_rename_action_key(action) in self.engine.rename_manual_targets]
        if selected == "Sadece Teknik":
            return [action for action in plan if "teknik" in normalize_text(action.reason)]
        return plan

    def populate_rename_tree(self) -> None:
        self._populating_rename = True
        self.rename_tree.clear()
        self.rename_items.clear()
        groups: dict[str, list[RenameAction]] = defaultdict(list)
        for action in self.filter_rename_plan():
            groups[action.group_name or "Diger"].append(action)
        for group, actions in groups.items():
            group_item = QTreeWidgetItem([f"{group} ({len(actions)})", "", "", "", "", "", ""])
            group_item.setFirstColumnSpanned(False)
            group_item.setExpanded(True)
            group_item.setFlags(group_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            group_item.setCheckState(0, Qt.CheckState.Unchecked)
            self.rename_tree.addTopLevelItem(group_item)
            for action in actions:
                key = self.engine.get_rename_action_key(action)
                selectable = self.engine.is_rename_action_selectable(action)
                if key not in self.engine.rename_selection:
                    self.engine.rename_selection[key] = selectable
                item = QTreeWidgetItem([
                    "",
                    action.status,
                    action.stock_code,
                    action.source.name,
                    self.format_rename_target_display(action),
                    action.reason,
                    self.format_relative_rename_path(action.source),
                ])
                flags = item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEditable
                item.setFlags(flags)
                item.setCheckState(0, Qt.CheckState.Checked if selectable and self.engine.rename_selection.get(key, True) else Qt.CheckState.Unchecked)
                if not selectable:
                    item.setDisabled(True)
                group_item.addChild(item)
                self.rename_items[id(item)] = action
            self.sync_rename_group_state(group_item)
        self.rename_tree.expandAll()
        self._populating_rename = False
        self.update_rename_summary()

    def sync_rename_group_state(self, group_item: QTreeWidgetItem) -> None:
        children = [group_item.child(index) for index in range(group_item.childCount())]
        selectable = [child for child in children if not child.isDisabled()]
        checked = [child for child in selectable if child.checkState(0) == Qt.CheckState.Checked]
        if not selectable or not checked:
            state = Qt.CheckState.Unchecked
        elif len(checked) == len(selectable):
            state = Qt.CheckState.Checked
        else:
            state = Qt.CheckState.PartiallyChecked
        old_populating = self._populating_rename
        self._populating_rename = True
        group_item.setCheckState(0, state)
        self._populating_rename = old_populating

    def format_relative_rename_path(self, path: Path) -> str:
        root = self.rename_root.text().strip()
        if not root:
            return str(path)
        try:
            return f".../{path.relative_to(Path(root).expanduser()).as_posix()}"
        except ValueError:
            return str(path)

    def format_rename_target_display(self, action: RenameAction) -> str:
        if action.target is None:
            return "-"
        if action.operation != "export":
            return action.target.name
        output = self.rename_output.text().strip()
        if output:
            try:
                return f".../{action.target.relative_to(Path(output).expanduser()).as_posix()}"
            except ValueError:
                pass
        return str(action.target)

    def on_rename_item_changed(self, item: QTreeWidgetItem, column: int) -> None:
        if self._populating_rename:
            return
        if id(item) not in self.rename_items:
            if column == 0:
                state = item.checkState(0)
                for index in range(item.childCount()):
                    child = item.child(index)
                    if not child.isDisabled():
                        child.setCheckState(0, state)
            return
        action = self.rename_items[id(item)]
        if column == 0:
            self.engine.rename_selection[self.engine.get_rename_action_key(action)] = item.checkState(0) == Qt.CheckState.Checked
            self.engine.save_runtime_state()
            parent = item.parent()
            if parent is not None:
                self.sync_rename_group_state(parent)
        elif column == 4 and action.target is not None:
            self.engine.set_action_target_name(action, item.text(4), save_state=True)
            self._populating_rename = True
            item.setText(1, action.status)
            item.setText(4, self.format_rename_target_display(action))
            item.setText(5, action.reason)
            self._populating_rename = False
        self.update_rename_summary()

    def selected_rename_actions(self) -> list[RenameAction]:
        return [
            action
            for action in self.rename_plan
            if self.engine.is_rename_action_selectable(action)
            and self.engine.rename_selection.get(self.engine.get_rename_action_key(action), False)
        ]

    def select_all_rename(self) -> None:
        for action in self.rename_plan:
            if self.engine.is_rename_action_selectable(action):
                self.engine.rename_selection[self.engine.get_rename_action_key(action)] = True
        self.engine.save_runtime_state()
        self.populate_rename_tree()

    def clear_all_rename(self) -> None:
        for action in self.rename_plan:
            if self.engine.is_rename_action_selectable(action):
                self.engine.rename_selection[self.engine.get_rename_action_key(action)] = False
        self.engine.save_runtime_state()
        self.populate_rename_tree()

    def update_rename_summary(self) -> None:
        total = len(self.rename_plan)
        selected = len(self.selected_rename_actions())
        self.set_status(f"Rename plan: {total} satir, {selected} secili.")

    def on_rename_current_changed(self, current: QTreeWidgetItem | None, _previous: QTreeWidgetItem | None) -> None:
        if current is None or id(current) not in self.rename_items:
            self.rename_preview_image.setPixmap(QPixmap())
            self.rename_preview_image.setText("Secim yok")
            self.rename_preview_text.setText("Listeden bir gorsel sec.")
            return
        action = self.rename_items[id(current)]
        pixmap = load_scaled_pixmap(action.source, QSize(330, 240), "#eef5fb")
        if pixmap:
            self.rename_preview_image.setText("")
            self.rename_preview_image.setPixmap(pixmap)
        else:
            self.rename_preview_image.setPixmap(QPixmap())
            self.rename_preview_image.setText("Onizleme yok")
        self.rename_preview_text.setText(
            f"Mevcut: {action.source.name}\nYeni: {action.target.name if action.target else '-'}\nStok: {action.stock_code}\nIslem: {'Cikartma' if action.operation == 'export' else 'Rename'}\nKonum: {self.format_relative_rename_path(action.source)}"
        )

    def open_selected_rename_file(self) -> None:
        item = self.rename_tree.currentItem()
        if item is None or id(item) not in self.rename_items:
            QMessageBox.warning(self, "Secim Yok", "Bir dosya sec.")
            return
        reveal_path(self.rename_items[id(item)].source, self)

    def apply_rename(self) -> None:
        actions = self.selected_rename_actions()
        if not actions:
            QMessageBox.information(self, "Islem Yok", "Secili dosya yok.")
            return
        duplicates = self.engine.find_duplicate_targets(actions)
        if duplicates:
            QMessageBox.warning(self, "Hedef Cakismasi", "Ayni hedefe giden dosyalar var:\n\n" + "\n".join(duplicates[:8]))
            return
        if QMessageBox.question(self, "Onay", f"{len(actions)} dosya uygulanacak. Devam edelim mi?") != QMessageBox.StandardButton.Yes:
            return
        self.run_worker(lambda: self.engine.execute_rename_actions(actions), lambda count: (QMessageBox.information(self, "Tamamlandi", f"{count} dosya uygulandi."), self.preview_rename()), "Dosyalar uygulanÄ±yor...")

    def search_ydk(self) -> None:
        query = normalize_text(self.ydk_search.text())
        products = sorted(self.engine.ydk_products.values(), key=lambda item: item.code.casefold())
        if query:
            products = [
                product
                for product in products
                if query in normalize_text(" ".join([product.code, product.model, product.producer_code, product.description_tr, product.description_en, product.unit_barcode, product.carton_barcode]))
            ]
        self.populate_ydk_table(products[:500])

    def populate_ydk_table(self, products: list[YdkProduct]) -> None:
        self.ydk_products_list = products
        self.ydk_table.setRowCount(len(products))
        for row, product in enumerate(products):
            for col, value in enumerate([product.code, product.model, product.producer_code, product.unit_barcode or "-"]):
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.ydk_table.setItem(row, col, item)
        if products:
            self.ydk_table.selectRow(0)
            self.show_ydk_product(0, force=True)
        else:
            self.current_ydk_product = None
            self.ydk_image.setPixmap(QPixmap())
            self.ydk_image.setText("Gorsel yok")
            self.ydk_summary.clear()

    def on_ydk_selection_changed(self) -> None:
        rows = self.ydk_table.selectionModel().selectedRows() if self.ydk_table.selectionModel() else []
        if not rows:
            return
        self.show_ydk_product(rows[0].row())

    def show_ydk_product(self, row: int, force: bool = False) -> None:
        if row < 0 or row >= len(getattr(self, "ydk_products_list", [])):
            return
        product = self.ydk_products_list[row]
        current = getattr(self, "current_ydk_product", None)
        if not force and current is not None and normalize_text(current.code) == normalize_text(product.code):
            return
        self.current_ydk_product = product
        if self.ydk_table.currentRow() != row:
            self.ydk_table.selectRow(row)
        image_path = self.engine.get_cached_preview_image(product.code)
        if image_path is None:
            self.ydk_image.setPixmap(QPixmap())
            self.ydk_image.setText("Gorsel araniyor...")
            self.fetch_ydk_image_async(product)
        else:
            self.render_ydk_product_image(image_path)
        self.ydk_summary.setPlainText(self.engine.build_ydk_summary_text(product, image_path))

    def render_ydk_product_image(self, image_path: Path | None) -> None:
        pixmap = load_scaled_pixmap(image_path, QSize(430, 220), "#eef5fb")
        if pixmap:
            self.ydk_image.setText("")
            self.ydk_image.setPixmap(pixmap)
        else:
            self.ydk_image.setPixmap(QPixmap())
            self.ydk_image.setText("Gorsel yok")

    def fetch_ydk_image_async(self, product: YdkProduct) -> None:
        stock_key = normalize_text(product.code)
        if not stock_key or stock_key in self.pending_ydk_images:
            return
        self.pending_ydk_images.add(stock_key)

        def job() -> Path | None:
            try:
                return self.engine.find_ydk_product_image(product.code)
            except Exception as exc:  # noqa: BLE001
                self.write_log(f"YDK image lookup error for {product.code}: {exc}")
                return None

        def on_success(image_path: Path | None) -> None:
            self.pending_ydk_images.discard(stock_key)
            current = getattr(self, "current_ydk_product", None)
            if current is None or normalize_text(current.code) != stock_key:
                return
            self.render_ydk_product_image(image_path)
            self.ydk_summary.setPlainText(self.engine.build_ydk_summary_text(current, image_path))
            self.set_status("Etiket gorseli yuklendi." if image_path else "Etiket gorseli bulunamadi.")

        self.run_worker(job, on_success, "Etiket gorseli arka planda araniyor...")

    def export_current_ydk(self, label_type: str) -> None:
        product = getattr(self, "current_ydk_product", None)
        if product is None:
            QMessageBox.warning(self, "Secim Yok", "Once bir etiket urunu sec.")
            return
        self.run_worker(
            lambda: self.engine.export_ydk_label_pdf(product, label_type),
            lambda path: QMessageBox.information(self, "PDF Kaydedildi", str(path)),
            "YDK PDF hazirlaniyor...",
        )

    def export_both_current_ydk(self) -> None:
        product = getattr(self, "current_ydk_product", None)
        if product is None:
            QMessageBox.warning(self, "Secim Yok", "Once bir etiket urunu sec.")
            return

        def job() -> list[Path]:
            saved = [self.engine.export_ydk_label_pdf(product, "unit")]
            if ean13_full_code(product.carton_barcode):
                saved.append(self.engine.export_ydk_label_pdf(product, "carton"))
            return saved

        self.run_worker(
            job,
            lambda paths: QMessageBox.information(self, "PDF Kaydedildi", "\n".join(str(path) for path in paths)),
            "YDK PDF'leri hazirlaniyor...",
        )


def main() -> None:
    app = QApplication(sys.argv)
    window = ModernMainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()





