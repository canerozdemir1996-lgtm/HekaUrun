from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import tkinter as tk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageTk


APP_NAME = "UrunYonetimMasasi_v3"
APP_TITLE = "Urun Yonetim Masasi v3"

APP_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)


def choose_data_dir() -> Path:
    candidates: list[Path] = []
    appdata = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
    if sys.platform == "darwin":
        candidates.append(Path.home() / "Library" / "Application Support" / APP_NAME)
    else:
        candidates.append(APP_DIR)
    if appdata and sys.platform != "darwin":
        candidates.append(Path(appdata) / APP_NAME)
    xdg_data = os.environ.get("XDG_DATA_HOME")
    if xdg_data and sys.platform not in ("win32", "darwin"):
        candidates.append(Path(xdg_data) / APP_NAME)
    candidates.append(Path.home() / ".urun_yonetim_masasi_v3")
    for candidate in candidates:
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            probe = candidate / ".write_test"
            probe.write_text("", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return candidate
        except OSError:
            continue
    return APP_DIR


DATA_DIR = choose_data_dir()
SETTINGS_FILE = DATA_DIR / "settings.json"
INDEX_FILE = DATA_DIR / "product_index.sqlite"
RENAME_LOG_FILE = DATA_DIR / "rename_log.jsonl"
ASSETS_DIR = APP_DIR / "assets"
APP_ICON_PNG = ASSETS_DIR / "app_icon.png"
APP_ICON_ICO = ASSETS_DIR / "app_icon.ico"
APP_LOGO_PNG = ASSETS_DIR / "logo_white.png"
YDK_LOGO_PNG = ASSETS_DIR / "ydk_logo.png"
YDK_LABEL_LOGO_PNG = ASSETS_DIR / "ydk_label_logo.png"
YDK_LABEL_ICON_PNG = ASSETS_DIR / "ydk_label_icon.png"
YDK_LABEL_SIZE = (717, 858)
DEFAULT_YDK_LABEL_LAYOUT = {
    "logo_x": 0,
    "logo_y": 0,
    "logo_w": 350,
    "logo_h": 145,
    "photo_x": 360,
    "photo_y": 0,
    "photo_w": 330,
    "photo_h": 330,
    "model_x": 28,
    "model_y": 190,
    "model_font": 40,
    "tr_x": 31,
    "tr_y": 312,
    "tr_font": 32,
    "en_x": 31,
    "en_y": 431,
    "en_font": 29,
    "barcode_x": 0,
    "barcode_y": 562,
    "barcode_w": 390,
    "barcode_h": 138,
    "info_x": 407,
    "heading_y": 582,
    "code_y": 630,
    "producer_box_x": 398,
    "producer_box_y": 680,
    "producer_box_w": 314,
    "producer_box_h": 78,
    "producer_font": 38,
    "footer_y": 788,
    "footer_font": 24,
}
LOGODATA_SQL_PROFILE_VERSION = "LOGODATA_HRMS_122_2026_05_20_SAVED_CONNECTION"
LOGODATA_SQL_PASSWORD = os.environ.get("HEKA_LOGODATA_SQL_PASSWORD", "")
LOGODATA_SQL_DEFAULTS = {
    "data_source": "sql",
    "sql_connection_string": f"DRIVER={{SQL Server}};SERVER=192.168.10.12;UID=sa;PWD={LOGODATA_SQL_PASSWORD};APP=Microsoft Office;WSID=NOTEBOOKX;DATABASE=LOGODATA",
    "sql_table": "LOGODATA.dbo.HRMS_122_MALZEMEBILGILERI",
    "sql_query": "SELECT *\nFROM LOGODATA.dbo.HRMS_122_MALZEMEBILGILERI HRMS_122_MALZEMEBILGILERI",
    "family_column": "OZELKOD3",
    "breakdown_column": "ADI",
    "stock_column": "KODU",
    "feature_columns": "",
    "sheet_name": "",
    "header_row": "1",
}

DEFAULT_SETTINGS = {
    "sql_auto_profile_version": LOGODATA_SQL_PROFILE_VERSION,
    "data_source": LOGODATA_SQL_DEFAULTS["data_source"],
    "excel_path": "",
    "sql_connection_string": LOGODATA_SQL_DEFAULTS["sql_connection_string"],
    "sql_table": LOGODATA_SQL_DEFAULTS["sql_table"],
    "sql_query": LOGODATA_SQL_DEFAULTS["sql_query"],
    "sheet_name": LOGODATA_SQL_DEFAULTS["sheet_name"],
    "header_row": LOGODATA_SQL_DEFAULTS["header_row"],
    "family_column": LOGODATA_SQL_DEFAULTS["family_column"],
    "breakdown_column": LOGODATA_SQL_DEFAULTS["breakdown_column"],
    "stock_column": LOGODATA_SQL_DEFAULTS["stock_column"],
    "feature_columns": LOGODATA_SQL_DEFAULTS["feature_columns"],
    "feature_aliases": "",
    "search_root": "O:/HEKA GÜNCEL ÜRÜNLER 2026",
    "preview_image_root": "O:/HEKA/ÜRÜN GÖRSELLERİ",
    "image_extensions": ".jpg,.jpeg,.png,.webp,.bmp",
    "stock_regex": r"^(?=[A-Z0-9-]*[A-Z])(?=[A-Z0-9-]*\d)([A-Z0-9]+(?:-[A-Z0-9]+)*)\b",
    "window_geometry": "",
    "last_family": "",
    "last_breakdown": "",
    "last_rename_path": "",
    "rename_output_path": "",
    "rename_export_enabled": "0",
    "rename_scope": "all",
    "rename_flat_images_enabled": "0",
    "rename_flat_output_enabled": "0",
    "rename_selection_state": "{}",
    "rename_manual_targets": "{}",
    "open_keywords": "acik",
    "closed_keywords": "kapali",
    "technical_keywords": "teknik,teknik cizim,technical,cizim,olcu",
    "ydk_image_root": "O:/HEKA/ÜRÜN GÖRSELLERİ",
    "ydk_output_root": "L:/04_STUDYO/004_YURTDISI ETIKET",
    "ydk_label_overrides": "{}",
    "ydk_label_layout": json.dumps(DEFAULT_YDK_LABEL_LAYOUT, ensure_ascii=False),
    "bartender_exe_path": "C:/Program Files/Seagull/BarTender 11.8/BarTend.exe",
    "bartender_printer_name": "Argox CP-3140EX PPLB",
    "ce_bartender_template_path": "K:/HEKA/000_HEKATasarımlar/009_Üretim Tasarımları/CE BARKOD/Barkodlar/Excel DEneme.btw",
    "qr_bartender_template_path": "K:/HEKA/000_HEKATasarımlar/009_Üretim Tasarımları/QR KOD BARKOD/QR.btw",
    "bartender_print_timeout_seconds": "180",
    "label_reset_enabled": "1",
    "label_reset_before_each_label": "0",
    "label_reset_feed_count": "1",
    "label_reset_pause_ms": "1200",
    "label_reset_raw_command": "\\x0c",
    "ui_theme": "light",
}

ALL_BREAKDOWNS_LABEL = "Tum Kirilimlar"
DATA_SOURCE_OPTIONS = {
    "Excel": "excel",
    "SQL": "sql",
}
CHANNEL_FOLDERS = {
    "b2b": ("B2B", ("b2b",)),
    "web": ("WEB", ("web", "katalog_web", "etsy")),
    "instagram": ("INSTAGRAM", ("instagram", "insta")),
    "beymen": ("BEYMEN", ("beymen",)),
    "technical": ("TEKNIK", ("teknik", "teknik cizim", "technical", "cizim", "olcu")),
}
RENAME_FILTER_OPTIONS = (
    "Tum Satirlar",
    "Sadece Secili",
    "Sadece Hazir",
    "Sadece Cakismalar",
    "Sadece Manuel",
    "Sadece Teknik",
)


@dataclass
class ProductRecord:
    family: str
    breakdown: str
    stock_code: str
    features: dict[str, str]
    raw_values: dict[str, str]


@dataclass
class RenameAction:
    source: Path
    target: Path | None
    reason: str
    stock_code: str
    status: str
    group_name: str = ""
    operation: str = "rename"


@dataclass
class YdkProduct:
    code: str
    description_tr: str = ""
    description_en: str = ""
    model: str = ""
    producer_code: str = ""
    brand: str = ""
    unit: str = ""
    product_type: str = ""
    unit_barcode: str = ""
    carton_barcode: str = ""
    carton_quantity: str = ""
    image_url: str = ""


def normalize_text(value: str) -> str:
    lowered = str(value or "").strip().casefold().translate(
        str.maketrans(
            {
                "ı": "i",
                "İ": "i",
                "ş": "s",
                "Ş": "s",
                "ğ": "g",
                "Ğ": "g",
                "ü": "u",
                "Ü": "u",
                "ö": "o",
                "Ö": "o",
                "ç": "c",
                "Ç": "c",
            }
        )
    )
    normalized = unicodedata.normalize("NFKD", lowered)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def parse_csv(value: str) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def display_data_source(value: str) -> str:
    key = normalize_data_source(value)
    return next((label for label, mapped in DATA_SOURCE_OPTIONS.items() if mapped == key), "Excel")


def normalize_data_source(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "excel"
    lowered = text.casefold()
    for label, mapped in DATA_SOURCE_OPTIONS.items():
        if lowered in {label.casefold(), mapped.casefold()}:
            return mapped
    return "sql" if lowered == "sql" else "excel"


def apply_logodata_sql_profile(settings: dict[str, str]) -> dict[str, str]:
    if settings.get("sql_auto_profile_version") == LOGODATA_SQL_PROFILE_VERSION:
        return settings
    updated = settings.copy()
    updated.update(LOGODATA_SQL_DEFAULTS)
    updated["sql_auto_profile_version"] = LOGODATA_SQL_PROFILE_VERSION
    return updated


def split_multivalue_config(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[\n,;]+", str(value or "")) if item.strip()]


def stringify_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def ensure_prefixed_extensions(values: list[str]) -> set[str]:
    result: set[str] = set()
    for value in values:
        ext = value.strip().lower()
        if not ext:
            continue
        result.add(ext if ext.startswith(".") else f".{ext}")
    return result


def unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = normalize_text(value)
        if key and key not in seen:
            seen.add(key)
            result.append(value)
    return result


def unique_paths(values: list[Path]) -> list[Path]:
    seen: set[str] = set()
    result: list[Path] = []
    for value in values:
        key = str(value).casefold()
        if key not in seen:
            seen.add(key)
            result.append(value)
    return result


def column_letters_to_index(value: str) -> int | None:
    letters = value.strip().upper()
    if not letters.isalpha():
        return None
    index = 0
    for char in letters:
        index = index * 26 + ord(char) - 64
    return index - 1


def resolve_column_index(spec: str, headers: list[str]) -> int | None:
    cleaned = str(spec or "").strip()
    if not cleaned:
        return None
    if cleaned.isdigit():
        index = int(cleaned) - 1
        return index if 0 <= index < len(headers) else None
    letter_index = column_letters_to_index(cleaned)
    if letter_index is not None and letter_index < len(headers):
        return letter_index
    normalized = normalize_text(cleaned)
    for index, header in enumerate(headers):
        if normalize_text(header) == normalized:
            return index
    return None


def resolve_header_first_index(spec: str, headers: list[str]) -> int | None:
    normalized = normalize_text(spec)
    for index, header in enumerate(headers):
        if normalize_text(header) == normalized:
            return index
    return resolve_column_index(spec, headers)


def prettify_feature_name(value: str) -> str:
    text = stringify_cell(value)
    return re.sub(r"[_\s]+", " ", text).strip() if text else ""


def is_generic_header(value: str) -> bool:
    normalized = normalize_text(value)
    return not normalized or normalized.startswith("sutun ") or normalized.isdigit()


def safe_path_part(value: str, fallback: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    return cleaned or fallback


def safe_resolve(path: Path) -> Path:
    try:
        return path.expanduser().resolve()
    except (OSError, RuntimeError):
        return path.expanduser()


def path_exists(path: Path) -> bool:
    try:
        return path.expanduser().exists()
    except (OSError, RuntimeError):
        return False


def path_is_file(path: Path) -> bool:
    try:
        return path.expanduser().is_file()
    except (OSError, RuntimeError):
        return False


def path_is_dir(path: Path) -> bool:
    try:
        return path.expanduser().is_dir()
    except (OSError, RuntimeError):
        return False


def safe_walk(root: Path):
    try:
        yield from os.walk(root, onerror=lambda _error: None)
    except (OSError, RuntimeError):
        return


def safe_rglob_files(root: Path):
    try:
        for path in root.rglob("*"):
            if path_is_file(path):
                yield path
    except (OSError, RuntimeError):
        return


def digits_only(value: str) -> str:
    return "".join(char for char in str(value or "") if char.isdigit())


def ean13_full_code(value: str) -> str:
    digits = digits_only(value)
    if len(digits) >= 13:
        return digits[:13]
    if len(digits) == 12:
        return digits + str(ean13_check_digit(digits))
    return ""


def ean13_check_digit(first_12: str) -> int:
    digits = digits_only(first_12)[:12]
    if len(digits) != 12:
        return 0
    checksum = sum(int(digits[index]) for index in range(1, 12, 2)) * 3
    checksum += sum(int(digits[index]) for index in range(0, 11, 2))
    return (10 - checksum % 10) % 10


def ean13_font_text(first_12_or_13: str) -> str:
    digits = digits_only(first_12_or_13)
    if len(digits) >= 13:
        digits = digits[:12]
    if len(digits) != 12:
        return ""
    full = ean13_full_code(digits)
    first = int(full[0])
    result = full[0] + chr(65 + int(full[1]))
    table_a_by_position = {
        2: {0, 1, 2, 3},
        3: {0, 4, 7, 8},
        4: {0, 1, 4, 5, 9},
        5: {0, 2, 5, 6, 7},
        6: {0, 3, 6, 8, 9},
    }
    for index in range(2, 7):
        digit = int(full[index])
        result += chr((65 if first in table_a_by_position[index] else 75) + digit)
    result += "*"
    for index in range(7, 13):
        result += chr(97 + int(full[index]))
    return result + "+"


def first_matching_value(
    values: dict[str, str],
    patterns: list[tuple[str, ...]],
    excludes: tuple[str, ...] = (),
) -> str:
    normalized_excludes = tuple(normalize_text(item) for item in excludes)
    for key, value in values.items():
        text = stringify_cell(value)
        if not text:
            continue
        normalized_key = normalize_text(key)
        if any(excluded and excluded in normalized_key for excluded in normalized_excludes):
            continue
        for pattern in patterns:
            if all(normalize_text(part) in normalized_key for part in pattern):
                return text
    return ""


def first_value_matching_regex(values: dict[str, str], pattern: str) -> str:
    expression = re.compile(pattern, re.IGNORECASE)
    for value in values.values():
        text = stringify_cell(value)
        if expression.search(text):
            return text
    return ""


def first_ean_value(
    values: dict[str, str],
    prefer_patterns: list[tuple[str, ...]],
    excludes: tuple[str, ...] = (),
) -> str:
    preferred = first_matching_value(values, prefer_patterns, excludes)
    if ean13_full_code(preferred):
        return ean13_full_code(preferred)
    normalized_excludes = tuple(normalize_text(item) for item in excludes)
    for key, value in values.items():
        normalized_key = normalize_text(key)
        if any(excluded and excluded in normalized_key for excluded in normalized_excludes):
            continue
        code = ean13_full_code(value)
        if code:
            return code
    return ""


def open_path_with_default_app(target: Path) -> None:
    target = safe_resolve(target)
    if not path_exists(target):
        raise FileNotFoundError(f"Konum bulunamadi: {target}")
    if sys.platform.startswith("win"):
        os.startfile(str(target))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(target)])
    else:
        subprocess.Popen(["xdg-open", str(target)])


def open_in_explorer(target: Path) -> None:
    target = safe_resolve(target)
    if not path_exists(target):
        raise FileNotFoundError(f"Konum bulunamadi: {target}")
    if sys.platform.startswith("win"):
        if path_is_file(target):
            subprocess.Popen(["explorer.exe", f'/select,"{target}"'])
        else:
            os.startfile(str(target))
    elif sys.platform == "darwin":
        if path_is_file(target):
            subprocess.Popen(["open", "-R", str(target)])
        else:
            subprocess.Popen(["open", str(target)])
    else:
        subprocess.Popen(["xdg-open", str(target.parent if path_is_file(target) else target)])


def enable_mousewheel_scroll(
    trigger_widget: tk.Widget,
    y_target: tk.Widget,
    x_target: tk.Widget | None = None,
) -> None:
    def scroll_target(event: tk.Event) -> str:
        use_horizontal = bool(getattr(event, "state", 0) & 0x0001) and x_target is not None
        target = x_target if use_horizontal else y_target
        command = target.xview_scroll if use_horizontal else target.yview_scroll

        if getattr(event, "num", None) == 4:
            units = -3
        elif getattr(event, "num", None) == 5:
            units = 3
        else:
            delta = getattr(event, "delta", 0)
            units = int(-delta / 120) if delta else 0
            if units == 0 and delta:
                units = -1 if delta > 0 else 1
        command(units, "units")
        return "break"

    def bind_scroll(_event: tk.Event | None = None) -> None:
        trigger_widget.bind_all("<MouseWheel>", scroll_target)
        trigger_widget.bind_all("<Shift-MouseWheel>", scroll_target)
        trigger_widget.bind_all("<Button-4>", scroll_target)
        trigger_widget.bind_all("<Button-5>", scroll_target)

    def unbind_scroll(_event: tk.Event | None = None) -> None:
        trigger_widget.unbind_all("<MouseWheel>")
        trigger_widget.unbind_all("<Shift-MouseWheel>")
        trigger_widget.unbind_all("<Button-4>")
        trigger_widget.unbind_all("<Button-5>")

    trigger_widget.bind("<Enter>", bind_scroll, add="+")
    trigger_widget.bind("<Leave>", unbind_scroll, add="+")


def load_excel_headers(excel_path: Path, sheet_name: str, header_row: int) -> list[str]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        values = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        return [
            stringify_cell(value) or f"Sutun {index + 1}"
            for index, value in enumerate(values)
        ]
    finally:
        workbook.close()


class SearchableCombobox(ttk.Combobox):
    def __init__(self, master: tk.Widget, textvariable: tk.StringVar, **kwargs: Any):
        super().__init__(master, textvariable=textvariable, state="normal", **kwargs)
        self._all_values: list[str] = []
        self.bind("<KeyRelease>", self._filter_values)
        self.bind("<FocusIn>", self._restore_values)

    def set_values(self, values: list[str]) -> None:
        self._all_values = values
        self.configure(values=values)

    def _filter_values(self, _event: tk.Event) -> None:
        query = normalize_text(self.get())
        if not query:
            self.configure(values=self._all_values)
            return
        self.configure(
            values=[value for value in self._all_values if query in normalize_text(value)]
        )

    def _restore_values(self, _event: tk.Event) -> None:
        self.configure(values=self._all_values)


class SettingsDialog(tk.Toplevel):
    FIELDS = [
        ("data_source", "Veri Kaynagi", "choice"),
        ("excel_path", "Excel Dosyasi", "file"),
        ("sql_connection_string", "SQL Baglanti", "multi"),
        ("sql_table", "SQL Tablo veya Gorunum", "text"),
        ("sql_query", "SQL Sorgu (Opsiyonel)", "multi"),
        ("sheet_name", "Sayfa Adi", "text"),
        ("header_row", "Baslik Satiri", "text"),
        ("family_column", "Urun Ailesi Sutunu", "text"),
        ("breakdown_column", "Kirilim Sutunu", "text"),
        ("stock_column", "Stok Kodu Sutunu", "text"),
        ("feature_columns", "Ozellik Sutunlari", "text"),
        ("feature_aliases", "Ozellik Baslik Eslestirme", "multi"),
        ("search_root", "Explorer Arama Ana Klasoru", "dir"),
        ("preview_image_root", "Onizleme Fotograf Klasoru", "dir"),
        ("image_extensions", "Gorsel Uzantilari", "text"),
        ("stock_regex", "Stok Kodu Regex", "text"),
        ("open_keywords", "Acik Anahtar Kelimeleri", "text"),
        ("closed_keywords", "Kapali Anahtar Kelimeleri", "text"),
        ("technical_keywords", "Teknik Cizim Anahtar Kelimeleri", "text"),
        ("ydk_image_root", "YDK Urun Gorsel Klasoru", "dir"),
        ("ydk_output_root", "YDK PDF Cikis Klasoru", "dir"),
    ]

    def __init__(self, master: tk.Tk, settings: dict[str, str]):
        super().__init__(master)
        self.title("Ayarlar — Urun Yonetim Masasi")
        self.geometry("860x860")
        self.transient(master)
        self.grab_set()
        self.configure(bg="#F0F4F8")
        self.result: dict[str, str] | None = None
        self.variables = {key: tk.StringVar(value=settings.get(key, "")) for key, _, _ in self.FIELDS}
        self.variables["data_source"].set(display_data_source(settings.get("data_source", "excel")))
        self.text_widgets: dict[str, tk.Text] = {}
        self._build()

    def _build(self) -> None:
        # ── Dialog header ────────────────────────────────────────────────────────
        header = tk.Frame(self, bg="#0D1B2A", padx=24, pady=16)
        header.pack(fill="x")
        tk.Label(header, text="Ayarlar", bg="#0D1B2A", fg="#FFFFFF",
                 font=("Bahnschrift", 18)).pack(side="left")
        tk.Frame(header, height=2, bg="#FF6B35").pack(fill="x", side="bottom")

        # ── Scrollable body ──────────────────────────────────────────────────────
        shell = ttk.Frame(self, padding=20)
        shell.pack(fill="both", expand=True)
        canvas = tk.Canvas(shell, highlightthickness=0, bg="#F0F4F8")
        scroll = ttk.Scrollbar(shell, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>",
                  lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="left", fill="y")
        enable_mousewheel_scroll(canvas, canvas)
        enable_mousewheel_scroll(body, canvas)
        body.columnconfigure(1, weight=1)

        for row, (key, label, kind) in enumerate(self.FIELDS):
            ttk.Label(body, text=label,
                      style="SectionLabel.TLabel").grid(
                row=row, column=0, sticky="nw", pady=(8, 2), padx=(0, 12))
            if kind == "multi":
                text = tk.Text(body, width=62, height=5,
                               bg="#FFFFFF", fg="#0F172A",
                               font=("Aptos", 10),
                               relief="solid", bd=1)
                text.insert("1.0", self.variables[key].get())
                text.grid(row=row, column=1, sticky="ew", pady=(8, 2))
                self.text_widgets[key] = text
            elif kind == "choice":
                values = list(DATA_SOURCE_OPTIONS)
                combo = ttk.Combobox(body, textvariable=self.variables[key],
                                     values=values, state="readonly", width=30)
                combo.grid(row=row, column=1, sticky="w", pady=(8, 2))
            else:
                entry = ttk.Entry(body, textvariable=self.variables[key],
                                  width=64)
                entry.grid(row=row, column=1, sticky="ew", pady=(8, 2))
                if kind in {"file", "dir"}:
                    ttk.Button(
                        body, text="Sec",
                        command=lambda k=key, t=kind: self._browse(k, t),
                    ).grid(row=row, column=2, padx=(8, 0), pady=(8, 2))

        # ── Footer ────────────────────────────────────────────────────────────────
        tk.Frame(self, height=1, bg="#E2E8F0").pack(fill="x")
        footer = ttk.Frame(self, padding=(20, 12))
        footer.pack(fill="x")
        ttk.Button(footer, text="↩  Varsayilan",
                   command=self._defaults).pack(side="left")
        ttk.Button(footer, text="Iptal",
                   command=self.destroy).pack(side="right")
        ttk.Button(footer, text="✓  Kaydet",
                   style="Accent.TButton",
                   command=self._save).pack(side="right", padx=(0, 8))

    def _browse(self, key: str, kind: str) -> None:
        selected = (
            filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("Tum Dosyalar", "*.*")])
            if kind == "file"
            else filedialog.askdirectory()
        )
        if selected:
            self.variables[key].set(selected)

    def _defaults(self) -> None:
        for key, _label, _kind in self.FIELDS:
            default_value = DEFAULT_SETTINGS.get(key, "")
            self.variables[key].set(display_data_source(default_value) if key == "data_source" else default_value)
            if key in self.text_widgets:
                self.text_widgets[key].delete("1.0", "end")
                self.text_widgets[key].insert("1.0", default_value)

    def _save(self) -> None:
        for key, widget in self.text_widgets.items():
            self.variables[key].set(widget.get("1.0", "end").strip())
        header_row = self.variables["header_row"].get().strip() or "1"
        if not header_row.isdigit() or int(header_row) < 1:
            messagebox.showwarning("Gecersiz Deger", "Baslik satiri 1 veya daha buyuk olmali.", parent=self)
            return
        self.result = {key: var.get().strip() for key, var in self.variables.items()}
        self.result["data_source"] = normalize_data_source(self.result.get("data_source", "excel"))
        self.destroy()


class ProductDesktopApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.app_icon_ref: tk.PhotoImage | None = None
        self.ui_logo_ref: ImageTk.PhotoImage | None = None
        self.configure_app_icon()
        self.settings = self.load_settings()
        try:
            self.geometry(self.settings.get("window_geometry") or "1440x920")
        except tk.TclError:
            self.geometry("1440x920")
        self.minsize(1200, 780)

        self.records: list[ProductRecord] = []
        self.search_folder_cache: dict[str, Path | None] = {}
        self.channel_folder_cache: dict[str, Path | None] = {}
        self.preview_image_cache: dict[str, Path | None] = {}
        self.preview_search_roots_cache: list[Path] | None = None
        self.preview_thumbnail_cache: dict[str, ImageTk.PhotoImage | None] = {}
        self.result_image_refs: list[ImageTk.PhotoImage] = []
        self.ydk_products: dict[str, YdkProduct] = {}
        self.ydk_current_product: YdkProduct | None = None
        self.ydk_result_products: list[YdkProduct] = []
        self.ydk_preview_photo_ref: ImageTk.PhotoImage | None = None
        self.ydk_product_photo_ref: ImageTk.PhotoImage | None = None
        self.ydk_logo_image: Image.Image | None = None
        self.ydk_icon_image: Image.Image | None = None
        self.ydk_label_overrides = self.load_json_setting("ydk_label_overrides", {})
        self.ydk_editor_vars: dict[str, tk.StringVar] = {}
        self.ydk_editor_texts: dict[str, tk.Text] = {}
        self.ydk_layout_vars: dict[str, tk.StringVar] = {}
        self.ydk_editor_loading = False
        self.rename_plan: list[RenameAction] = []
        self.product_data_warning = ""
        self.rename_selection = self.load_json_setting("rename_selection_state", {})
        self.rename_manual_targets = self.load_json_setting("rename_manual_targets", {})
        self.rename_item_actions: dict[str, RenameAction] = {}
        self.rename_group_items: dict[str, list[str]] = {}
        self.rename_edit_widget: ttk.Entry | None = None
        self.rename_autosize_after_id: str | None = None
        self.rename_preview_image_ref: ImageTk.PhotoImage | None = None

        self.status_var = tk.StringVar(value="Hazir.")
        self.settings_summary_var = tk.StringVar()
        self.results_summary_var = tk.StringVar(value="Veri bekleniyor")
        self.results_context_var = tk.StringVar(value="Veri kaynagi yuklenince urunler burada listelenir.")
        self.family_var = tk.StringVar()
        self.breakdown_var = tk.StringVar()
        self.product_search_var = tk.StringVar()
        self.rename_path_var = tk.StringVar(value=self.settings.get("last_rename_path", ""))
        self.rename_output_path_var = tk.StringVar(value=self.settings.get("rename_output_path", ""))
        self.rename_export_enabled_var = tk.BooleanVar(
            value=self.settings.get("rename_export_enabled", "0") == "1"
        )
        self.rename_filter_var = tk.StringVar(value=RENAME_FILTER_OPTIONS[0])
        self.rename_preview_text_var = tk.StringVar(value="Listeden bir gorsel sec.")
        self.ydk_image_root_var = tk.StringVar(value=self.settings.get("ydk_image_root", ""))
        self.ydk_output_root_var = tk.StringVar(value=self.settings.get("ydk_output_root", ""))
        self.ydk_search_var = tk.StringVar()
        self.ydk_detail_var = tk.StringVar(value="Ana veri kaynagi yuklenince etiket bilgileri burada gorunur.")

        self._build_styles()
        self.ensure_index_schema()
        self._build_menu()
        self._build_layout()
        self.refresh_settings_summary()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.refresh_product_data(show_success=False)

    def _build_styles(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")

        # ── Design tokens (inspired by 21st.dev / magic MCP premium dashboard) ──
        self.ui = {
            # Main canvas
            "bg":           "#F0F4F8",   # cool grey canvas
            "surface":      "#FFFFFF",   # pure white panels
            "surface_2":    "#F8FAFC",   # off-white secondary
            "card":         "#FFFFFF",
            # Text
            "ink":          "#0F172A",   # deep slate
            "muted":        "#64748B",   # slate-500
            "line":         "#E2E8F0",   # slate-200 divider
            # Brand hero strip
            "hero":         "#0D1B2A",   # deep navy
            "hero_2":       "#132335",
            # Accent – vibrant orange (magic MCP: #FF6B35)
            "accent":       "#FF6B35",
            "accent_hot":   "#FF8552",
            "accent_deep":  "#E0511F",
            # Sidebar
            "sidebar":      "#0D1B2A",   # same deep navy
            "sidebar_2":    "#1A2F44",   # hover fill
            "sidebar_muted":"#94A3B8",   # slate-400
            # Semantic
            "mint":         "#10B981",   # emerald-500 success
            "selected":     "#DBEAFE",   # blue-100 selection
            "warning":      "#FEF3C7",   # amber-100
        }

        F = "Aptos"          # body font
        T = "Bahnschrift"    # title / heading font

        bg      = self.ui["bg"]
        surface = self.ui["surface"]
        card    = self.ui["card"]
        ink     = self.ui["ink"]
        muted   = self.ui["muted"]
        line    = self.ui["line"]
        hero    = self.ui["hero"]
        accent  = self.ui["accent"]
        sidebar = self.ui["sidebar"]

        # ── Global default ──────────────────────────────────────────────────────
        style.configure(".", font=(F, 10), background=bg, foreground=ink)

        # ── Structural frames ────────────────────────────────────────────────────
        for s in ("App.TFrame", "Chrome.TFrame", "Main.TFrame", "Page.TFrame"):
            style.configure(s, background=bg)
        style.configure("Sidebar.TFrame", background=sidebar)
        style.configure("Topbar.TFrame",  background=surface)
        style.configure("TFrame",         background=surface)

        # ── Sidebar labels ───────────────────────────────────────────────────────
        style.configure("SidebarBrand.TLabel",
                        background=sidebar, foreground="#FFFFFF", font=(T, 16))
        style.configure("SidebarText.TLabel",
                        background=sidebar, foreground=self.ui["sidebar_muted"],
                        font=(F, 9))
        style.configure("SidebarLogo.TLabel", background=sidebar)

        # ── Topbar labels ────────────────────────────────────────────────────────
        style.configure("TopbarTitle.TLabel",
                        background=surface, foreground=ink, font=(T, 21))
        style.configure("TopbarText.TLabel",
                        background=surface, foreground=muted, font=(F, 10))

        # ── Notebook tabs ────────────────────────────────────────────────────────
        style.configure("TNotebook", background=bg, borderwidth=0,
                        tabmargins=(0, 0, 0, 8))
        style.configure("TNotebook.Tab",
                        background="#DDE4EF", foreground="#475569",
                        padding=(18, 10), font=(T, 10), borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", surface), ("active", "#E8EFF8")],
                  foreground=[("selected", ink), ("active", ink)],
                  padding=[("selected", (20, 11))])

        # ── Content typography ───────────────────────────────────────────────────
        style.configure("Header.TLabel",      background=surface,
                        font=(T, 22), foreground=ink)
        style.configure("SubHeader.TLabel",   background=surface,
                        font=(T, 13), foreground=ink)
        style.configure("SectionLabel.TLabel",background=surface,
                        font=(T, 10), foreground="#0369A1")   # sky-700
        style.configure("Muted.TLabel",        background=surface, foreground=muted)
        style.configure("Status.TLabel",
                        background="#E8EEF7", foreground="#334155",
                        padding=(14, 7), font=(F, 9))

        # ── Panel (bordered card) ────────────────────────────────────────────────
        style.configure("Panel.TFrame",
                        background=surface, relief="solid", borderwidth=1,
                        bordercolor=line, lightcolor=line, darkcolor=line)

        # ── Hero / banner strip ──────────────────────────────────────────────────
        style.configure("Hero.TFrame", background=hero, relief="flat", borderwidth=0)
        style.configure("HeroTitle.TLabel",
                        background=hero, foreground="#F8FAFC", font=(T, 26))
        style.configure("HeroText.TLabel",
                        background=hero, foreground="#CBD5E1", font=(F, 10))
        style.configure("HeroMeta.TLabel",
                        background=hero, foreground="#38BDF8", font=(T, 10))
        style.configure("HeroBadge.TLabel",
                        background="#1E3A52", foreground="#F8FAFC",
                        padding=(10, 5), font=(T, 9))
        style.configure("HeroLogo.TLabel", background=hero)

        # ── Cards ────────────────────────────────────────────────────────────────
        for s, border in (("Card.TFrame", line), ("ProductCard.TFrame", "#CBD5E1")):
            style.configure(s, background=card, relief="solid", borderwidth=1,
                            bordercolor=border, lightcolor=border, darkcolor=border)

        style.configure("Card.TLabel",        background=card, foreground=ink)
        style.configure("CardTitle.TLabel",   background=card,
                        foreground=ink, font=(T, 16))
        style.configure("CardSubTitle.TLabel",background=card,
                        foreground="#475569", font=(T, 10))
        style.configure("CardMuted.TLabel",   background=card,
                        foreground=muted, font=(F, 9))
        style.configure("CardSection.TLabel", background=card,
                        foreground="#0369A1", font=(T, 10))

        # ── Feature rows (product attributes) ───────────────────────────────────
        style.configure("FeatureRow.TFrame",   background="#F0F7FF")
        style.configure("FeatureName.TLabel",  background="#F0F7FF",
                        foreground="#0369A1", font=(T, 9))
        style.configure("FeatureValue.TLabel", background="#F0F7FF",
                        foreground=ink, font=(F, 9))

        # ── Image preview stage ──────────────────────────────────────────────────
        style.configure("Preview.TLabel",
                        background="#E8EEF5", foreground="#475566")
        style.configure("ImageStage.TLabel",
                        background="#E8EEF5", foreground="#475566", font=(T, 11))
        style.configure("Summary.TLabel",
                        background=card, foreground=ink, font=(T, 11))
        style.configure("Pill.TLabel",
                        background="#DCFCE7", foreground="#15803D",
                        padding=(10, 4), font=(T, 9))

        # ── Buttons ──────────────────────────────────────────────────────────────
        style.configure("TButton",
                        background="#E2EBF6", foreground=ink,
                        padding=(12, 8), borderwidth=0, focusthickness=0,
                        font=(T, 9))
        style.map("TButton",
                  background=[("active", "#C7D9EE"), ("pressed", "#B5CCE6"),
                               ("disabled", "#E6EDF4")],
                  foreground=[("disabled", "#94A3B8")])

        style.configure("Accent.TButton",
                        background=accent, foreground="#FFFFFF", padding=(16, 9))
        style.map("Accent.TButton",
                  background=[("active",   self.ui["accent_hot"]),
                               ("pressed",  self.ui["accent_deep"]),
                               ("disabled", "#CBD5E1")],
                  foreground=[("active",   "#FFFFFF"), ("pressed", "#FFFFFF"),
                               ("disabled","#94A3B8")])

        style.configure("Ghost.TButton",
                        background=surface, foreground="#0369A1", padding=(12, 8))
        style.map("Ghost.TButton",
                  background=[("active", "#EFF6FF"), ("pressed", "#DBEAFE")])

        style.configure("Carousel.TButton",
                        background="#0D1B2A", foreground="#F8FAFC",
                        padding=(8, 5), font=(T, 10))
        style.map("Carousel.TButton",
                  background=[("active", "#1A2F44"), ("pressed", "#060F18")],
                  foreground=[("disabled", "#7E9EB3")])

        # ── Inputs ───────────────────────────────────────────────────────────────
        _entry_kw = dict(fieldbackground="#FFFFFF", foreground=ink,
                         bordercolor="#CBD5E1", lightcolor="#CBD5E1",
                         darkcolor="#CBD5E1", padding=(9, 7))
        style.configure("TEntry",    **_entry_kw)
        style.configure("TCombobox", **_entry_kw)

        # ── Treeview ─────────────────────────────────────────────────────────────
        style.configure("Treeview",
                        rowheight=34,
                        fieldbackground="#FFFFFF", background="#FFFFFF",
                        foreground=ink,
                        bordercolor=line, lightcolor=line, darkcolor=line,
                        font=(F, 10))
        style.configure("Treeview.Heading",
                        font=(T, 10), padding=(10, 9),
                        background="#F1F5F9", foreground=ink)
        style.map("Treeview",
                  background=[("selected", "#DBEAFE")],
                  foreground=[("selected", "#1E3A8A")])

        # ── Scrollbars ───────────────────────────────────────────────────────────
        style.configure("Vertical.TScrollbar",
                        background="#CBD5E1", troughcolor=bg,
                        bordercolor=bg, arrowcolor="#64748B")
        style.configure("Horizontal.TScrollbar",
                        background="#CBD5E1", troughcolor=bg,
                        bordercolor=bg, arrowcolor="#64748B")

        self.configure(bg=bg)

    def _build_menu(self) -> None:
        menu = tk.Menu(self)
        settings_menu = tk.Menu(menu, tearoff=0)
        settings_menu.add_command(label="Ayarlar", command=self.open_settings_dialog)
        settings_menu.add_command(label="Veriyi Yenile", command=self.refresh_product_data)
        settings_menu.add_command(label="Gorsel Indeksini Yenile", command=self.rebuild_index_from_ui)
        settings_menu.add_command(label="Kanal Raporu", command=self.open_channel_report)
        menu.add_cascade(label="Ayarlar", menu=settings_menu)
        self.config(menu=menu)

    def configure_app_icon(self) -> None:
        if path_exists(APP_ICON_ICO):
            try:
                self.iconbitmap(str(APP_ICON_ICO))
            except tk.TclError:
                pass
        if path_exists(APP_ICON_PNG):
            try:
                self.app_icon_ref = tk.PhotoImage(file=str(APP_ICON_PNG))
                self.iconphoto(True, self.app_icon_ref)
            except tk.TclError:
                self.app_icon_ref = None

    def load_ui_logo(self) -> ImageTk.PhotoImage | None:
        if self.ui_logo_ref is not None:
            return self.ui_logo_ref
        if not path_exists(APP_LOGO_PNG):
            return None
        try:
            image = Image.open(APP_LOGO_PNG).convert("RGBA")
            self.ui_logo_ref = ImageTk.PhotoImage(image)
            return self.ui_logo_ref
        except Exception:  # noqa: BLE001
            return None

    def _build_layout(self) -> None:
        # ── Root shell (no outer padding – sidebar sits flush) ──────────────────
        shell = ttk.Frame(self, style="Chrome.TFrame")
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(1, weight=1)
        shell.rowconfigure(0, weight=1)

        # ── Sidebar ──────────────────────────────────────────────────────────────
        sidebar = ttk.Frame(shell, style="Sidebar.TFrame")
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.configure(width=268)
        sidebar.grid_propagate(False)

        # Orange top accent stripe
        tk.Frame(sidebar, height=3, bg=self.ui["accent"]).pack(fill="x")

        # Brand header
        brand_wrap = ttk.Frame(sidebar, style="Sidebar.TFrame")
        brand_wrap.pack(fill="x", padx=20, pady=(18, 0))
        logo = self.load_ui_logo()
        if logo is not None:
            ttk.Label(brand_wrap, image=logo, style="SidebarLogo.TLabel").pack(side="left", padx=(0, 10))
        brand_text = ttk.Frame(brand_wrap, style="Sidebar.TFrame")
        brand_text.pack(side="left")
        ttk.Label(brand_text, text="HEKA", style="SidebarBrand.TLabel").pack(anchor="w")
        ttk.Label(brand_text, text="Urun Operasyon Paneli", style="SidebarText.TLabel").pack(anchor="w", pady=(1, 0))

        # Divider
        tk.Frame(sidebar, height=1, bg="#1E3348").pack(fill="x", padx=20, pady=(18, 16))

        # Nav section label
        ttk.Label(sidebar, text="ANA MODÜLLER", style="SidebarText.TLabel").pack(anchor="w", padx=20, pady=(0, 8))

        # ── Navigation buttons ──────────────────────────────────────────────────
        # Icons (Unicode symbols that render well in Bahnschrift)
        self.nav_specs = {
            "product": ("◈  Urun Vitrini",     "Stok, gorsel ve kanal arama"),
            "rename":  ("⇄  Toplu Isim",        "Kontrollu dosya cikisi"),
            "ydk":     ("▤  Etiket Atolyesi",   "Yurtdisi PDF tasarimi"),
        }
        self.nav_buttons: dict[str, tk.Button] = {}
        for page_key, (title, hint) in self.nav_specs.items():
            btn = tk.Button(
                sidebar,
                text=f"{title}\n   {hint}",
                command=lambda key=page_key: self.show_page(key),
                anchor="w",
                justify="left",
                padx=16,
                pady=11,
                bd=0,
                relief="flat",
                cursor="hand2",
                font=("Bahnschrift", 10),
                bg=self.ui["sidebar"],
                fg="#C8D8E8",
                activebackground=self.ui["sidebar_2"],
                activeforeground="#FFFFFF",
            )
            btn.pack(fill="x", padx=10, pady=(0, 4))
            self.nav_buttons[page_key] = btn

        # Divider + Quick actions
        tk.Frame(sidebar, height=1, bg="#1E3348").pack(fill="x", padx=20, pady=(16, 14))
        ttk.Label(sidebar, text="HIZLI İŞLEMLER", style="SidebarText.TLabel").pack(anchor="w", padx=20, pady=(0, 8))

        _quick = [
            ("⚙  Ayarlar",         self.open_settings_dialog),
            ("↻  Veriyi Yenile",   self.refresh_product_data),
            ("⊕  Indeksi Yenile",  self.rebuild_index_from_ui),
            ("≡  Kanal Raporu",    self.open_channel_report),
        ]
        for label, cmd in _quick:
            tk.Button(
                sidebar,
                text=label,
                command=cmd,
                anchor="w",
                padx=16,
                pady=7,
                bd=0,
                relief="flat",
                cursor="hand2",
                font=("Bahnschrift", 9),
                bg="#0D1B2A",
                fg="#94A3B8",
                activebackground="#1A2F44",
                activeforeground="#FFFFFF",
            ).pack(fill="x", padx=10, pady=(0, 2))

        # Bottom status / data source info
        tk.Frame(sidebar, height=1, bg="#1E3348").pack(fill="x", padx=20, pady=(14, 10), side="bottom")
        ttk.Label(sidebar, textvariable=self.settings_summary_var,
                  style="SidebarText.TLabel", wraplength=228).pack(
            side="bottom", anchor="w", padx=20, pady=(0, 8))

        # ── Main content area ────────────────────────────────────────────────────
        main = ttk.Frame(shell, style="Main.TFrame")
        main.grid(row=0, column=1, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        # Topbar
        self.page_title_var = tk.StringVar()
        self.page_hint_var  = tk.StringVar()
        topbar = ttk.Frame(main, padding=(24, 13), style="Topbar.TFrame")
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.columnconfigure(0, weight=1)

        title_block = ttk.Frame(topbar, style="Topbar.TFrame")
        title_block.grid(row=0, column=0, sticky="w")
        ttk.Label(title_block, textvariable=self.page_title_var,
                  style="TopbarTitle.TLabel").pack(anchor="w")
        ttk.Label(title_block, textvariable=self.page_hint_var,
                  style="TopbarText.TLabel").pack(anchor="w", pady=(2, 0))

        top_actions = ttk.Frame(topbar, style="Topbar.TFrame")
        top_actions.grid(row=0, column=1, sticky="e")
        ttk.Button(top_actions, text="Ayarlar",
                   style="Ghost.TButton",
                   command=self.open_settings_dialog).pack(side="left")
        ttk.Button(top_actions, text="Veriyi Yenile",
                   style="Ghost.TButton",
                   command=self.refresh_product_data).pack(side="left", padx=(8, 0))
        ttk.Button(top_actions, text="⊕  Indeksi Yenile",
                   style="Accent.TButton",
                   command=self.rebuild_index_from_ui).pack(side="left", padx=(8, 0))

        # Thin separator line under topbar
        tk.Frame(main, height=1, bg=self.ui["line"]).grid(
            row=0, column=0, sticky="sew")

        # Page content frames
        content = ttk.Frame(main, style="Page.TFrame", padding=(0, 14, 0, 0))
        content.grid(row=1, column=0, sticky="nsew")
        content.rowconfigure(0, weight=1)
        content.columnconfigure(0, weight=1)

        product_tab = ttk.Frame(content, padding=0, style="Page.TFrame")
        rename_tab  = ttk.Frame(content, padding=0, style="Page.TFrame")
        ydk_tab     = ttk.Frame(content, padding=0, style="Page.TFrame")
        self.page_frames = {
            "product": product_tab,
            "rename":  rename_tab,
            "ydk":     ydk_tab,
        }
        for frame in self.page_frames.values():
            frame.grid(row=0, column=0, sticky="nsew")

        self._build_product_tab(product_tab)
        self._build_rename_tab(rename_tab)
        self._build_ydk_tab(ydk_tab)
        self.show_page("product")

        # Status bar
        ttk.Label(main, textvariable=self.status_var,
                  style="Status.TLabel").grid(row=2, column=0, sticky="ew",
                                              pady=(10, 0))

    def show_page(self, page_key: str) -> None:
        frame = getattr(self, "page_frames", {}).get(page_key)
        if frame is None:
            return
        frame.tkraise()
        title_icon, hint = self.nav_specs.get(page_key, ("Urun Yonetim Masasi", ""))
        # Strip the icon prefix for the topbar title
        clean_title = title_icon.split("  ", 1)[-1] if "  " in title_icon else title_icon
        self.page_title_var.set(clean_title)
        self.page_hint_var.set(hint)
        for key, button in self.nav_buttons.items():
            is_active = key == page_key
            t_icon, t_hint = self.nav_specs[key]
            if is_active:
                button.configure(
                    bg=self.ui["accent"],
                    fg="#FFFFFF",
                    activebackground=self.ui["accent_hot"],
                    activeforeground="#FFFFFF",
                    font=("Bahnschrift", 10),
                    text=f"{t_icon}\n   {t_hint}",
                )
            else:
                button.configure(
                    bg=self.ui["sidebar"],
                    fg="#C8D8E8",
                    activebackground=self.ui["sidebar_2"],
                    activeforeground="#FFFFFF",
                    font=("Bahnschrift", 10),
                    text=f"{t_icon}\n   {t_hint}",
                )

    def _build_product_tab(self, parent: ttk.Frame) -> None:
        shell = ttk.Frame(parent, style="Page.TFrame")
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(1, weight=1)
        shell.rowconfigure(1, weight=1)

        # ── Hero banner ─────────────────────────────────────────────────────────
        hero = ttk.Frame(shell, padding=(28, 20), style="Hero.TFrame")
        hero.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 0))
        hero.columnconfigure(0, weight=1)

        brand_row = ttk.Frame(hero, style="Hero.TFrame")
        brand_row.grid(row=0, column=0, sticky="w")
        logo = self.load_ui_logo()
        if logo is not None:
            ttk.Label(brand_row, image=logo,
                      style="HeroLogo.TLabel").pack(side="left", padx=(0, 16))
        title_block = ttk.Frame(brand_row, style="Hero.TFrame")
        title_block.pack(side="left")
        ttk.Label(title_block, text="Urun Vitrini",
                  style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(title_block,
                  text="SQL veri · Gorsel indeks · Kanal klasorleri",
                  style="HeroMeta.TLabel").pack(anchor="w", pady=(3, 0))

        badge_row = ttk.Frame(hero, style="Hero.TFrame")
        badge_row.grid(row=0, column=1, sticky="e")
        for lbl in ("SQL Bagli", "Gorsel Indeks", "Explorer Hazir"):
            ttk.Label(badge_row, text=lbl,
                      style="HeroBadge.TLabel").pack(side="left", padx=(8, 0))

        # Thin accent stripe at bottom of hero
        tk.Frame(shell, height=2,
                 bg=self.ui["accent"]).grid(row=0, column=0, columnspan=2,
                                            sticky="sew")

        # ── Left control panel ───────────────────────────────────────────────────
        controls = ttk.Frame(shell, padding=(18, 20), style="Panel.TFrame")
        controls.grid(row=1, column=0, sticky="nsew", pady=(14, 0))

        ttk.Label(controls, text="Filtre & Arama",
                  style="SubHeader.TLabel").pack(anchor="w")
        ttk.Label(controls, text="Aktif veri kaynagi ve secim kriterleri",
                  style="Muted.TLabel").pack(anchor="w", pady=(2, 14))

        # Data source summary pill
        pill_row = ttk.Frame(controls, style="FeatureRow.TFrame",
                             padding=(10, 6))
        pill_row.pack(fill="x", pady=(0, 18))
        ttk.Label(pill_row, textvariable=self.settings_summary_var,
                  style="FeatureValue.TLabel", wraplength=295).pack(anchor="w")

        # Search
        ttk.Label(controls, text="GENEL ARAMA",
                  style="SectionLabel.TLabel").pack(anchor="w")
        search_row = ttk.Frame(controls)
        search_row.pack(fill="x", pady=(6, 16))
        search_entry = ttk.Entry(search_row,
                                 textvariable=self.product_search_var)
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.bind("<Return>", lambda _e: self.search_products())
        ttk.Button(search_row, text="Ara",
                   style="Accent.TButton",
                   command=self.search_products,
                   width=8).pack(side="left", padx=(8, 0))

        # Family / breakdown selectors
        ttk.Label(controls, text="ÜRÜN AİLESİ",
                  style="SectionLabel.TLabel").pack(anchor="w")
        self.family_combo = SearchableCombobox(
            controls, textvariable=self.family_var, width=35)
        self.family_combo.pack(fill="x", pady=(6, 14))
        self.family_combo.bind("<<ComboboxSelected>>", self.on_family_changed)

        ttk.Label(controls, text="AİLE İÇİ KIRILIM",
                  style="SectionLabel.TLabel").pack(anchor="w")
        self.breakdown_combo = SearchableCombobox(
            controls, textvariable=self.breakdown_var, width=35)
        self.breakdown_combo.pack(fill="x", pady=(6, 16))
        self.breakdown_combo.bind("<<ComboboxSelected>>",
                                  lambda _e: self.persist_runtime_state())

        # Action buttons
        btn_row = ttk.Frame(controls)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="▶  Getir",
                   style="Accent.TButton",
                   command=self.fetch_selected_products).pack(side="left")
        ttk.Button(btn_row, text="Tum Aileyi Listele",
                   command=self.fetch_family_products).pack(
            side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Kanal Raporu",
                   command=self.open_channel_report).pack(
            side="left", padx=(8, 0))

        # ── Right result panel ───────────────────────────────────────────────────
        result_panel = ttk.Frame(shell, padding=(18, 20),
                                 style="Panel.TFrame")
        result_panel.grid(row=1, column=1, sticky="nsew",
                          padx=(14, 0), pady=(14, 0))
        result_panel.rowconfigure(1, weight=1)
        result_panel.columnconfigure(0, weight=1)

        # Summary strip
        summary = ttk.Frame(result_panel, padding=(14, 12), style="Card.TFrame")
        summary.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        summary.columnconfigure(0, weight=1)
        ttk.Label(summary, textvariable=self.results_summary_var,
                  style="Header.TLabel").pack(anchor="w")
        ttk.Label(summary, textvariable=self.results_context_var,
                  style="Muted.TLabel").pack(anchor="w", pady=(4, 0))

        holder = ttk.Frame(result_panel)
        holder.grid(row=1, column=0, sticky="nsew")
        holder.rowconfigure(0, weight=1)
        holder.columnconfigure(0, weight=1)
        self.results_canvas = tk.Canvas(holder, highlightthickness=0, bg=self.ui["surface"])
        self.results_frame = ttk.Frame(self.results_canvas, style="Page.TFrame")
        self.results_window = self.results_canvas.create_window((0, 0), window=self.results_frame, anchor="nw")
        scroll = ttk.Scrollbar(holder, orient="vertical", command=self.results_canvas.yview)
        self.results_canvas.configure(yscrollcommand=scroll.set)
        self.results_frame.bind("<Configure>", lambda _event: self.results_canvas.configure(scrollregion=self.results_canvas.bbox("all")))
        self.results_canvas.bind("<Configure>", lambda event: self.results_canvas.itemconfigure(self.results_window, width=event.width))
        self.results_canvas.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        enable_mousewheel_scroll(self.results_canvas, self.results_canvas)
        enable_mousewheel_scroll(self.results_frame, self.results_canvas)
        self.clear_results("Bir urun ailesi secip kirilim getirerek sonuclari burada gorebilirsin.")

    def _build_rename_tab(self, parent: ttk.Frame) -> None:
        # ── Hero banner ─────────────────────────────────────────────────────────
        hero = ttk.Frame(parent, padding=(28, 20), style="Hero.TFrame")
        hero.pack(fill="x")
        hero.columnconfigure(0, weight=1)

        brand_row = ttk.Frame(hero, style="Hero.TFrame")
        brand_row.pack(anchor="w")
        logo = self.load_ui_logo()
        if logo is not None:
            ttk.Label(brand_row, image=logo,
                      style="HeroLogo.TLabel").pack(side="left", padx=(0, 16))
        title_block = ttk.Frame(brand_row, style="Hero.TFrame")
        title_block.pack(side="left")
        ttk.Label(title_block, text="Toplu Isim Degistirme",
                  style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(title_block,
                  text="Onizle · Manuel duzenle · Klasor agacini koruyarak cikart",
                  style="HeroMeta.TLabel").pack(anchor="w", pady=(3, 0))

        # Accent stripe
        tk.Frame(parent, height=2,
                 bg=self.ui["accent"]).pack(fill="x")

        card = ttk.Frame(parent, padding=(18, 14), style="Panel.TFrame")
        card.pack(fill="x", pady=(12, 0))
        path_row = ttk.Frame(card)
        path_row.pack(fill="x")
        ttk.Label(path_row, text="TARAMA KLASÖRÜ", style="SectionLabel.TLabel").pack(side="left")
        ttk.Entry(path_row, textvariable=self.rename_path_var, width=82).pack(side="left", fill="x", expand=True, padx=(10, 0))
        ttk.Button(path_row, text="Sec", width=8, command=self.browse_rename_root).pack(side="left", padx=(8, 0))

        export_row = ttk.Frame(card)
        export_row.pack(fill="x", pady=(12, 0))
        ttk.Checkbutton(
            export_row,
            text="Farkli klasore cikart - orijinal dosyalara dokunma",
            variable=self.rename_export_enabled_var,
            command=self.on_rename_export_mode_changed,
        ).pack(side="left")

        output_row = ttk.Frame(card)
        output_row.pack(fill="x", pady=(8, 0))
        ttk.Label(output_row, text="Cikis Klasoru", style="SectionLabel.TLabel").pack(side="left")
        ttk.Entry(output_row, textvariable=self.rename_output_path_var, width=82).pack(side="left", fill="x", expand=True, padx=(10, 0))
        ttk.Button(output_row, text="Sec", width=8, command=self.browse_rename_output_root).pack(side="left", padx=(8, 0))

        actions = ttk.Frame(card)
        actions.pack(fill="x", pady=(14, 0))
        ttk.Button(actions, text="▶  Onizleme Olustur", style="Accent.TButton", command=self.preview_rename_plan).pack(side="left")
        ttk.Button(actions, text="Tumunu Sec", command=self.select_all_rename_actions).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="Tumunu Kaldir", command=self.clear_all_rename_actions).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="↕  Ac", command=self.expand_all_rename_groups).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="↕  Kapat", command=self.collapse_all_rename_groups).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="↩  Geri Al", command=self.undo_last_rename_batch).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="✓  Secilenleri Uygula", style="Accent.TButton", command=self.apply_rename_plan).pack(side="left", padx=(8, 0))

        tools = ttk.Frame(card)
        tools.pack(fill="x", pady=(10, 0))
        ttk.Label(tools, text="FİLTRE", style="SectionLabel.TLabel").pack(side="left")
        combo = ttk.Combobox(tools, textvariable=self.rename_filter_var, values=RENAME_FILTER_OPTIONS, state="readonly", width=20)
        combo.pack(side="left", padx=(8, 0))
        combo.bind("<<ComboboxSelected>>", lambda _event: self.populate_rename_tree(self.rename_plan))

        self.rename_summary_var = tk.StringVar(value="Onizleme olusturuldugunda ozet burada gorunur.")
        ttk.Label(parent, textvariable=self.rename_summary_var, style="Muted.TLabel").pack(anchor="w", pady=(14, 8))

        workspace = ttk.Frame(parent)
        workspace.pack(fill="both", expand=True)
        workspace.columnconfigure(0, weight=1)
        workspace.rowconfigure(0, weight=1)
        table = ttk.Frame(workspace)
        table.grid(row=0, column=0, sticky="nsew")
        table.columnconfigure(0, weight=1)
        table.rowconfigure(0, weight=1)
        columns = ("selected", "status", "stock", "current", "target", "reason", "relative")
        self.rename_tree = ttk.Treeview(table, columns=columns, show="tree headings", height=18)
        headings = {
            "#0": "Aile / Dosya",
            "selected": "Sec",
            "status": "Durum",
            "stock": "Stok",
            "current": "Mevcut",
            "target": "Yeni / Hedef",
            "reason": "Kural",
            "relative": "Kokten Sonra",
        }
        for col, text in headings.items():
            self.rename_tree.heading(col, text=text)
        self.rename_tree.column("#0", width=260, minwidth=180)
        self.rename_tree.column("selected", width=60, anchor="center", stretch=False)
        self.rename_tree.column("status", width=120, anchor="center")
        self.rename_tree.column("stock", width=150, anchor="center")
        self.rename_tree.column("current", width=260)
        self.rename_tree.column("target", width=320)
        self.rename_tree.column("reason", width=180)
        self.rename_tree.column("relative", width=360)
        self.rename_tree.tag_configure("group", background="#EFF6FF", foreground="#1E40AF")
        self.rename_tree.tag_configure("disabled", foreground="#94A3B8")
        self.rename_tree.bind("<ButtonRelease-1>", self.on_rename_tree_click)
        self.rename_tree.bind("<Double-1>", self.on_rename_tree_double_click)
        self.rename_tree.bind("<<TreeviewSelect>>", self.on_rename_selection_changed)
        sy = ttk.Scrollbar(table, orient="vertical", command=self.rename_tree.yview)
        sx = ttk.Scrollbar(table, orient="horizontal", command=self.rename_tree.xview)
        self.rename_tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.rename_tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        enable_mousewheel_scroll(self.rename_tree, self.rename_tree, self.rename_tree)
        table.bind("<Configure>", lambda _event: self.schedule_rename_autosize())

        preview = ttk.Frame(workspace, padding=16, style="ProductCard.TFrame")
        preview.grid(row=0, column=1, sticky="ns", padx=(12, 0))
        ttk.Label(preview, text="Gorsel Onizleme", style="CardSection.TLabel").pack(anchor="w")
        self.rename_preview_label = ttk.Label(preview, text="Secim yok", style="ImageStage.TLabel", anchor="center", width=36)
        self.rename_preview_label.pack(fill="x", pady=(10, 8))
        ttk.Label(preview, textvariable=self.rename_preview_text_var, style="CardMuted.TLabel", wraplength=310, justify="left").pack(anchor="w")
        ttk.Button(preview, text="Explorer'da Goster", command=self.reveal_selected_rename_file).pack(anchor="w", pady=(12, 0))

    def _build_ydk_tab(self, parent: ttk.Frame) -> None:
        # ── Hero banner ─────────────────────────────────────────────────────────
        hero = ttk.Frame(parent, padding=(28, 20), style="Hero.TFrame")
        hero.pack(fill="x")

        brand_row = ttk.Frame(hero, style="Hero.TFrame")
        brand_row.pack(anchor="w")
        logo = self.load_ui_logo()
        if logo is not None:
            ttk.Label(brand_row, image=logo,
                      style="HeroLogo.TLabel").pack(side="left", padx=(0, 16))
        title_block = ttk.Frame(brand_row, style="Hero.TFrame")
        title_block.pack(side="left")
        ttk.Label(title_block, text="Etiket Atolyesi",
                  style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(title_block,
                  text="Ana veri · Barkod · Tasarim onizleme · PDF cikisi",
                  style="HeroMeta.TLabel").pack(anchor="w", pady=(3, 0))

        # Accent stripe
        tk.Frame(parent, height=2,
                 bg=self.ui["accent"]).pack(fill="x")

        config = ttk.Frame(parent, padding=(18, 14), style="Panel.TFrame")
        config.pack(fill="x", pady=(12, 0))
        for label, variable, command in [
            ("GORSEL KLASÖRÜ", self.ydk_image_root_var, self.browse_ydk_image_root),
            ("PDF ÇIKIŞ",      self.ydk_output_root_var, self.browse_ydk_output_root),
        ]:
            row = ttk.Frame(config)
            row.pack(fill="x", pady=4)
            ttk.Label(row, text=label, style="SectionLabel.TLabel", width=16).pack(side="left")
            ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True, padx=(8, 0))
            ttk.Button(row, text="Sec", width=8, command=command).pack(side="left", padx=(8, 0))
        actions = ttk.Frame(config)
        actions.pack(fill="x", pady=(10, 0))
        ttk.Button(actions, text="↻  Ana Veriden Yenile", style="Accent.TButton", command=lambda: self.load_ydk_workbook(True)).pack(side="left")
        ttk.Label(actions, text="STOK / AÇIKLAMA ARA", style="SectionLabel.TLabel").pack(side="left", padx=(18, 8))
        search = ttk.Entry(actions, textvariable=self.ydk_search_var, width=34)
        search.pack(side="left")
        search.bind("<Return>", lambda _event: self.search_ydk_products())
        ttk.Button(actions, text="Ara", command=self.search_ydk_products).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="Temizle", command=self.clear_ydk_search).pack(side="left", padx=(8, 0))

        body = ttk.Frame(parent)
        body.pack(fill="both", expand=True, pady=(14, 0))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        list_card = ttk.Frame(body, padding=14, style="ProductCard.TFrame")
        list_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        list_card.rowconfigure(1, weight=1)
        list_card.columnconfigure(0, weight=1)
        ttk.Label(list_card, text="Urunler", style="SubHeader.TLabel").grid(row=0, column=0, sticky="w")
        columns = ("code", "model", "producer", "barcode")
        self.ydk_tree = ttk.Treeview(list_card, columns=columns, show="headings", height=18)
        for column, heading, width in [
            ("code", "Stok", 170),
            ("model", "Model", 120),
            ("producer", "Uretici Kodu", 130),
            ("barcode", "Barkod", 145),
        ]:
            self.ydk_tree.heading(column, text=heading)
            self.ydk_tree.column(column, width=width, anchor="w")
        self.ydk_tree.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        yscroll = ttk.Scrollbar(list_card, orient="vertical", command=self.ydk_tree.yview)
        self.ydk_tree.configure(yscrollcommand=yscroll.set)
        yscroll.grid(row=1, column=1, sticky="ns", pady=(8, 0))
        self.ydk_tree.bind("<<TreeviewSelect>>", self.on_ydk_selection_changed)
        enable_mousewheel_scroll(self.ydk_tree, self.ydk_tree)

        detail = ttk.Frame(body, padding=14, style="ProductCard.TFrame")
        detail.grid(row=0, column=1, sticky="nsew")
        detail.columnconfigure(0, weight=1)
        detail.rowconfigure(4, weight=1)
        ttk.Label(detail, text="Etiket Onizleme", style="SubHeader.TLabel").grid(row=0, column=0, sticky="w")
        preview_row = ttk.Frame(detail, style="Card.TFrame")
        preview_row.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        self.ydk_product_image_label = ttk.Label(preview_row, text="Urun gorseli", style="ImageStage.TLabel", anchor="center", width=28)
        self.ydk_product_image_label.pack(side="left", padx=(0, 10))
        self.ydk_label_preview = ttk.Label(preview_row, text="Etiket onizleme", style="ImageStage.TLabel", anchor="center", width=46)
        self.ydk_label_preview.pack(side="left", fill="x", expand=True)
        ttk.Label(detail, textvariable=self.ydk_detail_var, style="CardMuted.TLabel", wraplength=620, justify="left").grid(row=2, column=0, sticky="ew", pady=(12, 0))
        buttons = ttk.Frame(detail, style="Card.TFrame")
        buttons.grid(row=3, column=0, sticky="w", pady=(14, 0))
        ttk.Button(buttons, text="⎘  Bilgiyi Kopyala", command=self.copy_ydk_summary).pack(side="left")
        ttk.Button(buttons, text="Urun PDF", command=lambda: self.export_ydk_label_pdf("unit")).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="Koli PDF", command=lambda: self.export_ydk_label_pdf("carton")).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="▼  Ikisini Kaydet", style="Accent.TButton", command=self.export_both_ydk_labels).pack(side="left", padx=(8, 0))

        editor_shell = ttk.Frame(detail, style="Card.TFrame")
        editor_shell.grid(row=4, column=0, sticky="nsew", pady=(14, 0))
        editor_shell.rowconfigure(0, weight=1)
        editor_shell.columnconfigure(0, weight=1)
        editor_canvas = tk.Canvas(editor_shell, highlightthickness=0, background="#ffffff", height=275)
        editor_scroll = ttk.Scrollbar(editor_shell, orient="vertical", command=editor_canvas.yview)
        editor_body = ttk.Frame(editor_canvas, padding=(0, 0, 8, 0), style="Card.TFrame")
        editor_window = editor_canvas.create_window((0, 0), window=editor_body, anchor="nw")
        editor_body.bind("<Configure>", lambda _event: editor_canvas.configure(scrollregion=editor_canvas.bbox("all")))
        editor_canvas.bind("<Configure>", lambda event: editor_canvas.itemconfigure(editor_window, width=event.width))
        editor_canvas.configure(yscrollcommand=editor_scroll.set)
        editor_canvas.grid(row=0, column=0, sticky="nsew")
        editor_scroll.grid(row=0, column=1, sticky="ns")
        enable_mousewheel_scroll(editor_canvas, editor_canvas)
        enable_mousewheel_scroll(editor_body, editor_canvas)
        self._build_ydk_editor(editor_body)

    def _build_ydk_editor(self, parent: ttk.Frame) -> None:
        ttk.Label(parent, text="Etiket Duzenle",
                  style="SubHeader.TLabel").grid(row=0, column=0, columnspan=4,
                                                  sticky="w", pady=(0, 8))
        field_specs = [
            ("model", "Model", "entry"),
            ("producer_code", "HKA / Uretici Kodu", "entry"),
            ("unit_barcode", "Urun Barkodu", "entry"),
            ("carton_barcode", "Koli Barkodu", "entry"),
            ("carton_quantity", "Koli Ici Adet", "entry"),
            ("description_tr", "TR Aciklama", "text"),
            ("description_en", "EN Aciklama", "text"),
        ]
        row = 1
        for key, label, kind in field_specs:
            ttk.Label(parent, text=label, style="SectionLabel.TLabel").grid(row=row, column=0, sticky="nw", pady=4)
            if kind == "text":
                widget = tk.Text(parent, height=3, width=42, wrap="word")
                widget.grid(row=row, column=1, columnspan=3, sticky="ew", pady=4)
                self.ydk_editor_texts[key] = widget
            else:
                variable = tk.StringVar()
                ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, columnspan=3, sticky="ew", pady=4)
                self.ydk_editor_vars[key] = variable
            row += 1

        actions = ttk.Frame(parent, style="Card.TFrame")
        actions.grid(row=row, column=0, columnspan=4, sticky="w", pady=(8, 14))
        ttk.Button(actions, text="✓  Onizlemeyi Guncelle / Kaydet",
                   style="Accent.TButton",
                   command=self.apply_ydk_editor_changes).pack(side="left")
        ttk.Button(actions, text="↩  Kaynak Degerine Don",
                   command=self.reset_ydk_editor_changes).pack(side="left", padx=(8, 0))
        row += 1

        ttk.Label(parent, text="Tasarim Ayarlari",
                  style="SubHeader.TLabel").grid(row=row, column=0, columnspan=4,
                                                  sticky="w", pady=(0, 8))
        row += 1
        layout = self.get_ydk_label_layout(include_ui=False)
        layout_specs = [
            ("photo_x", "Foto X"), ("photo_y", "Foto Y"), ("photo_w", "Foto W"), ("photo_h", "Foto H"),
            ("model_y", "Model Y"), ("tr_y", "TR Y"), ("en_y", "EN Y"), ("barcode_y", "Barkod Y"),
            ("info_x", "Sag Blok X"), ("producer_box_y", "HKA Kutu Y"), ("tr_font", "TR Font"), ("en_font", "EN Font"),
            ("producer_font", "HKA Font"), ("footer_y", "Alt Bilgi Y"),
        ]
        for index, (key, label) in enumerate(layout_specs):
            column = 0 if index % 2 == 0 else 2
            if index % 2 == 0 and index:
                row += 1
            ttk.Label(parent, text=label, style="SectionLabel.TLabel").grid(row=row, column=column, sticky="w", pady=3, padx=(0 if column == 0 else 12, 4))
            variable = tk.StringVar(value=str(layout.get(key, DEFAULT_YDK_LABEL_LAYOUT[key])))
            ttk.Entry(parent, textvariable=variable, width=10).grid(row=row, column=column + 1, sticky="w", pady=3)
            self.ydk_layout_vars[key] = variable
        row += 1
        design_actions = ttk.Frame(parent, style="Card.TFrame")
        design_actions.grid(row=row, column=0, columnspan=4, sticky="w",
                            pady=(8, 0))
        ttk.Button(design_actions, text="✓  Tasarimi Uygula / Kaydet",
                   style="Accent.TButton",
                   command=self.apply_ydk_layout_changes).pack(side="left")
        ttk.Button(design_actions, text="↩  Varsayilan Tasarim",
                   command=self.reset_ydk_layout).pack(side="left", padx=(8, 0))
        parent.columnconfigure(1, weight=1)
        parent.columnconfigure(3, weight=1)

    def load_settings(self) -> dict[str, str]:
        if not path_exists(SETTINGS_FILE):
            self.save_settings(DEFAULT_SETTINGS.copy())
            return DEFAULT_SETTINGS.copy()
        try:
            content = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return DEFAULT_SETTINGS.copy()
        settings = DEFAULT_SETTINGS.copy()
        settings.update({key: str(value) for key, value in content.items()})
        settings = apply_logodata_sql_profile(settings)
        self.save_settings(settings)
        return settings

    def save_settings(self, settings: dict[str, str]) -> None:
        persisted = DEFAULT_SETTINGS.copy()
        persisted.update({key: str(value) for key, value in settings.items()})
        SETTINGS_FILE.write_text(json.dumps(persisted, ensure_ascii=False, indent=2), encoding="utf-8")

    def load_json_setting(self, key: str, default: dict[str, Any]) -> dict[str, Any]:
        try:
            loaded = json.loads(self.settings.get(key, "{}") or "{}")
        except json.JSONDecodeError:
            return default.copy()
        return loaded if isinstance(loaded, dict) else default.copy()

    def save_rename_state(self) -> None:
        self.settings["rename_selection_state"] = json.dumps(self.rename_selection, ensure_ascii=False)
        self.settings["rename_manual_targets"] = json.dumps(self.rename_manual_targets, ensure_ascii=False)
        self.save_settings(self.settings)

    def save_ydk_state(self) -> None:
        self.settings["ydk_label_overrides"] = json.dumps(self.ydk_label_overrides, ensure_ascii=False)
        self.settings["ydk_label_layout"] = json.dumps(self.get_ydk_label_layout(), ensure_ascii=False)
        self.save_settings(self.settings)

    def persist_runtime_state(self) -> None:
        self.settings["last_family"] = self.family_var.get().strip()
        self.settings["last_breakdown"] = self.breakdown_var.get().strip()
        self.settings["last_rename_path"] = self.rename_path_var.get().strip()
        self.settings["rename_output_path"] = self.rename_output_path_var.get().strip()
        self.settings["rename_export_enabled"] = "1" if self.rename_export_enabled_var.get() else "0"
        self.settings["ydk_image_root"] = self.ydk_image_root_var.get().strip()
        self.settings["ydk_output_root"] = self.ydk_output_root_var.get().strip()
        if hasattr(self, "ydk_label_overrides"):
            self.settings["ydk_label_overrides"] = json.dumps(self.ydk_label_overrides, ensure_ascii=False)
        if hasattr(self, "ydk_layout_vars"):
            self.settings["ydk_label_layout"] = json.dumps(self.get_ydk_label_layout(), ensure_ascii=False)
        self.settings["window_geometry"] = self.geometry()
        self.save_settings(self.settings)

    def on_close(self) -> None:
        self.save_rename_state()
        self.save_ydk_state()
        self.persist_runtime_state()
        self.destroy()

    def set_status(self, text: str) -> None:
        self.status_var.set(text)

    def run_in_background(self, job: Callable[[], Any], on_success: Callable[[Any], None], description: str) -> None:
        self.set_status(description)

        def worker() -> None:
            try:
                result = job()
            except Exception as exc:  # noqa: BLE001
                self.after(0, lambda error=exc: self._handle_background_error(description, error))
                return
            self.after(0, lambda: on_success(result))

        threading.Thread(target=worker, daemon=True).start()

    def _handle_background_error(self, description: str, exc: Exception) -> None:
        self.set_status(f"Hata: {exc}")
        messagebox.showerror("Islem Hatasi", f"{description}\n\n{exc}")

    def summarize_config_path(self, key: str, empty_label: str, expected: str = "any") -> str:
        text = self.settings.get(key, "").strip()
        if not text:
            return empty_label
        path = Path(text).expanduser()
        available = path_is_file(path) if expected == "file" else path_is_dir(path) if expected == "dir" else path_exists(path)
        label = path.name or str(path)
        return label if available else f"{label} (erisilemiyor)"

    def get_data_source_key(self) -> str:
        return normalize_data_source(self.settings.get("data_source", "excel"))

    def get_data_source_label(self) -> str:
        return display_data_source(self.get_data_source_key())

    def summarize_sql_source(self) -> str:
        connection = self.settings.get("sql_connection_string", "").strip()
        table = self.settings.get("sql_table", "").strip()
        query = self.settings.get("sql_query", "").strip()
        if not connection:
            return "SQL baglantisi yok"
        if table and query:
            return f"{table} + sorgu"
        if table:
            return table
        if query:
            return "Ozel sorgu"
        return "Baglanti hazir"

    def refresh_settings_summary(self) -> None:
        source_key = self.get_data_source_key()
        source_label = self.get_data_source_label()
        source_detail = (
            self.summarize_config_path("excel_path", "Excel secilmedi", "file")
            if source_key == "excel"
            else self.summarize_sql_source()
        )
        preview = self.summarize_config_path("preview_image_root", "Fotograf klasoru yok", "dir")
        root = self.summarize_config_path("search_root", "Arama klasoru yok", "dir")
        self.settings_summary_var.set(
            f"Kaynak: {source_label} - {source_detail}\nFotograf: {preview}\nExplorer: {root}\nIndeks: {self.get_index_last_rebuild_text()}"
        )

    def open_settings_dialog(self) -> None:
        dialog = SettingsDialog(self, self.settings)
        self.wait_window(dialog)
        if dialog.result is None:
            return
        self.settings.update(dialog.result)
        self.save_settings(self.settings)
        self.search_folder_cache.clear()
        self.channel_folder_cache.clear()
        self.preview_image_cache.clear()
        self.preview_search_roots_cache = None
        self.preview_thumbnail_cache.clear()
        self.ydk_image_root_var.set(self.settings.get("ydk_image_root", ""))
        self.ydk_output_root_var.set(self.settings.get("ydk_output_root", ""))
        self.refresh_settings_summary()
        self.refresh_product_data()

    def browse_ydk_image_root(self) -> None:
        selected = filedialog.askdirectory(title="YDK urun gorsel klasoru sec")
        if selected:
            self.ydk_image_root_var.set(selected)
            self.persist_runtime_state()
            if self.ydk_current_product is not None:
                self.set_current_ydk_product(self.ydk_current_product)

    def browse_ydk_output_root(self) -> None:
        selected = filedialog.askdirectory(title="YDK PDF cikis klasoru sec")
        if selected:
            self.ydk_output_root_var.set(selected)
            self.persist_runtime_state()

    def get_ydk_label_layout(self, include_ui: bool = True) -> dict[str, int]:
        layout = DEFAULT_YDK_LABEL_LAYOUT.copy()
        try:
            saved = json.loads(self.settings.get("ydk_label_layout", "{}") or "{}")
        except json.JSONDecodeError:
            saved = {}
        if isinstance(saved, dict):
            for key, value in saved.items():
                if key in layout:
                    try:
                        layout[key] = int(float(value))
                    except (TypeError, ValueError):
                        continue
        if include_ui:
            for key, variable in getattr(self, "ydk_layout_vars", {}).items():
                if key in layout:
                    try:
                        layout[key] = int(float(variable.get().strip()))
                    except ValueError:
                        continue
        return layout

    def apply_ydk_layout_changes(self) -> None:
        self.settings["ydk_label_layout"] = json.dumps(self.get_ydk_label_layout(), ensure_ascii=False)
        self.save_settings(self.settings)
        if self.ydk_current_product is not None:
            self.refresh_ydk_preview(self.ydk_current_product)
        self.set_status("Etiket tasarim ayarlari kaydedildi.")

    def reset_ydk_layout(self) -> None:
        for key, value in DEFAULT_YDK_LABEL_LAYOUT.items():
            if key in self.ydk_layout_vars:
                self.ydk_layout_vars[key].set(str(value))
        self.apply_ydk_layout_changes()

    def ydk_product_to_dict(self, product: YdkProduct) -> dict[str, str]:
        return {
            key: stringify_cell(getattr(product, key))
            for key in YdkProduct.__dataclass_fields__
        }

    def apply_ydk_product_override(self, product: YdkProduct) -> YdkProduct:
        key = normalize_text(product.code)
        override = self.ydk_label_overrides.get(key, {})
        if not isinstance(override, dict):
            return product
        values = self.ydk_product_to_dict(product)
        for field_name in values:
            if field_name in override:
                values[field_name] = stringify_cell(override[field_name])
        return YdkProduct(**values)

    def set_ydk_editor_product(self, product: YdkProduct) -> None:
        if not hasattr(self, "ydk_editor_vars"):
            return
        self.ydk_editor_loading = True
        try:
            values = self.ydk_product_to_dict(product)
            for key, variable in self.ydk_editor_vars.items():
                variable.set(values.get(key, ""))
            for key, widget in self.ydk_editor_texts.items():
                widget.delete("1.0", "end")
                widget.insert("1.0", values.get(key, ""))
        finally:
            self.ydk_editor_loading = False

    def collect_ydk_editor_product(self) -> YdkProduct | None:
        if self.ydk_current_product is None:
            return None
        values = self.ydk_product_to_dict(self.ydk_current_product)
        for key, variable in self.ydk_editor_vars.items():
            values[key] = variable.get().strip()
        for key, widget in self.ydk_editor_texts.items():
            values[key] = widget.get("1.0", "end").strip()
        values["unit_barcode"] = ean13_full_code(values.get("unit_barcode", "")) or values.get("unit_barcode", "")
        values["carton_barcode"] = ean13_full_code(values.get("carton_barcode", "")) or values.get("carton_barcode", "")
        return YdkProduct(**values)

    def apply_ydk_editor_changes(self) -> None:
        product = self.collect_ydk_editor_product()
        if product is None:
            messagebox.showwarning("Secim Yok", "Once bir urun sec.")
            return
        key = normalize_text(product.code)
        self.ydk_current_product = product
        self.ydk_products[key] = product
        self.ydk_label_overrides[key] = self.ydk_product_to_dict(product)
        for index, existing in enumerate(self.ydk_result_products):
            if normalize_text(existing.code) == key:
                self.ydk_result_products[index] = product
                break
        self.save_ydk_state()
        self.update_ydk_tree_row(product)
        self.refresh_ydk_preview(product)
        self.set_status(f"{product.code} etiket duzenlemesi kaydedildi.")

    def reset_ydk_editor_changes(self) -> None:
        if self.ydk_current_product is None:
            messagebox.showwarning("Secim Yok", "Once bir urun sec.")
            return
        code = self.ydk_current_product.code
        self.ydk_label_overrides.pop(normalize_text(code), None)
        self.save_ydk_state()
        self.load_ydk_workbook(show_success=False)
        if code:
            self.ydk_search_var.set(code)
            self.search_ydk_products()
        self.set_status(f"{code} icin kaynak degerlerine donuldu.")

    def update_ydk_tree_row(self, product: YdkProduct) -> None:
        if not hasattr(self, "ydk_tree"):
            return
        for item in self.ydk_tree.get_children():
            try:
                index = int(item)
            except ValueError:
                continue
            if index < len(self.ydk_result_products) and normalize_text(self.ydk_result_products[index].code) == normalize_text(product.code):
                self.ydk_tree.item(item, values=(product.code, product.model, product.producer_code, product.unit_barcode or "-"))
                return

    def refresh_ydk_preview(self, product: YdkProduct) -> None:
        image_path = self.find_ydk_product_image(product.code)
        self.update_ydk_product_image(image_path)
        label = self.render_ydk_label_image(product, "unit", image_path).resize((260, 312), Image.Resampling.LANCZOS)
        self.ydk_preview_photo_ref = ImageTk.PhotoImage(label)
        self.ydk_label_preview.configure(image=self.ydk_preview_photo_ref, text="")
        self.ydk_detail_var.set(self.build_ydk_summary_text(product, image_path))

    def load_ydk_workbook(self, show_success: bool = True) -> None:
        if not hasattr(self, "ydk_tree"):
            return
        self.persist_runtime_state()
        if not self.records:
            self.ydk_products = {}
            self.populate_ydk_tree([])
            message = self.product_data_warning or "Ana veri kaynagi henuz yuklenmedi. Ayarlar menusunden Excel veya SQL kaynagini sec."
            self.ydk_detail_var.set(message)
            self.set_status("Yurtdisi etiket icin ana veri kaynagi bekleniyor.")
            if show_success:
                messagebox.showwarning("Etiket Verisi Yok", message)
            return

        products = self.parse_ydk_records(self.records)
        self.ydk_products = products
        listed = sorted(products.values(), key=lambda item: item.code.casefold())[:500]
        self.populate_ydk_tree(listed)
        self.ydk_detail_var.set(f"{len(products)} urun ana veri kaynagindan etiket icin hazirlandi. Stok kodu veya aciklama arayabilirsin.")
        self.set_status(f"Yurtdisi etiket verisi hazir: {len(products)} urun.")
        if show_success:
            messagebox.showinfo("Etiket Verisi Hazir", f"{len(products)} urun ana veri kaynagindan hazirlandi.")

    def parse_ydk_records(self, records: list[ProductRecord]) -> dict[str, YdkProduct]:
        products: dict[str, YdkProduct] = {}
        for record in records:
            product = self.build_ydk_product_from_record(record)
            if product.code:
                product = self.apply_ydk_product_override(product)
                products[normalize_text(product.code)] = product
        return {key: value for key, value in products.items() if value.code}

    def build_ydk_product_from_record(self, record: ProductRecord) -> YdkProduct:
        values: dict[str, str] = {}
        values.update(record.raw_values)
        values.update(record.features)
        values.setdefault("Urun Ailesi", record.family)
        values.setdefault("Kirilim", record.breakdown)
        values.setdefault("Stok Kodu", record.stock_code)

        description_tr = first_matching_value(
            values,
            [("adi",), ("tr", "aciklama"), ("turkce",), ("aciklama",), ("urun", "adi"), ("kirilim",)],
            excludes=("adi2", "ingilizce", "english", "description", "barkod", "ean", "koli"),
        )
        description_en = first_matching_value(
            values,
            [("adi2",), ("ad2",), ("name2",), ("en", "aciklama"), ("ingilizce",), ("english",), ("description",)],
            excludes=("tr", "turkce", "barkod", "ean", "koli"),
        )
        model = first_matching_value(values, [("ozelkod3",), ("model",), ("koleksiyon",), ("aile",)], excludes=("kirilim",)) or record.family
        producer_code = (
            first_matching_value(
                values,
                [("uretici_kodu",), ("uretici", "kod"), ("ukod",), ("u_kod",), ("hka",), ("producer", "code"), ("supplier", "code"), ("tedarikci", "kod")],
                excludes=("stok", "urun", "barkod", "ean", "koli"),
            )
            or first_value_matching_regex(values, r"\bHKA[-\s]?\d+\b")
        )
        unit_barcode = first_ean_value(
            values,
            [("urun_barkod",), ("urun", "barkod"), ("unit", "barcode"), ("ean",), ("gtin",), ("barkod",), ("barcode",)],
            excludes=("koli", "carton", "case"),
        )
        carton_barcode = ean13_full_code(
            first_matching_value(
                values,
                [("koli_barkod",), ("koli", "barkod"), ("carton", "barcode"), ("case", "barcode"), ("box", "barcode")],
            )
        )
        carton_quantity = first_matching_value(
            values,
            [("koli", "adet"), ("koli", "ici"), ("carton", "quantity"), ("case", "qty"), ("box", "qty")],
            excludes=("barkod", "barcode", "ean"),
        )
        return YdkProduct(
            code=record.stock_code,
            description_tr=description_tr or record.breakdown,
            description_en=description_en,
            model=model or record.family,
            producer_code=producer_code,
            brand=first_matching_value(values, [("marka",), ("brand",)]),
            unit=first_matching_value(values, [("birim",), ("unit",)]),
            product_type=record.breakdown,
            unit_barcode=unit_barcode,
            carton_barcode=carton_barcode,
            carton_quantity=carton_quantity,
        )

    def search_ydk_products(self) -> None:
        query = normalize_text(self.ydk_search_var.get())
        if not self.ydk_products:
            messagebox.showwarning("Etiket Verisi Yok", "Once ana veri kaynagini yukle.")
            return
        if not query:
            matched = sorted(self.ydk_products.values(), key=lambda item: item.code.casefold())[:500]
        else:
            matched = [
                product
                for product in self.ydk_products.values()
                if query in normalize_text(
                    " ".join([
                        product.code,
                        product.model,
                        product.producer_code,
                        product.description_tr,
                        product.description_en,
                        product.unit_barcode,
                        product.carton_barcode,
                    ])
                )
            ][:500]
        self.populate_ydk_tree(matched)
        self.set_status(f"YDK arama: {len(matched)} sonuc.")

    def clear_ydk_search(self) -> None:
        self.ydk_search_var.set("")
        self.search_ydk_products()

    def populate_ydk_tree(self, products: list[YdkProduct]) -> None:
        self.ydk_result_products = products
        for item in self.ydk_tree.get_children():
            self.ydk_tree.delete(item)
        for index, product in enumerate(products):
            self.ydk_tree.insert(
                "",
                "end",
                iid=str(index),
                values=(product.code, product.model, product.producer_code, product.unit_barcode or "-"),
            )
        if products:
            first = "0"
            self.ydk_tree.selection_set(first)
            self.ydk_tree.focus(first)
            self.set_current_ydk_product(products[0])
        else:
            self.ydk_current_product = None
            self.ydk_preview_photo_ref = None
            self.ydk_product_photo_ref = None
            if hasattr(self, "ydk_label_preview"):
                self.ydk_label_preview.configure(image="", text="Etiket onizleme")
                self.ydk_product_image_label.configure(image="", text="Urun gorseli")

    def on_ydk_selection_changed(self, _event: tk.Event | None = None) -> None:
        selection = self.ydk_tree.selection()
        if not selection:
            return
        try:
            product = self.ydk_result_products[int(selection[0])]
        except (ValueError, IndexError):
            return
        self.set_current_ydk_product(product)

    def set_current_ydk_product(self, product: YdkProduct) -> None:
        self.ydk_current_product = product
        self.set_ydk_editor_product(product)
        self.refresh_ydk_preview(product)

    def find_ydk_product_image(self, stock_code: str) -> Path | None:
        roots = [
            Path(self.ydk_image_root_var.get().strip()).expanduser(),
            Path(self.settings.get("preview_image_root", "")).expanduser(),
        ]
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        stock_key = normalize_text(stock_code)
        for root in unique_paths([path for path in roots if path_is_dir(path)]):
            for extension in valid:
                direct = root / f"{stock_code}{extension}"
                if path_is_file(direct):
                    return direct
            best: Path | None = None
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if path.suffix.lower() not in valid:
                        continue
                    text = normalize_text(str(path))
                    if normalize_text(path.stem).startswith(stock_key) or stock_key in text:
                        best = path
                        break
                if best is not None:
                    return best
        return None

    def update_ydk_product_image(self, image_path: Path | None) -> None:
        if image_path is None or not path_is_file(image_path):
            self.ydk_product_photo_ref = None
            self.ydk_product_image_label.configure(image="", text="Gorsel yok")
            return
        try:
            image = Image.open(image_path)
            image = ImageOps.exif_transpose(image)
            image.thumbnail((220, 160))
            canvas = Image.new("RGB", (230, 170), "#eef3f7")
            canvas.paste(image.convert("RGB"), ((230 - image.width) // 2, (170 - image.height) // 2))
            self.ydk_product_photo_ref = ImageTk.PhotoImage(canvas)
            self.ydk_product_image_label.configure(image=self.ydk_product_photo_ref, text="")
        except Exception:  # noqa: BLE001
            self.ydk_product_photo_ref = None
            self.ydk_product_image_label.configure(image="", text="Gorsel acilamadi")

    def build_ydk_summary_text(self, product: YdkProduct, image_path: Path | None = None) -> str:
        return "\n".join([
            f"Stok Kodu: {product.code}",
            f"Model: {product.model or '-'}",
            f"Uretici Kodu: {product.producer_code or '-'}",
            f"TR Aciklama: {product.description_tr or '-'}",
            f"EN Description: {product.description_en or '-'}",
            f"Urun Barkodu: {product.unit_barcode or '-'}",
            f"Koli Barkodu: {product.carton_barcode or '-'}",
            f"Koli Ici Adet: {product.carton_quantity or '-'}",
            f"Gorsel: {image_path if image_path else 'Bulunamadi'}",
            f"EAN Font Metni: {ean13_font_text(product.unit_barcode) or '-'}",
        ])

    def copy_ydk_summary(self) -> None:
        if self.ydk_current_product is None:
            messagebox.showwarning("Secim Yok", "Once bir YDK urunu sec.")
            return
        self.copy_to_clipboard(self.build_ydk_summary_text(self.ydk_current_product, self.find_ydk_product_image(self.ydk_current_product.code)), "YDK bilgisi")

    def load_ydk_logo(self) -> Image.Image | None:
        if self.ydk_logo_image is not None:
            return self.ydk_logo_image
        logo_path = YDK_LABEL_LOGO_PNG if path_is_file(YDK_LABEL_LOGO_PNG) else YDK_LOGO_PNG
        if not path_is_file(logo_path):
            return None
        try:
            self.ydk_logo_image = Image.open(logo_path).convert("RGBA")
        except Exception:  # noqa: BLE001
            self.ydk_logo_image = None
        return self.ydk_logo_image

    def load_ydk_icon(self) -> Image.Image | None:
        if self.ydk_icon_image is not None:
            return self.ydk_icon_image
        if not path_is_file(YDK_LABEL_ICON_PNG):
            return None
        try:
            self.ydk_icon_image = Image.open(YDK_LABEL_ICON_PNG).convert("RGBA")
        except Exception:  # noqa: BLE001
            self.ydk_icon_image = None
        return self.ydk_icon_image

    def get_ydk_font(self, size: int, bold: bool = False, italic: bool = False) -> ImageFont.ImageFont:
        if bold and italic:
            file_names = ("arialbi.ttf", "segoeuiz.ttf")
        elif bold:
            file_names = ("arialbd.ttf", "segoeuib.ttf")
        elif italic:
            file_names = ("ariali.ttf", "segoeuii.ttf")
        else:
            file_names = ("arial.ttf", "segoeui.ttf")
        font_candidates = [
            Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / file_name
            for file_name in file_names
        ]
        for candidate in font_candidates:
            if path_is_file(candidate):
                try:
                    return ImageFont.truetype(str(candidate), size)
                except OSError:
                    continue
        return ImageFont.load_default()

    def wrap_ydk_text(self, draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int, max_lines: int = 2) -> list[str]:
        words = str(text or "").split()
        lines: list[str] = []
        current = ""
        for word in words:
            trial = f"{current} {word}".strip()
            if draw.textlength(trial, font=font) <= max_width or not current:
                current = trial
            else:
                lines.append(current)
                current = word
            if len(lines) >= max_lines:
                break
        if current and len(lines) < max_lines:
            lines.append(current)
        if len(lines) == max_lines and words:
            while draw.textlength(lines[-1] + "...", font=font) > max_width and len(lines[-1]) > 3:
                lines[-1] = lines[-1][:-1].rstrip()
            if " ".join(lines).casefold() != " ".join(words).casefold():
                lines[-1] += "..."
        return lines or [""]

    def render_ydk_label_image(self, product: YdkProduct, label_type: str, image_path: Path | None = None) -> Image.Image:
        width, height = YDK_LABEL_SIZE
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)
        layout = self.get_ydk_label_layout()

        def fitted_font(text: str, size: int, max_width: int, bold: bool = False, italic: bool = False) -> ImageFont.ImageFont:
            font = self.get_ydk_font(size, bold=bold, italic=italic)
            while size > 12 and draw.textlength(text, font=font) > max_width:
                size -= 2
                font = self.get_ydk_font(size, bold=bold, italic=italic)
            return font

        def ellipsized_text(text: str, font: ImageFont.ImageFont, max_width: int) -> str:
            clean_text = str(text or "-")
            if draw.textlength(clean_text, font=font) <= max_width:
                return clean_text
            suffix = "..."
            while len(clean_text) > 3 and draw.textlength(clean_text.rstrip() + suffix, font=font) > max_width:
                clean_text = clean_text[:-1].rstrip()
            return clean_text.rstrip() + suffix

        logo = self.load_ydk_logo()
        if logo is not None:
            logo_copy = logo.copy()
            logo_copy.thumbnail((layout["logo_w"], layout["logo_h"]))
            image.paste(logo_copy.convert("RGB"), (layout["logo_x"], layout["logo_y"]))
        else:
            draw.text((layout["logo_x"] + 12, layout["logo_y"] + 24), "HEKA", fill="black", font=self.get_ydk_font(70, True))

        product_box = (
            layout["photo_x"],
            layout["photo_y"],
            layout["photo_x"] + layout["photo_w"],
            layout["photo_y"] + layout["photo_h"],
        )
        def safe_label_text_width(x: int, y: int, fallback_width: int = 660) -> int:
            full_width = min(fallback_width, width - x - 24)
            photo_guard_bottom = product_box[3] + 26
            if x < product_box[0] and y < photo_guard_bottom:
                return max(120, min(full_width, product_box[0] - x - 22))
            return max(120, full_width)

        def full_label_text_width(x: int, fallback_width: int = 660) -> int:
            return max(120, min(fallback_width, width - x - 24))

        if image_path is not None and path_is_file(image_path):
            try:
                product_image = Image.open(image_path)
                product_image = ImageOps.exif_transpose(product_image)
                product_image = ImageOps.fit(product_image.convert("RGB"), (product_box[2] - product_box[0], product_box[3] - product_box[1]), method=Image.Resampling.LANCZOS)
                image.paste(product_image, (product_box[0], product_box[1]))
            except Exception:  # noqa: BLE001
                draw.rectangle(product_box, outline="#d8dee6", width=2)
        else:
            draw.rectangle(product_box, outline="#d8dee6", width=2)
            placeholder = "URUN GORSELI"
            placeholder_font = self.get_ydk_font(24, True)
            draw.text(
                (product_box[0] + ((product_box[2] - product_box[0]) - draw.textlength(placeholder, font=placeholder_font)) / 2, 150),
                placeholder,
                fill="#8892a0",
                font=placeholder_font,
            )

        model_text = (product.model or product.code or "-").upper()
        model_font = fitted_font(
            model_text,
            layout["model_font"],
            min(300, safe_label_text_width(layout["model_x"], layout["model_y"], 300)),
            bold=True,
        )
        draw.text((layout["model_x"], layout["model_y"]), model_text, fill="black", font=model_font)

        tr_text = (product.description_tr or product.product_type or "-").upper()
        tr_y = max(layout["tr_y"], product_box[3] + 34)
        tr_width = full_label_text_width(layout["tr_x"])
        tr_font = fitted_font(tr_text, layout["tr_font"], tr_width, bold=True)
        draw.text((layout["tr_x"], tr_y), ellipsized_text(tr_text, tr_font, tr_width), fill="black", font=tr_font)

        en_text = (product.description_en or "").upper()
        en_y = max(layout["en_y"], tr_y + 72)
        en_width = full_label_text_width(layout["en_x"])
        en_font = fitted_font(en_text, layout["en_font"], en_width, italic=True)
        draw.text((layout["en_x"], en_y), ellipsized_text(en_text, en_font, en_width), fill="black", font=en_font)

        barcode = product.unit_barcode if label_type == "unit" else product.carton_barcode
        barcode_image = self.create_ean13_barcode_image(barcode, layout["barcode_w"], layout["barcode_h"])
        if barcode_image is not None:
            image.paste(barcode_image, (layout["barcode_x"], layout["barcode_y"]))
        else:
            draw.rectangle(
                (
                    layout["barcode_x"],
                    layout["barcode_y"],
                    layout["barcode_x"] + layout["barcode_w"],
                    layout["barcode_y"] + layout["barcode_h"],
                ),
                outline="#b8c2cf",
                width=2,
            )
            draw.text((layout["barcode_x"] + 95, layout["barcode_y"] + 55), "BARKOD YOK", fill="#9b1c1c", font=self.get_ydk_font(30, True))

        heading = "ÜRÜN KODU" if label_type == "unit" else "KOLİ KODU"
        heading_font = self.get_ydk_font(30, True)
        heading_x, heading_y = layout["info_x"], layout["heading_y"]
        draw.text((heading_x, heading_y), heading, fill="black", font=heading_font)
        heading_width = draw.textlength(heading, font=heading_font)
        draw.line((heading_x, heading_y + 36, heading_x + heading_width, heading_y + 36), fill="black", width=3)

        code_font = fitted_font(product.code, 29, 295, bold=True)
        draw.text((layout["info_x"], layout["code_y"]), product.code, fill="black", font=code_font)

        producer = product.producer_code or "-"
        box = (
            layout["producer_box_x"],
            layout["producer_box_y"],
            layout["producer_box_x"] + layout["producer_box_w"],
            layout["producer_box_y"] + layout["producer_box_h"],
        )
        draw.rectangle(box, outline="black", width=4)
        producer_font = fitted_font(producer, layout["producer_font"], box[2] - box[0] - 24, bold=True)
        producer_bbox = draw.textbbox((0, 0), producer, font=producer_font)
        producer_x = box[0] + ((box[2] - box[0]) - (producer_bbox[2] - producer_bbox[0])) / 2
        producer_y = box[1] + ((box[3] - box[1]) - (producer_bbox[3] - producer_bbox[1])) / 2 - 3
        draw.text((producer_x, producer_y), producer, fill="black", font=producer_font)

        icon = self.load_ydk_icon()
        bottom_font = self.get_ydk_font(layout["footer_font"])
        if icon is not None:
            left_icon = icon.copy()
            left_icon.thumbnail((35, 45))
            image.paste(left_icon, (36, layout["footer_y"] - 7), left_icon)
            right_icon = icon.copy()
            right_icon.thumbnail((35, 45))
            image.paste(right_icon, (437, layout["footer_y"] - 7), right_icon)
        draw.text((84, layout["footer_y"]), "www.hekalighting.com", fill="black", font=bottom_font)
        draw.text((480, layout["footer_y"]), "+90 850 711 47 45", fill="black", font=bottom_font)
        return image

    def create_ean13_barcode_image(self, value: str, width: int, height: int) -> Image.Image | None:
        code = ean13_full_code(value)
        if len(code) != 13:
            return None
        l_codes = {"0": "0001101", "1": "0011001", "2": "0010011", "3": "0111101", "4": "0100011", "5": "0110001", "6": "0101111", "7": "0111011", "8": "0110111", "9": "0001011"}
        g_codes = {"0": "0100111", "1": "0110011", "2": "0011011", "3": "0100001", "4": "0011101", "5": "0111001", "6": "0000101", "7": "0010001", "8": "0001001", "9": "0010111"}
        r_codes = {"0": "1110010", "1": "1100110", "2": "1101100", "3": "1000010", "4": "1011100", "5": "1001110", "6": "1010000", "7": "1000100", "8": "1001000", "9": "1110100"}
        parity = {
            "0": "LLLLLL",
            "1": "LLGLGG",
            "2": "LLGGLG",
            "3": "LLGGGL",
            "4": "LGLLGG",
            "5": "LGGLLG",
            "6": "LGGGLL",
            "7": "LGLGLG",
            "8": "LGLGGL",
            "9": "LGGLGL",
        }
        sequence = "101"
        for digit, side in zip(code[1:7], parity[code[0]], strict=False):
            sequence += l_codes[digit] if side == "L" else g_codes[digit]
        sequence += "01010"
        for digit in code[7:]:
            sequence += r_codes[digit]
        sequence += "101"
        quiet_left = 40
        quiet_right = 18
        module_width = (width - quiet_left - quiet_right) / len(sequence)
        barcode = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(barcode)
        bar_height = height - 44
        guard_indexes = set(range(3)) | set(range(45, 50)) | set(range(92, 95))
        x = float(quiet_left)
        for index, bit in enumerate(sequence):
            if bit == "1":
                extra = 18 if index in guard_indexes else 0
                draw.rectangle((round(x), 0, round(x + module_width), bar_height + extra), fill="black")
            x += module_width
        font = self.get_ydk_font(35)
        text_y = height - 43
        draw.text((2, text_y), code[0], fill="black", font=font)
        left_digits = code[1:7]
        right_digits = code[7:]
        left_start = quiet_left + 3 * module_width
        left_width = 42 * module_width
        right_start = quiet_left + 50 * module_width
        right_width = 42 * module_width
        draw.text((left_start + (left_width - draw.textlength(left_digits, font=font)) / 2, text_y), left_digits, fill="black", font=font)
        draw.text((right_start + (right_width - draw.textlength(right_digits, font=font)) / 2, text_y), right_digits, fill="black", font=font)
        return barcode

    def get_ydk_output_root(self) -> Path:
        text = self.ydk_output_root_var.get().strip()
        return Path(text).expanduser() if text else APP_DIR / "ydk_exports"

    def build_ydk_pdf_path(self, product: YdkProduct, label_type: str) -> Path:
        date_folder = "YURTDISI ETIKET_" + datetime.now().strftime("%d.%m.%Y")
        folder = self.get_ydk_output_root() / date_folder
        suffix = "_KOLI_BARKOD" if label_type == "carton" else ""
        name = safe_path_part(f"{product.producer_code or product.code} - {product.code}_{product.model}{suffix}.pdf", "ydk_etiket.pdf")
        return folder / name

    def export_ydk_label_pdf(self, label_type: str) -> Path | None:
        if self.ydk_current_product is None:
            messagebox.showwarning("Secim Yok", "Once bir YDK urunu sec.")
            return None
        product = self.collect_ydk_editor_product() or self.ydk_current_product
        key = normalize_text(product.code)
        self.ydk_current_product = product
        self.ydk_products[key] = product
        self.ydk_label_overrides[key] = self.ydk_product_to_dict(product)
        self.save_ydk_state()
        barcode = product.unit_barcode if label_type == "unit" else product.carton_barcode
        if not ean13_full_code(barcode):
            messagebox.showwarning("Barkod Yok", "Bu urun icin secilen etiket tipinde barkod bulunamadi.")
            return None
        image_path = self.find_ydk_product_image(product.code)
        label_image = self.render_ydk_label_image(product, label_type, image_path)
        pdf_path = self.build_ydk_pdf_path(product, label_type)
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        label_image.save(pdf_path, "PDF", resolution=300.0)
        self.set_status(f"YDK PDF kaydedildi: {pdf_path.name}")
        messagebox.showinfo("PDF Kaydedildi", str(pdf_path))
        return pdf_path

    def export_both_ydk_labels(self) -> None:
        if self.ydk_current_product is None:
            messagebox.showwarning("Secim Yok", "Once bir YDK urunu sec.")
            return
        saved: list[Path] = []
        unit = self.export_ydk_label_pdf("unit")
        if unit is not None:
            saved.append(unit)
        if self.ydk_current_product and ean13_full_code(self.ydk_current_product.carton_barcode):
            carton = self.export_ydk_label_pdf("carton")
            if carton is not None:
                saved.append(carton)
        if saved:
            self.set_status(f"{len(saved)} YDK PDF kaydedildi.")

    def parse_feature_aliases(self, headers: list[str]) -> dict[int, str]:
        alias_map: dict[int, str] = {}
        for item in split_multivalue_config(self.settings.get("feature_aliases", "")):
            separator = "=" if "=" in item else ":" if ":" in item else None
            if separator is None:
                continue
            left, right = [part.strip() for part in item.split(separator, 1)]
            left_index = resolve_header_first_index(left, headers)
            right_index = resolve_header_first_index(right, headers)
            if left_index is not None and right:
                alias_map[left_index] = prettify_feature_name(right)
            elif right_index is not None and left:
                alias_map[right_index] = prettify_feature_name(left)
        return alias_map

    def build_feature_label(self, column_index: int, headers: list[str], aliases: dict[int, str]) -> str:
        alias = prettify_feature_name(aliases.get(column_index, ""))
        if alias:
            return alias
        header = prettify_feature_name(headers[column_index])
        return header if header and not is_generic_header(header) else f"Alan {column_index + 1}"

    def refresh_product_data(self, show_success: bool = True) -> None:
        self.product_data_warning = ""

        def job() -> list[ProductRecord]:
            try:
                return self.load_product_records()
            except (FileNotFoundError, PermissionError, OSError) as exc:
                self.product_data_warning = f"Veri kaynagina erisilemiyor: {exc}"
                return []

        def on_success(records: list[ProductRecord]) -> None:
            self.records = records
            families = unique_preserve_order([record.family for record in records if record.family])
            self.family_combo.set_values(families)
            if families:
                preferred = self.settings.get("last_family", "")
                family = next((item for item in families if normalize_text(item) == normalize_text(preferred)), families[0])
                self.family_var.set(family)
                self.update_breakdown_options(family, self.settings.get("last_breakdown", ""))
            else:
                self.family_var.set("")
                self.breakdown_var.set("")
                self.breakdown_combo.set_values([])
            if self.product_data_warning:
                self.results_summary_var.set("Ayar gerekiyor")
                self.results_context_var.set(self.product_data_warning)
                self.clear_results("Veri kaynagi veya klasor yolu bu bilgisayarda erisilebilir degilse Ayarlar'dan yeni yolu sec.")
                self.set_status(self.product_data_warning)
                if show_success:
                    messagebox.showwarning("Veri Kaynagi", self.product_data_warning)
                self.load_ydk_workbook(show_success=False)
                return
            summary = "Veri kaynagi hazir" if records else "Veri kaynagi bekleniyor"
            context = f"{len(records)} kayit, {len(families)} aile yuklendi." if records else "Ayarlar'dan Excel veya SQL kaynagini secerek baslayabilirsin."
            self.results_summary_var.set(summary)
            self.results_context_var.set(context)
            self.clear_results("Bir urun ailesi secip kirilim getirerek sonuclari burada gorebilirsin.")
            self.set_status(f"{len(records)} kayit yuklendi." if records else "Veri kaynagi ayari bekleniyor.")
            self.load_ydk_workbook(show_success=False)
            if show_success and records:
                messagebox.showinfo("Veri Yenilendi", f"{len(records)} kayit yuklendi.")

        self.run_in_background(job, on_success, "Veri kaynagi yukleniyor...")

    def load_product_records(self) -> list[ProductRecord]:
        return self.load_sql_product_records() if self.get_data_source_key() == "sql" else self.load_excel_product_records()

    def load_excel_product_records(self) -> list[ProductRecord]:
        excel_text = self.settings.get("excel_path", "").strip()
        if not excel_text:
            return []
        excel_path = Path(excel_text).expanduser()
        if not path_is_file(excel_path):
            self.product_data_warning = f"Excel dosyasina erisilemiyor: {excel_path}"
            return []
        header_row = int(self.settings.get("header_row", "1") or "1")
        try:
            workbook = load_workbook(excel_path, data_only=True, read_only=True)
        except (FileNotFoundError, PermissionError, OSError) as exc:
            self.product_data_warning = f"Excel dosyasina erisilemiyor: {exc}"
            return []
        try:
            sheet_name = self.settings.get("sheet_name", "").strip()
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            headers = load_excel_headers(excel_path, sheet_name, header_row)
            family_index = resolve_column_index(self.settings["family_column"], headers)
            breakdown_index = resolve_column_index(self.settings["breakdown_column"], headers)
            stock_index = resolve_column_index(self.settings["stock_column"], headers)
            if family_index is None or breakdown_index is None or stock_index is None:
                raise ValueError("Aile, kirilim veya stok sutunu cozulmedi.")
            reserved = {family_index, breakdown_index, stock_index}
            feature_specs = parse_csv(self.settings.get("feature_columns", ""))
            if feature_specs:
                feature_indexes = [
                    resolved
                    for spec in feature_specs
                    if (resolved := resolve_column_index(spec, headers)) is not None and resolved not in reserved
                ]
            else:
                feature_indexes = [index for index in range(len(headers)) if index not in reserved]
            aliases = self.parse_feature_aliases(headers)
            records: list[ProductRecord] = []
            for row_values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                row = [stringify_cell(value) for value in row_values]
                if not any(row):
                    continue
                while len(row) < len(headers):
                    row.append("")
                stock = row[stock_index].strip()
                if not stock:
                    continue
                features = {
                    self.build_feature_label(index, headers, aliases): row[index]
                    for index in feature_indexes
                    if index < len(row) and row[index]
                }
                raw = {
                    (self.build_feature_label(index, headers, aliases) if index in feature_indexes else headers[index]): row[index]
                    for index in range(len(headers))
                    if index < len(row) and row[index]
                }
                records.append(ProductRecord(row[family_index].strip(), row[breakdown_index].strip(), stock, features, raw))
            return records
        finally:
            workbook.close()

    def open_sql_connection(self) -> tuple[Any, str]:
        connection_text = self.settings.get("sql_connection_string", "").strip()
        if not connection_text:
            raise ValueError("SQL baglanti bilgisi girilmedi.")
        lowered = connection_text.casefold()
        looks_like_sqlite = lowered.startswith("file:") or lowered.startswith("sqlite:///") or lowered.endswith((".sqlite", ".sqlite3", ".db", ".db3"))
        if looks_like_sqlite:
            sqlite_target = connection_text[10:] if lowered.startswith("sqlite:///") else connection_text
            sqlite_uri = sqlite_target.startswith("file:")
            if not sqlite_uri:
                sqlite_target = str(Path(sqlite_target).expanduser())
            return sqlite3.connect(sqlite_target, uri=sqlite_uri), "sqlite"
        try:
            import pyodbc  # type: ignore[import-not-found]
        except ImportError as exc:  # noqa: F401
            raise RuntimeError("ODBC baglantilari icin pyodbc kurulu degil. SQLite dosya yolu ya da URI kullanabilir veya pyodbc ekleyebilirsin.") from exc
        try:
            return pyodbc.connect(connection_text), "odbc"
        except Exception as exc:
            fallback = self.build_trusted_odbc_fallback(connection_text)
            if fallback:
                try:
                    return pyodbc.connect(fallback), "odbc"
                except Exception:
                    pass
            raise exc

    def build_trusted_odbc_fallback(self, connection_text: str) -> str:
        parts = [part.strip() for part in connection_text.split(";") if part.strip()]
        lowered_keys = {part.split("=", 1)[0].strip().casefold() for part in parts if "=" in part}
        if "pwd" in lowered_keys or "password" in lowered_keys or "trusted_connection" in lowered_keys:
            return ""
        filtered = [
            part
            for part in parts
            if part.split("=", 1)[0].strip().casefold() not in {"uid", "user", "user id"}
        ]
        filtered.append("Trusted_Connection=yes")
        return ";".join(filtered)

    def detect_sql_table(self, cursor: Any, driver_kind: str) -> str:
        preferred = self.settings.get("sql_table", "").strip()
        if preferred:
            return preferred
        if driver_kind == "sqlite":
            row = cursor.execute(
                "SELECT name FROM sqlite_master WHERE type IN ('table', 'view') AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchone()
            return str(row[0]).strip() if row and row[0] else ""
        if hasattr(cursor, "tables"):
            for row in cursor.tables():
                table_type = normalize_text(getattr(row, "table_type", "") or (row[3] if len(row) > 3 else ""))
                name = str(getattr(row, "table_name", "") or (row[2] if len(row) > 2 else "")).strip()
                if name and table_type in {"table", "view"}:
                    return name
        return ""

    def resolve_sql_query(self, cursor: Any, driver_kind: str) -> str:
        query = self.settings.get("sql_query", "").strip()
        if query:
            return query
        table_name = self.detect_sql_table(cursor, driver_kind)
        if table_name:
            if driver_kind == "sqlite":
                escaped = table_name.replace('"', '""')
                return f'SELECT * FROM "{escaped}"'
            escaped = table_name.replace("]", "]]")
            return f"SELECT * FROM [{escaped}]"
        raise ValueError("SQL sorgusu bossa tablo veya gorunum secilmeli.")

    def load_sql_product_records(self) -> list[ProductRecord]:
        connection, driver_kind = self.open_sql_connection()
        try:
            cursor = connection.cursor()
            query = self.resolve_sql_query(cursor, driver_kind)
            cursor.execute(query)
            description = cursor.description or []
            headers = [stringify_cell(column[0]) or f"Alan {index + 1}" for index, column in enumerate(description)]
            if not headers:
                return []
            family_index = resolve_column_index(self.settings["family_column"], headers)
            breakdown_index = resolve_column_index(self.settings["breakdown_column"], headers)
            stock_index = resolve_column_index(self.settings["stock_column"], headers)
            if family_index is None or breakdown_index is None or stock_index is None:
                raise ValueError("Aile, kirilim veya stok sutunu SQL sonucunda cozulmedi.")
            reserved = {family_index, breakdown_index, stock_index}
            feature_specs = parse_csv(self.settings.get("feature_columns", ""))
            if feature_specs:
                feature_indexes = [
                    resolved
                    for spec in feature_specs
                    if (resolved := resolve_column_index(spec, headers)) is not None and resolved not in reserved
                ]
            else:
                feature_indexes = [index for index in range(len(headers)) if index not in reserved]
            aliases = self.parse_feature_aliases(headers)
            records: list[ProductRecord] = []
            for row_values in cursor.fetchall():
                row = [stringify_cell(value) for value in row_values]
                if not any(row):
                    continue
                while len(row) < len(headers):
                    row.append("")
                stock = row[stock_index].strip()
                if not stock:
                    continue
                features = {
                    self.build_feature_label(index, headers, aliases): row[index]
                    for index in feature_indexes
                    if index < len(row) and row[index]
                }
                raw = {
                    (self.build_feature_label(index, headers, aliases) if index in feature_indexes else headers[index]): row[index]
                    for index in range(len(headers))
                    if index < len(row) and row[index]
                }
                records.append(ProductRecord(row[family_index].strip(), row[breakdown_index].strip(), stock, features, raw))
            return records
        except sqlite3.Error as exc:
            self.product_data_warning = f"SQL baglantisi acilamadi: {exc}"
            return []
        except Exception as exc:
            self.product_data_warning = f"SQL verisi okunamadi: {exc}"
            return []
        finally:
            try:
                connection.close()
            except Exception:  # noqa: BLE001
                pass

    def on_family_changed(self, _event: tk.Event | None = None) -> None:
        self.update_breakdown_options(self.family_var.get())

    def update_breakdown_options(self, family: str, preferred_breakdown: str = "") -> None:
        breakdowns = unique_preserve_order([
            record.breakdown
            for record in self.records
            if normalize_text(record.family) == normalize_text(family) and record.breakdown
        ])
        if breakdowns:
            breakdowns = [ALL_BREAKDOWNS_LABEL, *breakdowns]
        self.breakdown_combo.set_values(breakdowns)
        selected = next(
            (item for item in breakdowns if normalize_text(item) == normalize_text(preferred_breakdown)),
            breakdowns[0] if breakdowns else "",
        )
        self.breakdown_var.set(selected)
        self.persist_runtime_state()

    def fetch_family_products(self) -> None:
        if not self.family_var.get().strip():
            messagebox.showwarning("Eksik Secim", "Lutfen once urun ailesi sec.")
            return
        self.breakdown_var.set(ALL_BREAKDOWNS_LABEL)
        self.fetch_selected_products()

    def fetch_selected_products(self) -> None:
        family = self.family_var.get().strip()
        breakdown = self.breakdown_var.get().strip()
        if not family or not breakdown:
            messagebox.showwarning("Eksik Secim", "Lutfen aile ve kirilim sec.")
            return
        matched = [
            record
            for record in self.records
            if normalize_text(record.family) == normalize_text(family)
            and (
                normalize_text(breakdown) == normalize_text(ALL_BREAKDOWNS_LABEL)
                or normalize_text(record.breakdown) == normalize_text(breakdown)
            )
        ]
        self.persist_runtime_state()
        self.render_matched_products(matched, f"Secim: {family} / {breakdown}")

    def search_products(self) -> None:
        query = self.product_search_var.get().strip()
        if not query:
            messagebox.showwarning("Arama Bos", "Aramak icin bir deger yaz.")
            return
        normalized = normalize_text(query)
        matched = [
            record
            for record in self.records
            if normalized in normalize_text(record.family)
            or normalized in normalize_text(record.breakdown)
            or normalized in normalize_text(record.stock_code)
            or any(normalized in normalize_text(f"{key} {value}") for key, value in record.raw_values.items())
        ]
        self.render_matched_products(matched[:250], f"Arama: {query}")

    def render_matched_products(self, matched: list[ProductRecord], context: str) -> None:
        if not matched:
            self.clear_results("Kayit bulunamadi.")
            self.results_summary_var.set("Sonuc yok")
            self.results_context_var.set(context)
            return

        def job() -> list[tuple[ProductRecord, Path | None, Path | None]]:
            stock_codes = [record.stock_code for record in matched]
            images = self.find_preview_images(stock_codes)
            folders = self.find_search_folders(stock_codes)
            return [
                (record, images.get(normalize_text(record.stock_code)), folders.get(normalize_text(record.stock_code)))
                for record in matched
            ]

        def on_success(items: list[tuple[ProductRecord, Path | None, Path | None]]) -> None:
            self.render_product_cards(items)
            self.results_summary_var.set(f"{len(items)} urun getirildi")
            self.results_context_var.set(context)
            self.set_status(f"{len(items)} urun listelendi.")

        self.run_in_background(job, on_success, "Kayitlar ve gorseller araniyor...")

    def clear_results(self, message: str) -> None:
        for child in self.results_frame.winfo_children():
            child.destroy()
        self.result_image_refs.clear()
        # Empty-state panel (magic-MCP style centered card)
        empty = tk.Frame(self.results_frame, bg="#F8FAFC",
                         padx=40, pady=40)
        empty.pack(anchor="center", padx=60, pady=60)
        tk.Label(empty, text="◈", bg="#F8FAFC", fg="#CBD5E1",
                 font=("Bahnschrift", 36)).pack()
        tk.Label(empty, text=message, bg="#F8FAFC", fg="#64748B",
                 font=("Aptos", 11), wraplength=320,
                 justify="center").pack(pady=(12, 0))

    def collect_product_images(self, stock_code: str, image_path: Path | None, folder_path: Path | None) -> list[Path]:
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        stock_key = normalize_text(stock_code)
        roots = unique_paths(
            [
                path
                for path in [folder_path, image_path.parent if image_path is not None else None]
                if isinstance(path, Path) and path_is_dir(path)
            ]
        )
        matches: list[Path] = []
        seen: set[str] = set()

        def add_match(path: Path) -> None:
            key = str(path).casefold()
            if key not in seen and path_is_file(path):
                seen.add(key)
                matches.append(path)

        if image_path is not None:
            add_match(image_path)
        for root in roots:
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if valid and path.suffix.lower() not in valid:
                        continue
                    norm = normalize_text(str(path))
                    stem = normalize_text(path.stem)
                    if stem == stock_key or stem.startswith(stock_key) or stock_key in norm:
                        add_match(path)
        return sorted(matches, key=lambda path: self.preview_image_rank(path, stock_key))

    def render_product_cards(self, items: list[tuple[ProductRecord, Path | None, Path | None]]) -> None:
        for child in self.results_frame.winfo_children():
            child.destroy()
        self.result_image_refs.clear()
        if not items:
            self.clear_results("Sonuc bulunamadi.")
            return
        columns = max(1, min(3, max(self.results_canvas.winfo_width(), 760) // 430))
        for index, (record, image_path, folder_path) in enumerate(items):
            card = ttk.Frame(self.results_frame, padding=0, style="ProductCard.TFrame")
            card.grid(row=index // columns, column=index % columns,
                      sticky="nsew", padx=10, pady=10)
            self.render_product_card(card, record, image_path, folder_path)

    def render_product_card(self, card: ttk.Frame, record: ProductRecord, image_path: Path | None, folder_path: Path | None) -> None:
        card_images = self.collect_product_images(record.stock_code, image_path, folder_path)
        current_image = {"path": card_images[0] if card_images else image_path}
        image_index  = {"value": 0}
        image_var    = tk.StringVar()
        image_count_var = tk.StringVar()

        # ── Header strip (tinted, magic-MCP style) ───────────────────────────────
        header = tk.Frame(card, bg="#F0F7FF", padx=14, pady=12)
        header.pack(fill="x")

        # Hash + stock code on one line
        stock_row = tk.Frame(header, bg="#F0F7FF")
        stock_row.pack(fill="x")
        tk.Label(stock_row, text="#", bg="#F0F7FF", fg="#94A3B8",
                 font=("Bahnschrift", 11)).pack(side="left")
        tk.Label(stock_row, text=record.stock_code or "-",
                 bg="#F0F7FF", fg="#0F172A",
                 font=("Bahnschrift", 11, "bold")).pack(side="left", padx=(2, 0))

        # Family title
        tk.Label(header, text=record.family or "-",
                 bg="#F0F7FF", fg="#0F172A",
                 font=("Bahnschrift", 15, "bold"),
                 anchor="w", justify="left",
                 wraplength=340).pack(fill="x", pady=(6, 0))

        # Breakdown badge row
        badge_row = tk.Frame(header, bg="#F0F7FF")
        badge_row.pack(fill="x", pady=(6, 0))
        # Orange family badge
        tk.Label(badge_row,
                 text=f"  {record.family or 'Aile yok'}  ",
                 bg="#FF6B35", fg="#FFFFFF",
                 font=("Bahnschrift", 8),
                 padx=2, pady=2).pack(side="left")
        # Breakdown outline badge
        tk.Label(badge_row,
                 text=f"  {record.breakdown or 'Kirilim yok'}  ",
                 bg="#E2E8F0", fg="#475569",
                 font=("Bahnschrift", 8),
                 padx=2, pady=2).pack(side="left", padx=(6, 0))

        # Thin orange accent line under header
        tk.Frame(card, height=2, bg="#FF6B35").pack(fill="x")

        # ── Image carousel ────────────────────────────────────────────────────────
        image_label = ttk.Label(card,
                                text="Gorsel bulunamadi" if current_image["path"] is None else "",
                                width=36, anchor="center",
                                style="ImageStage.TLabel")
        image_label.pack(fill="x", ipady=6)

        # Carousel nav row
        carousel_row = ttk.Frame(card, style="Card.TFrame")
        carousel_row.pack(fill="x", padx=12, pady=(6, 0))
        prev_button = ttk.Button(carousel_row, text="◀", width=3,
                                 style="Carousel.TButton")
        prev_button.pack(side="left")
        ttk.Label(carousel_row, textvariable=image_count_var,
                  style="Pill.TLabel").pack(side="left", padx=8)
        next_button = ttk.Button(carousel_row, text="▶", width=3,
                                 style="Carousel.TButton")
        next_button.pack(side="left")
        # Image filename hint
        ttk.Label(carousel_row, textvariable=image_var,
                  style="CardMuted.TLabel",
                  wraplength=190).pack(side="left", padx=(10, 0))

        def get_current_image() -> Path | None:
            return current_image["path"]

        def update_card_image(step: int = 0) -> None:
            if card_images:
                image_index["value"] = (image_index["value"] + step) % len(card_images)
                current_image["path"] = card_images[image_index["value"]]
                image_count_var.set(f"{image_index['value'] + 1} / {len(card_images)}")
            else:
                current_image["path"] = image_path
                image_count_var.set("0 / 0")
            photo = self.build_preview_image(current_image["path"])
            if photo is not None:
                image_label.configure(image=photo, text="")
                image_label.image = photo
                self.result_image_refs.append(photo)
            else:
                image_label.configure(image="", text="Gorsel bulunamadi")
                image_label.image = None
            image_var.set(
                current_image["path"].name if current_image["path"] else "—"
            )

        prev_button.configure(command=lambda upd=update_card_image: upd(-1))
        next_button.configure(command=lambda upd=update_card_image: upd(1))
        if len(card_images) <= 1:
            prev_button.state(["disabled"])
            next_button.state(["disabled"])
        update_card_image()

        # ── Action buttons row ────────────────────────────────────────────────────
        tk.Frame(card, height=1, bg="#E2E8F0").pack(fill="x", padx=12, pady=(10, 0))
        action_row = ttk.Frame(card, style="Card.TFrame")
        action_row.pack(fill="x", padx=12, pady=(8, 0))

        ttk.Button(
            action_row, text="⎘  Stok Kopyala",
            command=lambda t=record.stock_code: self.copy_to_clipboard(t, "Stok kodu"),
        ).pack(side="left", padx=(0, 5))
        ttk.Button(
            action_row, text="Karti Kopyala",
            command=lambda r=record, ci=get_current_image, f=folder_path: self.copy_to_clipboard(
                self.build_product_copy_text(r, ci(), f), "Urun karti"),
        ).pack(side="left", padx=(0, 5))
        ttk.Button(
            action_row, text="Yollari Kopyala",
            command=lambda ci=get_current_image, f=folder_path: self.copy_to_clipboard(
                self.build_product_paths_copy_text(ci(), f), "Yollar"),
        ).pack(side="left")

        # ── Folder / file access row ──────────────────────────────────────────────
        file_row = ttk.Frame(card, style="Card.TFrame")
        file_row.pack(fill="x", padx=12, pady=(6, 0))
        ttk.Button(file_row, text="📁  Urun Klasoru",
                   style="Ghost.TButton",
                   command=lambda p=folder_path: self.reveal_path(p)).pack(side="left", padx=(0, 5))
        ttk.Button(file_row, text="Gorseli Ac",
                   command=lambda ci=get_current_image: self.open_path_directly(ci()),
        ).pack(side="left")

        # ── Channel buttons ────────────────────────────────────────────────────────
        channel_row = ttk.Frame(card, style="Card.TFrame")
        channel_row.pack(fill="x", padx=12, pady=(6, 0))
        for channel_key, (lbl, _kw) in CHANNEL_FOLDERS.items():
            ttk.Button(
                channel_row, text=lbl,
                command=lambda s=record.stock_code, c=channel_key, h=folder_path:
                    self.reveal_channel_path(s, c, h),
            ).pack(side="left", padx=(0, 5))

        # ── Features grid (2-column, magic-MCP card pattern) ─────────────────────
        features = list(record.features.items())[:8]
        if features:
            tk.Frame(card, height=1, bg="#E2E8F0").pack(fill="x", padx=12, pady=(12, 0))
            feat_header = ttk.Frame(card, style="Card.TFrame")
            feat_header.pack(fill="x", padx=12, pady=(8, 6))
            ttk.Label(feat_header, text="ÖZELLIKLER",
                      style="CardSection.TLabel").pack(side="left")
            grid_frame = ttk.Frame(card, style="Card.TFrame")
            grid_frame.pack(fill="x", padx=12, pady=(0, 12))
            grid_frame.columnconfigure(0, weight=1)
            grid_frame.columnconfigure(1, weight=1)
            for i, (fname, fval) in enumerate(features):
                cell = ttk.Frame(grid_frame, padding=(8, 5),
                                 style="FeatureRow.TFrame")
                cell.grid(row=i // 2, column=i % 2,
                          sticky="nsew", padx=(0 if i % 2 == 0 else 4, 0),
                          pady=3)
                ttk.Label(cell, text=fname,
                          style="FeatureName.TLabel").pack(anchor="w")
                ttk.Label(cell, text=fval, wraplength=145,
                          style="FeatureValue.TLabel").pack(anchor="w")

    def copy_to_clipboard(self, text: str, label: str = "Metin") -> None:
        content = str(text or "").strip()
        if not content:
            messagebox.showwarning("Kopyalanacak Veri Yok", "Bu kartta kopyalanacak bilgi bulunamadi.")
            return
        self.clipboard_clear()
        self.clipboard_append(content)
        self.update_idletasks()
        self.set_status(f"{label} panoya kopyalandi.")

    def build_product_paths_copy_text(self, image_path: Path | None, folder_path: Path | None) -> str:
        lines: list[str] = []
        if image_path is not None:
            lines.append(f"Fotograf: {image_path}")
        if folder_path is not None:
            lines.append(f"Klasor: {folder_path}")
        return "\n".join(lines)

    def build_product_copy_text(self, record: ProductRecord, image_path: Path | None, folder_path: Path | None) -> str:
        lines = [
            f"Urun Ailesi: {record.family or '-'}",
            f"Kirilim: {record.breakdown or '-'}",
            f"Stok Kodu: {record.stock_code or '-'}",
            f"Fotograf: {image_path if image_path else 'Bulunamadi'}",
            f"Klasor: {folder_path if folder_path else 'Bulunamadi'}",
        ]
        if record.features:
            lines.append("")
            lines.append("Ozellikler:")
            lines.extend(f"{name}: {value}" for name, value in record.features.items())
        return "\n".join(lines)

    def describe_path(self, path: Path | None, max_parts: int = 3) -> str:
        if path is None:
            return "Bulunamadi"
        parts = path.parts[-max_parts:]
        return ".../" + "/".join(parts) if len(path.parts) > max_parts else str(path)

    def build_preview_image(self, image_path: Path | None) -> ImageTk.PhotoImage | None:
        if image_path is None or not path_is_file(image_path):
            return None
        key = str(image_path).casefold()
        if key in self.preview_thumbnail_cache:
            return self.preview_thumbnail_cache[key]
        try:
            image = Image.open(image_path)
            image = ImageOps.exif_transpose(image)
            image.thumbnail((320, 240))
            padded = Image.new("RGB", (340, 260), "#e9edf2")
            padded.paste(image.convert("RGB"), ((340 - image.width) // 2, (260 - image.height) // 2))
            photo = ImageTk.PhotoImage(padded)
            if len(self.preview_thumbnail_cache) > 300:
                self.preview_thumbnail_cache.clear()
            self.preview_thumbnail_cache[key] = photo
            return photo
        except Exception:  # noqa: BLE001
            self.preview_thumbnail_cache[key] = None
            return None

    def open_path_directly(self, path: Path | None) -> None:
        if path is None or not path_exists(path):
            messagebox.showwarning("Gorsel Yok", "Gorsel bulunamadi.")
            return
        open_path_with_default_app(path)

    def reveal_path(self, path: Path | None) -> None:
        if path is None:
            messagebox.showwarning("Konum Yok", "Konum bulunamadi.")
            return
        try:
            open_in_explorer(path)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Explorer Hatasi", str(exc))

    def reveal_channel_path(self, stock_code: str, channel_key: str, folder_hint: Path | None = None) -> None:
        label = CHANNEL_FOLDERS[channel_key][0]

        def job() -> Path | None:
            return self.find_channel_folder(stock_code, channel_key, folder_hint)

        def on_success(path: Path | None) -> None:
            if path is None:
                self.set_status(f"{stock_code} icin {label} klasoru bulunamadi.")
                messagebox.showwarning("Kanal Yok", f"{stock_code} icin {label} klasoru bulunamadi.")
                return
            self.set_status(f"{label} klasoru aciliyor: {path.name}")
            self.reveal_path(path)

        self.run_in_background(job, on_success, f"{stock_code} icin {label} klasoru araniyor...")

    def should_skip_fast_scan_folder(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return compact in {"2d", "3d", "arsiv", "archive", "backup", "yedek", "temp", "tmp", "nodemodules"}

    def get_preview_search_roots(self) -> list[Path]:
        if self.preview_search_roots_cache is not None:
            return self.preview_search_roots_cache
        configured = [
            Path(self.settings.get("preview_image_root", "")).expanduser(),
            Path(self.settings.get("search_root", "")).expanduser(),
        ]
        self.preview_search_roots_cache = unique_paths([path for path in configured if path_is_dir(path)])
        return self.preview_search_roots_cache

    def preview_image_rank(self, path: Path, stock_key: str, root_order: int = 0) -> tuple[int, int, int, str]:
        stem = normalize_text(path.stem)
        full = normalize_text(str(path))
        parts = [normalize_text(part) for part in path.parts]
        if stem == stock_key:
            stock_score = 0
        elif stem.startswith(stock_key):
            stock_score = 1
        elif any(part.startswith(stock_key) for part in parts):
            stock_score = 2
        elif stock_key in stem:
            stock_score = 3
        elif stock_key in full:
            stock_score = 4
        else:
            stock_score = 9
        channel_score = 0 if any(word in full for word in ("katalog", "web", "etsy")) else 1
        if any(word in full for word in ("teknik", "cizim", "technical")):
            channel_score += 4
        if "b2b" in full:
            channel_score += 1
        return (root_order, stock_score + channel_score, len(path.parts), str(path).casefold())

    def find_preview_images(self, stock_codes: list[str]) -> dict[str, Path | None]:
        requested = {normalize_text(code): code for code in stock_codes if normalize_text(code)}
        results: dict[str, Path | None] = {}
        missing = set(requested)
        indexed = self.lookup_index_paths("preview", missing)
        for key, path in indexed.items():
            self.preview_image_cache[key] = path
            results[key] = path
        missing -= set(indexed)
        for key in list(missing):
            if key in self.preview_image_cache:
                results[key] = self.preview_image_cache[key]
                missing.remove(key)
        if not missing:
            return results
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        candidates: dict[str, tuple[tuple[int, int, int, str], Path]] = {}
        for root_order, root in enumerate(self.get_preview_search_roots()):
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if path.suffix.lower() not in valid:
                        continue
                    norm = normalize_text(str(path))
                    stem = normalize_text(path.stem)
                    for key in missing:
                        if stem == key or stem.startswith(key) or key in norm:
                            rank = self.preview_image_rank(path, key, root_order)
                            current_best = candidates.get(key)
                            if current_best is None or rank < current_best[0]:
                                candidates[key] = (rank, path)
                if all(key in candidates and candidates[key][0][1] <= 2 for key in missing):
                    break
        for key in missing:
            image = candidates[key][1] if key in candidates else None
            self.preview_image_cache[key] = image
            results[key] = image
        return results

    def find_search_folders(self, stock_codes: list[str]) -> dict[str, Path | None]:
        requested = {normalize_text(code): code for code in stock_codes if normalize_text(code)}
        results: dict[str, Path | None] = {}
        missing = set(requested)
        indexed = self.lookup_index_paths("folder", missing)
        for key, path in indexed.items():
            self.search_folder_cache[key] = path
            results[key] = path
        missing -= set(indexed)
        for key in list(missing):
            if key in self.search_folder_cache:
                results[key] = self.search_folder_cache[key]
                missing.remove(key)
        if not missing:
            return results
        root = Path(self.settings.get("search_root", "")).expanduser()
        found: dict[str, tuple[tuple[int, int, int, str], Path]] = {}
        if path_is_dir(root):
            for current, dirs, files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for directory in dirs:
                    path = Path(current) / directory
                    norm = normalize_text(str(path))
                    name_norm = normalize_text(directory)
                    for key in missing:
                        if name_norm.startswith(key) or key in norm:
                            rank = self.folder_path_rank(path, key)
                            current_best = found.get(key)
                            if current_best is None or rank < current_best[0]:
                                found[key] = (rank, path)
                for file_name in files:
                    path = Path(current) / file_name
                    norm = normalize_text(str(path))
                    for key in missing:
                        if key in norm:
                            parent = Path(current)
                            rank = self.folder_path_rank(parent, key)
                            current_best = found.get(key)
                            if current_best is None or rank < current_best[0]:
                                found[key] = (rank, parent)
                if all(key in found and found[key][0][0] == 0 and found[key][0][1] <= 1 for key in missing):
                    break
        for key in missing:
            folder = found[key][1] if key in found else None
            self.search_folder_cache[key] = folder
            results[key] = folder
        return results

    def path_matches_channel(self, path: Path, channel_key: str) -> bool:
        text = normalize_text(str(path))
        return any(keyword in text for keyword in CHANNEL_FOLDERS[channel_key][1])

    def path_contains_stock(self, path: Path, stock_key: str) -> bool:
        return any(normalize_text(part).startswith(stock_key) for part in path.parts)

    def folder_path_rank(self, path: Path, stock_key: str) -> tuple[int, int, int, str]:
        parts = [normalize_text(part) for part in path.parts]
        name = normalize_text(path.name)
        media_score = 0 if any(self.is_product_images_folder_name(part) for part in path.parts) else 2
        stock_score = 0 if name.startswith(stock_key) else 1 if any(part.startswith(stock_key) for part in parts) else 3
        channel_score = 0 if any(word in normalize_text(str(path)) for word in ("katalog", "web", "etsy", "b2b")) else 1
        return (media_score, stock_score + channel_score, len(path.parts), str(path).casefold())

    def channel_path_rank(self, path: Path, stock_key: str, channel_key: str) -> tuple[int, int, int, str]:
        parts = [normalize_text(part) for part in path.parts]
        name = normalize_text(path.name)
        channel_exact = 0 if any(keyword in name for keyword in CHANNEL_FOLDERS[channel_key][1]) else 1
        stock_score = 0 if name.startswith(stock_key) else 1 if any(part.startswith(stock_key) for part in parts) else 4
        media_score = 0 if any(self.is_product_images_folder_name(part) for part in path.parts) else 2
        return (media_score, stock_score, channel_exact, len(path.parts), str(path).casefold())

    def get_product_images_ancestor(self, path: Path | None) -> Path | None:
        if path is None:
            return None
        base = path if path_is_dir(path) else path.parent
        for candidate in [base, *base.parents]:
            if self.is_product_images_folder_name(candidate.name):
                return candidate
        return None

    def collect_channel_candidates(self, root: Path, stock_key: str, channel_key: str) -> list[Path]:
        candidates: list[Path] = []
        if not path_is_dir(root):
            return candidates
        for current, dirs, _files in safe_walk(root):
            dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
            current_path = Path(current)
            if self.path_matches_channel(current_path, channel_key) and self.path_contains_stock(current_path, stock_key):
                candidates.append(current_path)
            for directory in dirs:
                path = current_path / directory
                if self.path_matches_channel(path, channel_key) and self.path_contains_stock(path, stock_key):
                    candidates.append(path)
        return candidates

    def find_channel_folder_near_hint(self, stock_key: str, channel_key: str, folder_hint: Path | None) -> Path | None:
        if folder_hint is None:
            return None
        hint = folder_hint if path_is_dir(folder_hint) else folder_hint.parent
        media_root = self.get_product_images_ancestor(hint)
        roots = unique_paths([path for path in [media_root, hint.parent if path_is_dir(hint) else hint] if path is not None and path_is_dir(path)])
        candidates: list[Path] = []
        for root in roots:
            candidates.extend(self.collect_channel_candidates(root, stock_key, channel_key))
        return min(candidates, key=lambda path: self.channel_path_rank(path, stock_key, channel_key)) if candidates else None

    def find_channel_folder(self, stock_code: str, channel_key: str, folder_hint: Path | None = None) -> Path | None:
        key = f"{normalize_text(stock_code)}::{channel_key}"
        stock_key = normalize_text(stock_code)
        hinted = self.find_channel_folder_near_hint(stock_key, channel_key, folder_hint)
        if hinted is not None:
            self.channel_folder_cache[key] = hinted
            return hinted
        if key in self.channel_folder_cache:
            return self.channel_folder_cache[key]
        indexed = self.lookup_index_paths("channel", {stock_key}, channel_key)
        if stock_key in indexed:
            self.channel_folder_cache[key] = indexed[stock_key]
            return indexed[stock_key]
        root = Path(self.settings.get("search_root", "")).expanduser()
        candidates: list[Path] = []
        if path_is_dir(root):
            for current, dirs, _files in safe_walk(root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                current_path = Path(current)
                root_channel = self.path_matches_channel(current_path, channel_key)
                root_stock = self.path_contains_stock(current_path, stock_key)
                for directory in dirs:
                    path = current_path / directory
                    dir_stock = normalize_text(directory).startswith(stock_key)
                    dir_channel = self.path_matches_channel(path, channel_key)
                    if dir_stock and (root_channel or dir_channel):
                        candidates.append(path)
                    elif dir_channel and root_stock:
                        candidates.append(path)
                if root_channel and root_stock:
                    candidates.append(current_path)
        result = min(candidates, key=lambda path: self.channel_path_rank(path, stock_key, channel_key)) if candidates else None
        self.channel_folder_cache[key] = result
        return result

    def ensure_index_schema(self) -> None:
        with sqlite3.connect(INDEX_FILE) as connection:
            connection.execute(
                "CREATE TABLE IF NOT EXISTS product_index (kind TEXT, stock_key TEXT, channel_key TEXT DEFAULT '', path TEXT, updated_at TEXT, PRIMARY KEY(kind, stock_key, channel_key, path))"
            )
            connection.execute("CREATE TABLE IF NOT EXISTS index_meta (key TEXT PRIMARY KEY, value TEXT)")
            connection.commit()

    def lookup_index_paths(self, kind: str, stock_keys: set[str], channel_key: str = "") -> dict[str, Path | None]:
        if not stock_keys or not path_exists(INDEX_FILE):
            return {}
        found: dict[str, Path | None] = {}
        with sqlite3.connect(INDEX_FILE) as connection:
            for key in stock_keys:
                rows = connection.execute(
                    "SELECT path FROM product_index WHERE kind=? AND stock_key=? AND channel_key=?",
                    (kind, key, channel_key),
                ).fetchall()
                candidates = [Path(row[0]) for row in rows if row and path_exists(Path(row[0]))]
                if not candidates:
                    continue
                if kind == "preview":
                    found[key] = min(candidates, key=lambda path: self.preview_image_rank(path, key))
                elif kind == "channel":
                    found[key] = min(candidates, key=lambda path: self.channel_path_rank(path, key, channel_key))
                else:
                    found[key] = min(candidates, key=lambda path: self.folder_path_rank(path, key))
        return found

    def get_index_last_rebuild_text(self) -> str:
        if not path_exists(INDEX_FILE):
            return "Indeks yok"
        try:
            with sqlite3.connect(INDEX_FILE) as connection:
                row = connection.execute("SELECT value FROM index_meta WHERE key='last_rebuild'").fetchone()
            return row[0] if row else "Indeks hazir"
        except sqlite3.Error:
            return "Indeks okunamadi"

    def rebuild_index_from_ui(self) -> None:
        if not self.records:
            messagebox.showwarning("Veri Yok", "Indeks icin once veri kaynagi yuklenmeli.")
            return

        def job() -> dict[str, int]:
            return self.rebuild_product_index()

        def on_success(stats: dict[str, int]) -> None:
            self.preview_image_cache.clear()
            self.search_folder_cache.clear()
            self.channel_folder_cache.clear()
            self.refresh_settings_summary()
            self.set_status(f"Indeks yenilendi: {stats}")
            messagebox.showinfo("Indeks Yenilendi", f"Fotograf: {stats['preview']}\nKlasor: {stats['folder']}\nKanal: {stats['channel']}")

        self.run_in_background(job, on_success, "Gorsel ve kanal klasorleri indeksleniyor...")

    def rebuild_product_index(self) -> dict[str, int]:
        self.ensure_index_schema()
        stock_keys = sorted({normalize_text(record.stock_code) for record in self.records if normalize_text(record.stock_code)}, key=len, reverse=True)
        if not stock_keys:
            return {"preview": 0, "folder": 0, "channel": 0}
        lookup = re.compile("|".join(re.escape(key) for key in stock_keys))
        rows: set[tuple[str, str, str, str, str]] = set()
        now = datetime.now().isoformat(timespec="seconds")
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))

        preview_root = Path(self.settings.get("preview_image_root", "")).expanduser()
        if path_is_dir(preview_root):
            for current, dirs, files in safe_walk(preview_root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                for file_name in files:
                    path = Path(current) / file_name
                    if path.suffix.lower() in valid and (match := lookup.search(normalize_text(str(path)))):
                        rows.add(("preview", match.group(0), "", str(path), now))

        search_root = Path(self.settings.get("search_root", "")).expanduser()
        if path_is_dir(search_root):
            for current, dirs, files in safe_walk(search_root):
                dirs[:] = [item for item in dirs if not self.should_skip_fast_scan_folder(item)]
                current_path = Path(current)
                root_match = lookup.search(normalize_text(str(current_path)))
                for directory in dirs:
                    path = current_path / directory
                    match = lookup.search(normalize_text(str(path)))
                    if match:
                        rows.add(("folder", match.group(0), "", str(path), now))
                    channel_stock = match.group(0) if match else root_match.group(0) if root_match else ""
                    if channel_stock:
                        for channel_key in CHANNEL_FOLDERS:
                            if self.path_matches_channel(path, channel_key):
                                rows.add(("channel", channel_stock, channel_key, str(path), now))
                for file_name in files:
                    path = current_path / file_name
                    if path.suffix.lower() in valid and (match := lookup.search(normalize_text(str(path)))):
                        rows.add(("folder", match.group(0), "", str(current_path), now))

        with sqlite3.connect(INDEX_FILE) as connection:
            connection.execute("DELETE FROM product_index")
            connection.executemany("INSERT OR REPLACE INTO product_index VALUES (?, ?, ?, ?, ?)", list(rows))
            connection.execute("INSERT OR REPLACE INTO index_meta VALUES ('last_rebuild', ?)", (now,))
            connection.commit()
        return {
            "preview": len([row for row in rows if row[0] == "preview"]),
            "folder": len([row for row in rows if row[0] == "folder"]),
            "channel": len([row for row in rows if row[0] == "channel"]),
        }

    def get_current_product_scope_records(self) -> list[ProductRecord]:
        family = self.family_var.get().strip()
        breakdown = self.breakdown_var.get().strip()
        if not family:
            return self.records
        return [
            record
            for record in self.records
            if normalize_text(record.family) == normalize_text(family)
            and (
                not breakdown
                or normalize_text(breakdown) == normalize_text(ALL_BREAKDOWNS_LABEL)
                or normalize_text(record.breakdown) == normalize_text(breakdown)
            )
        ]

    def open_channel_report(self) -> None:
        records = self.get_current_product_scope_records()[:1000]
        if not records:
            messagebox.showwarning("Kayit Yok", "Rapor icin kayit bulunamadi.")
            return

        def job() -> list[tuple[str, str, str, dict[str, bool]]]:
            return [
                (
                    record.stock_code,
                    record.family,
                    record.breakdown,
                    {channel: self.find_channel_folder(record.stock_code, channel) is not None for channel in CHANNEL_FOLDERS},
                )
                for record in records
            ]

        def on_success(rows: list[tuple[str, str, str, dict[str, bool]]]) -> None:
            dialog = tk.Toplevel(self)
            dialog.title("Kanal Raporu")
            dialog.geometry("980x620")
            shell = ttk.Frame(dialog, padding=16)
            shell.pack(fill="both", expand=True)
            columns = ("stock", "family", "breakdown", *CHANNEL_FOLDERS.keys())
            tree = ttk.Treeview(shell, columns=columns, show="headings")
            for column in columns:
                tree.heading(column, text=CHANNEL_FOLDERS[column][0] if column in CHANNEL_FOLDERS else column.title())
                tree.column(column, width=130 if column in CHANNEL_FOLDERS else 180, anchor="center" if column in CHANNEL_FOLDERS or column == "stock" else "w")
            for stock, family, breakdown, status in rows:
                tree.insert("", "end", values=(stock, family, breakdown, *["Var" if status[key] else "Yok" for key in CHANNEL_FOLDERS]))
            scroll_y = ttk.Scrollbar(shell, orient="vertical", command=tree.yview)
            scroll_x = ttk.Scrollbar(shell, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            tree.grid(row=0, column=0, sticky="nsew")
            scroll_y.grid(row=0, column=1, sticky="ns")
            scroll_x.grid(row=1, column=0, sticky="ew")
            shell.rowconfigure(0, weight=1)
            shell.columnconfigure(0, weight=1)
            enable_mousewheel_scroll(tree, tree, tree)

        self.run_in_background(job, on_success, "Kanal raporu hazirlaniyor...")

    def browse_rename_root(self) -> None:
        selected = filedialog.askdirectory(title="Tarama baslangic klasoru sec")
        if selected:
            self.rename_path_var.set(selected)
            self.persist_runtime_state()

    def browse_rename_output_root(self) -> None:
        selected = filedialog.askdirectory(title="Cikis klasoru sec")
        if selected:
            self.rename_output_path_var.set(selected)
            self.rename_export_enabled_var.set(True)
            self.persist_runtime_state()
            if self.rename_plan:
                self.preview_rename_plan()

    def on_rename_export_mode_changed(self) -> None:
        self.persist_runtime_state()
        if self.rename_plan:
            self.preview_rename_plan()

    def get_rename_output_root(self) -> Path | None:
        if not self.rename_export_enabled_var.get():
            return None
        text = self.rename_output_path_var.get().strip()
        return Path(text).expanduser() if text else None

    def preview_rename_plan(self) -> None:
        root_path = Path(self.rename_path_var.get().strip()).expanduser()
        if not path_is_dir(root_path):
            messagebox.showwarning("Gecersiz Klasor", "Lutfen gecerli bir klasor sec.")
            return
        output_root = self.get_rename_output_root()
        if self.rename_export_enabled_var.get() and output_root is None:
            messagebox.showwarning("Cikis Klasoru Eksik", "Cikartma icin cikis klasoru sec.")
            return

        def job() -> list[RenameAction]:
            return self.build_rename_plan(root_path, output_root)

        def on_success(plan: list[RenameAction]) -> None:
            self.rename_plan = plan
            self.populate_rename_tree(plan)
            self.set_status("Toplu isim degistirme onizlemesi hazir.")

        self.run_in_background(job, on_success, "Klasorler taraniyor...")

    def infer_rename_group_name(self, root_path: Path, stock_folder: Path) -> str:
        try:
            parts = stock_folder.relative_to(root_path).parts
        except ValueError:
            parts = stock_folder.parts
        for index, part in enumerate(parts):
            if self.is_product_images_folder_name(part) and index > 0:
                return parts[index - 1]
        return parts[-2] if len(parts) >= 2 else parts[0] if parts else "Diger"

    def is_product_images_folder_name(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return "urungorselleri" in compact

    def is_under_product_images_folder(self, path: Path, root_path: Path) -> bool:
        if self.is_product_images_folder_name(path.name) or self.is_product_images_folder_name(root_path.name):
            return True
        try:
            parts = path.relative_to(root_path).parts
        except ValueError:
            parts = path.parts
        return any(self.is_product_images_folder_name(part) for part in parts)

    def is_non_product_media_folder(self, folder_name: str) -> bool:
        compact = re.sub(r"[^a-z0-9]+", "", normalize_text(folder_name))
        return compact in {"2d", "3d"} or compact.startswith(("2d", "3d"))

    def build_rename_plan(self, root_path: Path, output_root: Path | None = None) -> list[RenameAction]:
        pattern = re.compile(self.settings.get("stock_regex", DEFAULT_SETTINGS["stock_regex"]), re.IGNORECASE)
        valid = ensure_prefixed_extensions(parse_csv(self.settings.get("image_extensions", "")))
        open_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("open_keywords", ""))]
        closed_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("closed_keywords", ""))]
        technical_keywords = [normalize_text(item) for item in parse_csv(self.settings.get("technical_keywords", ""))]
        plan: list[RenameAction] = []
        planned_targets: set[str] = set()
        for current, dirs, _files in safe_walk(root_path):
            current_path = Path(current)
            if not self.is_under_product_images_folder(current_path, root_path):
                dirs[:] = [item for item in dirs if not self.is_non_product_media_folder(item)]
                continue
            match = pattern.match(current_path.name.strip())
            if not match:
                continue
            stock_code = match.group(1).upper()
            group_name = self.infer_rename_group_name(root_path, current_path)
            plan.extend(
                self.build_stock_folder_actions(
                    current_path,
                    root_path,
                    stock_code,
                    group_name,
                    output_root,
                    valid,
                    open_keywords,
                    closed_keywords,
                    technical_keywords,
                    planned_targets,
                )
            )
            dirs.clear()
        return sorted(plan, key=lambda item: (normalize_text(item.group_name), item.stock_code, str(item.source).casefold()))

    def build_rename_export_parent(self, output_root: Path | None, root_path: Path, source_file: Path) -> Path | None:
        if output_root is None:
            return None
        try:
            relative_parent = source_file.parent.relative_to(root_path)
        except ValueError:
            relative_parent = Path(safe_path_part(source_file.parent.name, "Diger"))
        return output_root / relative_parent

    def build_stock_folder_actions(
        self,
        stock_folder: Path,
        root_path: Path,
        stock_code: str,
        group_name: str,
        output_root: Path | None,
        valid_extensions: set[str],
        open_keywords: list[str],
        closed_keywords: list[str],
        technical_keywords: list[str],
        planned_targets: set[str] | None = None,
    ) -> list[RenameAction]:
        actions: list[RenameAction] = []
        planned = planned_targets if planned_targets is not None else set()
        sequence = 1
        for file_path in sorted(safe_rglob_files(stock_folder), key=lambda path: str(path).casefold()):
            extension = file_path.suffix.lower()
            path_norm = normalize_text(str(file_path))
            stem_norm = normalize_text(file_path.stem)
            is_technical = any(keyword in path_norm for keyword in technical_keywords if keyword)
            if is_technical:
                if extension == ".ai":
                    actions.append(RenameAction(file_path, None, "Teknik cizimde AI atlandi", stock_code, "Atlandi", group_name))
                    continue
                if extension not in {".jpg", ".jpeg"}:
                    continue
                base_name = f"{stock_code}_t"
                rule = "Teknik cizim"
            else:
                if extension not in valid_extensions:
                    continue
                if any(keyword in stem_norm for keyword in open_keywords if keyword):
                    base_name = f"{stock_code}_acik"
                    rule = "Acik gorsel"
                elif any(keyword in stem_norm for keyword in closed_keywords if keyword):
                    base_name = f"{stock_code}_kapali"
                    rule = "Kapali gorsel"
                else:
                    base_name = f"{stock_code}_{sequence}"
                    rule = "Sirali gorsel"
                    sequence += 1
            operation = "export" if output_root is not None else "rename"
            target_parent = self.build_rename_export_parent(output_root, root_path, file_path)
            target = self.make_unique_target(file_path, base_name, planned, target_parent)
            if target.stem != base_name:
                rule = f"{rule} - isim cakismasi"
            if operation == "export":
                rule = f"Cikartma - {rule}"
                status = "Cikarilacak"
            else:
                status = "Ayni Ad" if target == file_path else "Hazir"
            actions.append(RenameAction(file_path, target, rule, stock_code, status, group_name, operation))
            planned.add(str(target).casefold())
        return actions

    def make_unique_target(self, source: Path, base_name: str, planned_targets: set[str], target_parent: Path | None = None) -> Path:
        parent = target_parent or source.parent
        allow_same_source = target_parent is None
        candidate = parent / f"{base_name}{source.suffix}"
        if self.target_is_available(candidate, source, planned_targets, allow_same_source):
            return candidate
        counter = 2
        while True:
            candidate = parent / f"{base_name}_{counter}{source.suffix}"
            if self.target_is_available(candidate, source, planned_targets, allow_same_source):
                return candidate
            counter += 1

    def target_is_available(self, candidate: Path, source: Path, planned_targets: set[str], allow_same_source: bool = True) -> bool:
        key = str(candidate).casefold()
        if key in planned_targets:
            return False
        if allow_same_source and candidate == source:
            return True
        return not path_exists(candidate)

    def populate_rename_tree(self, plan: list[RenameAction]) -> None:
        for item in self.rename_tree.get_children():
            self.rename_tree.delete(item)
        self.rename_item_actions.clear()
        self.rename_group_items.clear()
        self.clear_rename_preview()
        if not plan:
            self.rename_summary_var.set("Stok kodlu klasor bulunamadi.")
            return
        for action in plan:
            self.apply_saved_manual_target(action)
        filtered = self.filter_rename_plan(plan)
        if not filtered:
            self.rename_summary_var.set("Bu filtreye uyan satir yok.")
            return
        groups: dict[str, list[RenameAction]] = {}
        for action in filtered:
            groups.setdefault(action.group_name or "Diger", []).append(action)
        for group, actions in groups.items():
            group_item = self.rename_tree.insert("", "end", text=f"{group} ({len(actions)})", values=("", "", "", "", "", "", ""), open=True, tags=("group",))
            self.rename_group_items[group_item] = []
            for action in actions:
                key = self.get_rename_action_key(action)
                selectable = self.is_rename_action_selectable(action)
                self.rename_selection[key] = self.rename_selection.get(key, True) if selectable else False
                child = self.rename_tree.insert(
                    group_item,
                    "end",
                    text=action.source.name,
                    values=(
                        self.get_rename_checkbox_text(action),
                        action.status,
                        action.stock_code,
                        action.source.name,
                        self.format_rename_target_display(action),
                        action.reason,
                        self.format_relative_rename_path(action.source),
                    ),
                    tags=() if selectable else ("disabled",),
                )
                self.rename_item_actions[child] = action
                self.rename_group_items[group_item].append(child)
            self.update_rename_group_state(group_item)
        self.update_rename_summary()
        self.schedule_rename_autosize()

    def filter_rename_plan(self, plan: list[RenameAction]) -> list[RenameAction]:
        selected = self.rename_filter_var.get() or RENAME_FILTER_OPTIONS[0]
        if selected == "Tum Satirlar":
            return plan
        if selected == "Sadece Secili":
            return [action for action in plan if self.is_rename_action_selectable(action) and self.rename_selection.get(self.get_rename_action_key(action), True)]
        if selected == "Sadece Hazir":
            return [action for action in plan if action.status in {"Hazir", "Cikarilacak"}]
        if selected == "Sadece Cakismalar":
            return [action for action in plan if action.status == "Cakisma" or "cakisma" in normalize_text(action.reason)]
        if selected == "Sadece Manuel":
            return [action for action in plan if normalize_text(action.reason) == "manuel" or self.get_rename_action_key(action) in self.rename_manual_targets]
        if selected == "Sadece Teknik":
            return [action for action in plan if "teknik" in normalize_text(action.reason)]
        return plan

    def get_rename_action_key(self, action: RenameAction) -> str:
        return str(action.source).casefold()

    def format_relative_rename_path(self, path: Path) -> str:
        root = self.rename_path_var.get().strip()
        if not root:
            return str(path)
        try:
            return f".../{path.relative_to(Path(root).expanduser()).as_posix()}"
        except ValueError:
            return str(path)

    def format_rename_target_display(self, action: RenameAction) -> str:
        if action.target is None:
            return "-"
        if action.operation != "export":
            return action.target.name
        output = self.rename_output_path_var.get().strip()
        if output:
            try:
                return f".../{action.target.relative_to(Path(output).expanduser()).as_posix()}"
            except ValueError:
                pass
        return str(action.target)

    def apply_saved_manual_target(self, action: RenameAction) -> None:
        saved = self.rename_manual_targets.get(self.get_rename_action_key(action))
        if saved and action.target is not None:
            self.set_action_target_name(action, saved, save_state=False)

    def is_rename_action_selectable(self, action: RenameAction) -> bool:
        return action.status in {"Hazir", "Cikarilacak"} and action.target is not None

    def get_rename_checkbox_text(self, action: RenameAction) -> str:
        if not self.is_rename_action_selectable(action):
            return "-"
        return "[x]" if self.rename_selection.get(self.get_rename_action_key(action), True) else "[ ]"

    def get_group_checkbox_text(self, children: list[str]) -> str:
        selectable = [item for item in children if self.is_rename_action_selectable(self.rename_item_actions[item])]
        if not selectable:
            return "-"
        selected = [item for item in selectable if self.rename_selection.get(self.get_rename_action_key(self.rename_item_actions[item]), False)]
        if not selected:
            return "[ ]"
        return "[x]" if len(selected) == len(selectable) else "[-]"

    def update_rename_group_state(self, group_item: str) -> None:
        values = list(self.rename_tree.item(group_item, "values"))
        while len(values) < 7:
            values.append("")
        values[0] = self.get_group_checkbox_text(self.rename_group_items.get(group_item, []))
        self.rename_tree.item(group_item, values=values)

    def update_rename_action_row(self, item_id: str) -> None:
        action = self.rename_item_actions[item_id]
        values = list(self.rename_tree.item(item_id, "values"))
        values[0] = self.get_rename_checkbox_text(action)
        values[1] = action.status
        values[4] = self.format_rename_target_display(action)
        values[5] = action.reason
        values[6] = self.format_relative_rename_path(action.source)
        self.rename_tree.item(item_id, values=values)

    def on_rename_tree_click(self, event: tk.Event) -> None:
        if self.rename_tree.identify("region", event.x, event.y) != "cell" or self.rename_tree.identify_column(event.x) != "#1":
            return
        item = self.rename_tree.identify_row(event.y)
        if item in self.rename_group_items:
            self.toggle_rename_group(item)
        elif item in self.rename_item_actions:
            self.toggle_rename_action(item)

    def toggle_rename_action(self, item_id: str) -> None:
        action = self.rename_item_actions[item_id]
        if not self.is_rename_action_selectable(action):
            return
        key = self.get_rename_action_key(action)
        self.rename_selection[key] = not self.rename_selection.get(key, True)
        self.update_rename_action_row(item_id)
        parent = self.rename_tree.parent(item_id)
        if parent:
            self.update_rename_group_state(parent)
        self.update_rename_summary()
        self.save_rename_state()
        self.refresh_rename_filter_view_if_needed()

    def toggle_rename_group(self, group_item: str) -> None:
        children = self.rename_group_items.get(group_item, [])
        selectable = [item for item in children if self.is_rename_action_selectable(self.rename_item_actions[item])]
        if not selectable:
            return
        should_select = any(not self.rename_selection.get(self.get_rename_action_key(self.rename_item_actions[item]), False) for item in selectable)
        for item in selectable:
            action = self.rename_item_actions[item]
            self.rename_selection[self.get_rename_action_key(action)] = should_select
            self.update_rename_action_row(item)
        self.update_rename_group_state(group_item)
        self.update_rename_summary()
        self.save_rename_state()
        self.refresh_rename_filter_view_if_needed()

    def select_all_rename_actions(self) -> None:
        for item, action in self.rename_item_actions.items():
            if self.is_rename_action_selectable(action):
                self.rename_selection[self.get_rename_action_key(action)] = True
                self.update_rename_action_row(item)
        for group in self.rename_group_items:
            self.update_rename_group_state(group)
        self.update_rename_summary()
        self.save_rename_state()

    def clear_all_rename_actions(self) -> None:
        for item, action in self.rename_item_actions.items():
            if self.is_rename_action_selectable(action):
                self.rename_selection[self.get_rename_action_key(action)] = False
                self.update_rename_action_row(item)
        for group in self.rename_group_items:
            self.update_rename_group_state(group)
        self.update_rename_summary()
        self.save_rename_state()

    def expand_all_rename_groups(self) -> None:
        for group in self.rename_group_items:
            self.rename_tree.item(group, open=True)

    def collapse_all_rename_groups(self) -> None:
        for group in self.rename_group_items:
            self.rename_tree.item(group, open=False)

    def refresh_rename_filter_view_if_needed(self) -> None:
        if self.rename_filter_var.get() != RENAME_FILTER_OPTIONS[0]:
            self.populate_rename_tree(self.rename_plan)

    def clear_rename_preview(self) -> None:
        if not hasattr(self, "rename_preview_label"):
            return
        self.rename_preview_image_ref = None
        self.rename_preview_label.configure(image="", text="Secim yok")
        self.rename_preview_text_var.set("Listeden bir gorsel sec.")

    def on_rename_selection_changed(self, _event: tk.Event | None = None) -> None:
        selected = self.rename_tree.selection()
        if not selected or selected[0] not in self.rename_item_actions:
            self.clear_rename_preview()
            return
        action = self.rename_item_actions[selected[0]]
        photo = self.build_preview_image(action.source)
        self.rename_preview_image_ref = photo
        self.rename_preview_label.configure(image=photo if photo else "", text="" if photo else "Onizleme yok")
        target = str(action.target) if action.target else "-"
        operation = "Cikartma kopyasi" if action.operation == "export" else "Yerinde isim degistirme"
        self.rename_preview_text_var.set(
            f"Mevcut: {action.source.name}\nYeni: {action.target.name if action.target else '-'}\n"
            f"Stok: {action.stock_code}\nIslem: {operation}\nHedef: {target}\nKonum: {self.format_relative_rename_path(action.source)}"
        )

    def reveal_selected_rename_file(self) -> None:
        selected = self.rename_tree.selection()
        if not selected or selected[0] not in self.rename_item_actions:
            messagebox.showwarning("Secim Yok", "Bir dosya sec.")
            return
        self.reveal_path(self.rename_item_actions[selected[0]].source)

    def get_selected_rename_actions(self) -> list[RenameAction]:
        return [
            action
            for action in self.rename_plan
            if self.is_rename_action_selectable(action)
            and self.rename_selection.get(self.get_rename_action_key(action), False)
        ]

    def update_rename_summary(self) -> None:
        total = len(self.rename_plan)
        actionable = len([action for action in self.rename_plan if self.is_rename_action_selectable(action)])
        selected = len(self.get_selected_rename_actions())
        visible = len(self.rename_item_actions)
        self.rename_summary_var.set(
            f"{total} kayit tarandi. {visible} satir gorunuyor. {actionable} dosya uygulanabilir, {selected} secili."
        )

    def on_rename_tree_double_click(self, event: tk.Event) -> None:
        if self.rename_tree.identify_column(event.x) != "#5":
            return
        item = self.rename_tree.identify_row(event.y)
        if item not in self.rename_item_actions:
            return
        action = self.rename_item_actions[item]
        if action.target is None:
            return
        bbox = self.rename_tree.bbox(item, "#5")
        if not bbox:
            return
        if self.rename_edit_widget is not None:
            self.rename_edit_widget.destroy()
        x, y, width, height = bbox
        editor = ttk.Entry(self.rename_tree)
        editor.insert(0, action.target.name)
        editor.select_range(0, len(action.target.stem))
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()
        self.rename_edit_widget = editor

        def commit(_event: tk.Event | None = None) -> None:
            if self.rename_edit_widget is not editor:
                return
            value = editor.get().strip()
            editor.destroy()
            self.rename_edit_widget = None
            self.apply_manual_target_name(item, value)

        def cancel(_event: tk.Event | None = None) -> None:
            if self.rename_edit_widget is editor:
                editor.destroy()
                self.rename_edit_widget = None

        editor.bind("<Return>", commit)
        editor.bind("<FocusOut>", commit)
        editor.bind("<Escape>", cancel)

    def apply_manual_target_name(self, item_id: str, new_name: str) -> None:
        action = self.rename_item_actions[item_id]
        if not self.set_action_target_name(action, new_name, True):
            return
        self.rename_selection[self.get_rename_action_key(action)] = self.is_rename_action_selectable(action)
        self.update_rename_action_row(item_id)
        parent = self.rename_tree.parent(item_id)
        if parent:
            self.update_rename_group_state(parent)
        self.update_rename_summary()
        self.save_rename_state()
        self.schedule_rename_autosize()
        self.on_rename_selection_changed()
        self.refresh_rename_filter_view_if_needed()

    def set_action_target_name(self, action: RenameAction, new_name: str, save_state: bool) -> bool:
        if action.target is None:
            return False
        clean = Path(new_name.strip()).name
        if not clean or clean in {".", ".."}:
            return False
        stem = Path(clean).stem if Path(clean).suffix else clean
        if not stem:
            return False
        target = action.target.parent / f"{stem}{action.source.suffix}"
        action.target = target
        action.reason = "Manuel"
        if action.operation == "export":
            action.status = "Cikarilacak" if not path_exists(target) else "Cakisma"
        elif target == action.source:
            action.status = "Ayni Ad"
        elif path_exists(target):
            action.status = "Cakisma"
        else:
            action.status = "Hazir"
        if save_state:
            self.rename_manual_targets[self.get_rename_action_key(action)] = target.name
        return True

    def schedule_rename_autosize(self) -> None:
        if self.rename_autosize_after_id is not None:
            self.after_cancel(self.rename_autosize_after_id)
        self.rename_autosize_after_id = self.after(120, self.autosize_rename_columns)

    def autosize_rename_columns(self) -> None:
        self.rename_autosize_after_id = None
        if not hasattr(self, "rename_tree"):
            return
        font = tkfont.nametofont("TkDefaultFont")
        limits = {
            "#0": (180, 360),
            "selected": (58, 70),
            "status": (90, 140),
            "stock": (120, 180),
            "current": (160, 360),
            "target": (180, 460),
            "reason": (140, 260),
            "relative": (220, 540),
        }
        headings = {"#0": "Aile / Dosya", "selected": "Sec", "status": "Durum", "stock": "Stok", "current": "Mevcut", "target": "Yeni / Hedef", "reason": "Kural", "relative": "Kokten Sonra"}
        items: list[str] = []

        def collect(item_id: str) -> None:
            items.append(item_id)
            for child in self.rename_tree.get_children(item_id):
                collect(child)

        for root_item in self.rename_tree.get_children(""):
            collect(root_item)
        for column in ["#0", *self.rename_tree["columns"]]:
            min_width, max_width = limits.get(column, (80, 260))
            measured = font.measure(headings.get(column, column)) + 28
            for item in items[:500]:
                if column == "#0":
                    value = self.rename_tree.item(item, "text")
                else:
                    values = self.rename_tree.item(item, "values")
                    index = list(self.rename_tree["columns"]).index(column)
                    value = values[index] if index < len(values) else ""
                measured = max(measured, font.measure(str(value)) + 28)
            self.rename_tree.column(column, width=max(min_width, min(measured, max_width)))

    def apply_rename_plan(self) -> None:
        if not self.rename_plan:
            messagebox.showwarning("Plan Yok", "Once onizleme olustur.")
            return
        actions = self.get_selected_rename_actions()
        if not actions:
            messagebox.showinfo("Islem Yok", "Secili dosya yok.")
            return
        self.ensure_unique_selected_targets(actions)
        duplicates = self.find_duplicate_targets(actions)
        if duplicates:
            messagebox.showwarning("Hedef Cakismasi", "Ayni hedefe giden dosyalar var:\n\n" + "\n".join(duplicates[:8]))
            return
        export_mode = any(action.operation == "export" for action in actions)
        action_text = "farkli klasore cikartilacak" if export_mode else "yeniden adlandirilacak"
        if not messagebox.askyesno("Onay", f"{len(actions)} secili dosya {action_text}. Devam edelim mi?"):
            return

        def job() -> int:
            return self.execute_rename_actions(actions)

        def on_success(count: int) -> None:
            self.preview_rename_plan()
            done = "cikartildi" if export_mode else "yeniden adlandirildi"
            self.set_status(f"{count} dosya {done}.")
            messagebox.showinfo("Tamamlandi", f"{count} dosya {done}.")

        self.run_in_background(job, on_success, "Dosyalar cikartiliyor..." if export_mode else "Dosyalar yeniden adlandiriliyor...")

    def ensure_unique_selected_targets(self, actions: list[RenameAction]) -> None:
        planned: set[str] = set()
        changed = False
        for action in actions:
            if action.target is None:
                continue
            original_target = action.target
            target_parent = action.target.parent
            target = self.make_unique_target(
                source=action.source,
                base_name=action.target.stem,
                planned_targets=planned,
                target_parent=target_parent,
            )
            action.target = target
            if target != original_target:
                if "cakismasi" not in normalize_text(action.reason):
                    action.reason = f"{action.reason} - isim cakismasi"
                if action.operation == "export":
                    action.status = "Cikarilacak"
                elif target == action.source:
                    action.status = "Ayni Ad"
                elif path_exists(target):
                    action.status = "Cakisma"
                else:
                    action.status = "Hazir"
                changed = True
            planned.add(str(action.target).casefold())
        if changed:
            for item_id, action in self.rename_item_actions.items():
                if action in actions:
                    self.update_rename_action_row(item_id)
            self.update_rename_summary()
            self.schedule_rename_autosize()

    def find_duplicate_targets(self, actions: list[RenameAction]) -> list[str]:
        seen: set[str] = set()
        duplicates: list[str] = []
        for action in actions:
            if action.target is None:
                continue
            key = str(action.target).casefold()
            if key in seen:
                duplicates.append(str(action.target))
            seen.add(key)
        return duplicates

    def execute_rename_actions(self, actions: list[RenameAction]) -> int:
        batch_id = datetime.now().strftime("%Y%m%d%H%M%S") + f"-{uuid.uuid4().hex[:8]}"
        applied = 0
        export_actions = [action for action in actions if action.operation == "export"]
        rename_actions = [action for action in actions if action.operation != "export"]
        for action in export_actions:
            if action.target is None:
                continue
            if path_exists(action.target):
                raise FileExistsError(f"Hedef dosya zaten var: {action.target}")
            action.target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(action.source, action.target)
            applied += 1
            self.append_rename_log([
                {
                    "batch_id": batch_id,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "source": str(action.source),
                    "target": str(action.target),
                    "operation": "copy",
                }
            ])
        temp_moves: list[tuple[Path, Path, Path]] = []
        for action in rename_actions:
            if action.target is None or action.source == action.target:
                continue
            if path_exists(action.target):
                raise FileExistsError(f"Hedef dosya zaten var: {action.target}")
            temp = action.source.with_name(f"__codex_tmp__{uuid.uuid4().hex}{action.source.suffix}")
            action.source.rename(temp)
            temp_moves.append((action.source, temp, action.target))
        for original, temp, target in temp_moves:
            target.parent.mkdir(parents=True, exist_ok=True)
            temp.rename(target)
            applied += 1
            self.append_rename_log([
                {
                    "batch_id": batch_id,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "source": str(original),
                    "target": str(target),
                    "operation": "rename",
                }
            ])
        return applied

    def append_rename_log(self, entries: list[dict[str, str]]) -> None:
        with RENAME_LOG_FILE.open("a", encoding="utf-8") as handle:
            for entry in entries:
                handle.write(json.dumps(entry, ensure_ascii=False) + "\n")

    def read_rename_log_entries(self) -> list[dict[str, str]]:
        if not path_exists(RENAME_LOG_FILE):
            return []
        entries: list[dict[str, str]] = []
        for line in RENAME_LOG_FILE.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(item, dict) and item.get("batch_id"):
                entries.append({key: str(value) for key, value in item.items()})
        return entries

    def get_last_rename_batch(self) -> list[dict[str, str]]:
        entries = self.read_rename_log_entries()
        if not entries:
            return []
        batch_id = entries[-1]["batch_id"]
        return [entry for entry in entries if entry.get("batch_id") == batch_id]

    def remove_rename_batch_from_log(self, batch_id: str) -> None:
        entries = [entry for entry in self.read_rename_log_entries() if entry.get("batch_id") != batch_id]
        if not entries:
            if path_exists(RENAME_LOG_FILE):
                RENAME_LOG_FILE.unlink()
            return
        with RENAME_LOG_FILE.open("w", encoding="utf-8") as handle:
            for entry in entries:
                handle.write(json.dumps(entry, ensure_ascii=False) + "\n")

    def undo_last_rename_batch(self) -> None:
        batch = self.get_last_rename_batch()
        if not batch:
            messagebox.showinfo("Geri Alinacak Islem Yok", "Kayitli islem bulunamadi.")
            return
        if not messagebox.askyesno("Son Islemi Geri Al", f"Son toplu islemdeki {len(batch)} dosya geri alinacak. Cikartma kopyalari silinir, rename dosyalari eski adina alinir. Devam edelim mi?"):
            return

        def job() -> int:
            return self.execute_rename_undo(batch)

        def on_success(count: int) -> None:
            self.preview_rename_plan()
            self.set_status(f"{count} dosya geri alindi.")
            messagebox.showinfo("Geri Alma Tamamlandi", f"{count} dosya geri alindi.")

        self.run_in_background(job, on_success, "Son islem geri aliniyor...")

    def execute_rename_undo(self, batch_entries: list[dict[str, str]]) -> int:
        problems: list[str] = []
        for entry in batch_entries:
            source = Path(entry["source"])
            target = Path(entry["target"])
            operation = entry.get("operation", "rename")
            if not path_exists(target):
                problems.append(f"Hedef bulunamadi: {target}")
            if operation != "copy" and path_exists(source):
                problems.append(f"Eski ad zaten var: {source}")
        if problems:
            raise RuntimeError("\n".join(problems[:10]))
        count = 0
        for entry in reversed(batch_entries):
            source = Path(entry["source"])
            target = Path(entry["target"])
            operation = entry.get("operation", "rename")
            if operation == "copy":
                target.unlink()
            else:
                target.rename(source)
            count += 1
        self.remove_rename_batch_from_log(batch_entries[-1]["batch_id"])
        return count


def main() -> None:
    app = ProductDesktopApp()
    app.mainloop()


if __name__ == "__main__":
    main()
